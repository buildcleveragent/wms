from __future__ import annotations

from datetime import date
from tempfile import SpooledTemporaryFile

from django.http import FileResponse
from django.utils import timezone
from openpyxl import Workbook
from rest_framework import permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView

from allapp.accounts.access import AccessScope
from allapp.accounts.audit import record_audit_event

from .services_operations import (
    OperationFilters,
    build_operations_detail_rows,  # noqa: F401 - compatibility for older test patches
    build_operations_summary,
    count_operations_details,
    iter_operations_details,
    paginate_operations_details,
)


VIEW_PERMISSIONS = (
    "reports.view_warehouse_operations",
    "reports.view_owner_operations",
    "reports.view_boss_dashboard",
)


class CanViewOperations(permissions.BasePermission):
    message = "No permission to view operational reports."

    def has_permission(self, request, view):
        user = request.user
        if not user or not user.is_authenticated:
            return False
        if user.is_superuser:
            return True
        scope = AccessScope.for_user(user)
        return bool(
            scope.is_valid and any(user.has_perm(code) for code in VIEW_PERMISSIONS)
        )


class CanExportOperations(CanViewOperations):
    message = "No permission to export operational reports."

    def has_permission(self, request, view):
        return super().has_permission(request, view) and (
            request.user.is_superuser
            or request.user.has_perm("reports.export_operations")
        )


def _value(request, name, default=""):
    aliases = {"start_date": "date_from", "end_date": "date_to"}
    if request.method == "POST" and name in request.data:
        return request.data.get(name, default)
    if request.method == "POST" and aliases.get(name) in request.data:
        return request.data.get(aliases[name], default)
    return request.query_params.get(
        name, request.query_params.get(aliases.get(name), default)
    )


def _parse_date(value, name):
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError):
        raise ValueError(f"{name} must use YYYY-MM-DD.")


def _parse_optional_int(value, name):
    if value in (None, ""):
        return None
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{name} must be an integer id.")
    if parsed <= 0:
        raise ValueError(f"{name} must be a positive integer id.")
    return parsed


def parse_operation_filters(request) -> OperationFilters:
    now = timezone.now()
    today = timezone.localtime(now).date() if timezone.is_aware(now) else now.date()
    month_start = today.replace(day=1)
    filters = OperationFilters(
        start_date=_parse_date(
            _value(request, "start_date", month_start.isoformat()), "start_date"
        ),
        end_date=_parse_date(
            _value(request, "end_date", today.isoformat()), "end_date"
        ),
        direction=str(_value(request, "direction", "all") or "all").strip().lower(),
        metric_basis=str(_value(request, "metric_basis", "actual") or "actual")
        .strip()
        .lower(),
        owner_id=_parse_optional_int(_value(request, "owner"), "owner"),
        warehouse_id=_parse_optional_int(_value(request, "warehouse"), "warehouse"),
        status=str(_value(request, "status") or "").strip().upper(),
        order_no=str(_value(request, "order_no") or "").strip(),
        source_no=str(_value(request, "source_no") or "").strip(),
        product=str(_value(request, "product") or "").strip(),
        lot_no=str(_value(request, "lot_no") or "").strip(),
        task_no=str(_value(request, "task_no") or "").strip(),
        operator=str(_value(request, "operator") or "").strip(),
        exception_type=str(_value(request, "exception_type") or "").strip().lower(),
    )
    filters.validate()
    scope = AccessScope.for_user(request.user)
    if "warehouse_operator" in scope.roles and filters.metric_basis == "plan":
        raise PermissionError(
            "Warehouse operators may view only their own actual work facts."
        )
    if filters.owner_id and scope.owner_ids and filters.owner_id not in scope.owner_ids:
        raise PermissionError("No access to the requested owner.")
    if (
        filters.warehouse_id
        and scope.warehouse_ids
        and filters.warehouse_id not in scope.warehouse_ids
    ):
        raise PermissionError("No access to the requested warehouse.")
    return filters


class OperationsApiMixin:
    permission_classes = [CanViewOperations]

    def _filters_or_response(self, request):
        try:
            return parse_operation_filters(request), None
        except PermissionError as exc:
            return None, Response(
                {"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN
            )
        except ValueError as exc:
            return None, Response(
                {"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST
            )


class OperationsSummaryApi(OperationsApiMixin, APIView):
    def get(self, request):
        filters, error = self._filters_or_response(request)
        if error:
            return error
        payload = build_operations_summary(user=request.user, filters=filters)
        record_audit_event(
            action="QUERY",
            module="reports.operations.summary",
            request=request,
            owner_id=filters.owner_id,
            warehouse_id=filters.warehouse_id,
            metadata={
                "direction": filters.direction,
                "metric_basis": filters.metric_basis,
            },
        )
        return Response(payload)


class OperationsDetailsApi(OperationsApiMixin, APIView):
    def get(self, request):
        filters, error = self._filters_or_response(request)
        if error:
            return error
        try:
            page = max(1, int(request.query_params.get("page", 1)))
            page_size = min(200, max(1, int(request.query_params.get("page_size", 50))))
        except (TypeError, ValueError):
            return Response(
                {"detail": "page and page_size must be integers."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        count, rows = paginate_operations_details(
            user=request.user,
            filters=filters,
            page=page,
            page_size=page_size,
        )
        start = (page - 1) * page_size
        end = start + page_size
        record_audit_event(
            action="QUERY_DETAIL",
            module="reports.operations.details",
            request=request,
            owner_id=filters.owner_id,
            warehouse_id=filters.warehouse_id,
            metadata={
                "count": count,
                "direction": filters.direction,
                "metric_basis": filters.metric_basis,
            },
        )
        scope_payload = AccessScope.for_user(request.user).as_dict()
        scope_payload["actor_only"] = bool(
            {"warehouse_operator", "owner_salesperson"}.intersection(
                scope_payload["roles"]
            )
        )
        return Response(
            {
                "metric_basis": filters.metric_basis,
                "data_as_of": timezone.now().isoformat(),
                "scope": scope_payload,
                "range": {
                    "start": filters.start_date.isoformat(),
                    "end": filters.end_date.isoformat(),
                },
                "count": count,
                "page": page,
                "page_size": page_size,
                "next": page + 1 if end < count else None,
                "previous": page - 1 if page > 1 else None,
                "results": rows,
            }
        )


class OperationsExportApi(OperationsApiMixin, APIView):
    permission_classes = [CanExportOperations]
    max_rows = 50000

    def post(self, request):
        filters, error = self._filters_or_response(request)
        if error:
            return error
        count = count_operations_details(user=request.user, filters=filters)
        if count > self.max_rows:
            return Response(
                {
                    "detail": f"Export is limited to {self.max_rows} rows; narrow the filters."
                },
                status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            )

        wb = Workbook(write_only=True)
        generated_at = timezone.now()
        info = wb.create_sheet("报告信息")
        info.append(["字段", "值"])
        info.append(["报告类型", "运营计划/实际/异常明细"])
        info.append(["仓库", filters.warehouse_id or "全部授权仓库"])
        info.append(["货主", filters.owner_id or "全部货主"])
        info.append(["日期范围", f"{filters.start_date} ~ {filters.end_date}"])
        info.append(["指标口径", filters.metric_basis])
        info.append(["方向", filters.direction])
        info.append(["数据截至时间", generated_at.isoformat()])
        info.append(["数据状态", "COMPLETE"])
        ws = wb.create_sheet("operations")
        headers = [
            "direction",
            "metric_basis",
            "event_at",
            "order_no",
            "source_no",
            "task_no",
            "owner",
            "warehouse",
            "product_code",
            "sku",
            "product_name",
            "lot_no",
            "status",
            "operator",
            "planned_qty",
            "actual_qty",
            "exception_type",
        ]
        ws.append(headers)
        for row in iter_operations_details(user=request.user, filters=filters):
            ws.append(
                [
                    row["direction"],
                    row["metric_basis"],
                    row["event_at"],
                    row["order_no"],
                    row["source_no"],
                    row["task_no"],
                    row["owner"]["name"],
                    row["warehouse"]["name"],
                    row["product"]["code"],
                    row["product"]["sku"],
                    row["product"]["name"],
                    row["lot_no"],
                    row["status"],
                    row["operator"],
                    row["planned_qty"],
                    row["actual_qty"],
                    row["exception_type"],
                ]
            )
        output = SpooledTemporaryFile(max_size=8 * 1024 * 1024, mode="w+b")
        wb.save(output)
        output.seek(0)
        response = FileResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        generated_token = generated_at.strftime("%Y%m%dT%H%M%S")
        warehouse_token = filters.warehouse_id or "all-authorized"
        response["Content-Disposition"] = (
            f'attachment; filename="operations-{warehouse_token}-'
            f'{filters.start_date}-{filters.end_date}-{generated_token}.xlsx"'
        )
        response["X-Report-Data-As-Of"] = generated_at.isoformat()
        response["X-Report-Metric-Basis"] = filters.metric_basis
        record_audit_event(
            action="EXPORT",
            module="reports.operations",
            request=request,
            owner_id=filters.owner_id,
            warehouse_id=filters.warehouse_id,
            metadata={
                "rows": count,
                "direction": filters.direction,
                "metric_basis": filters.metric_basis,
            },
        )
        return response
