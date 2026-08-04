from __future__ import annotations

import datetime
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

import pytest
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from openpyxl import load_workbook
from rest_framework.test import APIClient

from allapp.accounts.models import UserRoleScope
from allapp.accounts.roles import ROLE_GROUP_TEMPLATES
from allapp.baseinfo.models import Customer, Owner, Supplier
from allapp.core.choices import InvTxType
from allapp.inbound.models import InboundOrder, InboundOrderLine
from allapp.inventory.models import InventoryTransaction
from allapp.locations.models import Location, Subwarehouse, Warehouse
from allapp.outbound.models import OutboundOrder, OutboundOrderLine
from allapp.products.models import Product, ProductUom
from allapp.tasking.models import WmsTask


pytestmark = [pytest.mark.api, pytest.mark.integration]


class OperationsV2ApiTests(TestCase):
    """Contract fixtures covering the five production roles and actual facts."""

    day = datetime.date(2026, 7, 15)

    @classmethod
    def setUpTestData(cls):
        call_command("sync_wms_role_groups", stdout=StringIO())

        cls.owner_a = Owner.objects.create(code="OVA", name="Owner V2 A")
        cls.owner_b = Owner.objects.create(code="OVB", name="Owner V2 B")
        cls.wh_1 = Warehouse.objects.create(code="V2WH1", name="V2 Warehouse 1")
        cls.wh_2 = Warehouse.objects.create(code="V2WH2", name="V2 Warehouse 2")

        sw_1 = Subwarehouse.objects.create(
            warehouse=cls.wh_1, code="V2S1", name="V2 Subwarehouse 1"
        )
        sw_2 = Subwarehouse.objects.create(
            warehouse=cls.wh_2, code="V2S2", name="V2 Subwarehouse 2"
        )
        cls.loc_1 = Location.objects.create(
            warehouse=cls.wh_1,
            subwarehouse=sw_1,
            code="V2S1-01-01-01",
            name="V2 Location 1",
        )
        cls.loc_2 = Location.objects.create(
            warehouse=cls.wh_2,
            subwarehouse=sw_2,
            code="V2S2-01-01-01",
            name="V2 Location 2",
        )
        uom = ProductUom.objects.create(code="V2EA", name="V2 Each")
        cls.product_a = Product.objects.create(
            owner=cls.owner_a,
            code="V2-A",
            sku="V2-SKU-A",
            name="V2 Product A",
            base_uom=uom,
            weight=Decimal("1.000"),
            volume=Decimal("0.001000"),
        )
        cls.product_b = Product.objects.create(
            owner=cls.owner_b,
            code="V2-B",
            sku="V2-SKU-B",
            name="V2 Product B",
            base_uom=uom,
            weight=Decimal("1.000"),
            volume=Decimal("0.001000"),
        )

        cls.operator = cls._role_user(
            "v2-operator",
            UserRoleScope.Role.WAREHOUSE_OPERATOR,
            warehouse=cls.wh_1,
        )
        cls.wh_manager = cls._role_user(
            "v2-wh-manager",
            UserRoleScope.Role.WAREHOUSE_MANAGER,
            warehouse=cls.wh_1,
        )
        cls.boss = cls._role_user(
            "v2-boss",
            UserRoleScope.Role.WAREHOUSE_BOSS,
            warehouse=cls.wh_1,
        )
        UserRoleScope.objects.create(
            user=cls.boss,
            role=UserRoleScope.Role.WAREHOUSE_BOSS,
            warehouse=cls.wh_2,
        )
        cls.owner_manager = cls._role_user(
            "v2-owner-manager",
            UserRoleScope.Role.OWNER_MANAGER,
            owner=cls.owner_a,
        )
        cls.salesperson = cls._role_user(
            "v2-salesperson",
            UserRoleScope.Role.OWNER_SALESPERSON,
            owner=cls.owner_a,
        )
        cls.other_actor = get_user_model().objects.create_user(username="v2-other-actor")

        cls.customer_a = Customer.objects.create(
            owner=cls.owner_a,
            salesperson=cls.salesperson,
            code="V2-CA",
            name="V2 Customer A",
        )
        cls.customer_b = Customer.objects.create(
            owner=cls.owner_b,
            salesperson=cls.other_actor,
            code="V2-CB",
            name="V2 Customer B",
        )
        cls.supplier_a = Supplier.objects.create(
            owner=cls.owner_a, code="V2-SA", name="V2 Supplier A"
        )
        cls.supplier_b = Supplier.objects.create(
            owner=cls.owner_b, code="V2-SB", name="V2 Supplier B"
        )

        # owner A / warehouse 1 / actor-owned work
        cls._actual_pair(
            suffix="A1-SELF",
            owner=cls.owner_a,
            warehouse=cls.wh_1,
            location=cls.loc_1,
            product=cls.product_a,
            customer=cls.customer_a,
            supplier=cls.supplier_a,
            creator=cls.salesperson,
            worker=cls.operator,
            received="6",
            shipped="9",
        )
        # owner A / warehouse 1 / another actor
        cls._actual_pair(
            suffix="A1-OTHER",
            owner=cls.owner_a,
            warehouse=cls.wh_1,
            location=cls.loc_1,
            product=cls.product_a,
            customer=cls.customer_a,
            supplier=cls.supplier_a,
            creator=cls.other_actor,
            worker=cls.other_actor,
            received="2",
            shipped="3",
        )
        # another owner in the same warehouse
        cls._actual_pair(
            suffix="B1-OTHER",
            owner=cls.owner_b,
            warehouse=cls.wh_1,
            location=cls.loc_1,
            product=cls.product_b,
            customer=cls.customer_b,
            supplier=cls.supplier_b,
            creator=cls.other_actor,
            worker=cls.other_actor,
            received="8",
            shipped="7",
        )
        # same owner in another explicitly boss-authorized warehouse
        cls._actual_pair(
            suffix="A2-SELF",
            owner=cls.owner_a,
            warehouse=cls.wh_2,
            location=cls.loc_2,
            product=cls.product_a,
            customer=cls.customer_a,
            supplier=cls.supplier_a,
            creator=cls.salesperson,
            worker=cls.salesperson,
            received="5",
            shipped="4",
        )

        # Draft and cancelled demand must never become default actual throughput.
        cls._unposted_order_pair("DRAFT", "DRAFT", "OWNER_PENDING", Decimal("99"))
        cls._unposted_order_pair("CANCEL", "SUBMITTED", "CANCELLED", Decimal("88"))

    @classmethod
    def _role_user(cls, username, role, *, owner=None, warehouse=None):
        user = get_user_model().objects.create_user(
            username=username,
            owner=owner,
            warehouse=warehouse,
        )
        user.groups.add(
            __import__("django.contrib.auth.models", fromlist=["Group"])
            .Group.objects.get(name=ROLE_GROUP_TEMPLATES[role].group_name)
        )
        UserRoleScope.objects.create(
            user=user,
            role=role,
            owner=owner if role in UserRoleScope.OWNER_ROLES else None,
            warehouse=warehouse if role in UserRoleScope.WAREHOUSE_ROLES else None,
        )
        return user

    @classmethod
    def _actual_pair(
        cls,
        *,
        suffix,
        owner,
        warehouse,
        location,
        product,
        customer,
        supplier,
        creator,
        worker,
        received,
        shipped,
    ):
        received = Decimal(received)
        shipped = Decimal(shipped)
        inbound = InboundOrder.objects.create(
            owner=owner,
            supplier=supplier,
            warehouse=warehouse,
            created_by=creator,
            order_no=f"V2-IN-{suffix}",
            src_bill_no=f"V2-ASN-{suffix}",
            biz_date=cls.day,
            submit_status="SUBMITTED",
            approval_status="WHS_APPROVED",
        )
        inbound_line = InboundOrderLine.objects.create(
            order=inbound,
            product=product,
            base_uom="V2EA",
            base_qty=received + Decimal("1"),
            base_price=Decimal("1"),
            line_no=10,
            lot_no=f"LOT-{suffix}",
        )
        receive_at = datetime.datetime(2026, 7, 15, 9, 0)
        receive = WmsTask.objects.create(
            owner=owner,
            warehouse=warehouse,
            task_no=f"V2-RCV-{suffix}",
            task_type=WmsTask.TaskType.RECEIVE,
            status=WmsTask.Status.COMPLETED,
            review_status=WmsTask.ReviewStatus.APPROVED,
            posting_status=WmsTask.PostingStatus.POSTED,
            posted_by=worker,
            posted_at=receive_at,
            finished_at=receive_at,
            ref_no=inbound.order_no,
            source_app="inbound",
            source_model="InboundOrder",
            source_pk=str(inbound.pk),
        )
        receive.lines.create(
            product=product,
            qty_plan=inbound_line.base_qty,
            qty_done=received,
            status=WmsTask.Status.COMPLETED,
            finished_by=worker,
            src_model="InboundOrderLine",
            src_id=inbound_line.pk,
            plan_meta={"lot_no": inbound_line.lot_no},
        )
        InventoryTransaction.objects.create(
            tx_type=InvTxType.RECEIVE,
            owner=owner,
            product=product,
            warehouse=warehouse,
            location=location,
            qty_delta=received,
            batch_no=inbound_line.lot_no,
            src_model="WmsTask",
            src_id=receive.pk,
            src_line_id=inbound_line.pk,
            src_no=inbound.order_no,
            posted_at=receive_at,
            posting_batch=receive.task_no,
        )

        outbound = OutboundOrder.objects.create(
            owner=owner,
            customer=customer,
            warehouse=warehouse,
            created_by=creator,
            order_no=f"V2-OUT-{suffix}",
            src_bill_no=f"V2-SO-{suffix}",
            biz_date=cls.day,
            etd=datetime.datetime(2026, 7, 15, 18, 0),
            submit_status="SUBMITTED",
            approval_status="WHS_APPROVED",
        )
        outbound_line = OutboundOrderLine.objects.create(
            order=outbound,
            product=product,
            base_uom=product.base_uom,
            base_qty=shipped + Decimal("1"),
            base_price=Decimal("1"),
            line_no=10,
            lot_no=f"LOT-{suffix}",
        )
        dispatch_at = datetime.datetime(2026, 7, 15, 16, 0)
        dispatch = WmsTask.objects.create(
            owner=owner,
            warehouse=warehouse,
            task_no=f"V2-DSP-{suffix}",
            task_type=WmsTask.TaskType.DISPATCH,
            status=WmsTask.Status.COMPLETED,
            review_status=WmsTask.ReviewStatus.APPROVED,
            posting_status=WmsTask.PostingStatus.POSTED,
            posted_by=worker,
            posted_at=dispatch_at,
            finished_at=dispatch_at,
            ref_no=outbound.order_no,
            source_app="outbound",
            source_model="OutboundOrder",
            source_pk=str(outbound.pk),
        )
        dispatch.lines.create(
            product=product,
            qty_plan=outbound_line.base_qty,
            qty_done=shipped,
            status=WmsTask.Status.COMPLETED,
            finished_at=dispatch_at,
            finished_by=worker,
            src_model="OutboundOrderLine",
            src_id=outbound_line.pk,
            plan_meta={"lot_no": outbound_line.lot_no},
        )

    @classmethod
    def _unposted_order_pair(cls, suffix, submit_status, approval_status, qty):
        inbound = InboundOrder.objects.create(
            owner=cls.owner_a,
            supplier=cls.supplier_a,
            warehouse=cls.wh_1,
            created_by=cls.salesperson,
            order_no=f"V2-IN-{suffix}",
            biz_date=cls.day,
            submit_status=submit_status,
            approval_status=("NOT_READY" if submit_status == "DRAFT" else approval_status),
        )
        InboundOrderLine.objects.create(
            order=inbound,
            product=cls.product_a,
            base_uom="V2EA",
            base_qty=qty,
            base_price=Decimal("1"),
            line_no=10,
        )
        outbound = OutboundOrder.objects.create(
            owner=cls.owner_a,
            customer=cls.customer_a,
            warehouse=cls.wh_1,
            created_by=cls.salesperson,
            order_no=f"V2-OUT-{suffix}",
            biz_date=cls.day,
            submit_status=submit_status,
            approval_status=approval_status,
        )
        OutboundOrderLine.objects.create(
            order=outbound,
            product=cls.product_a,
            base_uom=cls.product_a.base_uom,
            base_qty=qty,
            base_price=Decimal("1"),
            line_no=10,
        )

    def _get(self, user, endpoint="summary", **params):
        client = APIClient()
        client.force_authenticate(user)
        common = {
            "start_date": self.day.isoformat(),
            "end_date": self.day.isoformat(),
            "direction": "all",
            "metric_basis": "actual",
        }
        common.update(params)
        return client.get(f"/api/reports/v2/operations/{endpoint}/", common)

    def test_actual_summary_enforces_five_role_scopes_and_actor_only(self):
        cases = [
            (self.operator, "6", "9", True),
            (self.wh_manager, "16", "19", False),
            (self.boss, "21", "23", False),
            (self.owner_manager, "13", "16", False),
            (self.salesperson, "11", "13", True),
        ]
        for user, inbound_qty, outbound_qty, actor_only in cases:
            with self.subTest(user=user.username):
                response = self._get(user)
                self.assertEqual(response.status_code, 200, response.data)
                self.assertEqual(response.data["summary"]["inbound"]["metric_basis"], "inventory")
                self.assertEqual(response.data["summary"]["outbound"]["metric_basis"], "shipment")
                self.assertEqual(response.data["summary"]["inbound"]["qty"], inbound_qty)
                self.assertEqual(response.data["summary"]["outbound"]["qty"], outbound_qty)
                self.assertEqual(response.data["scope"]["actor_only"], actor_only)
                self.assertIn("data_as_of", response.data)

    def test_draft_cancelled_and_unshipped_orders_are_not_actual_throughput(self):
        actual = self._get(self.owner_manager)
        self.assertEqual(actual.status_code, 200)
        self.assertEqual(actual.data["summary"]["outbound"]["orders"], 3)
        self.assertEqual(actual.data["summary"]["outbound"]["qty"], "16")

        plan = self._get(
            self.owner_manager,
            direction="outbound",
            metric_basis="plan",
        )
        self.assertEqual(plan.status_code, 200, plan.data)
        # Three live actual-order plans (10 + 4 + 5) and a draft (99);
        # cancelled demand (88) is excluded from plan as well.
        self.assertEqual(plan.data["summary"]["outbound"]["qty"], "118")

    def test_outbound_inventory_basis_excludes_internal_putaway_issues(self):
        """PUTAWAY writes an ISSUE leg, but it is not an outbound shipment."""

        inbound = InboundOrder.objects.get(order_no="V2-IN-A1-SELF")
        outbound = OutboundOrder.objects.get(order_no="V2-OUT-A1-SELF")
        putaway = WmsTask.objects.create(
            owner=self.owner_a,
            warehouse=self.wh_1,
            task_no="V2-PA-ISSUE",
            task_type=WmsTask.TaskType.PUTAWAY,
            status=WmsTask.Status.COMPLETED,
            review_status=WmsTask.ReviewStatus.APPROVED,
            posting_status=WmsTask.PostingStatus.POSTED,
            posted_by=self.operator,
            posted_at=datetime.datetime(2026, 7, 15, 12, 0),
            finished_at=datetime.datetime(2026, 7, 15, 12, 0),
            ref_no=inbound.order_no,
            source_app="inbound",
            source_model="InboundOrder",
            source_pk=str(inbound.pk),
        )
        pick = WmsTask.objects.create(
            owner=self.owner_a,
            warehouse=self.wh_1,
            task_no="V2-PICK-ISSUE",
            task_type=WmsTask.TaskType.PICK,
            status=WmsTask.Status.COMPLETED,
            review_status=WmsTask.ReviewStatus.APPROVED,
            posting_status=WmsTask.PostingStatus.POSTED,
            posted_by=self.operator,
            posted_at=datetime.datetime(2026, 7, 15, 13, 0),
            finished_at=datetime.datetime(2026, 7, 15, 13, 0),
            ref_no=outbound.order_no,
            source_app="outbound",
            source_model="OutboundOrder",
            source_pk=str(outbound.pk),
        )
        for task, quantity in ((putaway, Decimal("-100")), (pick, Decimal("-2"))):
            InventoryTransaction.objects.create(
                tx_type=InvTxType.ISSUE,
                owner=self.owner_a,
                product=self.product_a,
                warehouse=self.wh_1,
                location=self.loc_1,
                qty_delta=quantity,
                src_model="WmsTask",
                src_id=task.pk,
                src_no=task.ref_no,
                posted_at=task.posted_at,
                posting_batch=task.task_no,
            )

        response = self._get(
            self.wh_manager,
            direction="outbound",
            metric_basis="inventory",
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["summary"]["outbound"]["qty"], "2")

    def test_operator_cannot_request_plan_basis(self):
        response = self._get(
            self.operator,
            direction="outbound",
            metric_basis="plan",
        )
        self.assertEqual(response.status_code, 403)

    def test_shipment_basis_requires_outbound_direction(self):
        response = self._get(
            self.wh_manager,
            direction="all",
            metric_basis="shipment",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("direction=outbound", response.data["detail"])

    def test_details_have_stable_pagination_and_actor_scope(self):
        manager = self._get(self.wh_manager, endpoint="details", page=1, page_size=2)
        self.assertEqual(manager.status_code, 200, manager.data)
        self.assertEqual(manager.data["count"], 6)
        self.assertEqual(len(manager.data["results"]), 2)
        self.assertEqual(manager.data["next"], 2)
        self.assertIsNone(manager.data["previous"])
        self.assertEqual(len({row["detail_id"] for row in manager.data["results"]}), 2)

        operator = self._get(self.operator, endpoint="details", page=1, page_size=50)
        self.assertEqual(operator.status_code, 200, operator.data)
        self.assertEqual(operator.data["count"], 2)
        self.assertEqual(
            {row["order_no"] for row in operator.data["results"]},
            {"V2-IN-A1-SELF", "V2-OUT-A1-SELF"},
        )

    def test_detail_page_is_sorted_and_sliced_by_mysql(self):
        with CaptureQueriesContext(connection) as captured:
            response = self._get(
                self.wh_manager,
                endpoint="details",
                page=1,
                page_size=2,
            )

        self.assertEqual(response.status_code, 200, response.data)
        data_queries = [
            query["sql"]
            for query in captured
            if "UNION ALL" in query["sql"] and "LIMIT 2" in query["sql"]
        ]
        self.assertEqual(len(data_queries), 1, [query["sql"] for query in captured])
        self.assertIn("ORDER BY", data_queries[0])
        self.assertEqual(len(response.data["results"]), 2)

    def test_cross_scope_query_is_forbidden_but_in_scope_dimension_filter_works(self):
        forbidden = self._get(self.owner_manager, owner=self.owner_b.pk)
        self.assertEqual(forbidden.status_code, 403)

        # A warehouse-scoped manager may narrow to an owner inside that warehouse.
        narrowed = self._get(self.wh_manager, owner=self.owner_a.pk)
        self.assertEqual(narrowed.status_code, 200, narrowed.data)
        self.assertEqual(narrowed.data["summary"]["inbound"]["qty"], "8")
        self.assertEqual(narrowed.data["summary"]["outbound"]["qty"], "12")

        # An owner-scoped manager may narrow to a warehouse without gaining other owners.
        narrowed = self._get(self.owner_manager, warehouse=self.wh_1.pk)
        self.assertEqual(narrowed.status_code, 200, narrowed.data)
        self.assertEqual(narrowed.data["summary"]["inbound"]["qty"], "8")
        self.assertEqual(narrowed.data["summary"]["outbound"]["qty"], "12")

    def test_exports_require_permission_and_preserve_scope(self):
        payload = {
            "start_date": self.day.isoformat(),
            "end_date": self.day.isoformat(),
            "direction": "all",
            "metric_basis": "actual",
        }
        for user in (self.operator, self.salesperson):
            client = APIClient()
            client.force_authenticate(user)
            response = client.post(
                "/api/reports/v2/operations/exports/", payload, format="json"
            )
            self.assertEqual(response.status_code, 403)

        client = APIClient()
        client.force_authenticate(self.owner_manager)
        response = client.post(
            "/api/reports/v2/operations/exports/", payload, format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-Report-Metric-Basis"], "actual")
        workbook = load_workbook(
            BytesIO(b"".join(response.streaming_content)),
            read_only=True,
        )
        rows = list(workbook["operations"].iter_rows(values_only=True))
        self.assertEqual(len(rows), 7)  # header + three receive + three dispatch facts
        self.assertNotIn("Owner V2 B", {row[6] for row in rows[1:]})

    def test_export_rejects_result_sets_over_the_contract_limit(self):
        payload = {
            "start_date": self.day.isoformat(),
            "end_date": self.day.isoformat(),
            "direction": "all",
            "metric_basis": "actual",
        }
        client = APIClient()
        client.force_authenticate(self.owner_manager)

        with patch(
            "allapp.reports.views_operations.OperationsExportApi.max_rows",
            1,
        ), patch(
            "allapp.reports.views_operations.Workbook"
        ) as workbook:
            response = client.post(
                "/api/reports/v2/operations/exports/",
                payload,
                format="json",
            )

        self.assertEqual(response.status_code, 413)
        workbook.assert_not_called()
        self.assertEqual(
            response.data["detail"],
            "Export is limited to 1 rows; narrow the filters.",
        )

    def test_documented_slashless_paths_do_not_redirect_or_drop_post_bodies(self):
        client = APIClient()
        client.force_authenticate(self.owner_manager)
        params = {
            "start_date": self.day.isoformat(),
            "end_date": self.day.isoformat(),
            "direction": "outbound",
            "metric_basis": "actual",
        }
        summary = client.get("/api/reports/v2/operations/summary", params)
        self.assertEqual(summary.status_code, 200, summary.data)
        self.assertEqual(summary.data["metric_basis"], "actual")
        self.assertIn("scope", summary.data)
        self.assertIn("summary", summary.data)
        self.assertIn("trend", summary.data)

        details = client.get("/api/reports/v2/operations/details", params)
        self.assertEqual(details.status_code, 200, details.data)
        for field in ("metric_basis", "data_as_of", "scope", "range", "count", "page", "page_size", "results"):
            self.assertIn(field, details.data)

        exported = client.post(
            "/api/reports/v2/operations/exports", params, format="json"
        )
        self.assertEqual(exported.status_code, 200)
        self.assertEqual(exported["X-Report-Metric-Basis"], "actual")
