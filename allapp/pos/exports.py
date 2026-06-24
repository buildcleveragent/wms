from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook

from .models import PosSale, PosSaleLine


def _text(value):
    if value is None:
        return ""
    return str(value)


def _dt(value):
    return value.isoformat() if value else ""


def _money(value):
    return Decimal(value or 0)


def _append_stats_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    return sheet


def build_sales_export_workbook(sales):
    sales = list(
        sales.select_related(
            "warehouse",
            "cashier",
            "shift",
            "selected_customer",
            "payment",
        ).order_by("-created_at", "-id")
    )
    sale_ids = [sale.id for sale in sales]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(
        [
            "销售单号",
            "小票号",
            "状态",
            "仓库",
            "班次号",
            "收银员",
            "客户",
            "销售时间",
            "应收金额",
            "支付方式",
            "实收金额",
            "找零",
            "支付参考号",
            "作废时间",
            "作废原因",
            "备注",
        ]
    )
    for sale in sales:
        payment = getattr(sale, "payment", None)
        sheet.append(
            [
                sale.sale_no,
                sale.src_bill_no,
                sale.status,
                getattr(sale.warehouse, "name", ""),
                getattr(sale.shift, "shift_no", "") if sale.shift_id else "",
                getattr(sale.cashier, "username", "") if sale.cashier_id else "",
                (
                    getattr(sale.selected_customer, "name", "")
                    if sale.selected_customer_id
                    else ""
                ),
                _dt(sale.created_at),
                _money(sale.total_amount),
                payment.method if payment else "",
                _money(payment.amount_received) if payment else Decimal("0.00"),
                _money(payment.change_amount) if payment else Decimal("0.00"),
                payment.reference_no if payment else "",
                _dt(sale.voided_at),
                sale.void_reason,
                sale.remark,
            ]
        )

    line_sheet = workbook.create_sheet("Lines")
    line_sheet.append(
        [
            "销售单号",
            "小票号",
            "状态",
            "班次号",
            "货主",
            "商品编码",
            "SKU",
            "商品名称",
            "数量",
            "单价",
            "金额",
            "出库单号",
            "行号",
            "销售时间",
        ]
    )
    lines = (
        PosSaleLine.objects.filter(sale_id__in=sale_ids)
        .select_related(
            "sale",
            "sale__shift",
            "owner",
            "product",
            "outbound_order_line",
            "outbound_order_line__order",
        )
        .order_by("-sale__created_at", "-sale_id", "line_no")
    )
    for line in lines:
        order_line = line.outbound_order_line
        outbound_order = order_line.order if order_line else None
        line_sheet.append(
            [
                line.sale.sale_no,
                line.sale.src_bill_no,
                line.sale.status,
                getattr(line.sale.shift, "shift_no", "") if line.sale.shift_id else "",
                getattr(line.owner, "name", ""),
                getattr(line.product, "code", ""),
                getattr(line.product, "sku", ""),
                getattr(line.product, "name", ""),
                Decimal(line.qty),
                Decimal(line.price),
                Decimal(line.amount),
                getattr(outbound_order, "order_no", "") if outbound_order else "",
                line.line_no,
                _dt(line.sale.created_at),
            ]
        )

    return workbook


def build_pos_stats_export_workbook(payload):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["字段", "值"])
    for key, value in payload.get("summary", {}).items():
        summary.append([key, value])
    summary.append(["start_date", payload.get("period", {}).get("start_date", "")])
    summary.append(["end_date", payload.get("period", {}).get("end_date", "")])

    _append_stats_sheet(
        workbook,
        "Payments",
        ["支付方式", "支付方式名称", "销售单数", "金额"],
        [
            [
                row.get("method", ""),
                row.get("method_label", ""),
                row.get("sale_count", 0),
                row.get("amount", "0.00"),
            ]
            for row in payload.get("payments", [])
        ],
    )
    _append_stats_sheet(
        workbook,
        "Owners",
        ["货主ID", "货主编码", "货主名称", "销售单数", "行数", "数量", "金额"],
        [
            [
                row.get("owner_id", ""),
                row.get("owner_code", ""),
                row.get("owner_name", ""),
                row.get("sale_count", 0),
                row.get("line_count", 0),
                row.get("qty", "0.000"),
                row.get("amount", "0.00"),
            ]
            for row in payload.get("owners", [])
        ],
    )
    _append_stats_sheet(
        workbook,
        "Products",
        [
            "商品ID",
            "商品编码",
            "SKU",
            "商品名称",
            "货主",
            "销售单数",
            "行数",
            "数量",
            "金额",
        ],
        [
            [
                row.get("product_id", ""),
                row.get("product_code", ""),
                row.get("product_sku", ""),
                row.get("product_name", ""),
                row.get("owner_name", ""),
                row.get("sale_count", 0),
                row.get("line_count", 0),
                row.get("qty", "0.000"),
                row.get("amount", "0.00"),
            ]
            for row in payload.get("products", [])
        ],
    )
    _append_stats_sheet(
        workbook,
        "Cashiers",
        [
            "收银员ID",
            "收银员",
            "销售单数",
            "完成单数",
            "作废单数",
            "完成金额",
            "作废金额",
        ],
        [
            [
                row.get("cashier_id", ""),
                row.get("cashier_username", ""),
                row.get("sale_count", 0),
                row.get("completed_count", 0),
                row.get("voided_count", 0),
                row.get("completed_amount", "0.00"),
                row.get("voided_amount", "0.00"),
            ]
            for row in payload.get("cashiers", [])
        ],
    )
    return workbook


def build_shift_export_workbook(shift, payload):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Shift"
    summary.append(["字段", "值"])
    summary.append(["班次号", payload.get("shift_no", "")])
    summary.append(["状态", payload.get("status", "")])
    summary.append(["收银员", payload.get("cashier_username", "")])
    summary.append(["开班时间", payload.get("opened_at", "")])
    summary.append(["交班时间", payload.get("closed_at", "")])
    summary.append(["备注", payload.get("remark", "")])
    for key, value in payload.get("summary", {}).items():
        if key == "payments":
            continue
        summary.append([key, value])

    _append_stats_sheet(
        workbook,
        "Payments",
        ["支付方式", "支付方式名称", "销售单数", "系统金额", "实点金额", "差异"],
        [
            [
                row.get("method", ""),
                row.get("method_label", ""),
                row.get("sale_count", 0),
                row.get("expected_amount", "0.00"),
                row.get("actual_amount", "0.00"),
                row.get("difference", "0.00"),
            ]
            for row in payload.get("summary", {}).get("payments", [])
        ],
    )

    sales_workbook = build_sales_export_workbook(PosSale.objects.filter(shift=shift))
    for sheet in sales_workbook.worksheets:
        copied = workbook.create_sheet(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            copied.append(list(row))
    return workbook
