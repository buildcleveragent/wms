import datetime
import importlib
import io
from decimal import Decimal
from unittest.mock import patch

from django.apps import apps
from django.contrib import admin
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import PermissionDenied, ValidationError
from django.test import SimpleTestCase, TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from allapp.baseinfo.models import Customer, Owner
from allapp.billing.enums import CalcMethod, ChargeType
from allapp.billing.models import BillingAccrual, BillingEvent, BillingRule
from allapp.core.choices import InvTxType, ZoneType
from allapp.inventory.locking import InventoryConcurrencyError
from allapp.inventory.models import (
    InventoryDetail,
    InventorySummary,
    InventoryTransaction,
    PostingJournal,
)
from allapp.locations.models import Location, Subwarehouse, Warehouse
from allapp.outbound.models import OutboundOrder, OutboundOrderLine
from allapp.pos.accuracy import _parse_params
from allapp.pos.models import (
    PosAuditLog,
    PosCustomer,
    PosCustomerRepayment,
    PosPayment,
    PosPaymentLine,
    PosPrintLog,
    PosReceiptWarehouseInfo,
    PosRefund,
    PosReturn,
    PosReturnLine,
    PosSale,
    PosSaleLine,
    PosSaleOrder,
    PosShift,
    PosShiftPaymentSummary,
)
from allapp.pos.serializers import SafeDateTimeField
from allapp.products.models import Product, ProductPackage, ProductUom
from allapp.tasking.models import TaskScanLog, TaskStatusLog, WmsTask, WmsTaskLine


class PosAccuracyUtilityTests(SimpleTestCase):
    def test_parse_params_uses_safe_default_date(self):
        start_date, end_date = _parse_params({})

        self.assertIsInstance(start_date, datetime.date)
        self.assertEqual(start_date, end_date)

    def test_pos_safe_datetime_field_handles_naive_datetime(self):
        field = SafeDateTimeField()
        value = field.to_representation(datetime.datetime(2026, 1, 10, 10, 30))

        self.assertIn("2026-01-10T10:30", value)


class PosAdminRegistrationTests(SimpleTestCase):
    def test_pos_models_are_registered_in_admin(self):
        registered_models = set(admin.site._registry)

        for model in (
            PosSale,
            PosSaleLine,
            PosPayment,
            PosPaymentLine,
            PosShift,
            PosShiftPaymentSummary,
            PosCustomer,
            PosReturn,
            PosReturnLine,
            PosRefund,
            PosCustomerRepayment,
            PosSaleOrder,
            PosPrintLog,
            PosAuditLog,
        ):
            self.assertIn(model, registered_models)

    def test_pos_sale_admin_is_read_only(self):
        model_admin = admin.site._registry[PosSale]

        self.assertFalse(model_admin.has_add_permission(None))
        self.assertFalse(model_admin.has_change_permission(None))
        self.assertFalse(model_admin.has_delete_permission(None))


class PosApiTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(name="POS Owner", code="POSOWN")
        self.other_owner = Owner.objects.create(name="POS Other Owner", code="POSOTH")
        self.warehouse = Warehouse.objects.create(code="WHPOS", name="POS Warehouse")
        self.subwarehouse = Subwarehouse.objects.create(
            warehouse=self.warehouse,
            code="SWPOS",
            name="POS Subwarehouse",
        )
        self.location = Location.objects.create(
            warehouse=self.warehouse,
            code="SWPOS-01-01-01",
            name="POS Pick Location",
        )
        self.user = get_user_model().objects.create_user(
            username="pos-admin",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        self.user.user_permissions.add(
            Permission.objects.get(codename="add_possale"),
            Permission.objects.get(codename="change_possale"),
            Permission.objects.get(codename="view_possale"),
            Permission.objects.get(codename="add_posreturn"),
            Permission.objects.get(codename="view_posreturn"),
            Permission.objects.get(codename="add_posrefund"),
            Permission.objects.get(codename="view_posrefund"),
        )
        self.customer = Customer.objects.create(
            owner=self.owner,
            salesperson=self.user,
            code="POS-CUS",
            name="POS Customer",
            phone="021-10000001",
            mobile="13800000001",
            address="POS Customer Address",
        )
        self.other_customer = Customer.objects.create(
            owner=self.other_owner,
            salesperson=self.user,
            code="POS-OTH-CUS",
            name="POS Other Customer",
        )
        self.pos_customer = PosCustomer.objects.create(
            warehouse=self.warehouse,
            code="PC-POS-CUS",
            name="POS Customer",
            phone="021-10000001",
            mobile="13800000001",
            address="POS Customer Address",
        )
        self.uom = ProductUom.objects.create(code="PCS-POS", name="件", is_active=True)
        self.carton_uom = ProductUom.objects.create(
            code="CTN-POS", name="箱", is_active=True
        )
        self.product = Product.objects.create(
            owner=self.owner,
            code="POS-SKU",
            name="POS Product",
            sku="POS-SKU",
            gtin="6901234567892",
            unit_barcode="POS-UNIT-BAR",
            base_uom=self.uom,
            price=Decimal("10.00"),
            min_price=Decimal("8.00"),
            max_discount=Decimal("20.00"),
            batch_control=False,
            expiry_control=False,
        )
        self.other_product = Product.objects.create(
            owner=self.other_owner,
            code="POS-OTH-SKU",
            name="POS Other Product",
            sku="POS-OTH-SKU",
            gtin="6901234567893",
            unit_barcode="POS-OTH-UNIT-BAR",
            base_uom=self.uom,
            price=Decimal("20.00"),
            min_price=Decimal("15.00"),
            max_discount=Decimal("10.00"),
            batch_control=False,
            expiry_control=False,
        )
        ProductPackage.objects.create(
            product=self.product,
            uom=self.carton_uom,
            qty_in_base=12,
            barcode="POS-CTN-BAR",
            is_sales_default=True,
        )
        InventoryDetail.objects.create(
            owner=self.owner,
            product=self.product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("10.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )
        InventoryDetail.objects.create(
            owner=self.other_owner,
            product=self.other_product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("8.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        self.shift = self.open_shift(self.user, opening_cash_amount=Decimal("100.00"))

    def payment(self, amount, method="CASH"):
        return {"method": method, "amount_received": str(amount)}

    def create_legacy_issue_for_sale_line(self, sale_line, *, remove_task_issue=False):
        task_line = WmsTaskLine.objects.get(
            task__source_app="pos",
            task__ref_no=sale_line.sale.sale_no,
            plan_meta__pos_sale_line_id=sale_line.id,
        )
        task_issue = InventoryTransaction.objects.get(
            src_model="WmsTask",
            src_id=task_line.task_id,
            src_line_id=task_line.id,
            tx_type=InvTxType.ISSUE,
        )
        detail_id = task_line.plan_meta["source_inventory_detail_id"]
        legacy = InventoryTransaction.objects.create(
            tx_type=InvTxType.ISSUE,
            owner_id=task_issue.owner_id,
            product_id=task_issue.product_id,
            warehouse_id=task_issue.warehouse_id,
            location_id=task_issue.location_id,
            subwarehouse_id=task_issue.subwarehouse_id,
            zone_type=task_issue.zone_type,
            batch_no=task_issue.batch_no,
            production_date=task_issue.production_date,
            expiry_date=task_issue.expiry_date,
            serial_no=task_issue.serial_no,
            qty_delta=task_issue.qty_delta,
            src_model="PosSaleLine",
            src_id=sale_line.id,
            src_line_id=detail_id,
            src_no=sale_line.sale.sale_no,
            memo="POS_SALE",
            posted_at=task_issue.posted_at,
            posting_batch=sale_line.sale.sale_no[:40],
        )
        if remove_task_issue:
            task_issue.delete()
        return legacy, detail_id

    def test_pos_customer_api_creates_free_customer_without_baseinfo_customer(self):
        baseinfo_count = Customer.objects.count()

        response = self.client.post(
            "/api/pos/customers/",
            {
                "name": "自由赊账客户",
                "phone": "18900000001",
                "address": "POS自由客户地址",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertTrue(response.data["code"].startswith("PC"))
        self.assertEqual(response.data["warehouse_id"], self.warehouse.id)
        self.assertEqual(response.data["name"], "自由赊账客户")
        self.assertEqual(Customer.objects.count(), baseinfo_count)

        customer = PosCustomer.objects.get(pk=response.data["id"])
        self.assertEqual(customer.warehouse_id, self.warehouse.id)
        self.assertEqual(customer.created_by_id, self.user.id)

        search = self.client.get("/api/pos/customers/", {"search": "18900000001"})
        self.assertEqual(search.status_code, 200, search.data)
        rows = search.data["results"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["id"], customer.id)

    def open_shift(self, user=None, *, opening_cash_amount=Decimal("0.00")):
        user = user or self.user
        index = PosShift.objects.count() + 1
        return PosShift.objects.create(
            shift_no=f"SHIFT-POS-TEST-{index}",
            warehouse=self.warehouse,
            cashier=user,
            opened_by=user,
            opened_at=timezone.now(),
            opening_cash_amount=opening_cash_amount,
            expected_cash_amount=opening_cash_amount,
            actual_cash_amount=opening_cash_amount,
        )

    def test_checkout_requires_pos_permission(self):
        no_pos_user = get_user_model().objects.create_user(
            username="pos-no-perm",
            password="x",
            warehouse=self.warehouse,
        )
        self.client.force_authenticate(no_pos_user)

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-NO-PERM",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-NO-PERM").exists()
        )

    def test_checkout_maps_exhausted_inventory_retry_to_http_409(self):
        payload = {
            "src_bill_no": "POS-INVENTORY-BUSY",
            "payment": self.payment("9.00"),
            "items": [
                {
                    "product_id": self.product.id,
                    "qty": "1.000",
                    "price": "9.0000",
                }
            ],
        }

        with patch(
            "allapp.pos.serializers.create_pos_sale",
            side_effect=InventoryConcurrencyError(),
        ):
            response = self.client.post(
                "/api/pos/checkout/",
                payload,
                format="json",
            )

        self.assertEqual(response.status_code, 409, response.data)
        self.assertEqual(response.data["code"], "inventory_busy")
        self.assertTrue(response.data["retryable"])
        self.assertEqual(response["Retry-After"], "1")
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-INVENTORY-BUSY").exists()
        )

    def test_checkout_requires_open_shift(self):
        cashier = get_user_model().objects.create_user(
            username="pos-no-shift",
            password="x",
            warehouse=self.warehouse,
        )
        cashier.user_permissions.add(Permission.objects.get(codename="add_possale"))
        self.client.force_authenticate(cashier)

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-NO-SHIFT",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-NO-SHIFT").exists()
        )

    def test_shift_open_api_records_opening_cash_and_blocks_duplicate_open(self):
        cashier = get_user_model().objects.create_user(
            username="pos-open-api",
            password="x",
            warehouse=self.warehouse,
        )
        cashier.user_permissions.add(Permission.objects.get(codename="add_possale"))
        self.client.force_authenticate(cashier)

        response = self.client.post(
            "/api/pos/shifts/open/",
            {"opening_cash_amount": "123.45", "remark": "morning drawer"},
            format="json",
        )
        duplicate = self.client.post(
            "/api/pos/shifts/open/",
            {"opening_cash_amount": "10.00"},
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(duplicate.status_code, 400)
        shift = PosShift.objects.get(pk=response.data["shift"]["id"])
        self.assertEqual(shift.status, PosShift.Status.OPEN)
        self.assertEqual(shift.opening_cash_amount, Decimal("123.45"))
        self.assertEqual(shift.expected_cash_amount, Decimal("123.45"))
        self.assertEqual(shift.actual_cash_amount, Decimal("123.45"))
        self.assertEqual(shift.remark, "morning drawer")
        self.assertEqual(
            PosAuditLog.objects.filter(
                action=PosAuditLog.Action.SHIFT_OPEN, shift=shift
            ).count(),
            1,
        )
        self.assertEqual(
            PosShift.objects.filter(
                cashier=cashier, status=PosShift.Status.OPEN
            ).count(),
            1,
        )

    def test_checkout_requires_user_warehouse_without_side_effects(self):
        no_warehouse_user = get_user_model().objects.create_user(
            username="pos-no-warehouse",
            password="x",
        )
        no_warehouse_user.user_permissions.add(
            Permission.objects.get(codename="add_possale")
        )
        self.client.force_authenticate(no_warehouse_user)

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-NO-WAREHOUSE",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-NO-WAREHOUSE").exists()
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("10.0000"),
        )

    def test_product_lookup_by_package_barcode_returns_available_qty(self):
        response = self.client.get("/api/pos/products/", {"barcode": "POS-CTN-BAR"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        row = response.data["results"][0]
        self.assertEqual(row["id"], self.product.id)
        self.assertEqual(row["code"], self.product.code)
        self.assertEqual(Decimal(str(row["available_qty"])), Decimal("10.0000"))
        self.assertEqual(row["unit_options"][0]["kind"], "base")
        self.assertEqual(row["unit_options"][1]["kind"], "package")

    def test_product_lookup_ignores_inactive_and_soft_deleted_package_barcode(self):
        package = self.product.packages.get(barcode="POS-CTN-BAR")
        package.is_active = False
        package.save(update_fields=["is_active"])

        inactive = self.client.get("/api/pos/products/", {"barcode": "POS-CTN-BAR"})

        self.assertEqual(inactive.status_code, 200)
        self.assertEqual(inactive.data["count"], 0)

        package.is_active = True
        package.is_deleted = True
        package.save(update_fields=["is_active", "is_deleted"])
        deleted = self.client.get("/api/pos/products/", {"barcode": "POS-CTN-BAR"})

        self.assertEqual(deleted.status_code, 200)
        self.assertEqual(deleted.data["count"], 0)

    def test_product_lookup_does_not_require_user_owner(self):
        no_owner_user = get_user_model().objects.create_user(
            username="pos-no-owner",
            password="x",
            warehouse=self.warehouse,
        )
        self.client.force_authenticate(no_owner_user)

        response = self.client.get("/api/pos/products/", {"barcode": "POS-CTN-BAR"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        row = response.data["results"][0]
        self.assertEqual(row["id"], self.product.id)
        self.assertEqual(Decimal(str(row["available_qty"])), Decimal("10.0000"))

    def test_product_lookup_returns_other_owner_available_qty(self):
        response = self.client.get(
            "/api/pos/products/", {"barcode": "POS-OTH-UNIT-BAR"}
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["count"], 1)
        row = response.data["results"][0]
        self.assertEqual(row["id"], self.other_product.id)
        self.assertEqual(Decimal(str(row["available_qty"])), Decimal("8.0000"))

    def test_product_lookup_rejects_invalid_stock_zone_scope(self):
        response = self.client.get(
            "/api/pos/products/",
            {"barcode": "POS-CTN-BAR", "zone_type": 9999},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["detail"], "POS 商品库存范围参数无效。")

    def test_checkout_creates_submitted_sales_outbound_order(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-RECEIPT-001",
                "remark": "cashier sale",
                "idempotency_key": "idem-pos-001",
                "payment": self.payment("20.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["sale"]["src_bill_no"], "POS-RECEIPT-001")
        self.assertEqual(
            response.data["receipt"]["customer"]["id"], self.pos_customer.id
        )
        self.assertEqual(response.data["receipt"]["customer"]["name"], "POS Customer")
        self.assertEqual(response.data["receipt"]["customer"]["phone"], "021-10000001")
        self.assertEqual(
            response.data["receipt"]["customer"]["address"], "POS Customer Address"
        )
        self.assertEqual(response.data["payment"]["method"], "CASH")
        self.assertEqual(
            Decimal(str(response.data["payment"]["amount_due"])), Decimal("18.00")
        )
        self.assertEqual(
            Decimal(str(response.data["payment"]["change_amount"])), Decimal("2.00")
        )
        order = OutboundOrder.objects.get(src_bill_no="POS-RECEIPT-001")
        self.assertEqual(order.owner_id, self.owner.id)
        self.assertEqual(order.warehouse_id, self.warehouse.id)
        cash_customer = Customer.objects.get(owner=self.owner, code="CASH")
        self.assertEqual(order.customer_id, cash_customer.id)
        self.assertEqual(order.outbound_type, "SALES")
        self.assertEqual(order.delivery_method, "PICKUP")
        self.assertEqual(order.submit_status, "SUBMITTED")
        self.assertEqual(order.approval_status, "WHS_APPROVED")
        self.assertTrue(order.is_closed)
        self.assertTrue(order.memo.startswith("[POS]"))
        self.assertEqual(order.close_reason, "POS即时销售完成")
        self.assertEqual(order.final_order_amount, Decimal("18.00"))
        line = OutboundOrderLine.objects.get(order=order)
        self.assertEqual(line.product_id, self.product.id)
        self.assertEqual(line.base_qty, Decimal("2.000"))
        self.assertEqual(line.base_price, Decimal("9.0000"))
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-001")
        self.assertEqual(sale.total_amount, Decimal("18.00"))
        self.assertEqual(sale.pos_customer_id, self.pos_customer.id)
        self.assertIsNone(sale.selected_customer_id)
        self.assertEqual(sale.shift_id, self.shift.id)
        self.assertEqual(sale.payment.amount_received, Decimal("20.00"))
        detail_response = self.client.get(f"/api/pos/sales/{sale.id}/")
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(
            detail_response.data["receipt"]["customer"]["name"], "POS Customer"
        )
        self.assertEqual(
            PosSaleOrder.objects.get(sale=sale).outbound_order_id, order.id
        )
        self.assertEqual(
            PosSaleLine.objects.get(sale=sale).outbound_order_line_id, line.id
        )
        task = WmsTask.objects.get(
            source_app="pos",
            source_model="outboundorder",
            source_pk=str(order.id),
        )
        self.assertEqual(task.task_type, WmsTask.TaskType.PICK)
        self.assertEqual(task.status, WmsTask.Status.COMPLETED)
        self.assertEqual(task.review_status, WmsTask.ReviewStatus.APPROVED)
        self.assertEqual(task.posting_status, WmsTask.PostingStatus.POSTED)
        self.assertEqual(task.posted_by_id, self.user.id)
        task_line = WmsTaskLine.objects.get(task=task)
        self.assertEqual(task_line.src_model, "OutboundOrderLine")
        self.assertEqual(task_line.src_id, line.id)
        self.assertEqual(task_line.plan_meta["pos_sale_id"], sale.id)
        scan = TaskScanLog.objects.get(task_line=task_line)
        issue_tx = InventoryTransaction.objects.get(
            src_model="WmsTask",
            src_id=task.id,
            src_line_id=task_line.id,
            tx_type=InvTxType.ISSUE,
        )
        journal = PostingJournal.objects.get(
            src_model="WmsTask", src_id=task.id, tx_type="POST"
        )
        self.assertEqual(scan.posting_journal_id, journal.id)
        self.assertIsNotNone(scan.posted_at)
        self.assertEqual(scan.posting_batch, issue_tx.posting_batch)
        self.assertEqual(issue_tx.src_no, sale.sale_no)
        self.assertEqual(issue_tx.memo, "POS_SALE")
        self.assertTrue(
            TaskStatusLog.objects.filter(
                task=task,
                old_status=WmsTask.Status.RESERVED,
                new_status=WmsTask.Status.COMPLETED,
            ).exists()
        )
        self.assertFalse(
            InventoryTransaction.objects.filter(
                src_model="PosSaleLine", tx_type=InvTxType.ISSUE
            ).exists()
        )

    def test_pos_accuracy_api_reports_pass_for_consistent_sale(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-ACCURACY-PASS",
                "idempotency_key": "idem-pos-accuracy-pass",
                "payment": self.payment("10.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "10.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)

        today = _parse_params({})[0].isoformat()
        response = self.client.get(
            "/api/pos/accuracy/",
            {"start_date": today, "end_date": today},
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["status"], "passed")
        self.assertEqual(response.data["summary"]["issue_count"], 0)

    def test_checkout_keeps_repeated_sale_lines_separate_in_pos_pick(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-REPEATED-LINES",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    },
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    },
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-REPEATED-LINES")
        task = WmsTask.objects.get(source_app="pos", ref_no=sale.sale_no)
        task_lines = list(WmsTaskLine.objects.filter(task=task).order_by("id"))
        issues = list(
            InventoryTransaction.objects.filter(
                src_model="WmsTask",
                src_id=task.id,
                tx_type=InvTxType.ISSUE,
            ).order_by("src_line_id")
        )

        self.assertEqual(len(task_lines), 2)
        self.assertEqual(TaskScanLog.objects.filter(task=task).count(), 2)
        self.assertEqual(len(issues), 2)
        self.assertEqual(
            {issue.src_line_id for issue in issues},
            {line.id for line in task_lines},
        )
        self.assertEqual(
            {line.plan_meta["source_inventory_detail_id"] for line in task_lines},
            {InventoryDetail.objects.get(owner=self.owner, product=self.product).id},
        )
        self.assertFalse(
            InventoryTransaction.objects.filter(
                src_model="PosSaleLine", tx_type=InvTxType.ISSUE
            ).exists()
        )

    def test_checkout_survives_billing_integrity_error_inside_savepoint(self):
        def create_real_unique_conflict(*args, **kwargs):
            sale = PosSale.objects.get(src_bill_no="POS-BILLING-DB-ERROR")
            PosSale.objects.create(
                sale_no=sale.sale_no,
                warehouse=sale.warehouse,
                total_amount=Decimal("0.00"),
            )

        with patch.object(
            BillingEvent.objects,
            "get_or_create",
            side_effect=create_real_unique_conflict,
        ):
            response = self.client.post(
                "/api/pos/checkout/",
                {
                    "src_bill_no": "POS-BILLING-DB-ERROR",
                    "payment": self.payment("9.00"),
                    "items": [
                        {
                            "product_id": self.product.id,
                            "qty": "1.000",
                            "price": "9.0000",
                        }
                    ],
                },
                format="json",
            )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-BILLING-DB-ERROR")
        self.assertTrue(PosPayment.objects.filter(sale=sale).exists())
        task = WmsTask.objects.get(source_app="pos", ref_no=sale.sale_no)
        self.assertEqual(task.posting_status, WmsTask.PostingStatus.POSTED)
        journal = PostingJournal.objects.get(
            src_model="WmsTask", src_id=task.id, tx_type="POST"
        )
        self.assertIn("BILLING_FAILED", journal.message)
        self.assertEqual(BillingEvent.objects.filter(task=task).count(), 0)
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_checkout_accrues_pick_and_order_processing_fees(self):
        BillingRule.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.PICK,
            calc_method=CalcMethod.PER_QTY_ABSDEL,
            unit_price=Decimal("2.0000"),
        )
        BillingRule.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("5.0000"),
        )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-BILLING-ACCRUAL",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-BILLING-ACCRUAL")
        task = WmsTask.objects.get(source_app="pos", ref_no=sale.sale_no)

        pick_events = BillingEvent.objects.filter(
            task=task, charge_type=ChargeType.PICK
        )
        self.assertEqual(pick_events.count(), 1)
        self.assertEqual(pick_events.get().quantity, Decimal("2.0000"))
        pick_accruals = BillingAccrual.objects.filter(
            owner=self.owner, warehouse=self.warehouse, charge_type=ChargeType.PICK
        )
        self.assertEqual(pick_accruals.count(), 1)
        self.assertEqual(pick_accruals.get().amount, Decimal("4.00"))

        order = OutboundOrder.objects.get(src_bill_no="POS-BILLING-ACCRUAL")
        order_events = BillingEvent.objects.filter(
            owner=self.owner,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
        )
        self.assertEqual(order_events.count(), 1)
        order_accruals = BillingAccrual.objects.filter(
            owner=self.owner, warehouse=self.warehouse, charge_type=ChargeType.DISPATCH
        )
        self.assertEqual(order_accruals.count(), 1)
        self.assertEqual(order_accruals.get().amount, Decimal("5.00"))

        journal = PostingJournal.objects.get(
            src_model="WmsTask", src_id=task.id, tx_type="POST"
        )
        self.assertNotIn("BILLING_FAILED", journal.message or "")
        self.assertIsNotNone(order.id)

    def test_pos_cashier_cannot_use_public_posting_service(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-PERM-CHECK",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-PERM-CHECK")
        task = WmsTask.objects.get(source_app="pos", ref_no=sale.sale_no)
        self.assertEqual(task.posted_by_id, self.user.id)
        self.assertFalse(self.user.has_perm("tasking.taskconfirm_as_wh_manager"))

        from allapp.tasking.services_posting import post_task as wh_manager_post_task

        with self.assertRaises(PermissionDenied):
            wh_manager_post_task(task.id, by_user=self.user)

    def test_checkout_posts_serial_from_original_inventory_layer(self):
        serial_product = Product.objects.create(
            owner=self.owner,
            code="POS-SERIAL-SKU",
            name="POS Serial Product",
            sku="POS-SERIAL-SKU",
            unit_barcode="POS-SERIAL-BAR",
            base_uom=self.uom,
            price=Decimal("30.00"),
            min_price=Decimal("1.00"),
            serial_control=True,
        )
        source_detail = InventoryDetail.objects.create(
            owner=self.owner,
            product=serial_product,
            warehouse=self.warehouse,
            location=self.location,
            serial_no="POS-SN-0001",
            onhand_qty=Decimal("1.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-SERIAL-SALE",
                "payment": self.payment("30.00"),
                "items": [
                    {
                        "product_id": serial_product.id,
                        "qty": "1.000",
                        "price": "30.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-SERIAL-SALE")
        task_line = WmsTaskLine.objects.get(
            task__source_app="pos", task__ref_no=sale.sale_no
        )
        scan = TaskScanLog.objects.get(task_line=task_line)
        issue = InventoryTransaction.objects.get(
            src_model="WmsTask",
            src_line_id=task_line.id,
            tx_type=InvTxType.ISSUE,
        )
        self.assertEqual(
            task_line.plan_meta["source_inventory_detail_id"], source_detail.id
        )
        self.assertEqual(scan.barcode, "POS-SN-0001")
        self.assertIsNone(scan.label_key)
        self.assertEqual(issue.serial_no, "POS-SN-0001")

    def test_pos_pick_releases_only_its_own_reservation(self):
        detail = InventoryDetail.objects.get(owner=self.owner, product=self.product)
        detail.allocated_qty = Decimal("3.0000")
        detail.save()

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-KEEP-OTHER-ALLOCATION",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        detail.refresh_from_db()
        summary = InventorySummary.objects.get(owner=self.owner, product=self.product)
        self.assertEqual(detail.onhand_qty, Decimal("8.0000"))
        self.assertEqual(detail.allocated_qty, Decimal("3.0000"))
        self.assertEqual(detail.available_qty, Decimal("5.0000"))
        self.assertEqual(summary.allocated_qty, Decimal("3.0000"))
        self.assertEqual(summary.available_qty, Decimal("5.0000"))

    def test_posting_failure_rolls_back_sale_task_scan_and_reservation(self):
        with patch(
            "allapp.pos.services.execute_posting_handler",
            side_effect=ValidationError("forced POS posting failure"),
        ):
            response = self.client.post(
                "/api/pos/checkout/",
                {
                    "src_bill_no": "POS-POSTING-ROLLBACK",
                    "payment": self.payment("9.00"),
                    "items": [
                        {
                            "product_id": self.product.id,
                            "qty": "1.000",
                            "price": "9.0000",
                        }
                    ],
                },
                format="json",
            )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-POSTING-ROLLBACK").exists()
        )
        self.assertFalse(
            OutboundOrder.objects.filter(src_bill_no="POS-POSTING-ROLLBACK").exists()
        )
        self.assertFalse(WmsTask.objects.filter(source_app="pos").exists())
        self.assertFalse(TaskScanLog.objects.exists())
        self.assertFalse(InventoryTransaction.objects.exists())
        detail = InventoryDetail.objects.get(owner=self.owner, product=self.product)
        self.assertEqual(detail.onhand_qty, Decimal("10.0000"))
        self.assertEqual(detail.allocated_qty, Decimal("0.0000"))

    def test_pos_pick_with_invalid_source_model_never_falls_back_to_generic_posting(
        self,
    ):
        from allapp.tasking.posting_exec import execute_posting_handler

        def corrupt_source_then_post(task, **kwargs):
            WmsTask.objects.filter(pk=task.id).update(source_model="broken")
            task.source_model = "broken"
            return execute_posting_handler(task, **kwargs)

        with patch(
            "allapp.pos.services.execute_posting_handler",
            side_effect=corrupt_source_then_post,
        ):
            response = self.client.post(
                "/api/pos/checkout/",
                {
                    "src_bill_no": "POS-INVALID-TASK-SOURCE",
                    "payment": self.payment("9.00"),
                    "items": [
                        {
                            "product_id": self.product.id,
                            "qty": "1.000",
                            "price": "9.0000",
                        }
                    ],
                },
                format="json",
            )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("来源必须为 OutboundOrder", str(response.data))
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-INVALID-TASK-SOURCE").exists()
        )
        self.assertFalse(WmsTask.objects.filter(source_app="pos").exists())
        self.assertFalse(InventoryTransaction.objects.exists())
        detail = InventoryDetail.objects.get(owner=self.owner, product=self.product)
        self.assertEqual(detail.onhand_qty, Decimal("10.0000"))
        self.assertEqual(detail.allocated_qty, Decimal("0.0000"))

    def test_payment_failure_rolls_back_already_posted_pos_pick(self):
        with patch.object(
            PosPayment.objects,
            "create",
            side_effect=RuntimeError("forced POS payment failure"),
        ):
            with self.assertRaisesRegex(RuntimeError, "forced POS payment failure"):
                self.client.post(
                    "/api/pos/checkout/",
                    {
                        "src_bill_no": "POS-PAYMENT-ROLLBACK",
                        "payment": self.payment("9.00"),
                        "items": [
                            {
                                "product_id": self.product.id,
                                "qty": "1.000",
                                "price": "9.0000",
                            }
                        ],
                    },
                    format="json",
                )

        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-PAYMENT-ROLLBACK").exists()
        )
        self.assertFalse(
            OutboundOrder.objects.filter(src_bill_no="POS-PAYMENT-ROLLBACK").exists()
        )
        self.assertFalse(WmsTask.objects.filter(source_app="pos").exists())
        self.assertFalse(TaskScanLog.objects.exists())
        self.assertFalse(PostingJournal.objects.exists())
        self.assertFalse(InventoryTransaction.objects.exists())
        detail = InventoryDetail.objects.get(owner=self.owner, product=self.product)
        self.assertEqual(detail.onhand_qty, Decimal("10.0000"))
        self.assertEqual(detail.allocated_qty, Decimal("0.0000"))

    def test_pos_accuracy_api_uses_safe_default_date(self):
        response = self.client.get("/api/pos/accuracy/")

        self.assertEqual(response.status_code, 200, response.data)
        self.assertIn("period", response.data)
        self.assertIn("start_date", response.data["period"])
        self.assertIn("end_date", response.data["period"])

    def test_pos_accuracy_api_detects_amount_mismatch(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-ACCURACY-BAD",
                "idempotency_key": "idem-pos-accuracy-bad",
                "payment": self.payment("10.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "10.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        PosSale.objects.filter(src_bill_no="POS-ACCURACY-BAD").update(
            total_amount=Decimal("11.00")
        )

        today = _parse_params({})[0].isoformat()
        response = self.client.get(
            "/api/pos/accuracy/",
            {"start_date": today, "end_date": today},
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["status"], "failed")
        self.assertGreater(response.data["summary"]["issue_count"], 0)
        self.assertTrue(
            any(issue["code"] == "sale_amount" for issue in response.data["issues"])
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )
        self.assertEqual(
            InventorySummary.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_payment_line_backfill_migration_copies_legacy_payment_once(self):
        migration = importlib.import_module(
            "allapp.pos.migrations.0006_backfill_pos_payment_lines"
        )
        legacy_sale = PosSale.objects.create(
            sale_no="POS-LEGACY-PAYMENT",
            src_bill_no="POS-LEGACY-PAYMENT",
            warehouse=self.warehouse,
            cashier=self.user,
            shift=self.shift,
            total_amount=Decimal("18.00"),
        )
        PosPayment.objects.create(
            sale=legacy_sale,
            method=PosPayment.Method.CASH,
            amount_due=Decimal("18.00"),
            amount_received=Decimal("20.00"),
            change_amount=Decimal("2.00"),
            reference_no="LEGACY-CASH",
            status=PosPayment.Status.PAID,
        )
        existing_sale = PosSale.objects.create(
            sale_no="POS-HAS-PAYMENT-LINE",
            src_bill_no="POS-HAS-PAYMENT-LINE",
            warehouse=self.warehouse,
            cashier=self.user,
            shift=self.shift,
            total_amount=Decimal("9.00"),
        )
        PosPayment.objects.create(
            sale=existing_sale,
            method=PosPayment.Method.WECHAT,
            amount_due=Decimal("9.00"),
            amount_received=Decimal("9.00"),
            change_amount=Decimal("0.00"),
            reference_no="LEGACY-WECHAT",
            status=PosPayment.Status.PAID,
        )
        PosPaymentLine.objects.create(
            sale=existing_sale,
            method=PosPayment.Method.WECHAT,
            amount=Decimal("9.00"),
            amount_received=Decimal("9.00"),
            change_amount=Decimal("0.00"),
            reference_no="EXISTING-LINE",
            status=PosPayment.Status.PAID,
        )

        migration.backfill_payment_lines(apps, None)

        legacy_line = PosPaymentLine.objects.get(sale=legacy_sale)
        self.assertEqual(legacy_line.method, PosPayment.Method.CASH)
        self.assertEqual(legacy_line.amount, Decimal("18.00"))
        self.assertEqual(legacy_line.amount_received, Decimal("20.00"))
        self.assertEqual(legacy_line.change_amount, Decimal("2.00"))
        self.assertEqual(legacy_line.reference_no, "LEGACY-CASH")
        self.assertEqual(legacy_line.status, PosPayment.Status.PAID)
        self.assertEqual(PosPaymentLine.objects.filter(sale=existing_sale).count(), 1)

        migration.backfill_payment_lines(apps, None)

        self.assertEqual(PosPaymentLine.objects.filter(sale=legacy_sale).count(), 1)
        self.assertEqual(PosPaymentLine.objects.filter(sale=existing_sale).count(), 1)

    def test_checkout_without_customer_uses_cash_customer(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-CASH",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        cash_customer = Customer.objects.get(owner=self.owner, code="CASH")
        self.assertEqual(cash_customer.name, "散客")
        order = OutboundOrder.objects.get(src_bill_no="POS-RECEIPT-CASH")
        self.assertEqual(order.customer_id, cash_customer.id)

    def test_checkout_does_not_require_user_owner(self):
        no_owner_user = get_user_model().objects.create_user(
            username="pos-checkout-no-owner",
            password="x",
            warehouse=self.warehouse,
        )
        no_owner_user.user_permissions.add(
            Permission.objects.get(codename="add_possale")
        )
        self.client.force_authenticate(no_owner_user)
        self.open_shift(no_owner_user)

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-NO-OWNER",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data["order_count"], 1)
        order = OutboundOrder.objects.get(src_bill_no="POS-RECEIPT-NO-OWNER")
        self.assertEqual(order.owner_id, self.owner.id)

    def test_checkout_splits_products_by_owner(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-RECEIPT-MULTI",
                "payment": self.payment("50.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    },
                    {
                        "product_id": self.other_product.id,
                        "qty": "2.000",
                        "price": "18.0000",
                    },
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["order_count"], 2)
        self.assertEqual(
            Decimal(str(response.data["payment"]["amount_due"])), Decimal("45.00")
        )
        self.assertEqual(
            Decimal(str(response.data["payment"]["change_amount"])), Decimal("5.00")
        )
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-MULTI")
        self.assertEqual(sale.total_amount, Decimal("45.00"))
        self.assertEqual(
            PosPayment.objects.get(sale=sale).amount_received, Decimal("50.00")
        )
        orders = OutboundOrder.objects.filter(src_bill_no="POS-RECEIPT-MULTI").order_by(
            "owner_id"
        )
        self.assertEqual(orders.count(), 2)
        order_by_owner = {order.owner_id: order for order in orders}
        cash_customer = Customer.objects.get(owner=self.owner, code="CASH")
        self.assertEqual(order_by_owner[self.owner.id].customer_id, cash_customer.id)
        cash_customer = Customer.objects.get(owner=self.other_owner, code="CASH")
        self.assertEqual(
            order_by_owner[self.other_owner.id].customer_id, cash_customer.id
        )
        self.assertEqual(
            OutboundOrderLine.objects.get(
                order=order_by_owner[self.owner.id]
            ).product_id,
            self.product.id,
        )
        self.assertEqual(
            OutboundOrderLine.objects.get(
                order=order_by_owner[self.other_owner.id]
            ).product_id,
            self.other_product.id,
        )
        self.assertEqual(
            order_by_owner[self.owner.id].final_order_amount, Decimal("9.00")
        )
        self.assertEqual(
            order_by_owner[self.other_owner.id].final_order_amount, Decimal("36.00")
        )
        self.assertEqual(PosSaleOrder.objects.filter(sale=sale).count(), 2)
        sale_order_amounts = {
            link.owner_id: link.amount
            for link in PosSaleOrder.objects.filter(sale=sale)
        }
        self.assertEqual(sale_order_amounts[self.owner.id], Decimal("9.00"))
        self.assertEqual(sale_order_amounts[self.other_owner.id], Decimal("36.00"))
        self.assertEqual(PosSaleLine.objects.filter(sale=sale).count(), 2)
        tasks = WmsTask.objects.filter(source_app="pos", ref_no=sale.sale_no)
        self.assertEqual(tasks.count(), 2)
        self.assertEqual(
            set(tasks.values_list("owner_id", flat=True)),
            {self.owner.id, self.other_owner.id},
        )
        self.assertEqual(
            set(tasks.values_list("posting_status", flat=True)),
            {WmsTask.PostingStatus.POSTED},
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.other_owner, product=self.other_product
            ).available_qty,
            Decimal("6.0000"),
        )
        self.assertEqual(
            InventorySummary.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )
        self.assertEqual(
            InventorySummary.objects.get(
                owner=self.other_owner, product=self.other_product
            ).available_qty,
            Decimal("6.0000"),
        )

    def test_checkout_accepts_split_payments_and_shift_close_uses_payment_lines(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-SPLIT-PAY",
                "payments": [
                    {
                        "method": "CASH",
                        "amount": "10.00",
                        "amount_received": "10.00",
                    },
                    {
                        "method": "WECHAT",
                        "amount": "8.00",
                        "amount_received": "8.00",
                        "reference_no": "WX-001",
                    },
                ],
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-SPLIT-PAY")
        self.assertEqual(sale.payment.method, PosPayment.Method.OTHER)
        self.assertEqual(sale.payment.reference_no, "MULTI")
        payment_lines = {
            line.method: line for line in PosPaymentLine.objects.filter(sale=sale)
        }
        self.assertEqual(payment_lines[PosPayment.Method.CASH].amount, Decimal("10.00"))
        self.assertEqual(
            payment_lines[PosPayment.Method.WECHAT].amount, Decimal("8.00")
        )
        self.assertEqual(
            Decimal(str(response.data["receipt"]["payment_lines"][0]["amount"])),
            Decimal("10.00"),
        )

        close_response = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/close/",
            {
                "actual_cash_amount": "110.00",
                "payments": [{"method": "WECHAT", "actual_amount": "8.00"}],
            },
            format="json",
        )

        self.assertEqual(close_response.status_code, 200, close_response.data)
        summary = close_response.data["shift"]["summary"]
        self.assertEqual(summary["net_amount"], "18.00")
        self.assertEqual(summary["expected_cash_amount"], "110.00")
        payments = {
            row.method: row
            for row in PosShiftPaymentSummary.objects.filter(shift=self.shift)
        }
        self.assertEqual(
            payments[PosPayment.Method.CASH].expected_amount, Decimal("10.00")
        )
        self.assertEqual(
            payments[PosPayment.Method.WECHAT].expected_amount, Decimal("8.00")
        )

    def test_checkout_allows_customer_credit_and_reports_debt(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-CREDIT-FULL",
                "payment": {"method": "CREDIT", "amount_received": "0.00"},
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "10.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-CREDIT-FULL")
        self.assertEqual(sale.payment.method, PosPayment.Method.CREDIT)
        self.assertEqual(sale.payment.amount_received, Decimal("0.00"))
        self.assertEqual(
            PosPaymentLine.objects.get(
                sale=sale, method=PosPayment.Method.CREDIT
            ).amount,
            Decimal("10.00"),
        )
        self.assertEqual(response.data["receipt"]["credit_amount"], "10.00")
        self.assertEqual(response.data["receipt"]["cumulative_debt"], "10.00")

        today = timezone.now().date().isoformat()
        stats = self.client.get(
            "/api/pos/stats/", {"start_date": today, "end_date": today}
        )
        self.assertEqual(stats.status_code, 200, stats.data)
        self.assertEqual(stats.data["summary"]["credit_amount"], "10.00")
        self.assertEqual(stats.data["summary"]["received_amount"], "0.00")
        payments = {row["method"]: row for row in stats.data["payments"]}
        self.assertEqual(payments[PosPayment.Method.CREDIT]["amount"], "10.00")

    def test_credit_checkout_requires_customer_without_side_effects(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-CREDIT-NO-CUSTOMER",
                "payment": {"method": "CREDIT", "amount_received": "0.00"},
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "10.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-CREDIT-NO-CUSTOMER").exists()
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("10.0000"),
        )

    def test_partial_credit_and_repayment_reduce_customer_debt_and_shift_cash(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-CREDIT-PARTIAL",
                "payments": [
                    {
                        "method": "WECHAT",
                        "amount": "4.00",
                        "amount_received": "4.00",
                    },
                    {
                        "method": "CREDIT",
                        "amount": "6.00",
                        "amount_received": "0.00",
                    },
                ],
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "10.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        self.assertEqual(checkout.data["receipt"]["credit_amount"], "6.00")
        self.assertEqual(checkout.data["receipt"]["cumulative_debt"], "6.00")

        debt = self.client.get(f"/api/pos/customer-debts/{self.pos_customer.id}/")
        self.assertEqual(debt.status_code, 200, debt.data)
        self.assertEqual(debt.data["debt_balance"], "6.00")

        repayment = self.client.post(
            "/api/pos/repayments/",
            {
                "customer_id": self.pos_customer.id,
                "method": "CASH",
                "amount": "2.50",
                "reference_no": "CASH-REPAY-001",
                "remark": "customer paid later",
            },
            format="json",
        )
        self.assertEqual(repayment.status_code, 201, repayment.data)
        self.assertEqual(repayment.data["debt_before"], "6.00")
        self.assertEqual(repayment.data["debt_after"], "3.50")
        self.assertEqual(PosCustomerRepayment.objects.count(), 1)

        today = timezone.now().date().isoformat()
        stats = self.client.get(
            "/api/pos/stats/", {"start_date": today, "end_date": today}
        )
        self.assertEqual(stats.status_code, 200, stats.data)
        self.assertEqual(stats.data["summary"]["credit_amount"], "6.00")
        self.assertEqual(stats.data["summary"]["repayment_amount"], "2.50")
        self.assertEqual(stats.data["summary"]["received_amount"], "6.50")
        payment_rows = {row["method"]: row for row in stats.data["payments"]}
        self.assertEqual(
            payment_rows[PosPayment.Method.CASH]["repayment_amount"], "2.50"
        )
        self.assertEqual(payment_rows[PosPayment.Method.WECHAT]["sale_amount"], "4.00")
        self.assertEqual(payment_rows[PosPayment.Method.CREDIT]["sale_amount"], "6.00")

        close_response = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/close/",
            {
                "actual_cash_amount": "102.50",
                "payments": [{"method": "WECHAT", "actual_amount": "4.00"}],
            },
            format="json",
        )
        self.assertEqual(close_response.status_code, 200, close_response.data)
        summary = close_response.data["shift"]["summary"]
        self.assertEqual(summary["expected_cash_amount"], "102.50")
        self.assertEqual(summary["credit_amount"], "6.00")
        self.assertEqual(summary["repayment_amount"], "2.50")
        shift_payments = {
            row.method: row
            for row in PosShiftPaymentSummary.objects.filter(shift=self.shift)
        }
        self.assertEqual(
            shift_payments[PosPayment.Method.CASH].repayment_amount,
            Decimal("2.50"),
        )
        self.assertEqual(
            shift_payments[PosPayment.Method.CREDIT].expected_amount,
            Decimal("6.00"),
        )

    def test_checkout_rejects_unbalanced_split_payments_without_side_effects(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-BAD-SPLIT",
                "payments": [
                    {
                        "method": "CASH",
                        "amount": "10.00",
                        "amount_received": "10.00",
                    },
                    {
                        "method": "WECHAT",
                        "amount": "7.99",
                        "amount_received": "7.99",
                    },
                ],
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-BAD-SPLIT").exists()
        )
        self.assertFalse(
            OutboundOrder.objects.filter(src_bill_no="POS-RECEIPT-BAD-SPLIT").exists()
        )
        self.assertEqual(PosPaymentLine.objects.count(), 0)
        self.assertEqual(
            InventoryTransaction.objects.filter(src_model="PosSaleLine").count(), 0
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("10.0000"),
        )

    def test_shift_close_records_cash_and_non_cash_differences(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-SHIFT-DIFF",
                "payments": [
                    {
                        "method": "CASH",
                        "amount": "10.00",
                        "amount_received": "10.00",
                    },
                    {
                        "method": "WECHAT",
                        "amount": "8.00",
                        "amount_received": "8.00",
                    },
                ],
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)

        close_response = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/close/",
            {
                "actual_cash_amount": "109.00",
                "payments": [{"method": "WECHAT", "actual_amount": "7.50"}],
                "remark": "short drawer",
            },
            format="json",
        )

        self.assertEqual(close_response.status_code, 200, close_response.data)
        summary = close_response.data["shift"]["summary"]
        self.assertEqual(summary["expected_cash_amount"], "110.00")
        self.assertEqual(summary["actual_cash_amount"], "109.00")
        self.assertEqual(summary["cash_difference"], "-1.00")
        shift = PosShift.objects.get(pk=self.shift.id)
        self.assertEqual(shift.cash_difference, Decimal("-1.00"))
        summaries = {
            row.method: row
            for row in PosShiftPaymentSummary.objects.filter(shift=self.shift)
        }
        self.assertEqual(
            summaries[PosPayment.Method.CASH].expected_amount, Decimal("10.00")
        )
        self.assertEqual(
            summaries[PosPayment.Method.CASH].actual_amount, Decimal("9.00")
        )
        self.assertEqual(summaries[PosPayment.Method.CASH].difference, Decimal("-1.00"))
        self.assertEqual(
            summaries[PosPayment.Method.WECHAT].expected_amount, Decimal("8.00")
        )
        self.assertEqual(
            summaries[PosPayment.Method.WECHAT].actual_amount, Decimal("7.50")
        )
        self.assertEqual(
            summaries[PosPayment.Method.WECHAT].difference, Decimal("-0.50")
        )
        self.assertTrue(
            PosAuditLog.objects.filter(
                action=PosAuditLog.Action.SHIFT_CLOSE,
                shift=self.shift,
                reason="short drawer",
            ).exists()
        )

    def test_checkout_idempotency_key_rejects_different_payload(self):
        payload = {
            "src_bill_no": "POS-RECEIPT-IDEM-CONFLICT",
            "idempotency_key": "idem-conflict-sale",
            "payment": self.payment("9.00"),
            "items": [
                {
                    "product_id": self.product.id,
                    "qty": "1.000",
                    "price": "9.0000",
                }
            ],
        }

        first = self.client.post("/api/pos/checkout/", payload, format="json")
        conflicting = {
            **payload,
            "payment": self.payment("18.00"),
            "items": [
                {
                    "product_id": self.product.id,
                    "qty": "2.000",
                    "price": "9.0000",
                }
            ],
        }
        second = self.client.post("/api/pos/checkout/", conflicting, format="json")

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(second.status_code, 400)
        self.assertEqual(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-IDEM-CONFLICT").count(), 1
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="WmsTask", tx_type=InvTxType.ISSUE
            ).count(),
            1,
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_checkout_idempotency_key_includes_explicit_line_amount(self):
        payload = {
            "src_bill_no": "POS-RECEIPT-IDEM-AMOUNT",
            "idempotency_key": "idem-explicit-amount",
            "payment": self.payment("9.00"),
            "items": [
                {
                    "product_id": self.product.id,
                    "qty": "1.000",
                    "price": "9.0000",
                    "amount": "9.00",
                }
            ],
        }
        first = self.client.post("/api/pos/checkout/", payload, format="json")
        conflicting = {
            **payload,
            "payment": self.payment("8.50"),
            "items": [
                {
                    "product_id": self.product.id,
                    "qty": "1.000",
                    "price": "8.5000",
                    "amount": "8.50",
                }
            ],
        }
        second = self.client.post("/api/pos/checkout/", conflicting, format="json")

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(second.status_code, 400)
        self.assertEqual(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-IDEM-AMOUNT").count(),
            1,
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_checkout_idempotency_key_does_not_double_post_stock(self):
        payload = {
            "src_bill_no": "POS-RECEIPT-IDEM",
            "idempotency_key": "idem-repeat-sale",
            "payment": self.payment("9.00"),
            "items": [
                {
                    "product_id": self.product.id,
                    "qty": "1.000",
                    "price": "9.0000",
                }
            ],
        }

        first = self.client.post("/api/pos/checkout/", payload, format="json")
        second = self.client.post("/api/pos/checkout/", payload, format="json")

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(second.status_code, 201, second.data)
        self.assertEqual(first.data["sale"]["id"], second.data["sale"]["id"])
        self.assertEqual(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-IDEM").count(), 1
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="WmsTask", tx_type=InvTxType.ISSUE
            ).count(),
            1,
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_checkout_rejects_duplicate_receipt_without_second_stock_post(self):
        first = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-DUP",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        second = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-DUP",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(second.status_code, 400)
        self.assertEqual(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-DUP").count(), 1
        )
        self.assertEqual(
            OutboundOrder.objects.filter(src_bill_no="POS-RECEIPT-DUP").count(), 1
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="WmsTask", tx_type=InvTxType.ISSUE
            ).count(),
            1,
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_checkout_rounds_amounts_consistently_across_sale_order_and_stats(self):
        rounding_product = Product.objects.create(
            owner=self.owner,
            code="POS-ROUND-SKU",
            name="POS Rounding Product",
            sku="POS-ROUND-SKU",
            unit_barcode="POS-ROUND-BAR",
            base_uom=self.uom,
            price=Decimal("0.3333"),
            min_price=Decimal("0.0001"),
            batch_control=False,
            expiry_control=False,
        )
        InventoryDetail.objects.create(
            owner=self.owner,
            product=rounding_product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("10.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-ROUNDING",
                "payment": self.payment("1.00"),
                "items": [
                    {
                        "product_id": rounding_product.id,
                        "qty": "3.000",
                        "price": "0.3333",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-ROUNDING")
        line = PosSaleLine.objects.get(sale=sale)
        order = OutboundOrder.objects.get(src_bill_no="POS-RECEIPT-ROUNDING")
        self.assertEqual(sale.total_amount, Decimal("1.00"))
        self.assertEqual(line.amount, Decimal("1.00"))
        self.assertEqual(order.final_order_amount, Decimal("1.00"))
        self.assertEqual(sale.payment.amount_due, Decimal("1.00"))
        self.assertEqual(PosPaymentLine.objects.get(sale=sale).amount, Decimal("1.00"))
        self.assertEqual(
            InventoryDetail.objects.get(product=rounding_product).available_qty,
            Decimal("7.0000"),
        )

        today = timezone.now().date().isoformat()
        stats = self.client.get(
            "/api/pos/stats/",
            {"start_date": today, "end_date": today, "owner_id": self.owner.id},
        )

        self.assertEqual(stats.status_code, 200, stats.data)
        self.assertEqual(stats.data["summary"]["sales_amount"], "1.00")
        self.assertEqual(stats.data["summary"]["net_amount"], "1.00")
        products = {row["product_id"]: row for row in stats.data["products"]}
        self.assertEqual(products[rounding_product.id]["sale_amount"], "1.00")
        self.assertEqual(products[rounding_product.id]["qty"], "3.000")

    def test_checkout_uses_explicit_negotiated_line_amount(self):
        negotiated_product = Product.objects.create(
            owner=self.owner,
            code="POS-NEGOTIATED-SKU",
            name="POS Negotiated Product",
            sku="POS-NEGOTIATED-SKU",
            unit_barcode="POS-NEGOTIATED-BAR",
            base_uom=self.uom,
            price=Decimal("10.0000"),
            min_price=Decimal("1.0000"),
            batch_control=False,
            expiry_control=False,
        )
        InventoryDetail.objects.create(
            owner=self.owner,
            product=negotiated_product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("20.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-NEGOTIATED-AMOUNT",
                "payment": self.payment("95.00"),
                "items": [
                    {
                        "product_id": negotiated_product.id,
                        "qty": "12.000",
                        "price": "7.9167",
                        "amount": "95.00",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-NEGOTIATED-AMOUNT")
        sale_line = sale.lines.get()
        order = OutboundOrder.objects.get(src_bill_no="POS-NEGOTIATED-AMOUNT")
        order_line = order.lines.get()
        self.assertEqual(sale.total_amount, Decimal("95.00"))
        self.assertEqual(sale_line.price, Decimal("7.9167"))
        self.assertEqual(sale_line.amount, Decimal("95.00"))
        self.assertEqual(order.final_order_amount, Decimal("95.00"))
        self.assertEqual(order_line.base_price, Decimal("7.9167"))
        self.assertEqual(order_line.final_line_amount, Decimal("95.00"))
        self.assertEqual(sale.payment.amount_due, Decimal("95.00"))

    def test_checkout_preserves_explicit_amount_not_representable_by_price(self):
        negotiated_product = Product.objects.create(
            owner=self.owner,
            code="POS-LARGE-NEGOTIATED-SKU",
            name="POS Large Negotiated Product",
            sku="POS-LARGE-NEGOTIATED-SKU",
            unit_barcode="POS-LARGE-NEGOTIATED-BAR",
            base_uom=self.uom,
            price=Decimal("0.0100"),
            min_price=Decimal("0.0001"),
            batch_control=False,
            expiry_control=False,
        )
        InventoryDetail.objects.create(
            owner=self.owner,
            product=negotiated_product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("10000.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-LARGE-NEGOTIATED-AMOUNT",
                "payment": self.payment("95.00"),
                "items": [
                    {
                        "product_id": negotiated_product.id,
                        "qty": "9999.000",
                        "price": "0.0095",
                        "amount": "95.00",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        sale = PosSale.objects.get(src_bill_no="POS-LARGE-NEGOTIATED-AMOUNT")
        self.assertEqual(sale.total_amount, Decimal("95.00"))
        self.assertEqual(sale.lines.get().amount, Decimal("95.00"))
        self.assertEqual(
            sale.lines.get().qty * sale.lines.get().price, Decimal("94.9905")
        )

    def test_checkout_rejects_explicit_amount_price_mismatch(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-NEGOTIATED-MISMATCH",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "8.9999",
                        "amount": "18.00",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-NEGOTIATED-MISMATCH").exists()
        )
        self.assertFalse(
            OutboundOrder.objects.filter(src_bill_no="POS-NEGOTIATED-MISMATCH").exists()
        )

    def test_void_sale_restores_stock_and_cancels_outbound_orders(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-VOID",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        sale_id = checkout.data["sale"]["id"]

        response = self.client.post(
            f"/api/pos/sales/{sale_id}/void/",
            {"reason": "cashier mistake"},
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        sale = PosSale.objects.get(pk=sale_id)
        self.assertEqual(sale.status, PosSale.Status.VOIDED)
        self.assertEqual(sale.payment.status, PosPayment.Status.VOIDED)
        order = OutboundOrder.objects.get(src_bill_no="POS-RECEIPT-VOID")
        self.assertEqual(order.approval_status, "CANCELLED")
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("10.0000"),
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="WmsTask", tx_type=InvTxType.ISSUE
            ).count(),
            1,
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="PosSaleLine", tx_type=InvTxType.RECEIVE
            ).count(),
            1,
        )

    def test_void_restores_inactive_source_layer_and_accuracy_warns(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-INACTIVE-LAYER",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-INACTIVE-LAYER")
        task_line = WmsTaskLine.objects.get(
            task__source_app="pos", task__ref_no=sale.sale_no
        )
        detail_id = task_line.plan_meta["source_inventory_detail_id"]
        InventoryDetail.objects.filter(pk=detail_id).update(is_active=False)

        response = self.client.post(
            f"/api/pos/sales/{sale.id}/void/",
            {"reason": "restore inactive source"},
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        detail = InventoryDetail.all_objects.get(pk=detail_id)
        self.assertFalse(detail.is_active)
        self.assertEqual(detail.onhand_qty, Decimal("10.0000"))
        self.assertEqual(detail.available_qty, Decimal("10.0000"))
        self.assertEqual(
            InventorySummary.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("0.0000"),
        )

        today = timezone.now().date().isoformat()
        accuracy = self.client.get(
            "/api/pos/accuracy/",
            {"start_date": today, "end_date": today},
        )
        self.assertEqual(accuracy.status_code, 200, accuracy.data)
        self.assertEqual(accuracy.data["status"], "passed")
        self.assertEqual(accuracy.data["summary"]["warning_count"], 1)
        self.assertTrue(
            any(
                issue["code"] == "inactive_restore_layer"
                and issue["severity"] == "warning"
                for issue in accuracy.data["issues"]
            )
        )

    def test_void_remains_compatible_with_legacy_pos_sale_line_issue(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-LEGACY-VOID",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-LEGACY-VOID")
        sale_line = PosSaleLine.objects.get(sale=sale)
        _legacy, detail_id = self.create_legacy_issue_for_sale_line(
            sale_line,
            remove_task_issue=True,
        )

        response = self.client.post(
            f"/api/pos/sales/{sale.id}/void/",
            {"reason": "legacy compatibility"},
            format="json",
        )

        self.assertEqual(response.status_code, 200, response.data)
        detail = InventoryDetail.objects.get(pk=detail_id)
        self.assertEqual(detail.onhand_qty, Decimal("10.0000"))
        self.assertTrue(
            InventoryTransaction.objects.filter(
                src_model="PosSaleLine",
                src_id=sale_line.id,
                src_line_id=detail_id,
                tx_type=InvTxType.RECEIVE,
            ).exists()
        )

    def test_mixed_legacy_and_task_issues_are_rejected_and_reported(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-MIXED-ISSUES",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-MIXED-ISSUES")
        sale_line = PosSaleLine.objects.get(sale=sale)
        self.create_legacy_issue_for_sale_line(sale_line)

        pos_return = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "mixed trace should fail",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "9.00"}],
            },
            format="json",
        )
        void = self.client.post(
            f"/api/pos/sales/{sale.id}/void/",
            {"reason": "mixed trace should fail"},
            format="json",
        )

        self.assertEqual(pos_return.status_code, 400, pos_return.data)
        self.assertEqual(void.status_code, 400, void.data)
        self.assertIn("同时存在新旧出库流水", str(pos_return.data))
        self.assertIn("同时存在新旧出库流水", str(void.data))
        self.assertFalse(PosReturn.objects.filter(sale=sale).exists())
        self.assertFalse(
            InventoryTransaction.objects.filter(
                tx_type=InvTxType.RECEIVE,
                src_model__in=["PosSaleLine", "PosReturnLine"],
            ).exists()
        )
        today = timezone.now().date().isoformat()
        accuracy = self.client.get(
            "/api/pos/accuracy/",
            {"start_date": today, "end_date": today},
        )
        self.assertEqual(accuracy.status_code, 200, accuracy.data)
        self.assertEqual(accuracy.data["status"], "failed")
        self.assertTrue(
            any(
                "同时存在新旧出库流水" in issue["message"]
                for issue in accuracy.data["issues"]
            )
        )

    def test_soft_deleted_source_layer_rejects_return_and_void_without_residue(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-SOFT-DELETED-LAYER",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-SOFT-DELETED-LAYER")
        sale_line = PosSaleLine.objects.get(sale=sale)
        task_line = WmsTaskLine.objects.get(
            task__source_app="pos", task__ref_no=sale.sale_no
        )
        detail_id = task_line.plan_meta["source_inventory_detail_id"]
        InventoryDetail.all_objects.filter(pk=detail_id).update(
            is_deleted=True,
            deleted_at=timezone.now(),
            deleted_by=self.user,
        )

        pos_return = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "soft deleted source",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "9.00"}],
            },
            format="json",
        )
        void = self.client.post(
            f"/api/pos/sales/{sale.id}/void/",
            {"reason": "soft deleted source"},
            format="json",
        )

        self.assertEqual(pos_return.status_code, 400, pos_return.data)
        self.assertEqual(void.status_code, 400, void.data)
        for error_response in (pos_return, void):
            error_text = str(error_response.data)
            self.assertIn(sale.sale_no, error_text)
            self.assertIn(str(detail_id), error_text)
            self.assertIn("不存在或已软删除", error_text)
        sale.refresh_from_db()
        self.assertEqual(sale.status, PosSale.Status.COMPLETED)
        self.assertEqual(sale.payment.status, PosPayment.Status.PAID)
        self.assertFalse(PosReturn.objects.filter(sale=sale).exists())
        self.assertFalse(PosRefund.objects.filter(sale=sale).exists())
        self.assertFalse(
            InventoryTransaction.objects.filter(
                tx_type=InvTxType.RECEIVE,
                src_model__in=["PosSaleLine", "PosReturnLine"],
            ).exists()
        )
        self.assertFalse(
            PosAuditLog.objects.filter(
                sale=sale,
                action__in=[PosAuditLog.Action.RETURN, PosAuditLog.Action.VOID],
            ).exists()
        )
        today = timezone.now().date().isoformat()
        accuracy = self.client.get(
            "/api/pos/accuracy/",
            {"start_date": today, "end_date": today},
        )
        self.assertEqual(accuracy.status_code, 200, accuracy.data)
        self.assertEqual(accuracy.data["status"], "failed")
        self.assertTrue(
            any(
                str(detail_id) == str(issue["object_id"])
                and "不存在或已软删除" in issue["message"]
                for issue in accuracy.data["issues"]
            )
        )

    def test_void_sale_requires_pos_void_permission(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-VOID-NO-PERM",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        sale_id = checkout.data["sale"]["id"]
        no_void_user = get_user_model().objects.create_user(
            username="pos-no-void-perm",
            password="x",
            warehouse=self.warehouse,
        )
        no_void_user.user_permissions.add(
            Permission.objects.get(codename="add_possale")
        )
        self.client.force_authenticate(no_void_user)

        response = self.client.post(
            f"/api/pos/sales/{sale_id}/void/",
            {"reason": "cashier mistake"},
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        sale = PosSale.objects.get(pk=sale_id)
        self.assertEqual(sale.status, PosSale.Status.COMPLETED)
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="PosSaleLine", tx_type=InvTxType.RECEIVE
            ).count(),
            0,
        )

    def test_void_sale_requires_reason(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-VOID-REASON",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        sale_id = checkout.data["sale"]["id"]

        response = self.client.post(
            f"/api/pos/sales/{sale_id}/void/",
            {"reason": ""},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        sale = PosSale.objects.get(pk=sale_id)
        self.assertEqual(sale.status, PosSale.Status.COMPLETED)
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="PosSaleLine", tx_type=InvTxType.RECEIVE
            ).count(),
            0,
        )

    def test_void_split_payment_marks_all_payment_lines_void_and_excludes_cash(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-VOID-SPLIT-PAY",
                "payments": [
                    {
                        "method": "CASH",
                        "amount": "10.00",
                        "amount_received": "10.00",
                    },
                    {
                        "method": "WECHAT",
                        "amount": "8.00",
                        "amount_received": "8.00",
                    },
                ],
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale_id = checkout.data["sale"]["id"]

        void_response = self.client.post(
            f"/api/pos/sales/{sale_id}/void/",
            {"reason": "split payment cashier mistake"},
            format="json",
        )

        self.assertEqual(void_response.status_code, 200, void_response.data)
        sale = PosSale.objects.get(pk=sale_id)
        self.assertEqual(sale.status, PosSale.Status.VOIDED)
        self.assertEqual(sale.payment.status, PosPayment.Status.VOIDED)
        self.assertEqual(
            set(sale.payment_lines.values_list("status", flat=True)),
            {PosPayment.Status.VOIDED},
        )
        close_response = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/close/",
            {"actual_cash_amount": "100.00"},
            format="json",
        )
        self.assertEqual(close_response.status_code, 200, close_response.data)
        summary = close_response.data["shift"]["summary"]
        self.assertEqual(summary["sale_count"], 1)
        self.assertEqual(summary["completed_count"], 0)
        self.assertEqual(summary["voided_count"], 1)
        self.assertEqual(summary["net_amount"], "0.00")
        self.assertEqual(summary["voided_amount"], "18.00")
        self.assertEqual(summary["expected_cash_amount"], "100.00")
        self.assertEqual(
            PosShiftPaymentSummary.objects.filter(shift=self.shift).count(), 0
        )

    def test_pos_return_restores_stock_and_updates_shift_stats_and_payments(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-RETURN",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-RETURN")
        sale_line = PosSaleLine.objects.get(sale=sale)

        response = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "customer returned one item",
                "idempotency_key": "idem-return-one",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "9.00"}],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        return_order = PosReturn.objects.get(sale=sale)
        self.assertEqual(return_order.total_amount, Decimal("9.00"))
        self.assertEqual(
            PosReturnLine.objects.get(return_order=return_order).qty, Decimal("1.000")
        )
        self.assertEqual(
            PosRefund.objects.get(return_order=return_order).amount, Decimal("9.00")
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )
        self.assertEqual(
            InventorySummary.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="PosReturnLine", tx_type=InvTxType.RECEIVE
            ).count(),
            1,
        )

        detail_response = self.client.get(f"/api/pos/sales/{sale.id}/")
        self.assertEqual(detail_response.status_code, 200, detail_response.data)
        line_payload = detail_response.data["lines"][0]
        self.assertEqual(Decimal(str(line_payload["returned_qty"])), Decimal("1.000"))
        self.assertEqual(Decimal(str(line_payload["returnable_qty"])), Decimal("1.000"))

        over_return = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "too many",
                "lines": [{"sale_line_id": sale_line.id, "qty": "2.000"}],
                "refunds": [{"method": "CASH", "amount": "18.00"}],
            },
            format="json",
        )
        self.assertEqual(over_return.status_code, 400)
        self.assertEqual(PosReturn.objects.filter(sale=sale).count(), 1)
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

        void_after_return = self.client.post(
            f"/api/pos/sales/{sale.id}/void/",
            {"reason": "cannot void returned sale"},
            format="json",
        )
        self.assertEqual(void_after_return.status_code, 400)
        sale.refresh_from_db()
        self.assertEqual(sale.status, PosSale.Status.COMPLETED)

        today = timezone.now().date().isoformat()
        stats = self.client.get(
            "/api/pos/stats/",
            {"start_date": today, "end_date": today},
        )
        self.assertEqual(stats.status_code, 200, stats.data)
        self.assertEqual(stats.data["summary"]["sales_amount"], "18.00")
        self.assertEqual(stats.data["summary"]["return_amount"], "9.00")
        self.assertEqual(stats.data["summary"]["net_amount"], "9.00")
        payments = {row["method"]: row for row in stats.data["payments"]}
        self.assertEqual(payments["CASH"]["sale_amount"], "18.00")
        self.assertEqual(payments["CASH"]["refund_amount"], "9.00")
        self.assertEqual(payments["CASH"]["net_amount"], "9.00")
        owners = {row["owner_id"]: row for row in stats.data["owners"]}
        self.assertEqual(owners[self.owner.id]["sale_amount"], "18.00")
        self.assertEqual(owners[self.owner.id]["return_amount"], "9.00")
        self.assertEqual(owners[self.owner.id]["amount"], "9.00")
        self.assertEqual(owners[self.owner.id]["qty"], "1.000")

        sales_export = self.client.get(
            "/api/pos/sales/export/", {"search": "POS-RECEIPT-RETURN"}
        )
        self.assertEqual(sales_export.status_code, 200)
        workbook = load_workbook(io.BytesIO(sales_export.content))
        self.assertIn("PaymentLines", workbook.sheetnames)
        self.assertIn("Returns", workbook.sheetnames)
        self.assertIn("ReturnLines", workbook.sheetnames)
        self.assertIn("Refunds", workbook.sheetnames)
        self.assertEqual(Decimal(str(workbook["Returns"]["G2"].value)), Decimal("9"))
        self.assertEqual(
            Decimal(str(workbook["ReturnLines"]["K2"].value)), Decimal("9")
        )
        self.assertEqual(Decimal(str(workbook["Refunds"]["E2"].value)), Decimal("9"))

        close_response = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/close/",
            {"actual_cash_amount": "109.00"},
            format="json",
        )
        self.assertEqual(close_response.status_code, 200, close_response.data)
        summary = close_response.data["shift"]["summary"]
        self.assertEqual(summary["gross_sales_amount"], "18.00")
        self.assertEqual(summary["return_count"], 1)
        self.assertEqual(summary["return_amount"], "9.00")
        self.assertEqual(summary["net_amount"], "9.00")
        self.assertEqual(summary["expected_cash_amount"], "109.00")
        shift = PosShift.objects.get(pk=self.shift.id)
        self.assertEqual(shift.total_sales_amount, Decimal("9.00"))
        self.assertEqual(shift.total_return_amount, Decimal("9.00"))
        self.assertEqual(shift.return_count, 1)
        cash_summary = PosShiftPaymentSummary.objects.get(
            shift=self.shift, method=PosPayment.Method.CASH
        )
        self.assertEqual(cash_summary.expected_amount, Decimal("9.00"))
        self.assertEqual(cash_summary.refund_count, 1)
        self.assertEqual(cash_summary.refund_amount, Decimal("9.00"))
        self.assertTrue(
            PosAuditLog.objects.filter(
                action=PosAuditLog.Action.RETURN, return_order=return_order
            ).exists()
        )

    def test_negotiated_amount_partial_returns_absorb_rounding_remainder(self):
        rounding_product = Product.objects.create(
            owner=self.owner,
            code="POS-RETURN-ROUND-SKU",
            name="POS Return Rounding Product",
            sku="POS-RETURN-ROUND-SKU",
            unit_barcode="POS-RETURN-ROUND-BAR",
            base_uom=self.uom,
            price=Decimal("0.3333"),
            min_price=Decimal("0.0001"),
            batch_control=False,
            expiry_control=False,
        )
        InventoryDetail.objects.create(
            owner=self.owner,
            product=rounding_product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("3.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RETURN-ROUND-AMOUNT",
                "payment": self.payment("1.00"),
                "items": [
                    {
                        "product_id": rounding_product.id,
                        "qty": "3.000",
                        "price": "0.3333",
                        "amount": "1.00",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-RETURN-ROUND-AMOUNT")
        sale_line = sale.lines.get()

        for index, refund_amount in enumerate(("0.33", "0.34", "0.33"), start=1):
            response = self.client.post(
                "/api/pos/returns/",
                {
                    "sale_id": sale.id,
                    "reason": f"partial return {index}",
                    "idempotency_key": f"partial-return-{index}",
                    "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                    "refunds": [{"method": "CASH", "amount": refund_amount}],
                },
                format="json",
            )
            self.assertEqual(response.status_code, 201, response.data)

        return_amounts = list(
            PosReturn.objects.filter(sale=sale)
            .order_by("id")
            .values_list("total_amount", flat=True)
        )
        self.assertEqual(
            return_amounts,
            [Decimal("0.33"), Decimal("0.34"), Decimal("0.33")],
        )
        self.assertEqual(sum(return_amounts, Decimal("0.00")), Decimal("1.00"))

        detail_response = self.client.get(f"/api/pos/sales/{sale.id}/")
        self.assertEqual(detail_response.status_code, 200, detail_response.data)
        line_payload = detail_response.data["lines"][0]
        self.assertEqual(Decimal(str(line_payload["returned_amount"])), Decimal("1.00"))
        self.assertEqual(
            Decimal(str(line_payload["returnable_amount"])), Decimal("0.00")
        )

    def test_pos_return_idempotency_does_not_double_restore_stock(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-RETURN-IDEM",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-RETURN-IDEM")
        sale_line = PosSaleLine.objects.get(sale=sale)
        payload = {
            "sale_id": sale.id,
            "reason": "same return retry",
            "idempotency_key": "idem-return-retry",
            "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
            "refunds": [{"method": "CASH", "amount": "9.00"}],
        }

        first = self.client.post("/api/pos/returns/", payload, format="json")
        second = self.client.post("/api/pos/returns/", payload, format="json")

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(second.status_code, 201, second.data)
        self.assertEqual(first.data["return"]["id"], second.data["return"]["id"])
        self.assertEqual(PosReturn.objects.filter(sale=sale).count(), 1)
        self.assertEqual(
            InventoryTransaction.objects.filter(src_model="PosReturnLine").count(),
            1,
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_pos_return_requires_refund_permission(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-RETURN-NO-PERM",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-RETURN-NO-PERM")
        sale_line = PosSaleLine.objects.get(sale=sale)
        cashier = get_user_model().objects.create_user(
            username="pos-return-no-perm",
            password="x",
            warehouse=self.warehouse,
        )
        cashier.user_permissions.add(Permission.objects.get(codename="view_possale"))
        self.client.force_authenticate(cashier)

        response = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "no permission",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "9.00"}],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(PosReturn.objects.filter(sale=sale).exists())
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_pos_return_requires_open_shift_without_stock_restore(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RETURN-NO-SHIFT",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-RETURN-NO-SHIFT")
        sale_line = PosSaleLine.objects.get(sale=sale)
        cashier = get_user_model().objects.create_user(
            username="pos-return-no-shift",
            password="x",
            warehouse=self.warehouse,
        )
        cashier.user_permissions.add(
            Permission.objects.get(codename="add_posreturn"),
            Permission.objects.get(codename="add_posrefund"),
        )
        self.client.force_authenticate(cashier)

        response = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "no open shift",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "9.00"}],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(PosReturn.objects.filter(sale=sale).exists())
        self.assertEqual(
            InventoryTransaction.objects.filter(src_model="PosReturnLine").count(), 0
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("9.0000"),
        )

    def test_pos_return_rejects_refund_mismatch_without_stock_restore(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RETURN-BAD-REFUND",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-RETURN-BAD-REFUND")
        sale_line = PosSaleLine.objects.get(sale=sale)

        response = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "refund amount does not match",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "8.99"}],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(PosReturn.objects.filter(sale=sale).exists())
        self.assertEqual(PosReturnLine.objects.count(), 0)
        self.assertEqual(PosRefund.objects.count(), 0)
        self.assertEqual(
            InventoryTransaction.objects.filter(src_model="PosReturnLine").count(), 0
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("8.0000"),
        )

    def test_pos_return_restores_stock_to_original_issued_inventory_layers(self):
        layered_product = Product.objects.create(
            owner=self.owner,
            code="POS-LAYER-SKU",
            name="POS Layer Product",
            sku="POS-LAYER-SKU",
            unit_barcode="POS-LAYER-BAR",
            base_uom=self.uom,
            price=Decimal("5.00"),
            min_price=Decimal("1.00"),
            batch_control=False,
            expiry_control=True,
        )
        first_detail = InventoryDetail.objects.create(
            owner=self.owner,
            product=layered_product,
            warehouse=self.warehouse,
            location=self.location,
            expiry_date=datetime.date(2026, 1, 1),
            onhand_qty=Decimal("1.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )
        second_detail = InventoryDetail.objects.create(
            owner=self.owner,
            product=layered_product,
            warehouse=self.warehouse,
            location=self.location,
            expiry_date=datetime.date(2026, 2, 1),
            onhand_qty=Decimal("2.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RETURN-LAYERS",
                "payment": self.payment("10.00"),
                "items": [
                    {
                        "product_id": layered_product.id,
                        "qty": "2.000",
                        "price": "5.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        first_detail.refresh_from_db()
        second_detail.refresh_from_db()
        self.assertEqual(first_detail.available_qty, Decimal("0.0000"))
        self.assertEqual(second_detail.available_qty, Decimal("1.0000"))
        sale = PosSale.objects.get(src_bill_no="POS-RETURN-LAYERS")
        sale_line = PosSaleLine.objects.get(sale=sale)
        task = WmsTask.objects.get(source_app="pos", ref_no=sale.sale_no)
        task_lines = list(WmsTaskLine.objects.filter(task=task).order_by("id"))
        self.assertEqual(len(task_lines), 2)
        self.assertEqual(
            {line.plan_meta["source_inventory_detail_id"] for line in task_lines},
            {first_detail.id, second_detail.id},
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="WmsTask",
                src_id=task.id,
                tx_type=InvTxType.ISSUE,
            ).count(),
            2,
        )

        pos_return = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "return first issued layer",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "5.00"}],
            },
            format="json",
        )

        self.assertEqual(pos_return.status_code, 201, pos_return.data)
        first_detail.refresh_from_db()
        second_detail.refresh_from_db()
        self.assertEqual(first_detail.available_qty, Decimal("1.0000"))
        self.assertEqual(second_detail.available_qty, Decimal("1.0000"))
        self.assertEqual(
            InventorySummary.objects.get(product=layered_product).available_qty,
            Decimal("2.0000"),
        )
        return_line = PosReturnLine.objects.get(return_order__sale=sale)
        receive_tx = InventoryTransaction.objects.get(
            src_model="PosReturnLine",
            src_id=return_line.id,
            tx_type=InvTxType.RECEIVE,
        )
        self.assertEqual(receive_tx.src_line_id, first_detail.id)
        self.assertEqual(receive_tx.qty_delta, Decimal("1.0000"))

    def test_two_partial_returns_exhaust_sale_then_third_is_rejected(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-PARTIAL-TWICE",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-PARTIAL-TWICE")
        sale_line = PosSaleLine.objects.get(sale=sale)
        detail = InventoryDetail.objects.get(owner=self.owner, product=self.product)
        self.assertEqual(detail.onhand_qty, Decimal("8.0000"))

        for index in (1, 2):
            response = self.client.post(
                "/api/pos/returns/",
                {
                    "sale_id": sale.id,
                    "reason": f"partial return {index}",
                    "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                    "refunds": [{"method": "CASH", "amount": "9.00"}],
                },
                format="json",
            )
            self.assertEqual(response.status_code, 201, response.data)
            detail.refresh_from_db()
            self.assertEqual(detail.onhand_qty, Decimal(f"{8 + index}.0000"))

        restored = InventoryTransaction.objects.filter(
            src_model="PosReturnLine",
            tx_type=InvTxType.RECEIVE,
            src_line_id=detail.id,
        )
        self.assertEqual(restored.count(), 2)
        self.assertEqual(
            sum((tx.qty_delta for tx in restored), Decimal("0")), Decimal("2.0000")
        )

        third = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "over return",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "9.00"}],
            },
            format="json",
        )
        self.assertEqual(third.status_code, 400)
        detail.refresh_from_db()
        self.assertEqual(detail.onhand_qty, Decimal("10.0000"))

    def test_return_never_treats_task_line_id_as_inventory_detail_id(self):
        collision_product = Product.objects.create(
            owner=self.owner,
            code="POS-ID-COLLISION-SKU",
            name="POS ID Collision Product",
            sku="POS-ID-COLLISION-SKU",
            unit_barcode="POS-ID-COLLISION-BAR",
            base_uom=self.uom,
            price=Decimal("7.00"),
            min_price=Decimal("1.00"),
        )
        source_detail = InventoryDetail.objects.create(
            id=1_000_000,
            owner=self.owner,
            product=collision_product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("1.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-ID-COLLISION",
                "payment": self.payment("7.00"),
                "items": [
                    {
                        "product_id": collision_product.id,
                        "qty": "1.000",
                        "price": "7.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-ID-COLLISION")
        sale_line = PosSaleLine.objects.get(sale=sale)
        task_line = WmsTaskLine.objects.get(
            task__source_app="pos", task__ref_no=sale.sale_no
        )
        self.assertNotEqual(task_line.id, source_detail.id)

        decoy = InventoryDetail.all_objects.filter(pk=task_line.id).first()
        if decoy is None:
            decoy_product = Product.objects.create(
                owner=self.owner,
                code="POS-ID-DECOY-SKU",
                name="POS ID Decoy Product",
                sku="POS-ID-DECOY-SKU",
                unit_barcode="POS-ID-DECOY-BAR",
                base_uom=self.uom,
                price=Decimal("1.00"),
                min_price=Decimal("0.10"),
            )
            decoy = InventoryDetail.objects.create(
                id=task_line.id,
                owner=self.owner,
                product=decoy_product,
                warehouse=self.warehouse,
                location=self.location,
                onhand_qty=Decimal("4.0000"),
                allocated_qty=Decimal("0.0000"),
                locked_qty=Decimal("0.0000"),
                damaged_qty=Decimal("0.0000"),
                base_unit=self.uom.code,
            )
        decoy_onhand = decoy.onhand_qty

        pos_return = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "verify id semantics",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "7.00"}],
            },
            format="json",
        )

        self.assertEqual(pos_return.status_code, 201, pos_return.data)
        source_detail.refresh_from_db()
        decoy.refresh_from_db()
        self.assertEqual(source_detail.onhand_qty, Decimal("1.0000"))
        self.assertEqual(decoy.onhand_qty, decoy_onhand)
        receive = InventoryTransaction.objects.get(
            src_model="PosReturnLine", tx_type=InvTxType.RECEIVE
        )
        self.assertEqual(receive.src_line_id, source_detail.id)

    def test_return_ignores_unrelated_task_with_same_generic_source_pk(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-UNRELATED-TASK-PK",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-UNRELATED-TASK-PK")
        sale_line = PosSaleLine.objects.get(sale=sale)
        now = timezone.now()
        unrelated_task = WmsTask.objects.create(
            task_no=f"UNRELATED-{sale.id}",
            task_type=WmsTask.TaskType.PUTAWAY,
            owner=self.owner,
            warehouse=self.warehouse,
            source_app="inbound",
            source_model="inboundorder",
            source_pk=str(sale_line.outbound_order_line.order_id),
            status=WmsTask.Status.COMPLETED,
            review_status=WmsTask.ReviewStatus.APPROVED,
            posting_status=WmsTask.PostingStatus.POSTED,
            released_at=now,
            started_at=now,
            finished_at=now,
            approved_at=now,
            created_by=self.user,
        )
        unrelated_line = WmsTaskLine.objects.create(
            task=unrelated_task,
            product=self.product,
            from_location=self.location,
            to_location=self.location,
            qty_plan=Decimal("0.250"),
            qty_done=Decimal("0.250"),
            status=WmsTaskLine.Status.COMPLETED,
            src_model="InboundOrderLine",
            src_id=sale_line.outbound_order_line_id,
        )
        detail = InventoryDetail.objects.get(
            owner=self.owner,
            product=self.product,
            is_active=True,
        )
        InventoryTransaction.objects.create(
            tx_type=InvTxType.ISSUE,
            owner=self.owner,
            product=self.product,
            warehouse=self.warehouse,
            location=self.location,
            subwarehouse=detail.subwarehouse,
            zone_type=detail.zone_type,
            qty_delta=Decimal("-0.2500"),
            src_model="WmsTask",
            src_id=unrelated_task.id,
            src_line_id=unrelated_line.id,
            src_no="UNRELATED",
            memo="UNRELATED",
            posted_at=now,
            posting_batch="UNRELATED",
        )

        pos_return = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "ignore unrelated task source pk",
                "lines": [{"sale_line_id": sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "CASH", "amount": "9.00"}],
            },
            format="json",
        )

        self.assertEqual(pos_return.status_code, 201, pos_return.data)
        receive = InventoryTransaction.objects.get(
            src_model="PosReturnLine",
            tx_type=InvTxType.RECEIVE,
        )
        self.assertEqual(
            receive.src_line_id,
            WmsTaskLine.objects.get(
                task__source_app="pos",
                plan_meta__pos_sale_line_id=sale_line.id,
            ).plan_meta["source_inventory_detail_id"],
        )

    def test_checkout_rejects_price_below_min_price(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-RECEIPT-LOW",
                "payment": self.payment("8.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "7.9900",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            OutboundOrder.objects.filter(src_bill_no="POS-RECEIPT-LOW").count(),
            0,
        )

    def test_checkout_rejects_explicit_amount_below_minimum_total(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-RECEIPT-LOW-AMOUNT",
                "payment": self.payment("15.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "7.5000",
                        "amount": "15.00",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-LOW-AMOUNT").exists()
        )
        self.assertFalse(
            OutboundOrder.objects.filter(src_bill_no="POS-RECEIPT-LOW-AMOUNT").exists()
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("10.0000"),
        )

    def test_checkout_rejects_max_discount_breach_without_stock_post(self):
        discount_product = Product.objects.create(
            owner=self.owner,
            code="POS-DISCOUNT-SKU",
            name="POS Discount Product",
            sku="POS-DISCOUNT-SKU",
            unit_barcode="POS-DISCOUNT-BAR",
            base_uom=self.uom,
            price=Decimal("100.00"),
            min_price=Decimal("1.00"),
            max_discount=Decimal("10.00"),
            batch_control=False,
            expiry_control=False,
        )
        InventoryDetail.objects.create(
            owner=self.owner,
            product=discount_product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("1.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-RECEIPT-DISCOUNT",
                "payment": self.payment("89.99"),
                "items": [
                    {
                        "product_id": discount_product.id,
                        "qty": "1.000",
                        "price": "89.9900",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-DISCOUNT").exists()
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                product=discount_product, src_model="PosSaleLine"
            ).count(),
            0,
        )
        self.assertEqual(
            InventoryDetail.objects.get(product=discount_product).available_qty,
            Decimal("1.0000"),
        )

    def test_checkout_rejects_repeated_product_lines_above_available_without_posting(
        self,
    ):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-RECEIPT-AGG-STOCK",
                "payment": self.payment("99.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "6.000",
                        "price": "9.0000",
                    },
                    {
                        "product_id": self.product.id,
                        "qty": "5.000",
                        "price": "9.0000",
                    },
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-AGG-STOCK").exists()
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(src_model="PosSaleLine").count(), 0
        )
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.owner, product=self.product
            ).available_qty,
            Decimal("10.0000"),
        )

    def test_checkout_does_not_combine_sub_millith_inventory_tails(self):
        tail_product = Product.objects.create(
            owner=self.owner,
            code="POS-TAIL-SKU",
            name="POS Tail Product",
            sku="POS-TAIL-SKU",
            unit_barcode="POS-TAIL-BAR",
            base_uom=self.uom,
            price=Decimal("1000.00"),
            min_price=Decimal("1.00"),
        )
        second_location = Location.objects.create(
            warehouse=self.warehouse,
            code="SWPOS-01-01-02",
            name="POS Tail Location",
        )
        for location in (self.location, second_location):
            InventoryDetail.objects.create(
                owner=self.owner,
                product=tail_product,
                warehouse=self.warehouse,
                location=location,
                onhand_qty=Decimal("0.0006"),
                allocated_qty=Decimal("0.0000"),
                locked_qty=Decimal("0.0000"),
                damaged_qty=Decimal("0.0000"),
                base_unit=self.uom.code,
            )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-SUB-MILLITH-TAILS",
                "payment": self.payment("1.00"),
                "items": [
                    {
                        "product_id": tail_product.id,
                        "qty": "0.001",
                        "price": "1000.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-SUB-MILLITH-TAILS").exists()
        )
        self.assertFalse(WmsTask.objects.filter(source_app="pos").exists())
        self.assertFalse(
            InventoryTransaction.objects.filter(product=tail_product).exists()
        )
        self.assertEqual(
            list(
                InventoryDetail.objects.filter(product=tail_product)
                .order_by("id")
                .values_list("available_qty", flat=True)
            ),
            [Decimal("0.0006"), Decimal("0.0006")],
        )

    def test_checkout_rejects_qty_above_available_stock(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "customer_id": self.pos_customer.id,
                "src_bill_no": "POS-RECEIPT-STOCK",
                "payment": self.payment("99.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "11.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            OutboundOrder.objects.filter(src_bill_no="POS-RECEIPT-STOCK").count(),
            0,
        )

    def test_checkout_uses_requested_stock_zone_scope(self):
        pick_location = Location.objects.create(
            warehouse=self.warehouse,
            code="SWPOS-02-01-01",
            name="POS Zone Pick Location",
            zone_type=ZoneType.PICK,
        )
        zone_product = Product.objects.create(
            owner=self.owner,
            code="POS-ZONE-SKU",
            name="POS Zone Product",
            sku="POS-ZONE-SKU",
            unit_barcode="POS-ZONE-BAR",
            base_uom=self.uom,
            price=Decimal("5.00"),
            min_price=Decimal("1.00"),
            batch_control=False,
            expiry_control=False,
        )
        InventoryDetail.objects.create(
            owner=self.owner,
            product=zone_product,
            warehouse=self.warehouse,
            location=pick_location,
            zone_type=ZoneType.PICK,
            onhand_qty=Decimal("3.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )

        lookup = self.client.get(
            "/api/pos/products/",
            {"barcode": "POS-ZONE-BAR", "zone_type": ZoneType.STORAGE},
        )
        self.assertEqual(lookup.status_code, 200)
        self.assertEqual(
            Decimal(str(lookup.data["results"][0]["available_qty"])), Decimal("0")
        )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-ZONE",
                "stock_zone_type": ZoneType.STORAGE,
                "payment": self.payment("5.00"),
                "items": [
                    {
                        "product_id": zone_product.id,
                        "qty": "1.000",
                        "price": "5.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            PosSale.objects.filter(src_bill_no="POS-RECEIPT-ZONE").exists()
        )

        matching_zone = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-ZONE-OK",
                "stock_zone_type": ZoneType.PICK,
                "payment": self.payment("5.00"),
                "items": [
                    {
                        "product_id": zone_product.id,
                        "qty": "1.000",
                        "price": "5.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(matching_zone.status_code, 201, matching_zone.data)
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-ZONE-OK")
        task = WmsTask.objects.get(source_app="pos", ref_no=sale.sale_no)
        task_line = WmsTaskLine.objects.get(task=task)
        zone_detail = InventoryDetail.objects.get(
            owner=self.owner, product=zone_product
        )
        self.assertEqual(
            task_line.plan_meta["source_inventory_detail_id"], zone_detail.id
        )
        self.assertEqual(task_line.plan_meta["zone_type"], ZoneType.PICK)
        issue_tx = InventoryTransaction.objects.get(
            src_model="WmsTask",
            src_id=task.id,
            src_line_id=task_line.id,
            tx_type=InvTxType.ISSUE,
        )
        self.assertEqual(issue_tx.zone_type, ZoneType.PICK)

    def test_checkout_fefo_consumes_expiring_stock_before_no_expiry_stock(self):
        fefo_product = Product.objects.create(
            owner=self.owner,
            code="POS-FEFO-SKU",
            name="POS FEFO Product",
            sku="POS-FEFO-SKU",
            unit_barcode="POS-FEFO-BAR",
            base_uom=self.uom,
            price=Decimal("6.00"),
            min_price=Decimal("1.00"),
            batch_control=False,
            expiry_control=True,
        )
        expiring_detail = InventoryDetail.objects.create(
            owner=self.owner,
            product=fefo_product,
            warehouse=self.warehouse,
            location=self.location,
            expiry_date=datetime.date(2026, 1, 1),
            onhand_qty=Decimal("2.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )
        no_expiry_detail = InventoryDetail.objects.create(
            owner=self.owner,
            product=fefo_product,
            warehouse=self.warehouse,
            location=self.location,
            onhand_qty=Decimal("2.0000"),
            allocated_qty=Decimal("0.0000"),
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.uom.code,
        )

        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-FEFO",
                "payment": self.payment("6.00"),
                "items": [
                    {
                        "product_id": fefo_product.id,
                        "qty": "1.000",
                        "price": "6.0000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.data)
        expiring_detail.refresh_from_db()
        no_expiry_detail.refresh_from_db()
        self.assertEqual(expiring_detail.available_qty, Decimal("1.0000"))
        self.assertEqual(no_expiry_detail.available_qty, Decimal("2.0000"))
        sale = PosSale.objects.get(src_bill_no="POS-RECEIPT-FEFO")
        task = WmsTask.objects.get(source_app="pos", ref_no=sale.sale_no)
        task_line = WmsTaskLine.objects.get(task=task)
        self.assertEqual(
            task_line.plan_meta["source_inventory_detail_id"], expiring_detail.id
        )
        self.assertEqual(task_line.from_location_id, expiring_detail.location_id)
        issue_tx = InventoryTransaction.objects.get(
            src_model="WmsTask",
            src_id=task.id,
            src_line_id=task_line.id,
            tx_type=InvTxType.ISSUE,
        )
        self.assertEqual(issue_tx.expiry_date, datetime.date(2026, 1, 1))

    def test_sales_list_filters_by_user_warehouse_and_search(self):
        response = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-RECEIPT-LIST",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201, response.data)
        other_warehouse = Warehouse.objects.create(
            code="WHPOS2", name="Other POS Warehouse"
        )
        PosSale.objects.create(
            sale_no="POS-OTHER-WAREHOUSE",
            src_bill_no="POS-RECEIPT-OTHER-WAREHOUSE",
            warehouse=other_warehouse,
            total_amount=Decimal("1.00"),
        )

        list_response = self.client.get("/api/pos/sales/", {"search": "LIST"})

        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(list_response.data["count"], 1)
        row = list_response.data["results"][0]
        self.assertEqual(row["src_bill_no"], "POS-RECEIPT-LIST")
        self.assertEqual(row["warehouse_id"], self.warehouse.id)

        other_response = self.client.get(
            "/api/pos/sales/", {"search": "OTHER-WAREHOUSE"}
        )
        self.assertEqual(other_response.status_code, 200)
        self.assertEqual(other_response.data["count"], 0)

    def test_sales_list_requires_pos_view_permission(self):
        no_view_user = get_user_model().objects.create_user(
            username="pos-no-view",
            password="x",
            warehouse=self.warehouse,
        )
        self.client.force_authenticate(no_view_user)

        response = self.client.get("/api/pos/sales/")

        self.assertEqual(response.status_code, 403)

    def test_pos_stats_returns_accurate_totals_for_multi_owner_void_and_payments(
        self,
    ):
        multi_owner = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-STATS-MULTI",
                "payment": self.payment("45.00", method="WECHAT"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    },
                    {
                        "product_id": self.other_product.id,
                        "qty": "2.000",
                        "price": "18.0000",
                    },
                ],
            },
            format="json",
        )
        cash_sale = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-STATS-CASH",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        void_checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-STATS-VOID",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(multi_owner.status_code, 201, multi_owner.data)
        self.assertEqual(cash_sale.status_code, 201, cash_sale.data)
        self.assertEqual(void_checkout.status_code, 201, void_checkout.data)
        void_response = self.client.post(
            f"/api/pos/sales/{void_checkout.data['sale']['id']}/void/",
            {"reason": "stats check"},
            format="json",
        )
        self.assertEqual(void_response.status_code, 200, void_response.data)

        other_warehouse = Warehouse.objects.create(
            code="WHPOSSTAT", name="Other POS Stats Warehouse"
        )
        other_sale = PosSale.objects.create(
            sale_no="POS-STATS-OTHER-WAREHOUSE",
            src_bill_no="POS-STATS-OTHER-WAREHOUSE",
            warehouse=other_warehouse,
            cashier=self.user,
            total_amount=Decimal("100.00"),
        )
        PosPayment.objects.create(
            sale=other_sale,
            method=PosPayment.Method.CASH,
            amount_due=Decimal("100.00"),
            amount_received=Decimal("100.00"),
        )
        PosSaleLine.objects.create(
            sale=other_sale,
            owner=self.owner,
            product=self.product,
            line_no=10,
            qty=Decimal("1.000"),
            price=Decimal("100.0000"),
            amount=Decimal("100.00"),
        )

        today = timezone.now().date().isoformat()
        response = self.client.get(
            "/api/pos/stats/",
            {"start_date": today, "end_date": today, "top_n": 5},
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["summary"]["sale_count"], 3)
        self.assertEqual(response.data["summary"]["completed_count"], 2)
        self.assertEqual(response.data["summary"]["voided_count"], 1)
        self.assertEqual(response.data["summary"]["gross_amount"], "72.00")
        self.assertEqual(response.data["summary"]["net_amount"], "63.00")
        self.assertEqual(response.data["summary"]["voided_amount"], "9.00")

        payments = {row["method"]: row for row in response.data["payments"]}
        self.assertEqual(payments["WECHAT"]["sale_count"], 1)
        self.assertEqual(payments["WECHAT"]["amount"], "45.00")
        self.assertEqual(payments["CASH"]["sale_count"], 1)
        self.assertEqual(payments["CASH"]["amount"], "18.00")

        owners = {row["owner_id"]: row for row in response.data["owners"]}
        self.assertEqual(owners[self.owner.id]["sale_count"], 2)
        self.assertEqual(owners[self.owner.id]["qty"], "3.000")
        self.assertEqual(owners[self.owner.id]["amount"], "27.00")
        self.assertEqual(owners[self.other_owner.id]["sale_count"], 1)
        self.assertEqual(owners[self.other_owner.id]["qty"], "2.000")
        self.assertEqual(owners[self.other_owner.id]["amount"], "36.00")

        products = {row["product_id"]: row for row in response.data["products"]}
        self.assertEqual(products[self.product.id]["amount"], "27.00")
        self.assertEqual(products[self.product.id]["qty"], "3.000")
        self.assertEqual(products[self.other_product.id]["amount"], "36.00")
        self.assertEqual(products[self.other_product.id]["qty"], "2.000")

        cashiers = {row["cashier_id"]: row for row in response.data["cashiers"]}
        self.assertEqual(cashiers[self.user.id]["sale_count"], 3)
        self.assertEqual(cashiers[self.user.id]["completed_count"], 2)
        self.assertEqual(cashiers[self.user.id]["voided_count"], 1)
        self.assertEqual(cashiers[self.user.id]["completed_amount"], "63.00")
        self.assertEqual(cashiers[self.user.id]["voided_amount"], "9.00")

        owner_response = self.client.get(
            "/api/pos/stats/",
            {"start_date": today, "end_date": today, "owner_id": self.owner.id},
        )

        self.assertEqual(owner_response.status_code, 200, owner_response.data)
        self.assertEqual(owner_response.data["summary"]["sale_count"], 3)
        self.assertEqual(owner_response.data["summary"]["gross_amount"], "36.00")
        self.assertEqual(owner_response.data["summary"]["net_amount"], "27.00")
        self.assertEqual(owner_response.data["summary"]["voided_amount"], "9.00")
        owner_payments = {row["method"]: row for row in owner_response.data["payments"]}
        self.assertEqual(owner_payments["WECHAT"]["amount"], "9.00")
        self.assertEqual(owner_payments["CASH"]["amount"], "18.00")
        self.assertEqual(len(owner_response.data["owners"]), 1)
        self.assertEqual(owner_response.data["owners"][0]["owner_id"], self.owner.id)

    def test_pos_stats_allocates_split_payments_and_refunds_by_owner_scope(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-STATS-ALLOCATED-PAY",
                "payments": [
                    {
                        "method": "CASH",
                        "amount": "10.00",
                        "amount_received": "10.00",
                    },
                    {
                        "method": "WECHAT",
                        "amount": "35.00",
                        "amount_received": "35.00",
                    },
                ],
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    },
                    {
                        "product_id": self.other_product.id,
                        "qty": "2.000",
                        "price": "18.0000",
                    },
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale = PosSale.objects.get(src_bill_no="POS-STATS-ALLOCATED-PAY")
        other_sale_line = PosSaleLine.objects.get(sale=sale, product=self.other_product)
        pos_return = self.client.post(
            "/api/pos/returns/",
            {
                "sale_id": sale.id,
                "reason": "customer returned one other-owner item",
                "lines": [{"sale_line_id": other_sale_line.id, "qty": "1.000"}],
                "refunds": [{"method": "WECHAT", "amount": "18.00"}],
            },
            format="json",
        )
        self.assertEqual(pos_return.status_code, 201, pos_return.data)

        today = timezone.now().date().isoformat()
        response = self.client.get(
            "/api/pos/stats/",
            {"start_date": today, "end_date": today, "top_n": 5},
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["summary"]["sales_amount"], "45.00")
        self.assertEqual(response.data["summary"]["return_amount"], "18.00")
        self.assertEqual(response.data["summary"]["net_amount"], "27.00")
        payments = {row["method"]: row for row in response.data["payments"]}
        self.assertEqual(payments["CASH"]["sale_amount"], "10.00")
        self.assertEqual(payments["CASH"]["refund_amount"], "0.00")
        self.assertEqual(payments["CASH"]["net_amount"], "10.00")
        self.assertEqual(payments["WECHAT"]["sale_amount"], "35.00")
        self.assertEqual(payments["WECHAT"]["refund_amount"], "18.00")
        self.assertEqual(payments["WECHAT"]["net_amount"], "17.00")
        owners = {row["owner_id"]: row for row in response.data["owners"]}
        self.assertEqual(owners[self.owner.id]["sale_amount"], "9.00")
        self.assertEqual(owners[self.owner.id]["return_amount"], "0.00")
        self.assertEqual(owners[self.owner.id]["net_amount"], "9.00")
        self.assertEqual(owners[self.other_owner.id]["sale_amount"], "36.00")
        self.assertEqual(owners[self.other_owner.id]["return_amount"], "18.00")
        self.assertEqual(owners[self.other_owner.id]["net_amount"], "18.00")

        owner_response = self.client.get(
            "/api/pos/stats/",
            {"start_date": today, "end_date": today, "owner_id": self.owner.id},
        )
        other_owner_response = self.client.get(
            "/api/pos/stats/",
            {
                "start_date": today,
                "end_date": today,
                "owner_id": self.other_owner.id,
            },
        )

        self.assertEqual(owner_response.status_code, 200, owner_response.data)
        owner_payments = {row["method"]: row for row in owner_response.data["payments"]}
        self.assertEqual(owner_response.data["summary"]["net_amount"], "9.00")
        self.assertEqual(owner_payments["CASH"]["sale_amount"], "2.00")
        self.assertEqual(owner_payments["WECHAT"]["sale_amount"], "7.00")
        self.assertEqual(owner_payments["WECHAT"]["refund_amount"], "0.00")

        self.assertEqual(
            other_owner_response.status_code, 200, other_owner_response.data
        )
        other_payments = {
            row["method"]: row for row in other_owner_response.data["payments"]
        }
        self.assertEqual(other_owner_response.data["summary"]["sales_amount"], "36.00")
        self.assertEqual(other_owner_response.data["summary"]["return_amount"], "18.00")
        self.assertEqual(other_owner_response.data["summary"]["net_amount"], "18.00")
        self.assertEqual(other_payments["CASH"]["sale_amount"], "8.00")
        self.assertEqual(other_payments["CASH"]["net_amount"], "8.00")
        self.assertEqual(other_payments["WECHAT"]["sale_amount"], "28.00")
        self.assertEqual(other_payments["WECHAT"]["refund_amount"], "18.00")
        self.assertEqual(other_payments["WECHAT"]["net_amount"], "10.00")
        self.assertEqual(
            InventoryDetail.objects.get(
                owner=self.other_owner, product=self.other_product
            ).available_qty,
            Decimal("7.0000"),
        )

    def test_shift_close_summarizes_payments_and_blocks_late_void(self):
        multi_owner = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-SHIFT-MULTI",
                "payment": self.payment("45.00", method="WECHAT"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    },
                    {
                        "product_id": self.other_product.id,
                        "qty": "2.000",
                        "price": "18.0000",
                    },
                ],
            },
            format="json",
        )
        cash_sale = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-SHIFT-CASH",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        void_checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-SHIFT-VOID",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(multi_owner.status_code, 201, multi_owner.data)
        self.assertEqual(cash_sale.status_code, 201, cash_sale.data)
        self.assertEqual(void_checkout.status_code, 201, void_checkout.data)

        void_response = self.client.post(
            f"/api/pos/sales/{void_checkout.data['sale']['id']}/void/",
            {"reason": "shift accuracy"},
            format="json",
        )
        self.assertEqual(void_response.status_code, 200, void_response.data)

        close_response = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/close/",
            {
                "actual_cash_amount": "118.00",
                "payments": [{"method": "WECHAT", "actual_amount": "45.00"}],
                "remark": "end shift",
            },
            format="json",
        )

        self.assertEqual(close_response.status_code, 200, close_response.data)
        summary = close_response.data["shift"]["summary"]
        self.assertEqual(summary["sale_count"], 3)
        self.assertEqual(summary["completed_count"], 2)
        self.assertEqual(summary["voided_count"], 1)
        self.assertEqual(summary["net_amount"], "63.00")
        self.assertEqual(summary["voided_amount"], "9.00")
        self.assertEqual(summary["opening_cash_amount"], "100.00")
        self.assertEqual(summary["expected_cash_amount"], "118.00")
        self.assertEqual(summary["actual_cash_amount"], "118.00")
        self.assertEqual(summary["cash_difference"], "0.00")

        shift = PosShift.objects.get(pk=self.shift.id)
        self.assertEqual(shift.status, PosShift.Status.CLOSED)
        self.assertEqual(shift.total_sales_amount, Decimal("63.00"))
        self.assertEqual(shift.total_voided_amount, Decimal("9.00"))
        payments = {
            row.method: row
            for row in PosShiftPaymentSummary.objects.filter(shift=self.shift)
        }
        self.assertEqual(
            payments[PosPayment.Method.CASH].expected_amount, Decimal("18.00")
        )
        self.assertEqual(
            payments[PosPayment.Method.CASH].actual_amount, Decimal("18.00")
        )
        self.assertEqual(
            payments[PosPayment.Method.WECHAT].expected_amount, Decimal("45.00")
        )
        self.assertEqual(payments[PosPayment.Method.WECHAT].difference, Decimal("0.00"))

        late_void = self.client.post(
            f"/api/pos/sales/{cash_sale.data['sale']['id']}/void/",
            {"reason": "late void"},
            format="json",
        )
        self.assertEqual(late_void.status_code, 400)
        self.assertEqual(
            PosSale.objects.get(pk=cash_sale.data["sale"]["id"]).status,
            PosSale.Status.COMPLETED,
        )

    def test_print_and_excel_export_are_traceable_to_sale_data(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-EXPORT-PRINT",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale_id = checkout.data["sale"]["id"]

        first_print = self.client.get(f"/api/pos/sales/{sale_id}/print/")
        second_print = self.client.get(f"/api/pos/sales/{sale_id}/print/")
        self.assertEqual(first_print.status_code, 200)
        self.assertEqual(second_print.status_code, 200)
        self.assertContains(first_print, "POS-EXPORT-PRINT")
        logs = PosPrintLog.objects.filter(
            sale_id=sale_id,
            print_type=PosPrintLog.PrintType.RECEIPT,
        ).order_by("copy_no")
        self.assertEqual(list(logs.values_list("copy_no", flat=True)), [1, 2])

        sales_export = self.client.get(
            "/api/pos/sales/export/", {"search": "POS-EXPORT-PRINT"}
        )
        self.assertEqual(sales_export.status_code, 200)
        workbook = load_workbook(io.BytesIO(sales_export.content))
        self.assertIn("Sales", workbook.sheetnames)
        self.assertIn("Lines", workbook.sheetnames)
        sales_sheet = workbook["Sales"]
        lines_sheet = workbook["Lines"]
        self.assertEqual(sales_sheet["B2"].value, "POS-EXPORT-PRINT")
        self.assertEqual(Decimal(str(sales_sheet["I2"].value)), Decimal("9"))
        self.assertEqual(lines_sheet["B2"].value, "POS-EXPORT-PRINT")
        self.assertEqual(Decimal(str(lines_sheet["K2"].value)), Decimal("9"))

        stats_export = self.client.get("/api/pos/stats/export/")
        self.assertEqual(stats_export.status_code, 200)
        stats_workbook = load_workbook(io.BytesIO(stats_export.content))
        self.assertIn("Summary", stats_workbook.sheetnames)
        summary_rows = {
            row[0].value: row[1].value
            for row in stats_workbook["Summary"].iter_rows(min_row=2, max_col=2)
        }
        self.assertEqual(summary_rows["net_amount"], "9.00")

    def test_frontend_print_log_records_source_and_copy_number(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-FRONTEND-PRINT",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale_id = checkout.data["sale"]["id"]

        first = self.client.post(
            f"/api/pos/sales/{sale_id}/print-log/",
            {"remark": "frontend auto print"},
            format="json",
        )
        second = self.client.post(
            f"/api/pos/sales/{sale_id}/print-log/",
            {"remark": "frontend reprint"},
            format="json",
        )

        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(second.status_code, 201, second.data)
        self.assertEqual(first.data["copy_no"], 1)
        self.assertEqual(second.data["copy_no"], 2)
        logs = PosPrintLog.objects.filter(
            sale_id=sale_id,
            print_type=PosPrintLog.PrintType.RECEIPT,
        ).order_by("copy_no")
        self.assertEqual(
            list(logs.values_list("source", flat=True)),
            [PosPrintLog.Source.FRONTEND_HTML, PosPrintLog.Source.FRONTEND_HTML],
        )
        self.assertTrue(all(log.payload_hash for log in logs))

    def test_print_pages_accept_query_token_and_record_backend_source(self):
        checkout = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-TOKEN-PRINT",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(checkout.status_code, 201, checkout.data)
        sale_id = checkout.data["sale"]["id"]
        token = str(RefreshToken.for_user(self.user).access_token)
        anonymous = APIClient()

        sale_print = anonymous.get(f"/api/pos/sales/{sale_id}/print/?token={token}")
        shift_print = anonymous.get(
            f"/api/pos/shifts/{self.shift.id}/print/?token={token}"
        )

        self.assertEqual(sale_print.status_code, 200)
        self.assertContains(sale_print, "POS-TOKEN-PRINT")
        self.assertEqual(shift_print.status_code, 200)
        self.assertContains(shift_print, self.shift.shift_no)
        self.assertEqual(
            PosPrintLog.objects.get(
                sale_id=sale_id, print_type=PosPrintLog.PrintType.RECEIPT
            ).source,
            PosPrintLog.Source.BACKEND_HTML,
        )
        self.assertEqual(
            PosPrintLog.objects.get(
                shift_id=self.shift.id,
                print_type=PosPrintLog.PrintType.SHIFT_SUMMARY,
            ).source,
            PosPrintLog.Source.BACKEND_HTML,
        )

    def test_reopen_shift_allows_more_sales_and_recalculates_shift_totals(self):
        first = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-REOPEN-FIRST",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(first.status_code, 201, first.data)
        close_first = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/close/",
            {"actual_cash_amount": "109.00"},
            format="json",
        )
        self.assertEqual(close_first.status_code, 200, close_first.data)

        blocked_sale = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-REOPEN-BLOCKED",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(blocked_sale.status_code, 400)

        no_supervisor = get_user_model().objects.create_user(
            username="pos-no-reopen-perm",
            password="x",
            warehouse=self.warehouse,
        )
        no_supervisor.user_permissions.add(
            Permission.objects.get(codename="view_possale")
        )
        no_supervisor_client = APIClient()
        no_supervisor_client.force_authenticate(no_supervisor)
        denied = no_supervisor_client.post(
            f"/api/pos/shifts/{self.shift.id}/reopen/",
            {"reason": "no permission"},
            format="json",
        )
        self.assertEqual(denied.status_code, 403)

        reopen = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/reopen/",
            {"reason": "cashier found missed sale"},
            format="json",
        )
        self.assertEqual(reopen.status_code, 200, reopen.data)
        self.assertEqual(reopen.data["shift"]["status"], PosShift.Status.REOPENED)
        self.assertEqual(reopen.data["shift"]["reopen_count"], 1)
        current = self.client.get("/api/pos/shifts/current/")
        self.assertEqual(current.status_code, 200, current.data)
        self.assertEqual(current.data["shift"]["id"], self.shift.id)

        second = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-REOPEN-SECOND",
                "payment": self.payment("18.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "2.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(second.status_code, 201, second.data)
        self.assertEqual(second.data["sale"]["shift_id"], self.shift.id)

        close_second = self.client.post(
            f"/api/pos/shifts/{self.shift.id}/close/",
            {"actual_cash_amount": "127.00"},
            format="json",
        )
        self.assertEqual(close_second.status_code, 200, close_second.data)
        summary = close_second.data["shift"]["summary"]
        self.assertEqual(summary["sale_count"], 2)
        self.assertEqual(summary["completed_count"], 2)
        self.assertEqual(summary["net_amount"], "27.00")
        self.assertEqual(summary["expected_cash_amount"], "127.00")
        self.assertEqual(summary["actual_cash_amount"], "127.00")
        self.assertEqual(summary["cash_difference"], "0.00")

        shift = PosShift.objects.get(pk=self.shift.id)
        self.assertEqual(shift.status, PosShift.Status.CLOSED)
        self.assertEqual(shift.reopen_count, 1)
        self.assertEqual(shift.reopen_reason, "cashier found missed sale")
        self.assertEqual(shift.total_sales_amount, Decimal("27.00"))
        cash_summary = PosShiftPaymentSummary.objects.get(
            shift=self.shift, method=PosPayment.Method.CASH
        )
        self.assertEqual(cash_summary.expected_amount, Decimal("27.00"))
        self.assertEqual(cash_summary.actual_amount, Decimal("27.00"))

    def test_pos_stats_respects_date_range(self):
        inside = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-STATS-DATE-IN",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        outside = self.client.post(
            "/api/pos/checkout/",
            {
                "src_bill_no": "POS-STATS-DATE-OUT",
                "payment": self.payment("9.00"),
                "items": [
                    {
                        "product_id": self.product.id,
                        "qty": "1.000",
                        "price": "9.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(inside.status_code, 201, inside.data)
        self.assertEqual(outside.status_code, 201, outside.data)
        PosSale.objects.filter(pk=inside.data["sale"]["id"]).update(
            created_at=datetime.datetime(2026, 1, 10, 10)
        )
        PosSale.objects.filter(pk=outside.data["sale"]["id"]).update(
            created_at=datetime.datetime(2026, 1, 11, 10)
        )

        response = self.client.get(
            "/api/pos/stats/",
            {"start_date": "2026-01-10", "end_date": "2026-01-10"},
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["summary"]["sale_count"], 1)
        self.assertEqual(response.data["summary"]["net_amount"], "9.00")

    def test_pos_stats_requires_pos_view_permission(self):
        no_view_user = get_user_model().objects.create_user(
            username="pos-stats-no-view",
            password="x",
            warehouse=self.warehouse,
        )
        self.client.force_authenticate(no_view_user)

        response = self.client.get("/api/pos/stats/")

        self.assertEqual(response.status_code, 403)

    def test_pos_stats_rejects_invalid_date_range(self):
        response = self.client.get(
            "/api/pos/stats/",
            {"start_date": "2026-01-11", "end_date": "2026-01-10"},
        )

        self.assertEqual(response.status_code, 400)

    def test_receipt_warehouse_info_list_keeps_global_and_current_warehouse_only(self):
        other_warehouse = Warehouse.objects.create(
            code="WHPOSR2",
            name="Other Receipt Warehouse",
        )
        global_info = PosReceiptWarehouseInfo.objects.create(
            name="Global Receipt Header",
            is_default=True,
        )
        own_info = PosReceiptWarehouseInfo.objects.create(
            warehouse=self.warehouse,
            name="Own Receipt Header",
            is_default=True,
        )
        PosReceiptWarehouseInfo.objects.create(
            warehouse=other_warehouse,
            name="Other Receipt Header",
            is_default=True,
        )
        PosReceiptWarehouseInfo.objects.create(
            warehouse=self.warehouse,
            name="Inactive Receipt Header",
            is_active=False,
        )

        response = self.client.get("/api/pos/receipt-warehouse-infos/")

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(
            {row["id"] for row in response.data},
            {global_info.id, own_info.id},
        )

    def test_customer_detail_is_warehouse_scoped_and_warehouse_is_immutable(self):
        other_warehouse = Warehouse.objects.create(
            code="WHPOSC2",
            name="Other Customer Warehouse",
        )
        other_customer = PosCustomer.objects.create(
            warehouse=other_warehouse,
            code="PC-OTHER-WH",
            name="Other Warehouse Customer",
        )

        hidden = self.client.get(f"/api/pos/customers/{other_customer.id}/")
        updated = self.client.patch(
            f"/api/pos/customers/{self.pos_customer.id}/",
            {
                "name": "Updated Current Customer",
                "warehouse_id": other_warehouse.id,
            },
            format="json",
        )

        self.assertEqual(hidden.status_code, 404)
        self.assertEqual(updated.status_code, 200, updated.data)
        self.pos_customer.refresh_from_db()
        self.assertEqual(self.pos_customer.name, "Updated Current Customer")
        self.assertEqual(self.pos_customer.warehouse_id, self.warehouse.id)

    def test_shift_list_and_detail_do_not_expose_another_warehouse(self):
        other_warehouse = Warehouse.objects.create(
            code="WHPOSS2",
            name="Other Shift Warehouse",
        )
        other_user = get_user_model().objects.create_user(
            username="pos-other-shift-user",
            password="x",
            warehouse=other_warehouse,
        )
        other_shift = PosShift.objects.create(
            shift_no="SHIFT-POS-OTHER-WAREHOUSE",
            warehouse=other_warehouse,
            cashier=other_user,
            opened_by=other_user,
            opened_at=timezone.now(),
        )

        listing = self.client.get("/api/pos/shifts/")
        hidden = self.client.get(f"/api/pos/shifts/{other_shift.id}/")

        self.assertEqual(listing.status_code, 200, listing.data)
        self.assertEqual(
            {row["id"] for row in listing.data["results"]},
            {self.shift.id},
        )
        self.assertEqual(hidden.status_code, 404)

    def test_return_and_repayment_lists_reject_non_integer_filters(self):
        invalid_return = self.client.get("/api/pos/returns/", {"sale_id": "bad"})
        invalid_repayment = self.client.get(
            "/api/pos/repayments/",
            {"customer_id": "bad"},
        )

        self.assertEqual(invalid_return.status_code, 400, invalid_return.data)
        self.assertEqual(invalid_repayment.status_code, 400, invalid_repayment.data)
