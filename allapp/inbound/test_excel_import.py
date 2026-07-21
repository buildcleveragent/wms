import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import load_workbook
from rest_framework.test import APIClient

from allapp.accounts.models import UserRoleScope
from allapp.baseinfo.models import Owner
from allapp.inventory.models import InventoryDetail
from allapp.locations.models import Location, Subwarehouse, Warehouse
from allapp.products.models import Product, ProductPackage, ProductUom
from allapp.tasking.models import WmsTask


class NoOrderReceiveExcelImportTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(code="EXOWN", name="Excel Owner")
        self.other_owner = Owner.objects.create(code="OTHER", name="Other Owner")
        self.warehouse = Warehouse.objects.create(code="EXWH", name="Excel Warehouse")
        self.subwarehouse = Subwarehouse.objects.create(
            warehouse=self.warehouse,
            code="EXSW",
            name="Excel Subwarehouse",
        )
        self.location = Location.objects.create(
            warehouse=self.warehouse,
            subwarehouse=self.subwarehouse,
            code="EXSW-01-01-01",
            name="Excel Receiving",
        )
        self.each = ProductUom.objects.create(code="EA", name="件", is_active=True)
        self.carton = ProductUom.objects.create(code="CS", name="箱", is_active=True)
        self.product = Product.objects.create(
            owner=self.owner,
            code="EXSKU1",
            sku="EXSKU1",
            name="Excel Product",
            base_uom=self.each,
            batch_control=False,
            expiry_control=False,
            serial_control=False,
        )
        ProductPackage.objects.create(
            product=self.product,
            uom=self.carton,
            qty_in_base=12,
            is_active=True,
        )
        self.serial_product = Product.objects.create(
            owner=self.owner,
            code="SERIAL1",
            sku="SERIAL1",
            name="Serial Product",
            base_uom=self.each,
            batch_control=False,
            expiry_control=False,
            serial_control=True,
        )
        self.user = get_user_model().objects.create_user(
            username="excel-receiver",
            password="x",
            warehouse=self.warehouse,
        )
        UserRoleScope.objects.create(
            user=self.user,
            role=UserRoleScope.Role.WAREHOUSE_OPERATOR,
            warehouse=self.warehouse,
        )
        permission = Permission.objects.get(
            content_type__app_label="accounts",
            codename="receive_without_order",
        )
        self.user.user_permissions.add(permission)
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _template_response(self, owner=None):
        selected_owner = owner or self.owner
        return self.client.get(
            "/api/inbound/receive_without_order/import_template/",
            {"owner_id": selected_owner.id},
        )

    def _file_with_rows(self, rows, *, owner=None):
        response = self._template_response(owner=owner)
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook["无订单收货"]
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            "无订单收货.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _preview(self, rows, *, owner=None, template_owner=None):
        selected_owner = owner or self.owner
        upload = self._file_with_rows(rows, owner=template_owner or selected_owner)
        return self.client.post(
            "/api/inbound/receive_without_order/import_preview/",
            {"owner_id": selected_owner.id, "file": upload},
            format="multipart",
        )

    def test_template_contains_owner_metadata_reference_and_expected_columns(self):
        response = self._template_response()

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["填写说明", "无订单收货", "商品单位参考", "_lists", "_meta"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["无订单收货"][1]],
            [
                "商品编号",
                "商品名称",
                "收货数量",
                "收货单位代码",
                "批次号",
                "生产日期",
                "有效截止日期",
            ],
        )
        metadata = dict(
            workbook["_meta"].iter_rows(min_col=1, max_col=2, values_only=True)
        )
        self.assertEqual(metadata["owner_id"], self.owner.id)
        reference_rows = list(
            workbook["商品单位参考"].iter_rows(min_row=2, values_only=True)
        )
        self.assertIn(
            (
                "EXSKU1",
                "Excel Product",
                None,
                "CS",
                "箱",
                12,
                "EA",
                "否",
                "否",
                "MFG",
                "否",
            ),
            reference_rows,
        )

    def test_preview_converts_package_quantity_without_writing_inventory(self):
        response = self._preview([["EXSKU1", "仅供核对的名称", "2", "CS", "", "", ""]])

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["total_rows"], 1)
        self.assertEqual(response.data["product_count"], 1)
        self.assertEqual(response.data["rows"][0]["base_qty"], "24")
        self.assertEqual(response.data["rows"][0]["product_name"], "Excel Product")
        self.assertEqual(response.data["items"][0]["qty"], "24")
        self.assertTrue(response.data["preview_token"])
        self.assertFalse(InventoryDetail.objects.exists())
        self.assertFalse(WmsTask.objects.exists())

    def test_serial_controlled_product_rejects_the_whole_file(self):
        response = self._preview([["SERIAL1", "Serial Product", "1", "EA", "", "", ""]])

        self.assertEqual(response.status_code, 400, response.data)
        self.assertTrue(
            any(
                "序列号管理商品暂不支持" in error["message"]
                for error in response.data["errors"]
            )
        )
        self.assertFalse(WmsTask.objects.exists())

    def test_formula_cells_are_rejected(self):
        response = self._preview(
            [['=CONCAT("EX","SKU1")', "Excel Product", "1", "EA", "", "", ""]]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertTrue(
            any(
                error["message"] == "不允许使用公式"
                for error in response.data["errors"]
            )
        )

    def test_batch_and_expiry_control_require_tracking_fields(self):
        self.product.batch_control = True
        self.product.expiry_control = True
        self.product.expiry_basis = Product.ExpiryBasis.MFG
        self.product.shelf_life_days = 365
        self.product.save(
            update_fields=[
                "batch_control",
                "expiry_control",
                "expiry_basis",
                "shelf_life_days",
                "updated_at",
            ]
        )

        response = self._preview([["EXSKU1", "Excel Product", "1", "EA", "", "", ""]])

        self.assertEqual(response.status_code, 400, response.data)
        messages = {error["message"] for error in response.data["errors"]}
        self.assertTrue(any("批次号不能为空" in message for message in messages))
        self.assertTrue(any("有效截止日期不能为空" in message for message in messages))
        self.assertTrue(any("生产日期不能为空" in message for message in messages))

    def test_non_xlsx_and_missing_template_column_are_rejected(self):
        wrong_type = SimpleUploadedFile("无订单收货.xls", b"not-an-xlsx")
        wrong_type_response = self.client.post(
            "/api/inbound/receive_without_order/import_preview/",
            {"owner_id": self.owner.id, "file": wrong_type},
            format="multipart",
        )
        self.assertEqual(wrong_type_response.status_code, 400)
        self.assertIn(".xlsx", wrong_type_response.data["detail"])

        template = self._template_response()
        workbook = load_workbook(io.BytesIO(template.content))
        workbook["无订单收货"].delete_cols(7)
        output = io.BytesIO()
        workbook.save(output)
        missing_column = SimpleUploadedFile("无订单收货.xlsx", output.getvalue())
        missing_column_response = self.client.post(
            "/api/inbound/receive_without_order/import_preview/",
            {"owner_id": self.owner.id, "file": missing_column},
            format="multipart",
        )
        self.assertEqual(missing_column_response.status_code, 400)
        self.assertIn("模板缺少必要列", missing_column_response.data["detail"])

    def test_template_for_another_owner_is_rejected(self):
        response = self._preview(
            [["EXSKU1", "Excel Product", "1", "EA", "", "", ""]],
            owner=self.owner,
            template_owner=self.other_owner,
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("模板所属货主", response.data["detail"])

    def test_tampered_preview_items_are_rejected(self):
        preview = self._preview([["EXSKU1", "错误名称也应忽略", "2", "CS", "", "", ""]])
        items = [dict(item) for item in preview.data["items"]]
        items[0]["qty"] = "25"

        response = self.client.post(
            "/api/inbound/receive_without_order/import_confirm/",
            {
                "preview_token": preview.data["preview_token"],
                "request_id": preview.data["request_id"],
                "items": items,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(WmsTask.objects.exists())
        self.assertFalse(InventoryDetail.objects.exists())

    def test_confirm_posts_inventory_and_retry_is_idempotent(self):
        preview = self._preview([["EXSKU1", "Excel Product", "2", "CS", "", "", ""]])
        payload = {
            "preview_token": preview.data["preview_token"],
            "request_id": preview.data["request_id"],
            "items": preview.data["items"],
        }

        with override_settings(TASKING_DEFAULT_RECEIVE_LOCATION_ID=self.location.id):
            created = self.client.post(
                "/api/inbound/receive_without_order/import_confirm/",
                payload,
                format="json",
            )
            replayed = self.client.post(
                "/api/inbound/receive_without_order/import_confirm/",
                payload,
                format="json",
            )

        self.assertEqual(created.status_code, 201, created.data)
        self.assertEqual(replayed.status_code, 200, replayed.data)
        self.assertTrue(replayed.data["idempotent"])
        self.assertEqual(replayed.data["task_id"], created.data["task_id"])
        detail = InventoryDetail.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            location=self.location,
            product=self.product,
        )
        self.assertEqual(detail.onhand_qty, Decimal("24.0000"))
        self.assertEqual(
            WmsTask.objects.filter(task_type=WmsTask.TaskType.RECEIVE).count(), 1
        )
