# allapp/inbound/export_views.py
from django.http import HttpResponse
from django.db.models import Prefetch
from django.shortcuts import get_object_or_404
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

from allapp.inbound.permissions import can_view_receive_tasks, scoped_receive_tasks
from allapp.tasking.models import TaskScanLog, WmsTask, WmsTaskLine


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_receive_task_excel(request, task_id):
    if not can_view_receive_tasks(request.user):
        raise PermissionDenied("没有查看收货任务的权限")

    # 1) 拿到当前入库任务 + 明细
    task = get_object_or_404(
        scoped_receive_tasks(
            request.user,
            WmsTask.objects.select_related("owner", "warehouse", "created_by"),
        ),
        pk=task_id,
        task_type=WmsTask.TaskType.RECEIVE,
    )

    export_scans = (
        TaskScanLog.objects.filter(status=TaskScanLog.ScanStatus.OK)
        .exclude(review_status=TaskScanLog.ReviewStatus.REJECTED)
        .select_related("location", "product")
        .order_by("id")
    )
    lines = (
        WmsTaskLine.objects.filter(task_id=task.id)
        .select_related("product", "product__base_uom", "to_location")
        .prefetch_related(
            Prefetch("scan_logs", queryset=export_scans, to_attr="export_scans")
        )
        .order_by("id")
    )

    # 2) 创建工作簿/工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "入库单"

    # 一些样式
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(size=14, bold=True)
    header_font = Font(bold=True)

    # 列宽
    ws.column_dimensions["A"].width = 6  # 行号
    ws.column_dimensions["B"].width = 18  # SKU
    ws.column_dimensions["C"].width = 30  # 品名
    ws.column_dimensions["D"].width = 18  # 规格
    ws.column_dimensions["E"].width = 8  # 单位
    ws.column_dimensions["F"].width = 14  # 批号
    ws.column_dimensions["G"].width = 12  # 生产日期
    ws.column_dimensions["H"].width = 12  # 效期
    ws.column_dimensions["I"].width = 12  # 库位
    ws.column_dimensions["J"].width = 12  # 计划数量
    ws.column_dimensions["K"].width = 14  # 实收数量

    row = 1

    # 3) 标题行：入库单 + 单号
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    cell = ws.cell(row=row, column=1, value=f"入 库 单（{task.task_no}）")
    cell.alignment = center
    cell.font = title_font
    row += 2  # 空一行

    # 4) 任务头信息
    ws.cell(row=row, column=1, value="货主：").alignment = left
    ws.cell(
        row=row, column=2, value=task.owner.name if task.owner else ""
    ).alignment = left

    ws.cell(row=row, column=4, value="仓库：").alignment = left
    ws.cell(
        row=row, column=5, value=task.warehouse.name if task.warehouse else ""
    ).alignment = left

    ws.cell(row=row, column=7, value="日期：").alignment = left
    ws.cell(
        row=row, column=8, value=task.created_at.date().strftime("%Y-%m-%d")
    ).alignment = left
    row += 1

    ws.cell(row=row, column=1, value="类型：").alignment = left
    ws.cell(row=row, column=2, value=task.get_task_type_display()).alignment = left

    ws.cell(row=row, column=4, value="制单人：").alignment = left
    ws.cell(
        row=row, column=5, value=(task.created_by.username if task.created_by else "")
    ).alignment = left

    ws.cell(row=row, column=7, value="备注：").alignment = left
    ws.cell(row=row, column=8, value=task.posting_note or "").alignment = left
    row += 2  # 再空一行

    # 5) 表头
    headers = [
        "行号",
        "仓库SKU编码",
        "品名",
        "规格",
        "单位",
        "批号",
        "生产日期",
        "效期",
        "库位",
        "计划数量",
        "实际收货数量",
    ]
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=title)
        c.font = header_font
        c.alignment = center
        c.border = border
    row += 1

    # 6) 明细行
    display_index = 0
    for line in lines:
        scans = line.export_scans
        facts = scans or [None]
        for fact_index, scan in enumerate(facts):
            display_index += 1
            product = (
                scan.product if scan and scan.product_id else None
            ) or line.product
            uom = getattr(product, "base_uom", None)
            actual_qty = (
                scan.qty_base_delta
                if scan and scan.qty_base_delta is not None
                else scan.qty_base if scan else line.qty_done
            )
            location = scan.location if scan else line.to_location
            values = [
                display_index,
                getattr(product, "sku", "") or getattr(product, "code", ""),
                getattr(product, "name", ""),
                getattr(product, "spec", ""),
                (getattr(uom, "name", "") or getattr(uom, "code", "")),
                scan.lot_no if scan else "",
                scan.mfg_date.strftime("%Y-%m-%d") if scan and scan.mfg_date else "",
                scan.exp_date.strftime("%Y-%m-%d") if scan and scan.exp_date else "",
                (getattr(location, "code", "") or getattr(location, "name", "")),
                float(line.qty_plan) if fact_index == 0 else None,
                float(actual_qty or 0),
            ]
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=column, value=value)
                cell.alignment = (
                    right
                    if column in (10, 11)
                    else center if column in (1, 5, 7, 8) else left
                )
                cell.border = border
            row += 1

    # 7) 输出为 Excel 响应
    filename = f"receive_{task.task_no}.xlsx"
    resp = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp
