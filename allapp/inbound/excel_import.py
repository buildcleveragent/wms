import hashlib
import io
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core import signing
from django.db.models import Prefetch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from allapp.products.models import Product, ProductPackage

from .services import no_order_items_hash

TEMPLATE_VERSION = "1"
IMPORT_SHEET_NAME = "无订单收货"
REFERENCE_SHEET_NAME = "商品单位参考"
MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
PREVIEW_TOKEN_MAX_AGE = 30 * 60
PREVIEW_TOKEN_SALT = "inbound.no-order-excel-preview.v1"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADERS = (
    "货主商品编码",
    "商品名称",
    "收货数量",
    "收货单位代码",
    "批次号",
    "生产日期",
    "有效截止日期",
)
REQUIRED_VALUE_HEADERS = frozenset({"货主商品编码", "收货数量", "收货单位代码"})
IGNORED_VALUE_HEADERS = frozenset({"商品名称"})
REQUIRED_TEMPLATE_HEADERS = frozenset(HEADERS) - IGNORED_VALUE_HEADERS


class InboundExcelFileError(Exception):
    pass


def _active_products(owner_id):
    package_qs = ProductPackage.objects.filter(is_active=True).select_related("uom")
    return list(
        Product.objects.filter(owner_id=owner_id, is_active=True)
        .select_related("base_uom")
        .prefetch_related(Prefetch("packages", queryset=package_qs))
        .order_by("code")
    )


def build_no_order_receive_template(owner):
    products = _active_products(owner.id)
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "填写说明"
    data_sheet = workbook.create_sheet(IMPORT_SHEET_NAME)
    references = workbook.create_sheet(REFERENCE_SHEET_NAME)
    lists = workbook.create_sheet("_lists")
    metadata = workbook.create_sheet("_meta")

    _write_instructions(instructions, owner)
    _write_data_sheet(data_sheet)
    product_codes, uom_codes = _write_reference_sheet(references, products)
    _write_list_sheet(lists, product_codes, uom_codes)
    _write_metadata(metadata, owner)
    _add_validations(workbook, data_sheet, product_codes, uom_codes)

    lists.sheet_state = "hidden"
    metadata.sheet_state = "hidden"
    workbook.active = workbook.sheetnames.index(IMPORT_SHEET_NAME)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_instructions(sheet, owner):
    rows = (
        ("无订单收货 Excel 模板", f"模板版本：{TEMPLATE_VERSION}"),
        ("适用货主", f"{owner.code} - {owner.name}"),
        ("使用步骤", "填写“无订单收货”工作表，在 wmspda 上传校验，确认后一次性入库。"),
        ("货主商品编码", "必须填写“商品单位参考”中的货主商品编码，只匹配当前货主的已有商品。"),
        (
            "商品名称",
            "仅供人工核对，可留空；系统导入时忽略此列，仅按货主商品编码识别商品。",
        ),
        (
            "数量与单位",
            "收货数量必须大于 0；单位代码必须是该商品的基本单位或包装单位。",
        ),
        ("批次规则", "启用批次管理的商品必须填写批次号；未启用时请留空。"),
        (
            "效期规则",
            "启用效期管理的商品必须填写有效截止日期；生产日期基准商品还必须填写生产日期。",
        ),
        ("序列号商品", "首版不支持序列号管理商品，请改用人工收货。"),
        ("整批规则", "任意一行有错误时整批不能确认，修正后重新上传。"),
        ("注意", "编码和批次号请按文本填写以保留前导零；不要使用公式。"),
    )
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="2563EB")
    sheet["B1"].fill = PatternFill("solid", fgColor="2563EB")
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 88
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_data_sheet(sheet):
    sheet.append(list(HEADERS))
    for index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill(
            "solid",
            fgColor="DC2626" if header in REQUIRED_VALUE_HEADERS else "1D4ED8",
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = (
            30 if header == "商品名称" else 18
        )
    for row_number in range(2, MAX_IMPORT_ROWS + 2):
        for header in ("货主商品编码", "商品名称", "收货单位代码", "批次号"):
            column = HEADERS.index(header) + 1
            sheet.cell(row=row_number, column=column).number_format = "@"
        for header in ("生产日期", "有效截止日期"):
            column = HEADERS.index(header) + 1
            sheet.cell(row=row_number, column=column).number_format = "yyyy-mm-dd"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    sheet.row_dimensions[1].height = 28


def _write_reference_sheet(sheet, products):
    headers = (
        "货主商品编码",
        "商品名称",
        "规格",
        "收货单位代码",
        "单位名称",
        "换算基本数量",
        "基本单位代码",
        "批次管理",
        "效期管理",
        "效期基准",
        "序列号管理",
    )
    sheet.append(headers)
    product_codes = []
    uom_codes = set()
    row_number = 2
    for product in products:
        product_codes.append(product.code)
        unit_rows = [(product.base_uom, Decimal("1"))]
        unit_rows.extend(
            (package.uom, Decimal(package.qty_in_base))
            for package in product.packages.all()
        )
        seen_units = set()
        for uom, multiplier in unit_rows:
            code = (uom.code or "").strip().upper()
            if not code or code in seen_units or not uom.is_active:
                continue
            seen_units.add(code)
            uom_codes.add(code)
            sheet.append(
                [
                    product.code,
                    product.name,
                    product.spec or "",
                    code,
                    uom.name,
                    multiplier,
                    product.base_uom.code,
                    "是" if product.batch_control else "否",
                    "是" if product.expiry_control else "否",
                    product.expiry_basis or "",
                    "是" if product.serial_control else "否",
                ]
            )
            sheet.cell(row=row_number, column=1).number_format = "@"
            sheet.cell(row=row_number, column=4).number_format = "@"
            row_number += 1
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="475569")
        sheet.column_dimensions[get_column_letter(column)].width = 20
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{max(1, row_number - 1)}"
    )
    return product_codes, sorted(uom_codes)


def _write_list_sheet(sheet, product_codes, uom_codes):
    sheet.append(["货主商品编码", "单位代码"])
    for index in range(max(len(product_codes), len(uom_codes))):
        sheet.append(
            [
                product_codes[index] if index < len(product_codes) else None,
                uom_codes[index] if index < len(uom_codes) else None,
            ]
        )


def _write_metadata(sheet, owner):
    sheet.append(["schema", "no_order_receive_import"])
    sheet.append(["version", TEMPLATE_VERSION])
    sheet.append(["owner_id", owner.id])
    sheet.append(["owner_code", owner.code])
    sheet.append(["max_rows", MAX_IMPORT_ROWS])


def _add_validations(workbook, data_sheet, product_codes, uom_codes):
    def add_named_range(name, column, count):
        workbook.defined_names.add(
            DefinedName(name, attr_text=f"'_lists'!${column}$2:${column}${count + 1}")
        )

    def add_dropdown(header, formula):
        column = get_column_letter(HEADERS.index(header) + 1)
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = "请选择模板下拉列表中的有效值。"
        validation.errorTitle = "无效值"
        validation.showErrorMessage = True
        data_sheet.add_data_validation(validation)
        validation.add(f"{column}2:{column}{MAX_IMPORT_ROWS + 1}")

    if product_codes:
        add_named_range("InboundProductCodes", "A", len(product_codes))
        add_dropdown("货主商品编码", "InboundProductCodes")
    if uom_codes:
        add_named_range("InboundUomCodes", "B", len(uom_codes))
        add_dropdown("收货单位代码", "InboundUomCodes")


def _read_upload(uploaded_file):
    filename = (getattr(uploaded_file, "name", "") or "").lower()
    if not filename.endswith(".xlsx"):
        raise InboundExcelFileError("仅支持 .xlsx 文件。")
    size = getattr(uploaded_file, "size", None)
    if size is not None and size > MAX_IMPORT_FILE_SIZE:
        raise InboundExcelFileError("Excel 文件不能超过 5 MB。")
    data = uploaded_file.read(MAX_IMPORT_FILE_SIZE + 1)
    if len(data) > MAX_IMPORT_FILE_SIZE:
        raise InboundExcelFileError("Excel 文件不能超过 5 MB。")
    try:
        return data, load_workbook(io.BytesIO(data), data_only=False, read_only=False)
    except Exception as exc:
        raise InboundExcelFileError(
            "Excel 文件无法解析，请重新下载标准模板填写。"
        ) from exc


def _metadata(workbook):
    if "_meta" not in workbook.sheetnames:
        raise InboundExcelFileError("文件不是系统生成的无订单收货模板。")
    values = {}
    for key, value in workbook["_meta"].iter_rows(
        min_col=1, max_col=2, values_only=True
    ):
        if key:
            values[str(key)] = value
    if (
        values.get("schema") != "no_order_receive_import"
        or str(values.get("version")) != TEMPLATE_VERSION
    ):
        raise InboundExcelFileError("模板版本不匹配，请重新下载最新模板。")
    return values


def _text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_date(value, field):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for pattern in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    raise ValueError(f"{field}必须是 YYYY-MM-DD 格式")


def _row_error(errors, row, field, message):
    errors.append({"row": row, "field": field, "message": message})


def _fits_decimal_field(value, *, max_digits=18, decimal_places=4):
    normalized = value.normalize()
    exponent = normalized.as_tuple().exponent
    stored_decimal_places = max(-exponent, 0)
    whole_digits = max(len(normalized.as_tuple().digits) - stored_decimal_places, 0)
    return (
        stored_decimal_places <= decimal_places
        and whole_digits + stored_decimal_places <= max_digits
    )


def parse_no_order_receive_excel(uploaded_file, *, owner):
    data, workbook = _read_upload(uploaded_file)
    metadata = _metadata(workbook)
    try:
        template_owner_id = int(metadata.get("owner_id"))
    except (TypeError, ValueError) as exc:
        raise InboundExcelFileError("模板货主信息无效，请重新下载模板。") from exc
    if (
        template_owner_id != owner.id
        or _text(metadata.get("owner_code")).upper() != owner.code.upper()
    ):
        raise InboundExcelFileError("模板所属货主与当前选择货主不一致。")
    if IMPORT_SHEET_NAME not in workbook.sheetnames:
        raise InboundExcelFileError(f"Excel 中缺少“{IMPORT_SHEET_NAME}”工作表。")

    sheet = workbook[IMPORT_SHEET_NAME]
    raw_headers = [_text(cell.value) for cell in sheet[1]]
    nonempty_headers = [header for header in raw_headers if header]
    duplicate_headers = sorted(
        {header for header in nonempty_headers if nonempty_headers.count(header) > 1}
    )
    if duplicate_headers:
        raise InboundExcelFileError(f"模板存在重复列：{duplicate_headers}")
    missing_headers = sorted(REQUIRED_TEMPLATE_HEADERS - set(nonempty_headers))
    if missing_headers:
        raise InboundExcelFileError(f"模板缺少必要列：{missing_headers}")
    unknown_headers = [header for header in nonempty_headers if header not in HEADERS]
    if unknown_headers:
        raise InboundExcelFileError(f"模板存在未知列：{unknown_headers}")
    header_columns = {
        header: raw_headers.index(header) + 1 for header in nonempty_headers
    }

    products = _active_products(owner.id)
    product_map = {product.code.strip().upper(): product for product in products}
    rows = []
    normalized_items = []
    errors = []
    nonempty_count = 0

    for excel_row in range(2, sheet.max_row + 1):
        cells = {
            header: sheet.cell(excel_row, column)
            for header, column in header_columns.items()
        }
        import_cells = {
            header: cell
            for header, cell in cells.items()
            if header not in IGNORED_VALUE_HEADERS
        }
        if all(_text(cell.value) == "" for cell in import_cells.values()):
            continue
        nonempty_count += 1
        if nonempty_count > MAX_IMPORT_ROWS:
            raise InboundExcelFileError(f"一次最多导入 {MAX_IMPORT_ROWS} 条明细。")
        for header, cell in import_cells.items():
            if cell.data_type == "f":
                _row_error(errors, excel_row, header, "不允许使用公式")

        product_code = _text(cells["货主商品编码"].value).upper()
        uom_code = _text(cells["收货单位代码"].value).upper()
        lot_no = _text(cells.get("批次号").value if cells.get("批次号") else "").upper()
        product = product_map.get(product_code)
        if not product_code:
            _row_error(errors, excel_row, "货主商品编码", "不能为空")
        elif product is None:
            _row_error(errors, excel_row, "货主商品编码", "当前货主下不存在或商品已停用")

        qty = None
        raw_qty = cells["收货数量"].value
        try:
            if isinstance(raw_qty, bool) or raw_qty in (None, ""):
                raise InvalidOperation
            qty = Decimal(str(raw_qty))
            if not qty.is_finite() or qty <= 0:
                raise InvalidOperation
        except (InvalidOperation, ValueError):
            _row_error(errors, excel_row, "收货数量", "必须是大于 0 的数字")
            qty = None

        if not uom_code:
            _row_error(errors, excel_row, "收货单位代码", "不能为空")

        mfg_date = None
        exp_date = None
        try:
            mfg_date = _parse_date(
                cells.get("生产日期").value if cells.get("生产日期") else None,
                "生产日期",
            )
        except ValueError as exc:
            _row_error(errors, excel_row, "生产日期", str(exc))
        try:
            exp_date = _parse_date(
                cells.get("有效截止日期").value if cells.get("有效截止日期") else None,
                "有效截止日期",
            )
        except ValueError as exc:
            _row_error(errors, excel_row, "有效截止日期", str(exc))

        multiplier = None
        base_qty = None
        if product is not None:
            if product.serial_control:
                _row_error(
                    errors,
                    excel_row,
                    "货主商品编码",
                    "序列号管理商品暂不支持 Excel 入库，请改用人工收货",
                )
            unit_map = {product.base_uom.code.strip().upper(): Decimal("1")}
            for package in product.packages.all():
                if package.is_active and package.uom.is_active:
                    unit_map[package.uom.code.strip().upper()] = Decimal(
                        package.qty_in_base
                    )
            multiplier = unit_map.get(uom_code)
            if uom_code and multiplier is None:
                _row_error(
                    errors,
                    excel_row,
                    "收货单位代码",
                    "不是该商品的有效基本单位或包装单位",
                )
            if qty is not None and multiplier is not None:
                base_qty = qty * multiplier
                if not _fits_decimal_field(base_qty):
                    _row_error(
                        errors,
                        excel_row,
                        "收货数量",
                        "换算后的基本数量必须不超过 18 位且最多 4 位小数",
                    )

            if product.batch_control and not lot_no:
                _row_error(
                    errors, excel_row, "批次号", "该商品启用批次管理，批次号不能为空"
                )
            if not product.batch_control and lot_no:
                _row_error(errors, excel_row, "批次号", "该商品未启用批次管理，请留空")
            if product.expiry_control:
                if not exp_date:
                    _row_error(
                        errors,
                        excel_row,
                        "有效截止日期",
                        "该商品启用效期管理，有效截止日期不能为空",
                    )
                if product.expiry_basis == Product.ExpiryBasis.MFG and not mfg_date:
                    _row_error(
                        errors,
                        excel_row,
                        "生产日期",
                        "该商品按生产日期管理效期，生产日期不能为空",
                    )
            elif mfg_date or exp_date:
                _row_error(
                    errors,
                    excel_row,
                    "有效截止日期",
                    "该商品未启用效期管理，生产日期和有效截止日期必须留空",
                )
        if mfg_date and exp_date and exp_date < mfg_date:
            _row_error(errors, excel_row, "有效截止日期", "不得早于生产日期")

        if (
            product is not None
            and qty is not None
            and multiplier is not None
            and base_qty is not None
        ):
            row = {
                "row": excel_row,
                "product_id": product.id,
                "product_code": product.code,
                "product_name": product.name,
                "input_qty": format(qty, "f"),
                "uom_code": uom_code,
                "multiplier": format(multiplier, "f"),
                "base_qty": format(base_qty, "f"),
                "base_uom_code": product.base_uom.code,
                "lot_no": lot_no,
                "mfg_date": mfg_date.isoformat() if mfg_date else None,
                "exp_date": exp_date.isoformat() if exp_date else None,
            }
            rows.append(row)
            normalized_items.append(
                {
                    "product_id": product.id,
                    "qty": format(base_qty, "f"),
                    "lot_no": lot_no,
                    "mfg_date": row["mfg_date"],
                    "exp_date": row["exp_date"],
                }
            )

    if nonempty_count == 0:
        raise InboundExcelFileError("“无订单收货”工作表没有可导入的明细。")
    return {
        "file_sha256": hashlib.sha256(data).hexdigest(),
        "total_rows": nonempty_count,
        "product_count": len({row["product_id"] for row in rows}),
        "rows": rows,
        "normalized_items": normalized_items,
        "errors": errors,
        "error_count": len(errors),
    }


def create_preview_credentials(*, user_id, owner_id, warehouse_id, items, file_sha256):
    request_id = f"excel-{uuid.uuid4().hex}"
    payload = {
        "user_id": int(user_id),
        "owner_id": int(owner_id),
        "warehouse_id": int(warehouse_id),
        "request_id": request_id,
        "items_hash": no_order_items_hash(items),
        "file_sha256": file_sha256,
    }
    return request_id, signing.dumps(payload, salt=PREVIEW_TOKEN_SALT, compress=True)


def load_preview_credentials(token):
    try:
        return signing.loads(
            token,
            salt=PREVIEW_TOKEN_SALT,
            max_age=PREVIEW_TOKEN_MAX_AGE,
        )
    except signing.SignatureExpired as exc:
        raise InboundExcelFileError(
            "预览已超过 30 分钟，请重新上传 Excel 校验。"
        ) from exc
    except signing.BadSignature as exc:
        raise InboundExcelFileError(
            "预览凭证无效或已被篡改，请重新上传 Excel 校验。"
        ) from exc
