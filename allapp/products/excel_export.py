from __future__ import annotations

import io
from dataclasses import dataclass

from django.db.models import Q
from rest_framework.exceptions import PermissionDenied

from allapp.accounts.access import AccessScope
from allapp.baseinfo.models import Owner, OwnerWarehouseBinding

from .excel_import import (
    IMPORT_SHEET_NAME,
    PACKAGE_HEADERS,
    PACKAGE_SHEET_NAME,
    PRODUCT_HEADERS,
    _add_template_validations,
    _write_import_sheet,
    _write_instruction_sheet,
    _write_meta_sheet,
    _write_package_sheet,
    _write_reference_sheet,
)
from .models import Brand, Product, ProductCategory, ProductUom
from .permissions import can_view_all_owner_products


@dataclass(frozen=True)
class ProductExportAccess:
    allowed_owner_ids: frozenset[int] | None


def resolve_product_export_access(user) -> ProductExportAccess:
    if not user or not getattr(user, "is_authenticated", False):
        raise PermissionDenied("请先登录。")
    if not (
        getattr(user, "is_superuser", False) or user.has_perm("products.view_product")
    ):
        raise PermissionDenied("当前账号没有商品查看权限。")
    if can_view_all_owner_products(user):
        return ProductExportAccess(allowed_owner_ids=None)

    scope = AccessScope.for_user(user)
    if not scope.is_valid:
        raise PermissionDenied("当前账号没有有效的角色数据范围。")
    owner_ids = set(scope.owner_ids)
    if scope.warehouse_ids:
        owner_ids.update(
            OwnerWarehouseBinding.objects.filter(
                warehouse_id__in=scope.warehouse_ids,
                is_active=True,
                owner__is_active=True,
            ).values_list("owner_id", flat=True)
        )
    if not owner_ids:
        raise PermissionDenied("当前账号没有可导出的货主范围。")
    return ProductExportAccess(allowed_owner_ids=frozenset(owner_ids))


def can_export_products(user) -> bool:
    try:
        resolve_product_export_access(user)
    except PermissionDenied:
        return False
    return True


def export_owner_queryset(user, *, search=""):
    access = resolve_product_export_access(user)
    queryset = Owner.objects.filter(is_active=True).order_by("code", "id")
    if access.allowed_owner_ids is not None:
        queryset = queryset.filter(pk__in=access.allowed_owner_ids)
    search = (search or "").strip()
    if search:
        queryset = queryset.filter(
            Q(code__icontains=search) | Q(name__icontains=search)
        )
    return queryset


def resolve_export_owner(user, owner_id) -> Owner:
    try:
        owner_id = int(owner_id)
    except (TypeError, ValueError) as exc:
        raise PermissionDenied("请选择有效货主。") from exc
    owner = export_owner_queryset(user).filter(pk=owner_id).first()
    if owner is None:
        raise PermissionDenied("当前账号无权导出该货主的商品。")
    return owner


def build_product_export_workbook(owner: Owner) -> tuple[bytes, int, int]:
    from openpyxl import Workbook

    owners = [(owner.code, owner.name)]
    uoms = list(
        ProductUom.objects.filter(is_active=True)
        .order_by("code")
        .values_list("code", "name", "kind", "decimal_places")
    )
    categories = [
        (category.code, category.full_path)
        for category in ProductCategory.objects.filter(is_active=True)
        .select_related("parent__parent")
        .order_by("code")
        if category.has_active_path()
    ]
    brands = list(
        Brand.objects.filter(is_active=True)
        .order_by("code")
        .values_list("code", "name")
    )

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "填写说明"
    product_sheet = workbook.create_sheet(IMPORT_SHEET_NAME)
    package_sheet = workbook.create_sheet(PACKAGE_SHEET_NAME)
    refs = workbook.create_sheet("基础资料")
    meta = workbook.create_sheet("_meta")
    _write_instruction_sheet(instructions)
    _write_import_sheet(product_sheet)
    _write_package_sheet(package_sheet)
    _write_reference_sheet(refs, owners, uoms, categories, brands)
    _write_meta_sheet(meta)
    _add_template_validations(
        workbook, product_sheet, package_sheet, owners, uoms, categories, brands
    )

    products = list(
        Product.objects.filter(owner=owner)
        .select_related("owner", "category", "brand", "base_uom", "carton_package__uom")
        .prefetch_related("packages__uom")
        .order_by("code", "id")
    )
    package_count = 0
    package_row_number = 2
    for product_row_number, product in enumerate(products, start=2):
        _write_safe_row(
            product_sheet,
            product_row_number,
            PRODUCT_HEADERS,
            _product_values(product),
        )
        for package in product.packages.all().order_by("sort_order", "uom__code", "id"):
            _write_safe_row(
                package_sheet,
                package_row_number,
                PACKAGE_HEADERS,
                _package_values(product, package),
            )
            package_row_number += 1
            package_count += 1

    if len(products) > 1000:
        instructions.append(
            [
                "回导限制",
                f"本档案包含 {len(products)} 条商品；系统单次最多导入 1000 条，请拆分后回导。",
            ]
        )
    meta.sheet_state = "hidden"
    workbook.active = workbook.sheetnames.index(IMPORT_SHEET_NAME)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue(), len(products), package_count


def _write_safe_row(sheet, row_number, headers, values):
    for column, header in enumerate(headers, start=1):
        value = values.get(header)
        cell = sheet.cell(row=row_number, column=column, value=value)
        if isinstance(value, str):
            cell.data_type = "s"
        if header in {
            "货主编码",
            "货主商品编码",
            "分类",
            "品牌编码",
            "基本单位",
            "标准贸易条码",
            "零码",
            "箱码",
            "箱码对应包装单位编码",
            "外部系统商品编码",
            "包装单位编码",
            "包装条码",
        }:
            cell.number_format = "@"


def _yes_no(value):
    if value is None:
        return None
    return "是" if value else "否"


def _product_values(product):
    return {
        "货主编码": product.owner.code,
        "货主商品编码": product.code,
        "标准贸易条码": product.gtin,
        "商品名称": product.name,
        "分类": product.category.full_path if product.category_id else None,
        "基本单位": product.base_uom.name,
        "基本单位类型": product.base_uom.get_kind_display(),
        "单位小数位数": product.base_uom.decimal_places,
        "基本单位数量": None,
        "规格": product.spec,
        "品牌编码": product.brand.code if product.brand_id else None,
        "零码": product.unit_barcode,
        "箱码": product.carton_barcode,
        "箱码对应包装单位编码": (
            product.carton_package.uom.code if product.carton_package_id else None
        ),
        "外部系统商品编码": product.external_code,
        "默认价格": product.price,
        "最低价格": product.min_price,
        "最高折扣%": product.max_discount,
        "重量kg": product.weight,
        "体积m³": product.volume,
        "净含量g": product.net_content,
        "厂家": product.vender,
        "材质": product.material_quality,
        "描述": product.description,
        "最低库存": product.min_stock,
        "最高库存": product.max_stock,
        "序列号管理": _yes_no(product.serial_control),
        "批次管理": _yes_no(product.batch_control),
        "启用": _yes_no(product.is_active),
        "保质期管理": _yes_no(product.expiry_control),
        "效期基准": product.expiry_basis,
        "保质期天数": product.shelf_life_days,
        "入库有效天数": product.inbound_valid_days,
        "效期预警天数": product.expiry_warning_days,
        "FEFO": _yes_no(product.fefo_required),
    }


def _package_values(product, package):
    return {
        "货主编码": product.owner.code,
        "货主商品编码": product.code,
        "包装单位编码": package.uom.code,
        "包装换算数量": package.qty_in_base,
        "包装条码": package.barcode,
        "长cm": package.length_cm,
        "宽cm": package.width_cm,
        "高cm": package.height_cm,
        "毛重kg": package.gross_weight_kg,
        "体积m³": package.volume_m3,
        "体积自动计算": _yes_no(package.volume_auto),
        "可直接拣配": _yes_no(package.is_pickable),
        "采购默认": _yes_no(package.is_purchase_default),
        "销售默认": _yes_no(package.is_sales_default),
        "启用": _yes_no(package.is_active),
        "排序": package.sort_order,
    }
