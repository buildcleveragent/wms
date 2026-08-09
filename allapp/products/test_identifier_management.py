import io
from datetime import timedelta
from unittest import mock

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from allapp.accounts.models import UserRoleScope
from allapp.baseinfo.models import Owner

from .identifier_excel import (
    BARCODE_HEADERS,
    BARCODE_SHEET,
    EXTERNAL_HEADERS,
    EXTERNAL_SHEET,
    build_identifier_export,
    build_identifier_template,
)
from .identifier_services import IdentifierConcurrencyError, add_product_barcode
from .models import (
    Product,
    ProductBarcode,
    ProductExternalIdentifier,
    ProductIdentifierRegistry,
    ProductPackage,
    ProductUom,
)


class IdentifierManagementApiTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = Owner.objects.create(code="IDM", name="标识维护货主")
        cls.other_owner = Owner.objects.create(code="IDM2", name="其他标识货主")
        cls.each = ProductUom.objects.create(code="IDM-EA", name="件")
        cls.carton = ProductUom.objects.create(code="IDM-CTN", name="箱")
        cls.product = Product.objects.create(
            owner=cls.owner,
            code="IDM-P1",
            name="标识维护商品",
            base_uom=cls.each,
            expiry_control=False,
            expiry_basis=None,
        )
        cls.package = ProductPackage.objects.create(
            product=cls.product,
            uom=cls.carton,
            qty_in_base=24,
        )
        cls.other_product = Product.objects.create(
            owner=cls.other_owner,
            code="IDM-P2",
            name="其他商品",
            base_uom=cls.each,
            expiry_control=False,
            expiry_basis=None,
        )
        cls.user = get_user_model().objects.create_user(
            username="identifier-manager", password="x"
        )
        UserRoleScope.objects.create(
            user=cls.user,
            role=UserRoleScope.Role.OWNER_MANAGER,
            owner=cls.owner,
        )
        permissions = []
        for model in (Product, ProductBarcode, ProductExternalIdentifier):
            content_type = ContentType.objects.get_for_model(model)
            permissions.extend(
                Permission.objects.filter(
                    content_type=content_type,
                    codename__in=[
                        f"add_{model._meta.model_name}",
                        f"view_{model._meta.model_name}",
                        f"change_{model._meta.model_name}",
                        f"delete_{model._meta.model_name}",
                    ],
                )
            )
        cls.user.user_permissions.add(*permissions)

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def test_barcode_api_lifecycle_and_identity_fields(self):
        created = self.client.post(
            "/api/product-barcodes/",
            {
                "product": self.product.pk,
                "barcode": " idm-other-1 ",
                "barcode_type": "OTHER",
                "is_primary": True,
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201, created.data)
        record_id = created.data["id"]
        self.assertEqual(created.data["normalized_value"], "IDM-OTHER-1")

        immutable = self.client.patch(
            f"/api/product-barcodes/{record_id}/",
            {"barcode": "CHANGED"},
            format="json",
        )
        self.assertEqual(immutable.status_code, 400, immutable.data)

        retired = self.client.delete(f"/api/product-barcodes/{record_id}/")
        self.assertEqual(retired.status_code, 204)
        record = ProductBarcode.all_objects.get(pk=record_id)
        self.assertFalse(record.is_active)
        self.assertTrue(
            ProductIdentifierRegistry.objects.filter(
                owner=self.owner, normalized_value="IDM-OTHER-1"
            ).exists()
        )

        reactivated = self.client.post(f"/api/product-barcodes/{record_id}/reactivate/")
        self.assertEqual(reactivated.status_code, 200, reactivated.data)
        record.refresh_from_db()
        self.assertTrue(record.is_active)

    def test_set_primary_rejects_ineffective_records_without_demoting_current(self):
        current = add_product_barcode(
            product=self.product,
            barcode="IDM-CURRENT",
            barcode_type="OTHER",
            is_primary=True,
        )
        future = add_product_barcode(
            product=self.product,
            barcode="IDM-FUTURE",
            barcode_type="OTHER",
            valid_from=timezone.now() + timedelta(days=1),
        )

        response = self.client.post(f"/api/product-barcodes/{future.pk}/set-primary/")

        self.assertEqual(response.status_code, 400, response.data)
        current.refresh_from_db()
        future.refresh_from_db()
        self.assertTrue(current.is_primary)
        self.assertFalse(future.is_primary)

    def test_inactive_package_and_cross_owner_records_are_rejected_or_hidden(self):
        self.package.is_active = False
        self.package.save(update_fields=["is_active"])
        rejected = self.client.post(
            "/api/product-barcodes/",
            {
                "product": self.product.pk,
                "barcode": "IDM-INACTIVE-PACKAGE",
                "barcode_type": "PACKAGE",
                "package": self.package.pk,
            },
            format="json",
        )
        self.assertEqual(rejected.status_code, 400, rejected.data)

        foreign = add_product_barcode(
            product=self.other_product,
            barcode="IDM-FOREIGN",
            barcode_type="OTHER",
        )
        listing = self.client.get("/api/product-barcodes/")
        self.assertEqual(listing.status_code, 200)
        items = (
            listing.data["results"] if isinstance(listing.data, dict) else listing.data
        )
        ids = {item["id"] for item in items}
        self.assertNotIn(foreign.pk, ids)
        self.assertEqual(
            self.client.get(f"/api/product-barcodes/{foreign.pk}/").status_code,
            404,
        )

    @mock.patch(
        "allapp.products.serializers.add_product_barcode",
        side_effect=IdentifierConcurrencyError("并发占用"),
    )
    def test_concurrent_create_is_reported_as_409(self, _mocked_add):
        response = self.client.post(
            "/api/product-barcodes/",
            {
                "product": self.product.pk,
                "barcode": "IDM-CONCURRENT",
                "barcode_type": "OTHER",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 409, response.data)

    def test_external_identifier_lifecycle(self):
        created = self.client.post(
            "/api/product-external-identifiers/",
            {
                "product": self.product.pk,
                "source_system": "OMS",
                "external_code": " IDM-OMS-1 ",
                "is_primary": True,
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201, created.data)
        record_id = created.data["id"]
        self.assertEqual(created.data["normalized_value"], "IDM-OMS-1")

        retired = self.client.post(
            f"/api/product-external-identifiers/{record_id}/retire/"
        )
        self.assertEqual(retired.status_code, 200, retired.data)
        self.assertFalse(
            ProductExternalIdentifier.all_objects.get(pk=record_id).is_active
        )
        reactivated = self.client.post(
            f"/api/product-external-identifiers/{record_id}/reactivate/"
        )
        self.assertEqual(reactivated.status_code, 200, reactivated.data)


class IdentifierMaintenanceExcelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = Owner.objects.create(code="IDX", name="Excel 标识货主")
        cls.each = ProductUom.objects.create(code="IDX-EA", name="瓶")
        cls.carton = ProductUom.objects.create(code="IDX-CTN", name="箱")
        cls.product = Product.objects.create(
            owner=cls.owner,
            code="IDX-P1",
            name="Excel 标识商品",
            base_uom=cls.each,
            expiry_control=False,
            expiry_basis=None,
        )
        cls.package = ProductPackage.objects.create(
            product=cls.product,
            uom=cls.carton,
            qty_in_base=30,
        )
        cls.user = get_user_model().objects.create_user(
            username="identifier-excel", password="x"
        )
        UserRoleScope.objects.create(
            user=cls.user,
            role=UserRoleScope.Role.OWNER_MANAGER,
            owner=cls.owner,
        )
        cls.user.user_permissions.add(
            Permission.objects.get(
                content_type=ContentType.objects.get_for_model(Product),
                codename="add_product",
            ),
            Permission.objects.get(
                content_type=ContentType.objects.get_for_model(Product),
                codename="view_product",
            ),
        )

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _upload(self, workbook):
        stream = io.BytesIO()
        workbook.save(stream)
        uploaded = SimpleUploadedFile(
            "identifier-maintenance.xlsx",
            stream.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        return self.client.post(
            "/api/products/identifier-maintenance-import/",
            {"owner_id": self.owner.pk, "file": uploaded},
            format="multipart",
        )

    def test_template_import_actions_and_export_include_conversion_snapshot(self):
        workbook = load_workbook(io.BytesIO(build_identifier_template()))
        self.assertEqual(
            [cell.value for cell in workbook[BARCODE_SHEET][1]], BARCODE_HEADERS
        )
        self.assertEqual(
            [cell.value for cell in workbook[EXTERNAL_SHEET][1]], EXTERNAL_HEADERS
        )
        barcode_row = {header: "" for header in BARCODE_HEADERS}
        barcode_row.update(
            {
                "操作": "ADD",
                "货主商品编码": self.product.code,
                "条码": "IDX-BOX-30",
                "条码类型": "PACKAGE",
                "包装单位编码": self.carton.code,
                "是否主码": "是",
                "启用状态": "是",
            }
        )
        workbook[BARCODE_SHEET].append(
            [barcode_row[header] for header in BARCODE_HEADERS]
        )
        external_row = {header: "" for header in EXTERNAL_HEADERS}
        external_row.update(
            {
                "操作": "ADD",
                "货主商品编码": self.product.code,
                "来源系统": "OMS",
                "外部系统商品编码": "IDX-OMS-1",
                "是否主标识": "是",
                "启用状态": "是",
            }
        )
        workbook[EXTERNAL_SHEET].append(
            [external_row[header] for header in EXTERNAL_HEADERS]
        )

        imported = self._upload(workbook)

        self.assertEqual(imported.status_code, 200, imported.data)
        self.assertEqual(imported.data["changed_count"], 2)
        barcode = ProductBarcode.objects.get(normalized_value="IDX-BOX-30")
        self.assertEqual(barcode.qty_in_base, 30)
        self.assertTrue(
            ProductExternalIdentifier.objects.filter(
                normalized_value="IDX-OMS-1"
            ).exists()
        )
        exported = load_workbook(
            io.BytesIO(build_identifier_export(self.owner)), data_only=True
        )
        header_index = {cell.value: cell.column for cell in exported[BARCODE_SHEET][1]}
        barcode_rows = {
            exported[BARCODE_SHEET].cell(row, header_index["条码"]).value: row
            for row in range(2, exported[BARCODE_SHEET].max_row + 1)
        }
        self.assertEqual(
            exported[BARCODE_SHEET]
            .cell(
                barcode_rows["IDX-BOX-30"],
                header_index["基础单位换算快照"],
            )
            .value,
            30,
        )

    def test_workbook_failure_rolls_back_all_prior_rows(self):
        workbook = load_workbook(io.BytesIO(build_identifier_template()))
        for barcode, barcode_type in (
            ("IDX-ROLLBACK", "OTHER"),
            ("IDX-BAD-PACKAGE", "PACKAGE"),
        ):
            row = {header: "" for header in BARCODE_HEADERS}
            row.update(
                {
                    "操作": "ADD",
                    "货主商品编码": self.product.code,
                    "条码": barcode,
                    "条码类型": barcode_type,
                    "启用状态": "是",
                }
            )
            workbook[BARCODE_SHEET].append([row[header] for header in BARCODE_HEADERS])

        response = self._upload(workbook)

        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(
            ProductBarcode.all_objects.filter(normalized_value="IDX-ROLLBACK").exists()
        )
