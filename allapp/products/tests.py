# allapp/products/tests.py
# -*- coding: utf-8 -*-
import io
import json
import os
import tempfile
import unittest
from decimal import Decimal
from unittest.mock import patch

from django.apps import apps
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.db import IntegrityError
from django.test import RequestFactory, TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient, APIRequestFactory, force_authenticate

from allapp.accounts.models import AuditEvent, UserRoleScope

from .category_backfill import (
    CategoryBackfillError,
    build_category_backfill_workbook,
    import_category_backfill,
)
from .excel_import import (
    HEADERS,
    IMPORT_SHEET_NAME,
    MAX_IMPORT_FILE_SIZE,
    PACKAGE_HEADERS,
    PACKAGE_SHEET_NAME,
    PRODUCT_HEADERS,
)
from .identifier_services import add_product_barcode
from .models import ProductBarcode
from .views import ProductViewSet

# 业务模型
Owner = apps.get_model("baseinfo", "Owner")
ProductUom = apps.get_model("products", "ProductUom")
ProductCategory = apps.get_model("products", "ProductCategory")
Product = apps.get_model("products", "Product")
ProductPackage = apps.get_model("products", "ProductPackage")
Warehouse = apps.get_model("locations", "Warehouse")
Location = apps.get_model("locations", "Location")
Subwarehouse = apps.get_model("locations", "Subwarehouse")
InventoryDetail = apps.get_model("inventory", "InventoryDetail")
OwnerWarehouseBinding = apps.get_model("baseinfo", "OwnerWarehouseBinding")

# 可选：DAL 自动补全视图（存在则测试）
try:
    # 如果你把视图放在其他模块，请相应调整
    from .autocomplete import (
        ProductUomAutocomplete,
    )

    DAL_OK = True
except Exception:
    DAL_OK = False

# 依赖是否齐全
DEPENDENCIES_OK = all([Owner, ProductUom, Product])


class ProductCategoryHierarchyTests(TestCase):
    def setUp(self):
        self.root = ProductCategory.objects.create(
            code="FRESH", name="生鲜", sort_order=1
        )
        self.middle = ProductCategory.objects.create(
            code="FRESH-FRUIT", name="水果", parent=self.root
        )
        self.small = ProductCategory.objects.create(
            code="FRESH-BERRY", name="莓果", parent=self.middle
        )

    def test_levels_path_and_descendants_are_derived_from_parent_tree(self):
        self.assertEqual(self.root.depth, 1)
        self.assertEqual(self.middle.level_name, "中类")
        self.assertEqual(self.small.level_name, "小类")
        self.assertEqual(self.small.full_path, "生鲜 > 水果 > 莓果")
        self.assertEqual(
            set(self.root.descendant_ids()),
            {self.root.id, self.middle.id, self.small.id},
        )

    def test_fourth_level_duplicate_sibling_and_cycle_are_rejected(self):
        with self.assertRaises(ValidationError):
            ProductCategory.objects.create(
                code="FRESH-BERRY-RED", name="红莓", parent=self.small
            )
        with self.assertRaises(ValidationError):
            ProductCategory.objects.create(
                code="FRESH-FRUIT-2", name="水果", parent=self.root
            )
        self.root.parent = self.small
        with self.assertRaises(ValidationError):
            self.root.save()

    def test_reparenting_subtree_too_deep_and_inactive_parent_are_rejected(self):
        another_root = ProductCategory.objects.create(code="FOOD", name="食品")
        self.root.parent = another_root
        with self.assertRaises(ValidationError):
            self.root.save()

        with self.assertRaises(ValidationError):
            ProductCategory.objects.create(
                code="DISABLED-CHILD",
                name="停用父类的子类",
                parent=ProductCategory.objects.create(
                    code="DISABLED", name="停用分类", is_active=False
                ),
            )

    def test_new_product_requires_category_but_legacy_blank_can_converge(self):
        owner = Owner.objects.create(code="CAT-OWNER", name="分类货主")
        uom = ProductUom.objects.create(code="CAT-EA", name="个")
        product = Product(
            owner=owner,
            code="CAT-P1",
            name="新商品",
            base_uom=uom,
            expiry_control=False,
        )
        with self.assertRaises(ValidationError):
            product.full_clean()

        legacy = Product.objects.create(
            owner=owner,
            code="CAT-LEGACY",
            name="历史商品",
            base_uom=uom,
            expiry_control=False,
        )
        legacy.full_clean()
        legacy.category = self.root
        legacy.full_clean()
        legacy.save(update_fields=["category", "updated_at"])
        legacy.category = None
        with self.assertRaises(ValidationError):
            legacy.full_clean()


class ProductCategoryBackfillTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="category-backfill-admin",
            password="x",
            email="admin@example.com",
        )
        self.owner = Owner.objects.create(code="BACKFILL", name="补录货主")
        self.uom = ProductUom.objects.create(code="BACKFILL-EA", name="件")
        self.category = ProductCategory.objects.create(
            code="BACKFILL-FOOD", name="食品"
        )
        self.product = Product.objects.create(
            owner=self.owner,
            code="BACKFILL-P1",
            name="待补录商品",
            base_uom=self.uom,
        )

    def _upload(self, workbook):
        output = io.BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            "商品分类补录.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    def test_export_and_atomic_import_only_update_category(self):
        content = build_category_backfill_workbook(
            Product.objects.filter(pk=self.product.pk)
        )
        workbook = load_workbook(io.BytesIO(content))
        workbook["商品分类补录"]["E2"] = self.category.code

        result = import_category_backfill(self._upload(workbook), user=self.user)

        self.product.refresh_from_db()
        self.assertEqual(self.product.category_id, self.category.id)
        self.assertEqual(result, {"row_count": 1, "changed_count": 1})
        self.assertTrue(
            AuditEvent.objects.filter(action="products.category_backfill").exists()
        )

    def test_invalid_row_rolls_back_every_valid_row(self):
        content = build_category_backfill_workbook(
            Product.objects.filter(pk=self.product.pk)
        )
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook["商品分类补录"]
        sheet["E2"] = self.category.code
        sheet.append(
            [self.owner.code, "MISSING", "不存在", "未分类", self.category.code]
        )

        with self.assertRaises(CategoryBackfillError):
            import_category_backfill(self._upload(workbook), user=self.user)

        self.product.refresh_from_db()
        self.assertIsNone(self.product.category_id)


@unittest.skipUnless(
    DEPENDENCIES_OK, "缺少 baseinfo/products 依赖模型，跳过 products 测试"
)
class ProductViewSetTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        # 基础数据：两个货主、一个基础单位
        cls.owner_a = Owner.objects.create(code="OA", name="Owner-A")
        cls.owner_b = Owner.objects.create(code="OB", name="Owner-B")
        cls.uom = ProductUom.objects.create(code="PCS", name="件", is_active=True)
        cls.category = ProductCategory.objects.create(
            code="VIEWSET-CAT", name="接口商品"
        )

        # 用户：user_a 属于 owner_a；user_b 属于 owner_b
        User = get_user_model()
        cls.user_a = User.objects.create_user(
            username="a", password="a", owner=cls.owner_a, is_staff=True
        )
        cls.user_b = User.objects.create_user(
            username="b", password="b", owner=cls.owner_b, is_staff=True
        )

        # 给两位普通用户授予 Product 的内置增/改/查/删权限（DjangoModelPermissions 需要）。
        ct = ContentType.objects.get_for_model(Product)
        builtin_perms = Permission.objects.filter(
            content_type=ct,
            codename__in=[
                "add_product",
                "change_product",
                "delete_product",
                "view_product",
            ],
        )
        cls.user_a.user_permissions.add(*list(builtin_perms))
        cls.user_b.user_permissions.add(*list(builtin_perms))

        cls.view_all_user = User.objects.create_user(
            username="view-all",
            password="x",
            owner=cls.owner_a,
            is_staff=True,
        )
        cls.manage_all_user = User.objects.create_user(
            username="manage-all",
            password="x",
            owner=cls.owner_a,
            is_staff=True,
        )
        cls.view_all_user.user_permissions.add(*list(builtin_perms))
        cls.view_all_user.user_permissions.add(
            Permission.objects.get(content_type=ct, codename="view_all_owner_products")
        )
        cls.manage_all_user.user_permissions.add(*list(builtin_perms))
        cls.manage_all_user.user_permissions.add(
            Permission.objects.get(
                content_type=ct, codename="manage_all_owner_products"
            )
        )

        for user, owner in (
            (cls.user_a, cls.owner_a),
            (cls.user_b, cls.owner_b),
            (cls.view_all_user, cls.owner_a),
            (cls.manage_all_user, cls.owner_a),
        ):
            UserRoleScope.objects.create(
                user=user,
                role=UserRoleScope.Role.OWNER_MANAGER,
                owner=owner,
            )

        # 现存商品：A/B 各一条
        cls.prod_a = Product.objects.create(
            owner=cls.owner_a,
            code="SKU-A",
            name="商品A",
            base_uom=cls.uom,
            is_active=True,
        )
        cls.prod_b = Product.objects.create(
            owner=cls.owner_b,
            code="SKU-B",
            name="商品B",
            base_uom=cls.uom,
            is_active=True,
        )

        cls.factory = APIRequestFactory()

    def test_list_scoped_by_owner(self):
        """
        非超管仅能看到自己 owner 的商品
        """
        view = ProductViewSet.as_view({"get": "list"})
        req = self.factory.get("/products/")
        force_authenticate(req, user=self.user_a)
        resp = view(req)
        self.assertEqual(resp.status_code, 200)
        codes = [item["code"] for item in resp.data]
        self.assertIn("SKU-A", codes)
        self.assertNotIn("SKU-B", codes)

    def test_create_auto_bind_owner(self):
        """
        非超管创建商品时，若请求未显式传 owner，后台会自动绑定为当前用户的 owner
        """
        view = ProductViewSet.as_view({"post": "create"})
        payload = {
            "code": "SKU-NEW",
            "name": "新商品",
            "base_uom": self.uom.id,
            "category": self.category.id,
            "purchase_price": "8.25",
            "is_active": True,
        }
        req = self.factory.post("/products/", data=payload, format="json")
        force_authenticate(req, user=self.user_a)
        resp = view(req)
        self.assertEqual(resp.status_code, 201, resp.data)
        self.assertEqual(resp.data["owner"], self.owner_a.id)
        self.assertEqual(resp.data["purchase_price"], "8.25")
        self.assertEqual(
            Product.objects.get(code="SKU-NEW").purchase_price,
            Decimal("8.25"),
        )

        # 再次 list，应该能看到新商品
        view_list = ProductViewSet.as_view({"get": "list"})
        req2 = self.factory.get("/products/")
        force_authenticate(req2, user=self.user_a)
        resp2 = view_list(req2)
        codes = [i["code"] for i in resp2.data]
        self.assertIn("SKU-NEW", codes)

    def test_create_rejects_cross_field_identifier_with_field_error(self):
        add_product_barcode(
            product=self.prod_a,
            barcode="API-CROSS-001",
            barcode_type=ProductBarcode.BarcodeType.UNIT,
            is_primary=True,
        )
        view = ProductViewSet.as_view({"post": "create"})
        req = self.factory.post(
            "/products/",
            data={
                "code": "api-cross-001",
                "name": "冲突商品",
                "base_uom": self.uom.id,
                "category": self.category.id,
                "is_active": True,
            },
            format="json",
        )
        force_authenticate(req, user=self.user_a)

        resp = view(req)

        self.assertEqual(resp.status_code, 400, resp.data)
        self.assertIn("code", resp.data)
        self.assertIn("已被商品", str(resp.data["code"]))

    def test_regular_user_cannot_create_for_other_owner(self):
        view = ProductViewSet.as_view({"post": "create"})
        payload = {
            "owner": self.owner_b.id,
            "code": "SKU-LOCKED",
            "name": "锁回当前货主",
            "base_uom": self.uom.id,
            "category": self.category.id,
            "is_active": True,
        }
        req = self.factory.post("/products/", data=payload, format="json")
        force_authenticate(req, user=self.user_a)
        resp = view(req)

        self.assertEqual(resp.status_code, 201, resp.data)
        product = Product.objects.get(code="SKU-LOCKED")
        self.assertEqual(product.owner_id, self.owner_a.id)

    def test_view_all_owner_permission_can_list_other_owner_products(self):
        view = ProductViewSet.as_view({"get": "list"})
        req = self.factory.get("/products/")
        force_authenticate(req, user=self.view_all_user)
        resp = view(req)

        self.assertEqual(resp.status_code, 200)
        codes = [item["code"] for item in resp.data]
        self.assertIn("SKU-A", codes)
        self.assertIn("SKU-B", codes)

    def test_view_all_owner_permission_cannot_modify_other_owner_products(self):
        view = ProductViewSet.as_view({"patch": "partial_update"})
        req = self.factory.patch(
            f"/products/{self.prod_b.id}/",
            data={"name": "Should Not Update"},
            format="json",
        )
        force_authenticate(req, user=self.view_all_user)
        resp = view(req, pk=self.prod_b.id)

        self.assertEqual(resp.status_code, 404)

    def test_manage_all_owner_permission_can_modify_other_owner_products(self):
        view = ProductViewSet.as_view({"patch": "partial_update"})
        req = self.factory.patch(
            f"/products/{self.prod_b.id}/",
            data={"name": "Managed Update"},
            format="json",
        )
        force_authenticate(req, user=self.manage_all_user)
        resp = view(req, pk=self.prod_b.id)

        self.assertEqual(resp.status_code, 200, resp.data)
        self.prod_b.refresh_from_db()
        self.assertEqual(self.prod_b.name, "Managed Update")
        self.assertEqual(self.prod_b.owner_id, self.owner_b.id)

    def test_regular_user_cannot_reassign_product_owner_on_update(self):
        view = ProductViewSet.as_view({"patch": "partial_update"})
        req = self.factory.patch(
            f"/products/{self.prod_a.id}/",
            data={"owner": self.owner_b.id, "name": "Owner Guarded"},
            format="json",
        )
        force_authenticate(req, user=self.user_a)
        resp = view(req, pk=self.prod_a.id)

        self.assertEqual(resp.status_code, 200, resp.data)
        self.prod_a.refresh_from_db()
        self.assertEqual(self.prod_a.owner_id, self.owner_a.id)
        self.assertEqual(self.prod_a.name, "Owner Guarded")

    def test_barcode_action_returns_zpl(self):
        """
        /products/{id}/barcode/ 应返回 ZPL 文本，且包含 base_uom.code
        """
        view = ProductViewSet.as_view({"get": "barcode"})
        req = self.factory.get(f"/products/{self.prod_a.id}/barcode/")
        force_authenticate(req, user=self.user_a)
        resp = view(req, pk=self.prod_a.id)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data.get("type"), "zpl")
        self.assertIn("PCS", resp.data.get("content", ""))  # base_uom.code

    def test_template_download_xlsx_headers(self):
        """
        /products/template/ 默认返回 Excel 商品模板。
        """
        view = ProductViewSet.as_view({"get": "template"})
        req = self.factory.get("/products/template/")
        force_authenticate(req, user=self.user_a)
        resp = view(req)
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp["Content-Type"],
        )
        workbook = load_workbook(io.BytesIO(resp.content))
        self.assertIn(IMPORT_SHEET_NAME, workbook.sheetnames)
        headers = [cell.value for cell in workbook[IMPORT_SHEET_NAME][1]]
        self.assertIn("货主商品编码", headers)

    def test_template_download_keeps_csv_compatibility(self):
        view = ProductViewSet.as_view({"get": "template"})
        req = self.factory.get("/products/template/?format=csv")
        force_authenticate(req, user=self.user_a)
        resp = view(req)

        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp["Content-Type"])
        self.assertIn("owner_code", resp.content.decode("utf-8"))

    def test_bulk_activate_and_deactivate_are_owner_scoped(self):
        self.prod_a.is_active = False
        self.prod_a.save(update_fields=["is_active"])
        self.prod_b.is_active = False
        self.prod_b.save(update_fields=["is_active"])

        activate_view = ProductViewSet.as_view({"post": "bulk_activate"})
        req = self.factory.post(
            "/products/bulk-activate/",
            data={"ids": [self.prod_a.id, self.prod_b.id]},
            format="json",
        )
        force_authenticate(req, user=self.user_a)
        resp = activate_view(req)

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["updated"], 1)
        self.prod_a.refresh_from_db()
        self.prod_b.refresh_from_db()
        self.assertTrue(self.prod_a.is_active)
        self.assertFalse(self.prod_b.is_active)

        deactivate_view = ProductViewSet.as_view({"post": "bulk_deactivate"})
        req = self.factory.post(
            "/products/bulk-deactivate/",
            data={"ids": [self.prod_a.id, self.prod_b.id]},
            format="json",
        )
        force_authenticate(req, user=self.user_a)
        resp = deactivate_view(req)

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["updated"], 1)
        self.prod_a.refresh_from_db()
        self.prod_b.refresh_from_db()
        self.assertFalse(self.prod_a.is_active)
        self.assertFalse(self.prod_b.is_active)

    def test_import_requires_file_and_export_still_reports_missing_resource(self):
        import_view = ProductViewSet.as_view({"post": "import_file"})
        req = self.factory.post("/products/import/", data={}, format="multipart")
        force_authenticate(req, user=self.user_a)
        import_resp = import_view(req)

        export_view = ProductViewSet.as_view({"get": "export_file"})
        req = self.factory.get("/products/export/")
        force_authenticate(req, user=self.user_a)
        export_resp = export_view(req)

        self.assertEqual(import_resp.status_code, 400)
        self.assertIn("file", import_resp.data["detail"])
        self.assertEqual(export_resp.status_code, 501)
        self.assertIn("ProductResource", export_resp.data["detail"])

    def test_get_product_details_returns_base_and_package_uoms(self):
        pack_uom = ProductUom.objects.create(code="CTN", name="箱", is_active=True)
        self.prod_a.packages.create(uom=pack_uom, qty_in_base=12)

        request = RequestFactory().get(
            f"/products/get_product_details/{self.prod_a.id}/"
        )
        from .views import get_product_details

        response = get_product_details(request, self.prod_a.id)

        self.assertEqual(response.status_code, 200)
        payload = json.loads(response.content.decode("utf-8"))
        self.assertEqual(payload["base_uom"], self.uom.name)
        self.assertEqual(payload["pack_uoms"][0]["uom"], "箱")
        self.assertEqual(payload["pack_uoms"][0]["pack_qty"], 12)


@unittest.skipUnless(
    DEPENDENCIES_OK and DAL_OK, "缺少依赖（DAL 或模型），跳过 UOM 自动补全测试"
)
class ProductUomAutocompleteTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.uom1 = ProductUom.objects.create(
            code="PCS", name="件", is_active=True, kind="COUNT"
        )
        cls.uom2 = ProductUom.objects.create(
            code="KG", name="千克", is_active=True, kind="WEIGHT"
        )

    def test_autocomplete_filters_and_forwards(self):
        """
        基本搜索 + forwarded 参数 only_count=1 时仅返回 COUNT 类
        DAL 会把 forwarded 参数解析到 view.self.forwarded
        """
        from django.test import RequestFactory

        rf = RequestFactory()
        view = ProductUomAutocomplete.as_view()

        # 不带 forward：按关键字
        req1 = rf.get("/autocomplete/uom/?q=K")
        resp1 = view(req1)
        self.assertEqual(resp1.status_code, 200)
        content1 = resp1.content.decode("utf-8")
        self.assertIn("KG", content1)

        # 带 forward：only_count=1
        req2 = rf.get(
            "/autocomplete/uom/?q=&forward=%7B%22only_count%22%3A%20%221%22%7D"
        )
        resp2 = view(req2)
        self.assertEqual(resp2.status_code, 200)
        content2 = resp2.content.decode("utf-8")
        self.assertIn("PCS", content2)  # COUNT 类
        self.assertNotIn("KG", content2)  # 非 COUNT 类不应出现


@unittest.skipUnless(
    DEPENDENCIES_OK, "缺少 baseinfo/products 依赖模型，跳过商品 Excel 导入测试"
)
class ProductExcelImportApiTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = Owner.objects.create(code="PXIA", name="Product Excel Owner A")
        cls.other_owner = Owner.objects.create(
            code="PXIB", name="Product Excel Owner B"
        )
        cls.uom = ProductUom.objects.create(code="EA-X", name="个", is_active=True)
        cls.carton_uom = ProductUom.objects.create(
            code="CTN-X", name="箱", is_active=True
        )
        cls.category = apps.get_model("products", "ProductCategory").objects.create(
            code="FOOD-X", name="食品", is_active=True
        )
        cls.brand = apps.get_model("products", "Brand").objects.create(
            code="BRAND-X", name="测试品牌", is_active=True
        )

        User = get_user_model()
        cls.user = User.objects.create_user(
            username="product-excel-owner", password="x"
        )
        cls.cross_owner_user = User.objects.create_user(
            username="product-excel-global", password="x"
        )
        cls.no_permission_user = User.objects.create_user(
            username="product-excel-denied", password="x"
        )
        for user in (cls.user, cls.cross_owner_user, cls.no_permission_user):
            UserRoleScope.objects.create(
                user=user,
                role=UserRoleScope.Role.OWNER_MANAGER,
                owner=cls.owner,
            )

        cls.warehouse = Warehouse.objects.create(
            code="PXWH",
            name="Product Excel Warehouse",
        )
        cls.subwarehouse = Subwarehouse.objects.create(
            warehouse=cls.warehouse,
            code="PXSW",
            name="Product Excel Subwarehouse",
        )
        cls.receive_location = Location.objects.create(
            warehouse=cls.warehouse,
            subwarehouse=cls.subwarehouse,
            code="PXSW-01-01-01",
            name="默认收货库位",
        )
        cls.warehouse.default_receive_location = cls.receive_location
        cls.warehouse.save(update_fields=["default_receive_location", "updated_at"])
        cls.warehouse_user = User.objects.create_user(
            username="product-excel-warehouse-global",
            password="x",
        )
        cls.warehouse_denied_user = User.objects.create_user(
            username="product-excel-warehouse-denied",
            password="x",
        )
        for user in (cls.warehouse_user, cls.warehouse_denied_user):
            UserRoleScope.objects.create(
                user=user,
                role=UserRoleScope.Role.WAREHOUSE_MANAGER,
                warehouse=cls.warehouse,
            )

        product_ct = ContentType.objects.get_for_model(Product)
        add_permission = Permission.objects.get(
            content_type=product_ct, codename="add_product"
        )
        view_permission = Permission.objects.get(
            content_type=product_ct, codename="view_product"
        )
        cls.user.user_permissions.add(add_permission, view_permission)
        cls.cross_owner_user.user_permissions.add(add_permission, view_permission)
        cls.cross_owner_user.user_permissions.add(
            Permission.objects.get(
                content_type=product_ct,
                codename="manage_all_owner_products",
            )
        )
        cls.warehouse_user.user_permissions.add(add_permission, view_permission)
        cls.warehouse_user.user_permissions.add(
            Permission.objects.get(
                content_type=product_ct,
                codename="manage_all_owner_products",
            )
        )
        cls.warehouse_denied_user.user_permissions.add(add_permission, view_permission)

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _workbook_file(self, rows, *, headers=HEADERS, filename="products.xlsx"):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = IMPORT_SHEET_NAME
        worksheet.append(list(headers))
        for row in rows:
            worksheet.append([row.get(header) for header in headers])
        output = io.BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            filename,
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _post_rows(
        self, rows, *, client=None, headers=HEADERS, url=None, warehouse_id=None
    ):
        data = {"file": self._workbook_file(rows, headers=headers)}
        if warehouse_id is not None:
            data["warehouse_id"] = warehouse_id
        return (client or self.client).post(
            url or "/api/products/import-excel/",
            data,
            format="multipart",
        )

    def _valid_row(self, code="PDA-XLSX-1", **overrides):
        row = {
            "货主编码": self.owner.code,
            "货主商品编码": code,
            "商品名称": f"导入商品 {code}",
            "基本单位": self.uom.name,
            "分类": self.category.full_path,
            "品牌编码": self.brand.code,
            "批次管理": "",
            "保质期管理": "",
        }
        row.update(overrides)
        return row

    def test_template_contains_scoped_references_and_metadata(self):
        response = self.client.get("/api/products/import-template/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("filename*=UTF-8", response["Content-Disposition"])
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ["填写说明", "商品导入", "商品包装", "基础资料", "_meta"],
        )
        self.assertEqual(workbook["_meta"].sheet_state, "hidden")
        self.assertEqual(workbook["_meta"]["B2"].value, "8")
        headers = [cell.value for cell in workbook[IMPORT_SHEET_NAME][1]]
        self.assertEqual(tuple(headers), PRODUCT_HEADERS)
        self.assertTrue({"包装单位编码", "包装换算数量", "包装条码"}.issubset(headers))
        package_headers = [cell.value for cell in workbook[PACKAGE_SHEET_NAME][1]]
        self.assertEqual(tuple(package_headers), PACKAGE_HEADERS)
        self.assertEqual(
            tuple(headers[:12]),
            (
                "货主编码",
                "货主商品编码",
                "默认价格",
                "进价",
                "标准贸易条码",
                "商品名称",
                "规格",
                "基本单位",
                "基本单位数量",
                "分类",
                "基本单位类型",
                "单位小数位数",
            ),
        )
        self.assertEqual(
            tuple(package_headers[:4]),
            ("货主编码", "货主商品编码", "包装单位编码", "包装换算数量"),
        )
        instructions = {
            row[0].value: row[1].value
            for row in workbook["填写说明"].iter_rows(min_col=1, max_col=2)
            if row[0].value and row[1].value
        }
        self.assertIn("货主编码", instructions["必填字段"])
        self.assertNotIn("货主商品编码", instructions["必填字段"])
        self.assertIn("标准贸易条码", instructions["货主商品编码规则"])
        self.assertIn("仓库SKU编码", instructions["货主商品编码规则"])
        self.assertIn("系统按", instructions["仓库SKU编码规则"])
        self.assertIn("默认价格和进价均可留空", instructions["价格规则"])
        self.assertIn("批次、序列号和保质期管理默认否", instructions["布尔值"])
        self.assertIn("整批不写入", instructions["重复规则"])
        owner_codes = {
            workbook["基础资料"].cell(row=row, column=1).value
            for row in range(2, workbook["基础资料"].max_row + 1)
        }
        self.assertIn(self.owner.code, owner_codes)
        self.assertNotIn(self.other_owner.code, owner_codes)
        self.assertIn("ProductImportUomCodes", workbook.defined_names)
        self.assertIn("ProductImportUomNames", workbook.defined_names)
        self.assertIn("ProductImportCategoryNames", workbook.defined_names)
        self.assertNotIn("仓库SKU编码", headers)
        code_column = PRODUCT_HEADERS.index("货主商品编码") + 1
        owner_column = PRODUCT_HEADERS.index("货主编码") + 1
        barcode_column = PRODUCT_HEADERS.index("标准贸易条码") + 1
        self.assertNotEqual(
            workbook[IMPORT_SHEET_NAME].cell(1, owner_column).fill.fgColor.rgb,
            workbook[IMPORT_SHEET_NAME].cell(1, code_column).fill.fgColor.rgb,
        )
        self.assertEqual(
            workbook[IMPORT_SHEET_NAME].cell(2, code_column).number_format, "@"
        )
        self.assertEqual(
            workbook[IMPORT_SHEET_NAME].cell(2, barcode_column).number_format,
            "@",
        )

    def test_legacy_product_code_header_is_rejected_with_new_header_guidance(self):
        legacy_headers = tuple(
            "商品编号" if header == "货主商品编码" else header for header in HEADERS
        )

        response = self._post_rows(
            [self._valid_row()],
            headers=legacy_headers,
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertIn("v8 模板不一致", response.data["detail"])

    def test_old_and_reordered_headers_are_rejected_with_v8_guidance(self):
        v6_headers = (
            "货主编码",
            "货主商品编码",
            "标准贸易条码",
            "商品名称",
            "分类",
            "基本单位",
            "基本单位类型",
            "单位小数位数",
            "基本单位数量",
            "规格",
            "品牌编码",
            "零码",
            "箱码",
            "箱码对应包装单位编码",
            "外部系统商品编码",
            "默认价格",
            "最低价格",
            "最高折扣%",
            "重量kg",
            "体积m³",
            "净含量g",
            "厂家",
            "材质",
            "描述",
            "最低库存",
            "最高库存",
            "序列号管理",
            "批次管理",
            "启用",
            "保质期管理",
            "效期基准",
            "保质期天数",
            "入库有效天数",
            "效期预警天数",
            "FEFO",
            "包装单位编码",
            "包装换算数量",
            "包装条码",
            "采购默认",
            "销售默认",
        )
        reordered_headers = list(HEADERS)
        reordered_headers[2], reordered_headers[3] = (
            reordered_headers[3],
            reordered_headers[2],
        )
        v7_headers = list(HEADERS)
        quantity_header = v7_headers.pop(8)
        v7_headers.insert(11, quantity_header)
        without_default_price = tuple(
            header for header in HEADERS if header != "默认价格"
        )
        without_purchase_price = tuple(header for header in HEADERS if header != "进价")

        v6_response = self._post_rows(
            [self._valid_row("PDA-V6-HEADER")], headers=v6_headers
        )
        v7_response = self._post_rows(
            [self._valid_row("PDA-V7-HEADER")], headers=tuple(v7_headers)
        )
        reordered_response = self._post_rows(
            [self._valid_row("PDA-REORDERED-HEADER")],
            headers=tuple(reordered_headers),
        )
        missing_default_response = self._post_rows(
            [self._valid_row("PDA-MISSING-DEFAULT-PRICE")],
            headers=without_default_price,
        )
        missing_purchase_response = self._post_rows(
            [self._valid_row("PDA-MISSING-PURCHASE-PRICE")],
            headers=without_purchase_price,
        )

        self.assertEqual(v6_response.status_code, 400, v6_response.data)
        self.assertIn("v8 模板不一致", v6_response.data["detail"])
        self.assertEqual(v7_response.status_code, 400, v7_response.data)
        self.assertIn("v8 模板不一致", v7_response.data["detail"])
        self.assertEqual(
            reordered_response.status_code, 400, reordered_response.data
        )
        self.assertIn("v8 模板不一致", reordered_response.data["detail"])
        self.assertEqual(
            missing_default_response.status_code, 400, missing_default_response.data
        )
        self.assertIn("v8 模板不一致", missing_default_response.data["detail"])
        self.assertEqual(
            missing_purchase_response.status_code,
            400,
            missing_purchase_response.data,
        )
        self.assertIn("v8 模板不一致", missing_purchase_response.data["detail"])

    def test_category_and_base_uom_names_resolve_or_create_with_pinyin_codes(self):
        ProductCategory.objects.create(code="YINLIAOXINPIN", name="占用分类代码")
        ProductUom.objects.create(code="DAORUCESHIPING", name="占用单位代码")

        response = self._post_rows(
            [
                self._valid_row(
                    "PDA-NAME-DICT",
                    **{
                        "分类": "饮料新品",
                        "基本单位": "导入测试瓶",
                        "基本单位类型": "计数",
                        "单位小数位数": 0,
                    },
                )
            ]
        )

        self.assertEqual(response.status_code, 200, response.data)
        product = Product.objects.get(code="PDA-NAME-DICT")
        self.assertEqual(product.category.name, "饮料新品")
        self.assertEqual(product.category.code, "YINLIAOXINPIN_2")
        self.assertEqual(product.base_uom.name, "导入测试瓶")
        self.assertEqual(product.base_uom.code, "DAORUCESHIPING_2")
        self.assertEqual(
            response.data["created_categories"][0]["code"], "YINLIAOXINPIN_2"
        )
        self.assertEqual(response.data["created_uoms"][0]["code"], "DAORUCESHIPING_2")

    def test_category_full_path_and_legacy_codes_remain_accepted(self):
        child = ProductCategory.objects.create(
            code="FOOD-X-DRINK", name="饮料", parent=self.category
        )
        responses = [
            self._post_rows(
                [self._valid_row("PDA-CAT-PATH", **{"分类": child.full_path})]
            ),
            self._post_rows(
                [
                    self._valid_row(
                        "PDA-CODE-ALIASES",
                        **{"分类": self.category.code, "基本单位": self.uom.code},
                    )
                ]
            ),
        ]

        self.assertTrue(all(response.status_code == 200 for response in responses))
        self.assertEqual(Product.objects.get(code="PDA-CAT-PATH").category_id, child.id)

    def test_auto_created_dictionaries_roll_back_with_later_row_error(self):
        response = self._post_rows(
            [
                self._valid_row(
                    "PDA-DICT-ROLLBACK",
                    **{
                        "分类": "事务回滚分类",
                        "基本单位": "事务回滚单位",
                        "基本单位类型": "计数",
                        "单位小数位数": 0,
                    },
                ),
                self._valid_row("PDA-DICT-BAD", **{"标准贸易条码": "123"}),
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(ProductCategory.objects.filter(name="事务回滚分类").exists())
        self.assertFalse(ProductUom.objects.filter(name="事务回滚单位").exists())
        self.assertFalse(Product.objects.filter(code="PDA-DICT-ROLLBACK").exists())

    def test_import_warehouse_options_and_positive_quantity_auto_receive(self):
        OwnerWarehouseBinding.objects.create(owner=self.owner, warehouse=self.warehouse)
        self.user.user_permissions.add(
            Permission.objects.get(codename="receive_without_order")
        )

        options = self.client.get("/api/products/import-warehouses/")
        response = self._post_rows(
            [self._valid_row("PDA-AUTO-RECEIVE", **{"基本单位数量": "3"})],
            warehouse_id=self.warehouse.pk,
        )

        self.assertEqual(options.status_code, 200, options.data)
        self.assertTrue(options.data["results"][0]["receipt_ready"])
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["received_product_count"], 1)
        self.assertEqual(len(response.data["receipts"]), 1)
        product = Product.objects.get(code="PDA-AUTO-RECEIVE")
        inventory = InventoryDetail.objects.get(
            product=product,
            warehouse=self.warehouse,
            location=self.receive_location,
        )
        self.assertEqual(inventory.onhand_qty, 3)

    def test_positive_quantities_create_one_receipt_per_owner(self):
        OwnerWarehouseBinding.objects.create(
            owner=self.owner, warehouse=self.warehouse
        )
        OwnerWarehouseBinding.objects.create(
            owner=self.other_owner, warehouse=self.warehouse
        )
        self.cross_owner_user.user_permissions.add(
            Permission.objects.get(codename="receive_without_order")
        )
        client = APIClient()
        client.force_authenticate(self.cross_owner_user)

        response = self._post_rows(
            [
                self._valid_row("PDA-RECEIVE-A", **{"基本单位数量": 2}),
                self._valid_row(
                    "PDA-RECEIVE-B",
                    **{
                        "货主编码": self.other_owner.code,
                        "基本单位数量": 4,
                    },
                ),
            ],
            client=client,
            warehouse_id=self.warehouse.pk,
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["received_product_count"], 2)
        self.assertEqual(
            {receipt["owner_code"] for receipt in response.data["receipts"]},
            {self.owner.code, self.other_owner.code},
        )
        self.assertEqual(len(response.data["receipts"]), 2)
        self.assertEqual(
            InventoryDetail.objects.filter(
                product__code__in=["PDA-RECEIVE-A", "PDA-RECEIVE-B"],
                warehouse=self.warehouse,
                location=self.receive_location,
            ).count(),
            2,
        )

    def test_second_owner_receipt_failure_rolls_back_entire_import(self):
        from . import excel_import

        OwnerWarehouseBinding.objects.create(
            owner=self.owner, warehouse=self.warehouse
        )
        OwnerWarehouseBinding.objects.create(
            owner=self.other_owner, warehouse=self.warehouse
        )
        self.cross_owner_user.user_permissions.add(
            Permission.objects.get(codename="receive_without_order")
        )
        client = APIClient()
        client.force_authenticate(self.cross_owner_user)
        real_receive = excel_import.receive_goods_without_order
        call_count = 0

        def fail_second_receipt(**kwargs):
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise ValidationError("第二个货主收货过账失败")
            return real_receive(**kwargs)

        with patch(
            "allapp.products.excel_import.receive_goods_without_order",
            side_effect=fail_second_receipt,
        ):
            response = self._post_rows(
                [
                    self._valid_row("PDA-ROLLBACK-A", **{"基本单位数量": 2}),
                    self._valid_row(
                        "PDA-ROLLBACK-B",
                        **{
                            "货主编码": self.other_owner.code,
                            "基本单位数量": 4,
                        },
                    ),
                ],
                client=client,
                warehouse_id=self.warehouse.pk,
            )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertFalse(Product.objects.filter(code__startswith="PDA-ROLLBACK-").exists())
        self.assertFalse(
            InventoryDetail.objects.filter(
                warehouse=self.warehouse,
                location=self.receive_location,
            ).exists()
        )

    def test_positive_quantity_requires_context_and_rejects_controlled_goods(self):
        OwnerWarehouseBinding.objects.create(owner=self.owner, warehouse=self.warehouse)
        missing_warehouse = self._post_rows(
            [self._valid_row("PDA-QTY-NO-WH", **{"基本单位数量": 1})]
        )
        controlled = self._post_rows(
            [
                self._valid_row(
                    "PDA-QTY-CONTROLLED",
                    **{"基本单位数量": 1, "批次管理": "是"},
                )
            ],
            warehouse_id=self.warehouse.pk,
        )

        self.assertEqual(missing_warehouse.status_code, 400, missing_warehouse.data)
        self.assertEqual(controlled.status_code, 400, controlled.data)
        self.assertFalse(Product.objects.filter(code__startswith="PDA-QTY-").exists())

    def test_zero_quantity_needs_no_warehouse_or_receive_permission(self):
        response = self._post_rows(
            [self._valid_row("PDA-QTY-ZERO", **{"基本单位数量": 0})]
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["received_product_count"], 0)
        self.assertEqual(response.data["receipts"], [])

    def test_happy_path_creates_product_package_and_audit_event(self):
        response = self._post_rows(
            [
                self._valid_row(
                    code=" pda-xlsx-happy ",
                    **{
                        "默认价格": "12.50",
                        "进价": "8.25",
                        "最低库存": 2,
                        "最高库存": 20,
                        "序列号管理": "否",
                        "包装单位编码": self.carton_uom.code,
                        "包装换算数量": 12,
                        "包装条码": "000123456789",
                        "采购默认": "是",
                        "销售默认": "是",
                    },
                )
            ]
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["created_count"], 1)
        product = Product.objects.get(owner=self.owner, code="PDA-XLSX-HAPPY")
        self.assertEqual(product.sku, "PXIA1")
        self.assertEqual(product.price, Decimal("12.50"))
        self.assertEqual(product.purchase_price, Decimal("8.25"))
        self.assertEqual(product.created_by_id, self.user.id)
        self.assertFalse(product.expiry_control)
        self.assertFalse(product.serial_control)
        self.assertFalse(product.batch_control)
        self.assertTrue(product.is_active)
        package = product.packages.get()
        self.assertEqual(package.uom_id, self.carton_uom.id)
        self.assertEqual(package.qty_in_base, 12)
        self.assertTrue(package.is_purchase_default)
        self.assertTrue(
            AuditEvent.objects.filter(
                actor=self.user,
                action="products.import_excel",
                object_type="",
            ).exists()
        )

    def test_blank_prices_remain_null_and_do_not_backfill_each_other(self):
        response = self._post_rows([self._valid_row("PDA-BLANK-PRICES")])

        self.assertEqual(response.status_code, 200, response.data)
        product = Product.objects.get(owner=self.owner, code="PDA-BLANK-PRICES")
        self.assertIsNone(product.price)
        self.assertIsNone(product.purchase_price)

    def test_invalid_default_and_purchase_prices_reject_the_whole_batch(self):
        response = self._post_rows(
            [
                self._valid_row("PDA-BAD-DEFAULT-PRICE", **{"默认价格": "-1"}),
                self._valid_row("PDA-BAD-PURCHASE-PRICE", **{"进价": "invalid"}),
                self._valid_row("PDA-BAD-PRICE-SCALE", **{"进价": "1.234"}),
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        fields = {error["field"] for error in response.data["errors"]}
        self.assertTrue({"默认价格", "进价"}.issubset(fields), response.data)
        self.assertFalse(Product.objects.filter(code__startswith="PDA-BAD-").exists())

    def test_product_code_prefers_supplied_code_over_gtin(self):
        response = self._post_rows(
            [
                self._valid_row(
                    " OWNER-CODE-1 ",
                    **{"标准贸易条码": "12345678"},
                )
            ]
        )

        self.assertEqual(response.status_code, 200, response.data)
        product = Product.objects.get(owner=self.owner, code="OWNER-CODE-1")
        self.assertEqual(product.gtin, "12345678")
        self.assertEqual(response.data["created"][0]["code"], "OWNER-CODE-1")

    def test_blank_product_code_falls_back_to_gtin(self):
        response = self._post_rows(
            [self._valid_row("   ", **{"标准贸易条码": "12345678"})]
        )

        self.assertEqual(response.status_code, 200, response.data)
        product = Product.objects.get(owner=self.owner, code="12345678")
        self.assertEqual(product.gtin, "12345678")
        self.assertEqual(response.data["created"][0]["code"], "12345678")

    def test_gtin_fallback_still_validates_format_and_identifier_ownership(self):
        invalid = self._post_rows([self._valid_row("", **{"标准贸易条码": "123"})])

        self.assertEqual(invalid.status_code, 400, invalid.data)
        self.assertTrue(
            any(
                error["field"] == "标准贸易条码" and "8/12/13/14" in error["message"]
                for error in invalid.data["errors"]
            )
        )

        Product.objects.create(
            owner=self.owner,
            code="GTIN-OWNER",
            name="占用 GTIN 的商品",
            category=self.category,
            base_uom=self.uom,
            gtin="87654321",
            expiry_control=False,
            expiry_basis=None,
        )
        occupied = self._post_rows(
            [self._valid_row("", **{"标准贸易条码": "87654321"})]
        )

        self.assertEqual(occupied.status_code, 400, occupied.data)
        self.assertTrue(
            any(
                error["field"] in {"货主商品编码", "标准贸易条码"}
                and ("占用" in error["message"] or "已存在" in error["message"])
                for error in occupied.data["errors"]
            )
        )
        self.assertEqual(Product.objects.filter(gtin="87654321").count(), 1)

    def test_blank_product_code_and_gtin_use_generated_sku(self):
        response = self._post_rows(
            [
                self._valid_row("", **{"商品名称": "无码商品一"}),
                self._valid_row("   ", **{"商品名称": "无码商品二"}),
            ]
        )

        self.assertEqual(response.status_code, 200, response.data)
        products = list(
            Product.objects.filter(owner=self.owner)
            .order_by("sku")
            .values_list("code", "sku")
        )
        self.assertEqual(products, [("PXIA1", "PXIA1"), ("PXIA2", "PXIA2")])
        self.assertEqual(
            [item["code"] for item in response.data["created"]],
            ["PXIA1", "PXIA2"],
        )

    def test_generated_code_skips_owner_identifier_collision(self):
        existing = Product.objects.create(
            owner=self.owner,
            code="EXISTING-CODE",
            name="已有商品",
            category=self.category,
            base_uom=self.uom,
            expiry_control=False,
            expiry_basis=None,
        )
        add_product_barcode(
            product=existing,
            barcode="PXIA2",
            barcode_type=ProductBarcode.BarcodeType.OTHER,
            is_primary=False,
        )

        response = self._post_rows(
            [self._valid_row("", **{"商品名称": "跳号无码商品"})]
        )

        self.assertEqual(response.status_code, 200, response.data)
        product = Product.objects.get(name="跳号无码商品")
        self.assertEqual((product.code, product.sku), ("PXIA3", "PXIA3"))

    def test_gtin_fallback_can_be_referenced_by_package_sheet(self):
        workbook = Workbook()
        product_sheet = workbook.active
        product_sheet.title = IMPORT_SHEET_NAME
        product_sheet.append(list(HEADERS))
        row = self._valid_row("", **{"标准贸易条码": "12345678"})
        product_sheet.append([row.get(header) for header in HEADERS])
        package_sheet = workbook.create_sheet(PACKAGE_SHEET_NAME)
        package_sheet.append(list(PACKAGE_HEADERS))
        package_row = {
            "货主编码": self.owner.code,
            "货主商品编码": "12345678",
            "包装单位编码": self.carton_uom.code,
            "包装换算数量": 12,
            "包装条码": "GTIN-PACKAGE-12",
        }
        package_sheet.append([package_row.get(header) for header in PACKAGE_HEADERS])
        output = io.BytesIO()
        workbook.save(output)

        response = self.client.post(
            "/api/products/import-excel/",
            {"file": SimpleUploadedFile("gtin-package.xlsx", output.getvalue())},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200, response.data)
        product = Product.objects.get(owner=self.owner, code="12345678")
        self.assertTrue(
            product.packages.filter(
                uom=self.carton_uom, barcode="GTIN-PACKAGE-12"
            ).exists()
        )

    def test_generated_code_product_uses_inline_package_only(self):
        inline_response = self._post_rows(
            [
                self._valid_row(
                    "",
                    **{
                        "商品名称": "无码同行包装",
                        "包装单位编码": self.carton_uom.code,
                        "包装换算数量": 12,
                        "包装条码": "INLINE-AUTO-PACKAGE",
                    },
                )
            ]
        )

        self.assertEqual(inline_response.status_code, 200, inline_response.data)
        product = Product.objects.get(name="无码同行包装")
        self.assertEqual(product.code, product.sku)
        self.assertTrue(product.packages.filter(barcode="INLINE-AUTO-PACKAGE").exists())

        workbook = Workbook()
        product_sheet = workbook.active
        product_sheet.title = IMPORT_SHEET_NAME
        product_sheet.append(list(HEADERS))
        row = self._valid_row("", **{"商品名称": "无码包装表引用"})
        product_sheet.append([row.get(header) for header in HEADERS])
        package_sheet = workbook.create_sheet(PACKAGE_SHEET_NAME)
        package_sheet.append(list(PACKAGE_HEADERS))
        package_sheet.append(
            [self.owner.code, "", self.carton_uom.code, 12, "AUTO-SHEET-PACKAGE"]
        )
        output = io.BytesIO()
        workbook.save(output)

        package_response = self.client.post(
            "/api/products/import-excel/",
            {"file": SimpleUploadedFile("auto-package.xlsx", output.getvalue())},
            format="multipart",
        )

        self.assertEqual(package_response.status_code, 400, package_response.data)
        self.assertTrue(
            any(
                error["field"] == "货主商品编码"
                and "主表同行包装列" in error["message"]
                for error in package_response.data["errors"]
            )
        )
        self.assertFalse(Product.objects.filter(name="无码包装表引用").exists())

    def test_legacy_package_barcode_cannot_reuse_same_product_identifier(self):
        response = self._post_rows(
            [
                self._valid_row(
                    "PDA-PACKAGE-CONFLICT",
                    **{
                        "包装单位编码": self.carton_uom.code,
                        "包装换算数量": 12,
                        "包装条码": " pda-package-conflict ",
                    },
                )
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertTrue(
            any(
                error["field"] == "包装条码" and "货主商品编码" in error["message"]
                for error in response.data["errors"]
            )
        )

    def test_package_sheet_barcode_conflicts_with_product_sheet_identifier(self):
        workbook = Workbook()
        product_sheet = workbook.active
        product_sheet.title = IMPORT_SHEET_NAME
        product_sheet.append(list(HEADERS))
        product_row = self._valid_row(
            "PDA-PACKAGE-SHEET",
            **{"外部系统商品编码": "PACKAGE-SHEET-SHARED"},
        )
        product_sheet.append([product_row.get(header) for header in HEADERS])
        package_sheet = workbook.create_sheet(PACKAGE_SHEET_NAME)
        package_sheet.append(list(PACKAGE_HEADERS))
        package_row = {
            "货主编码": self.owner.code,
            "货主商品编码": "PDA-PACKAGE-SHEET",
            "包装单位编码": self.carton_uom.code,
            "包装换算数量": 12,
            "包装条码": " package-sheet-shared ",
        }
        package_sheet.append([package_row.get(header) for header in PACKAGE_HEADERS])
        output = io.BytesIO()
        workbook.save(output)
        uploaded = SimpleUploadedFile(
            "package-sheet-conflict.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        response = self.client.post(
            "/api/products/import-excel/",
            {"file": uploaded},
            format="multipart",
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertTrue(
            any(
                error["sheet"] == PACKAGE_SHEET_NAME
                and error["field"] == "包装条码"
                and "外部系统商品编码" in error["message"]
                for error in response.data["errors"]
            )
        )

    def test_single_owner_scope_rejects_other_owner_and_writes_nothing(self):
        response = self._post_rows(
            [
                self._valid_row("PDA-OWN-A"),
                self._valid_row("PDA-OWN-B", **{"货主编码": self.other_owner.code}),
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertTrue(
            any(error["field"] == "货主编码" for error in response.data["errors"])
        )
        self.assertFalse(Product.objects.filter(code="PDA-OWN-A").exists())

    def test_owner_code_is_required_even_for_single_owner_scope(self):
        response = self._post_rows(
            [self._valid_row("PDA-OWNER-REQUIRED", **{"货主编码": ""})]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertTrue(
            any(
                error["field"] == "货主编码" and "不能为空" in error["message"]
                for error in response.data["errors"]
            )
        )
        self.assertFalse(Product.objects.filter(code="PDA-OWNER-REQUIRED").exists())

    def test_cross_owner_permission_can_import_for_other_owner(self):
        client = APIClient()
        client.force_authenticate(self.cross_owner_user)
        response = self._post_rows(
            [self._valid_row("PDA-GLOBAL", **{"货主编码": self.other_owner.code})],
            client=client,
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(
            Product.objects.filter(owner=self.other_owner, code="PDA-GLOBAL").exists()
        )

    def test_warehouse_scope_requires_explicit_cross_owner_permission(self):
        denied_client = APIClient()
        denied_client.force_authenticate(self.warehouse_denied_user)
        denied = denied_client.get("/api/products/import-template/")

        allowed_client = APIClient()
        allowed_client.force_authenticate(self.warehouse_user)
        missing_owner = self._post_rows(
            [self._valid_row("PDA-WAREHOUSE-GLOBAL", **{"货主编码": ""})],
            client=allowed_client,
        )

        self.assertEqual(denied.status_code, 403)
        self.assertEqual(missing_owner.status_code, 400, missing_owner.data)
        self.assertTrue(
            any(error["field"] == "货主编码" for error in missing_owner.data["errors"])
        )

        allowed = self._post_rows(
            [
                self._valid_row(
                    "PDA-WAREHOUSE-GLOBAL",
                    **{"货主编码": self.owner.code},
                )
            ],
            client=allowed_client,
        )
        self.assertEqual(allowed.status_code, 200, allowed.data)

    def test_existing_product_code_rejects_whole_batch_without_update(self):
        existing = Product.objects.create(
            owner=self.owner,
            code="PDA-EXISTING",
            sku="PDA-EXISTING",
            name="原商品名称",
            base_uom=self.uom,
            expiry_control=False,
            expiry_basis=None,
        )

        response = self._post_rows(
            [
                self._valid_row("PDA-NEW-BEFORE-CONFLICT"),
                self._valid_row(
                    "PDA-EXISTING",
                    **{"商品名称": "尝试修改原商品"},
                ),
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertEqual(response.data["skipped_count"], 0)
        self.assertTrue(
            any(
                error["field"] == "货主商品编码" and "已存在" in error["message"]
                for error in response.data["errors"]
            )
        )
        self.assertFalse(
            Product.objects.filter(code="PDA-NEW-BEFORE-CONFLICT").exists()
        )
        existing.refresh_from_db()
        self.assertEqual(existing.name, "原商品名称")

    def test_soft_deleted_product_code_rejects_batch_with_recovery_guidance(self):
        product = Product.objects.create(
            owner=self.owner,
            code="PDA-DELETED",
            sku="PDA-DELETED",
            name="已删除商品",
            base_uom=self.uom,
            expiry_control=False,
            expiry_basis=None,
        )
        Product.all_objects.filter(pk=product.pk).update(is_deleted=True)

        response = self._post_rows([self._valid_row("PDA-DELETED")])

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["skipped_count"], 0)
        self.assertTrue(
            any(
                error["field"] == "货主商品编码" and "恢复旧商品" in error["message"]
                for error in response.data["errors"]
            )
        )
        self.assertTrue(Product.all_objects.get(pk=product.pk).is_deleted)

    def test_identifier_owned_by_other_product_is_an_error(self):
        Product.objects.create(
            owner=self.owner,
            code="PDA-IDENTIFIER-OLD",
            sku="PDA-IDENTIFIER-OLD",
            name="占用标识的商品",
            base_uom=self.uom,
            external_code="SHARED-EXTERNAL",
            expiry_control=False,
            expiry_basis=None,
        )

        response = self._post_rows(
            [
                self._valid_row(
                    "PDA-IDENTIFIER-NEW",
                    **{"外部系统商品编码": "SHARED-EXTERNAL"},
                )
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertTrue(
            any(
                error["field"] == "外部系统商品编码"
                for error in response.data["errors"]
            )
        )
        self.assertFalse(Product.objects.filter(code="PDA-IDENTIFIER-NEW").exists())

    def test_existing_cross_field_identifier_rejects_whole_batch(self):
        Product.objects.create(
            owner=self.owner,
            code="PDA-CROSS-OLD",
            name="跨字段占用商品",
            base_uom=self.uom,
            unit_barcode="CROSS-FIELD-001",
            expiry_control=False,
            expiry_basis=None,
        )

        response = self._post_rows([self._valid_row("cross-field-001")])

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertTrue(
            any(
                error["field"] == "货主商品编码" and "标识" in error["message"]
                for error in response.data["errors"]
            )
        )
        self.assertFalse(Product.objects.filter(code="CROSS-FIELD-001").exists())

    def test_invalid_row_makes_whole_batch_atomic(self):
        response = self._post_rows(
            [
                self._valid_row("PDA-ATOMIC-OK"),
                self._valid_row(
                    "PDA-ATOMIC-BAD",
                    **{
                        "基本单位": "新单位但缺少元数据",
                        "基本单位类型": "",
                        "单位小数位数": "",
                    },
                ),
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertFalse(Product.objects.filter(code="PDA-ATOMIC-OK").exists())

    def test_supplied_skus_are_ignored_and_generated_sequentially(self):
        response = self._post_rows(
            [
                self._valid_row("PDA-DUP-1"),
                self._valid_row("PDA-DUP-2"),
            ]
        )

        self.assertEqual(response.status_code, 200, response.data)
        products = list(
            Product.objects.filter(code__startswith="PDA-DUP")
            .order_by("code")
            .values_list("sku", flat=True)
        )
        self.assertEqual(products, ["PXIA1", "PXIA2"])

    def test_file_duplicates_are_checked_even_with_existing_code_conflict(self):
        Product.objects.create(
            owner=self.owner,
            code="PDA-DUP-EXISTING",
            sku="PDA-DUP-EXISTING",
            name="已存在商品",
            base_uom=self.uom,
            expiry_control=False,
            expiry_basis=None,
        )

        response = self._post_rows(
            [
                self._valid_row("PDA-DUP-EXISTING"),
                self._valid_row("PDA-DUP-EXISTING"),
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertTrue(
            any(
                error["row"] == 3 and error["field"] == "货主商品编码"
                for error in response.data["errors"]
            )
        )

    def test_file_cross_field_duplicates_report_both_rows_and_fields(self):
        response = self._post_rows(
            [
                self._valid_row(
                    "PDA-CROSS-ROW-1",
                    **{"箱码": "FILE-CROSS-001"},
                ),
                self._valid_row("file-cross-001"),
            ]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        self.assertTrue(
            any(
                error["row"] == 3
                and error["field"] == "货主商品编码"
                and "第 2 行的箱码" in error["message"]
                for error in response.data["errors"]
            )
        )

    def test_defaults_and_both_expiry_bases(self):
        response = self._post_rows(
            [
                self._valid_row(
                    "PDA-MFG",
                    **{
                        "批次管理": "",
                        "序列号管理": "",
                        "启用": "",
                        "保质期管理": "是",
                        "效期基准": "MFG",
                        "保质期天数": 365,
                        "效期预警天数": 30,
                    },
                ),
                self._valid_row(
                    "PDA-INBOUND",
                    **{
                        "保质期管理": "true",
                        "效期基准": "INBOUND",
                        "入库有效天数": 90,
                        "效期预警天数": 10,
                        "FEFO": "1",
                    },
                ),
            ]
        )

        self.assertEqual(response.status_code, 200, response.data)
        mfg = Product.objects.get(code="PDA-MFG")
        inbound = Product.objects.get(code="PDA-INBOUND")
        self.assertEqual(mfg.sku, "PXIA1")
        self.assertEqual(inbound.sku, "PXIA2")
        self.assertFalse(mfg.batch_control)
        self.assertFalse(mfg.serial_control)
        self.assertTrue(mfg.is_active)
        self.assertEqual(mfg.shelf_life_days, 365)
        self.assertEqual(inbound.inbound_valid_days, 90)
        self.assertTrue(inbound.fefo_required)

    def test_dictionary_and_model_validation_errors_are_all_reported(self):
        inactive_uom = ProductUom.objects.create(
            code="INACTIVE-X",
            name="停用单位",
            is_active=False,
        )
        rows = [
            self._valid_row(
                "PDA-BAD-UOM",
                **{
                    "基本单位": "坏单位",
                    "基本单位类型": "",
                    "单位小数位数": "",
                },
            ),
            self._valid_row(
                "PDA-BAD-INACTIVE-UOM",
                **{"基本单位": inactive_uom.name},
            ),
            self._valid_row("PDA-BAD-CATEGORY", **{"分类": "不存在 > 多级分类"}),
            self._valid_row("PDA-BAD-BRAND", **{"品牌编码": "NO-BRAND"}),
            self._valid_row("PDA-BAD-STOCK", **{"最低库存": 10, "最高库存": 5}),
            self._valid_row("PDA-BAD-GTIN", **{"标准贸易条码": "123"}),
            self._valid_row(
                "PDA-BAD-BOOLEAN",
                **{"批次管理": "不确定"},
            ),
            self._valid_row(
                "PDA-BAD-EXPIRY",
                **{"保质期管理": "是", "效期基准": "MFG"},
            ),
            self._valid_row(
                "PDA-BAD-PACKAGE",
                **{
                    "包装单位编码": self.uom.code,
                    "包装换算数量": 2,
                },
            ),
            self._valid_row(
                "PDA-BAD-PACKAGE-MODEL",
                **{
                    "包装单位编码": self.carton_uom.code,
                    "包装换算数量": 2,
                    "包装条码": "12",
                },
            ),
            self._valid_row(
                "PDA-BAD-PACKAGE-UOM",
                **{
                    "包装单位编码": "NO-PACKAGE-UOM",
                    "包装换算数量": 2,
                },
            ),
            self._valid_row(
                "PDA-BAD-PACKAGE-PARTIAL",
                **{"采购默认": "是"},
            ),
        ]

        response = self._post_rows(rows)

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["created_count"], 0)
        fields = {error["field"] for error in response.data["errors"]}
        self.assertTrue(
            {
                "基本单位",
                "分类",
                "品牌编码",
                "最低库存",
                "标准贸易条码",
                "保质期天数",
                "包装换算数量",
                "包装条码",
                "包装单位编码",
            }.issubset(fields),
            response.data,
        )
        self.assertFalse(Product.objects.filter(code__startswith="PDA-BAD-").exists())
        self.assertTrue(
            any(
                error["field"] == "基本单位" and inactive_uom.code in error["message"]
                for error in response.data["errors"]
            )
        )
        self.assertTrue(
            any(
                error["field"] == "批次管理" and "true/false" in error["message"]
                for error in response.data["errors"]
            )
        )

    def test_formula_cell_is_rejected(self):
        response = self._post_rows(
            [self._valid_row("PDA-FORMULA", **{"商品名称": '=CONCAT("商品","名称")'})]
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertEqual(response.data["errors"][0]["field"], "商品名称")
        self.assertIn("公式", response.data["errors"][0]["message"])

    def test_permission_and_profile_capability_are_fail_closed(self):
        denied_client = APIClient()
        denied_client.force_authenticate(self.no_permission_user)
        denied = denied_client.get("/api/products/import-template/")
        allowed_profile = self.client.get("/api/auth/profile/")
        denied_profile = denied_client.get("/api/auth/profile/")

        self.assertEqual(denied.status_code, 403)
        self.assertTrue(allowed_profile.data["capabilities"]["can_import_products"])
        self.assertFalse(denied_profile.data["capabilities"]["can_import_products"])

    def test_anonymous_and_missing_scope_are_denied(self):
        anonymous = APIClient().get("/api/products/import-template/")
        User = get_user_model()
        unscoped_user = User.objects.create_user(username="product-excel-unscoped")
        product_ct = ContentType.objects.get_for_model(Product)
        unscoped_user.user_permissions.add(
            Permission.objects.get(content_type=product_ct, codename="add_product")
        )
        unscoped_client = APIClient()
        unscoped_client.force_authenticate(unscoped_user)

        unscoped = unscoped_client.get("/api/products/import-template/")

        self.assertEqual(anonymous.status_code, 401)
        self.assertEqual(unscoped.status_code, 403)

    def test_invalid_extension_and_oversized_file_are_rejected(self):
        wrong_type = SimpleUploadedFile("products.xls", b"not-xlsx")
        wrong_response = self.client.post(
            "/api/products/import-excel/", {"file": wrong_type}, format="multipart"
        )
        oversized = SimpleUploadedFile(
            "products.xlsx", b"x" * (MAX_IMPORT_FILE_SIZE + 1)
        )
        oversized_response = self.client.post(
            "/api/products/import-excel/", {"file": oversized}, format="multipart"
        )

        self.assertEqual(wrong_response.status_code, 400)
        self.assertIn(".xlsx", wrong_response.data["detail"])
        self.assertEqual(oversized_response.status_code, 400)
        self.assertIn("5 MB", oversized_response.data["detail"])

    def test_empty_corrupt_and_header_errors_are_rejected(self):
        empty_response = self.client.post(
            "/api/products/import-excel/",
            {"file": SimpleUploadedFile("products.xlsx", b"")},
            format="multipart",
        )
        corrupt_response = self.client.post(
            "/api/products/import-excel/",
            {"file": SimpleUploadedFile("products.xlsx", b"not-a-zip")},
            format="multipart",
        )
        no_rows_response = self._post_rows([])
        missing_headers = tuple(header for header in HEADERS if header != "商品名称")
        missing_header_response = self._post_rows(
            [self._valid_row("PDA-MISSING-HEADER")],
            headers=missing_headers,
        )
        duplicate_header_response = self._post_rows(
            [self._valid_row("PDA-DUP-HEADER")],
            headers=HEADERS + ("货主商品编码",),
        )

        self.assertEqual(empty_response.status_code, 400)
        self.assertIn("为空", empty_response.data["detail"])
        self.assertEqual(corrupt_response.status_code, 400)
        self.assertIn("无法解析", corrupt_response.data["detail"])
        self.assertEqual(no_rows_response.status_code, 400)
        self.assertIn("没有数据行", no_rows_response.data["detail"])
        self.assertEqual(missing_header_response.status_code, 400)
        self.assertIn("v8 模板不一致", missing_header_response.data["detail"])
        self.assertEqual(duplicate_header_response.status_code, 400)
        self.assertIn("表头重复", duplicate_header_response.data["detail"])

    def test_more_than_one_thousand_rows_are_rejected(self):
        rows = [self._valid_row(f"PDA-LIMIT-{index}") for index in range(1001)]

        response = self._post_rows(rows)

        self.assertEqual(response.status_code, 400)
        self.assertIn("最多导入 1000", response.data["detail"])
        self.assertFalse(Product.objects.filter(code__startswith="PDA-LIMIT-").exists())

    @patch(
        "allapp.products.excel_import.ProductExcelImporter._persist",
        side_effect=IntegrityError("duplicate key"),
    )
    def test_concurrent_unique_conflict_returns_409(self, mocked_persist):
        response = self._post_rows([self._valid_row("PDA-CONCURRENT")])

        self.assertTrue(mocked_persist.called)
        self.assertEqual(response.status_code, 409)
        self.assertIn("整批已回滚", response.data["detail"])
        self.assertFalse(Product.objects.filter(code="PDA-CONCURRENT").exists())

    def test_legacy_import_action_uses_same_service(self):
        response = self._post_rows(
            [self._valid_row("", **{"标准贸易条码": "23456789"})],
            url="/products/products/import/",
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertTrue(Product.objects.filter(code="23456789").exists())

    def test_export_owner_scope_permission_and_profile_capability(self):
        OwnerWarehouseBinding.objects.create(owner=self.owner, warehouse=self.warehouse)
        warehouse_client = APIClient()
        warehouse_client.force_authenticate(self.warehouse_denied_user)
        owners = warehouse_client.get("/api/products/export-owners/")
        profile = warehouse_client.get("/api/auth/profile/")

        self.assertEqual(owners.status_code, 200, owners.data)
        self.assertEqual(
            [item["id"] for item in owners.data["results"]], [self.owner.id]
        )
        self.assertTrue(profile.data["capabilities"]["can_export_products"])

        denied_client = APIClient()
        denied_client.force_authenticate(self.no_permission_user)
        denied = denied_client.get("/api/products/export-owners/")
        denied_profile = denied_client.get("/api/auth/profile/")
        self.assertEqual(denied.status_code, 403)
        self.assertFalse(denied_profile.data["capabilities"]["can_export_products"])

    def test_export_rejects_missing_and_out_of_scope_owner(self):
        missing = self.client.get("/api/products/export-excel/")
        forbidden = self.client.get(
            f"/api/products/export-excel/?owner_id={self.other_owner.id}"
        )

        self.assertEqual(missing.status_code, 400)
        self.assertEqual(forbidden.status_code, 403)

    def test_export_v8_round_trip_preserves_prices_and_multiple_packages(self):
        product = Product.objects.create(
            owner=self.owner,
            code="PDA-EXPORT-001",
            name="=档案商品 001",
            category=self.category,
            brand=self.brand,
            base_uom=self.uom,
            gtin="00012345",
            batch_control=False,
            expiry_control=False,
            expiry_basis=None,
            is_active=False,
            price=Decimal("19.90"),
            purchase_price=Decimal("11.25"),
        )
        ProductPackage.objects.create(
            product=product,
            uom=self.uom,
            qty_in_base=1,
            barcode="000001",
            is_pickable=True,
            sort_order=1,
        )
        ProductPackage.objects.create(
            product=product,
            uom=self.carton_uom,
            qty_in_base=12,
            barcode="000012",
            length_cm=10,
            width_cm=20,
            height_cm=30,
            gross_weight_kg="2.500",
            is_purchase_default=True,
            is_sales_default=True,
            sort_order=2,
        )

        exported = self.client.get(
            f"/api/products/export-excel/?owner_id={self.owner.id}"
        )
        self.assertEqual(exported.status_code, 200)
        self.assertIn("filename*=UTF-8", exported["Content-Disposition"])
        workbook = load_workbook(io.BytesIO(exported.content), data_only=False)
        self.assertEqual(workbook["_meta"]["B2"].value, "8")
        self.assertEqual(workbook[IMPORT_SHEET_NAME]["A2"].value, self.owner.code)
        price_column = PRODUCT_HEADERS.index("默认价格") + 1
        purchase_price_column = PRODUCT_HEADERS.index("进价") + 1
        self.assertEqual(
            Decimal(str(workbook[IMPORT_SHEET_NAME].cell(2, price_column).value)),
            Decimal("19.90"),
        )
        self.assertEqual(
            Decimal(
                str(
                    workbook[IMPORT_SHEET_NAME]
                    .cell(2, purchase_price_column)
                    .value
                )
            ),
            Decimal("11.25"),
        )
        name_column = PRODUCT_HEADERS.index("商品名称") + 1
        self.assertEqual(
            workbook[IMPORT_SHEET_NAME].cell(2, name_column).data_type, "s"
        )
        package_codes = {
            workbook[PACKAGE_SHEET_NAME].cell(row, 3).value
            for row in range(2, workbook[PACKAGE_SHEET_NAME].max_row + 1)
            if workbook[PACKAGE_SHEET_NAME].cell(row, 3).value
        }
        self.assertEqual(package_codes, {self.uom.code, self.carton_uom.code})

        product_owner_column = PRODUCT_HEADERS.index("货主编码") + 1
        product_code_column = PRODUCT_HEADERS.index("货主商品编码") + 1
        for row in range(2, workbook[IMPORT_SHEET_NAME].max_row + 1):
            if workbook[IMPORT_SHEET_NAME].cell(row, product_code_column).value:
                workbook[IMPORT_SHEET_NAME].cell(
                    row, product_owner_column, self.other_owner.code
                )
        package_owner_column = PACKAGE_HEADERS.index("货主编码") + 1
        package_code_column = PACKAGE_HEADERS.index("货主商品编码") + 1
        for row in range(2, workbook[PACKAGE_SHEET_NAME].max_row + 1):
            if workbook[PACKAGE_SHEET_NAME].cell(row, package_code_column).value:
                workbook[PACKAGE_SHEET_NAME].cell(
                    row, package_owner_column, self.other_owner.code
                )
        round_trip = io.BytesIO()
        workbook.save(round_trip)
        uploaded = SimpleUploadedFile(
            "round-trip.xlsx",
            round_trip.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.force_authenticate(self.cross_owner_user)
        imported = self.client.post(
            "/api/products/import-excel/", {"file": uploaded}, format="multipart"
        )

        self.assertEqual(imported.status_code, 200, imported.data)
        recreated = Product.objects.get(owner=self.other_owner, code="PDA-EXPORT-001")
        self.assertFalse(recreated.is_active)
        self.assertEqual(recreated.name, "=档案商品 001")
        self.assertEqual(recreated.price, Decimal("19.90"))
        self.assertEqual(recreated.purchase_price, Decimal("11.25"))
        packages = list(recreated.packages.order_by("sort_order"))
        self.assertEqual(len(packages), 2)
        self.assertEqual(packages[1].qty_in_base, 12)
        self.assertTrue(packages[1].is_purchase_default)
        self.assertEqual(product.packages.count(), 2)

    def test_inline_and_package_sheet_data_can_mix_for_different_products(self):
        workbook = Workbook()
        product_sheet = workbook.active
        product_sheet.title = IMPORT_SHEET_NAME
        product_sheet.append(list(HEADERS))
        inline_row = self._valid_row(
            "PDA-MIXED-INLINE",
            **{
                "包装单位编码": self.carton_uom.code,
                "包装换算数量": 12,
            },
        )
        sheet_row = self._valid_row("PDA-MIXED-SHEET")
        product_sheet.append([inline_row.get(header) for header in HEADERS])
        product_sheet.append([sheet_row.get(header) for header in HEADERS])
        package_sheet = workbook.create_sheet(PACKAGE_SHEET_NAME)
        package_sheet.append(list(PACKAGE_HEADERS))
        package_sheet.append(
            [
                self.owner.code,
                "PDA-MIXED-SHEET",
                self.carton_uom.code,
                12,
            ]
        )
        output = io.BytesIO()
        workbook.save(output)

        response = self.client.post(
            "/api/products/import-excel/",
            {"file": SimpleUploadedFile("mixed.xlsx", output.getvalue())},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(
            Product.objects.get(code="PDA-MIXED-INLINE").packages.count(), 1
        )
        self.assertEqual(
            Product.objects.get(code="PDA-MIXED-SHEET").packages.count(), 1
        )

    def test_same_product_cannot_use_inline_and_package_sheet_data(self):
        workbook = Workbook()
        product_sheet = workbook.active
        product_sheet.title = IMPORT_SHEET_NAME
        product_sheet.append(list(HEADERS))
        row = self._valid_row(
            "PDA-MIXED-SAME",
            **{
                "包装单位编码": self.carton_uom.code,
                "包装换算数量": 12,
            },
        )
        product_sheet.append([row.get(header) for header in HEADERS])
        package_sheet = workbook.create_sheet(PACKAGE_SHEET_NAME)
        package_sheet.append(list(PACKAGE_HEADERS))
        package_sheet.append(
            [self.owner.code, "PDA-MIXED-SAME", self.carton_uom.code, 24]
        )
        output = io.BytesIO()
        workbook.save(output)

        response = self.client.post(
            "/api/products/import-excel/",
            {"file": SimpleUploadedFile("mixed-same.xlsx", output.getvalue())},
            format="multipart",
        )

        self.assertEqual(response.status_code, 400, response.data)
        self.assertTrue(
            any(
                error["field"] == "货主商品编码" and "不得同时" in error["message"]
                for error in response.data["errors"]
            )
        )
        self.assertFalse(Product.objects.filter(code="PDA-MIXED-SAME").exists())


@unittest.skipUnless(
    DEPENDENCIES_OK, "缺少 baseinfo/products 依赖模型，跳过 products 导入命令测试"
)
class ProductImportCommandTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(code="PIC", name="Product Import Command")
        self.category = ProductCategory.objects.create(
            code="PIC-CAT", name="命令导入分类"
        )

    def _write_workbook(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "sheet1"
        for row in rows:
            worksheet.append(row)
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook.save(handle.name)
        return handle.name

    def test_import_product_master_sheet_creates_product_and_uom(self):
        path = self._write_workbook(
            [
                ["owner", "货主商品编码", "name", "base_uom", "category"],
                [
                    self.owner.code,
                    "CMD-SKU-1",
                    "命令导入商品",
                    "件",
                    self.category.code,
                ],
            ]
        )

        try:
            call_command("import_product_master_sheet", "--file", path)
        finally:
            os.unlink(path)

        product = Product.objects.get(owner=self.owner, code="CMD-SKU-1")
        self.assertEqual(product.sku, "PIC1")
        self.assertEqual(product.name, "命令导入商品")
        self.assertEqual(product.base_uom.name, "件")

    def test_import_product_master_sheet_dry_run_does_not_persist_product(self):
        path = self._write_workbook(
            [
                ["owner", "货主商品编码", "name", "base_uom", "category"],
                [
                    self.owner.code,
                    "CMD-DRY-1",
                    "Dry Run 商品",
                    "件",
                    self.category.code,
                ],
            ]
        )

        try:
            call_command("import_product_master_sheet", "--file", path, "--dry-run")
        finally:
            os.unlink(path)

        self.assertFalse(
            Product.objects.filter(owner=self.owner, code="CMD-DRY-1").exists()
        )
