from __future__ import annotations

import io
import zipfile
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from allapp.accounts.access import AccessScope
from allapp.accounts.audit import record_audit_event

from .excel_import import (
    MAX_IMPORT_FILE_SIZE,
    MAX_IMPORT_ROWS,
    MAX_XLSX_ENTRIES,
    MAX_XLSX_UNCOMPRESSED_SIZE,
)
from .models import Product, ProductCategory
from .permissions import can_manage_all_owner_products


SHEET_NAME = "商品分类补录"
HEADERS = ("货主编码", "商品编号", "商品名称", "当前分类路径", "目标分类编码")


class CategoryBackfillError(ValueError):
    pass


def allowed_owner_ids(user):
    if can_manage_all_owner_products(user):
        return None
    scope = AccessScope.for_user(user)
    if not scope.is_valid or not scope.single_owner_id:
        return set()
    return {scope.single_owner_id}


def scoped_products(user, queryset=None):
    queryset = queryset if queryset is not None else Product.objects.all()
    owner_ids = allowed_owner_ids(user)
    if owner_ids is None:
        return queryset
    if not owner_ids:
        return queryset.none()
    return queryset.filter(owner_id__in=owner_ids)


def active_categories():
    return [
        category
        for category in ProductCategory.objects.select_related(
            "parent", "parent__parent"
        ).order_by("sort_order", "code")
        if category.has_active_path()
    ]


def build_category_backfill_workbook(products) -> bytes:
    categories = active_categories()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    refs = workbook.create_sheet("分类参考")

    sheet.append(list(HEADERS))
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="0F766E")
    for product in products.select_related("owner", "category__parent__parent"):
        sheet.append(
            [
                product.owner.code,
                product.code,
                product.name,
                product.category.full_path if product.category_id else "未分类",
                product.category.code if product.category_id else "",
            ]
        )
    widths = (18, 22, 34, 38, 22)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{max(sheet.max_row, 1)}"

    refs.append(["分类编码", "分类路径", "层级"])
    for category in categories:
        refs.append([category.code, category.full_path, category.level_name])
    for cell in refs[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="2563EB")
    refs.column_dimensions["A"].width = 22
    refs.column_dimensions["B"].width = 48
    refs.column_dimensions["C"].width = 12

    if categories:
        workbook.defined_names.add(
            DefinedName(
                "CategoryBackfillCodes",
                attr_text=f"'分类参考'!$A$2:$A${len(categories) + 1}",
            )
        )
        validation = DataValidation(
            type="list", formula1="CategoryBackfillCodes", allow_blank=False
        )
        validation.error = "请选择分类参考中的有效分类编码。"
        validation.errorTitle = "无效分类"
        sheet.add_data_validation(validation)
        validation.add(f"E2:E{max(sheet.max_row, 2)}")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load_workbook(uploaded_file):
    name = Path(getattr(uploaded_file, "name", "") or "").name
    if Path(name).suffix.lower() != ".xlsx":
        raise CategoryBackfillError("仅支持 .xlsx 文件。")
    if getattr(uploaded_file, "size", 0) > MAX_IMPORT_FILE_SIZE:
        raise CategoryBackfillError("Excel 文件不能超过 5 MB。")
    data = uploaded_file.read(MAX_IMPORT_FILE_SIZE + 1)
    if not data:
        raise CategoryBackfillError("Excel 文件为空。")
    if len(data) > MAX_IMPORT_FILE_SIZE:
        raise CategoryBackfillError("Excel 文件不能超过 5 MB。")
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ENTRIES:
                raise CategoryBackfillError("Excel 文件结构异常。")
            if sum(row.file_size for row in entries) > MAX_XLSX_UNCOMPRESSED_SIZE:
                raise CategoryBackfillError("Excel 解压后内容过大。")
            if archive.testzip() is not None:
                raise CategoryBackfillError("Excel 文件已损坏。")
    except CategoryBackfillError:
        raise
    except (zipfile.BadZipFile, OSError) as exc:
        raise CategoryBackfillError("无法解析 Excel，请重新下载系统模板。") from exc
    try:
        return load_workbook(io.BytesIO(data), data_only=False)
    except Exception as exc:
        raise CategoryBackfillError(f"无法打开 Excel：{exc}") from exc


def import_category_backfill(uploaded_file, *, user, request=None):
    workbook = _load_workbook(uploaded_file)
    if SHEET_NAME not in workbook.sheetnames:
        raise CategoryBackfillError(f"Excel 中缺少“{SHEET_NAME}”工作表。")
    sheet = workbook[SHEET_NAME]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if tuple(headers[: len(HEADERS)]) != HEADERS:
        raise CategoryBackfillError("表头不匹配，请重新下载分类补录模板。")

    owner_ids = allowed_owner_ids(user)
    categories = {category.code.upper(): category for category in active_categories()}
    updates = []
    errors = []
    seen = set()
    for row_number in range(2, sheet.max_row + 1):
        cells = [sheet.cell(row_number, column) for column in range(1, 6)]
        if not any(str(cell.value or "").strip() for cell in cells):
            continue
        if len(updates) + len(errors) >= MAX_IMPORT_ROWS:
            raise CategoryBackfillError(f"一次最多处理 {MAX_IMPORT_ROWS} 条商品。")
        if any(cell.data_type == "f" for cell in cells):
            errors.append(f"第 {row_number} 行：业务单元格不能使用公式")
            continue
        owner_code = str(cells[0].value or "").strip().upper()
        product_code = str(cells[1].value or "").strip().upper()
        category_code = str(cells[4].value or "").strip().upper()
        key = (owner_code, product_code)
        if not owner_code or not product_code or not category_code:
            errors.append(f"第 {row_number} 行：货主、商品编号和目标分类编码不能为空")
            continue
        if key in seen:
            errors.append(f"第 {row_number} 行：商品重复 {owner_code}/{product_code}")
            continue
        seen.add(key)
        product = (
            Product.objects.select_related("owner", "category")
            .filter(
                owner__code__iexact=owner_code,
                code__iexact=product_code,
            )
            .first()
        )
        if product is None:
            errors.append(f"第 {row_number} 行：找不到商品 {owner_code}/{product_code}")
            continue
        if owner_ids is not None and product.owner_id not in owner_ids:
            errors.append(f"第 {row_number} 行：无权修改货主 {owner_code} 的商品")
            continue
        category = categories.get(category_code)
        if category is None:
            errors.append(
                f"第 {row_number} 行：分类不存在或分类链已停用 {category_code}"
            )
            continue
        updates.append((product, category, row_number))

    if not updates and not errors:
        raise CategoryBackfillError("补录工作表没有数据行。")
    if errors:
        raise CategoryBackfillError("；".join(errors[:20]))

    changed = 0
    with transaction.atomic():
        locked = {
            product.id: product
            for product in Product.objects.select_for_update().filter(
                id__in=[product.id for product, _category, _row in updates]
            )
        }
        for parsed_product, category, _row_number in updates:
            product = locked[parsed_product.id]
            if product.category_id == category.id:
                continue
            product.category = category
            product.updated_by = user
            # This workflow intentionally changes only the category. Historical
            # products may still contain unrelated legacy validation debt, which
            # must not prevent a safe category-only remediation.
            product.save(update_fields=["category", "updated_by", "updated_at"])
            changed += 1
        record_audit_event(
            action="products.category_backfill",
            module="products",
            request=request,
            user=user,
            metadata={
                "filename": Path(getattr(uploaded_file, "name", "") or "").name,
                "row_count": len(updates),
                "changed_count": changed,
                "completed_at": timezone.now().isoformat(),
            },
        )
    return {"row_count": len(updates), "changed_count": changed}
