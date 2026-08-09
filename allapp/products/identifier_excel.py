"""Excel import/export for append-only product identifiers."""

from __future__ import annotations

import io
from datetime import datetime

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .identifier_services import (
    IdentifierConcurrencyError,
    add_external_identifier,
    add_product_barcode,
    set_barcode_primary,
    set_external_primary,
    set_identifier_active,
)
from .models import Product, ProductBarcode, ProductExternalIdentifier, ProductPackage

BARCODE_SHEET = "商品条码维护"
EXTERNAL_SHEET = "外部标识维护"
BARCODE_HEADERS = [
    "操作",
    "记录ID",
    "货主商品编码",
    "条码",
    "条码类型",
    "包装单位编码",
    "基础单位换算快照",
    "是否主码",
    "生效时间",
    "失效时间",
    "启用状态",
]
EXTERNAL_HEADERS = [
    "操作",
    "记录ID",
    "货主商品编码",
    "来源系统",
    "外部系统商品编码",
    "是否主标识",
    "生效时间",
    "失效时间",
    "启用状态",
]


class IdentifierExcelError(ValueError):
    pass


class IdentifierExcelConflictError(IdentifierExcelError):
    pass


def _bool(value, default=False):
    if value in (None, ""):
        return default
    return str(value).strip().upper() in {"1", "Y", "YES", "TRUE", "是", "启用"}


def _datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        parsed = datetime.fromisoformat(str(value).strip())
    return timezone.make_aware(parsed) if timezone.is_naive(parsed) else parsed


def _excel_datetime(value):
    if value is None:
        return None
    return (
        timezone.localtime(value).replace(tzinfo=None)
        if timezone.is_aware(value)
        else value
    )


def build_identifier_template():
    workbook = Workbook()
    barcode = workbook.active
    barcode.title = BARCODE_SHEET
    barcode.append(BARCODE_HEADERS)
    external = workbook.create_sheet(EXTERNAL_SHEET)
    external.append(EXTERNAL_HEADERS)
    for sheet in (barcode, external):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_identifier_export(owner):
    workbook = Workbook()
    barcode_sheet = workbook.active
    barcode_sheet.title = BARCODE_SHEET
    barcode_sheet.append(BARCODE_HEADERS)
    for record in ProductBarcode.all_objects.filter(owner=owner).select_related(
        "product", "package__uom"
    ):
        barcode_sheet.append(
            [
                "",
                record.pk,
                record.product.code,
                record.barcode,
                record.barcode_type,
                record.package.uom.code if record.package_id else "",
                record.qty_in_base,
                record.is_primary,
                _excel_datetime(record.valid_from),
                _excel_datetime(record.valid_to),
                record.is_active,
            ]
        )
    external_sheet = workbook.create_sheet(EXTERNAL_SHEET)
    external_sheet.append(EXTERNAL_HEADERS)
    for record in ProductExternalIdentifier.all_objects.filter(
        owner=owner
    ).select_related("product"):
        external_sheet.append(
            [
                "",
                record.pk,
                record.product.code,
                record.source_system,
                record.external_code,
                record.is_primary,
                _excel_datetime(record.valid_from),
                _excel_datetime(record.valid_to),
                record.is_active,
            ]
        )
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _rows(sheet, headers):
    actual = [cell.value for cell in sheet[1]]
    if actual[: len(headers)] != headers:
        raise IdentifierExcelError(
            f"工作表“{sheet.title}”表头必须为：{'、'.join(headers)}"
        )
    for number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not any(value not in (None, "") for value in values):
            continue
        yield number, dict(zip(headers, values))


@transaction.atomic
def import_identifier_workbook(uploaded_file, *, owner):
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        raise IdentifierExcelError("无法读取标识维护 Excel 文件。") from exc
    missing = {BARCODE_SHEET, EXTERNAL_SHEET} - set(workbook.sheetnames)
    if missing:
        raise IdentifierExcelError("缺少工作表：" + "、".join(sorted(missing)))
    changed = 0
    try:
        for row_number, row in _rows(workbook[BARCODE_SHEET], BARCODE_HEADERS):
            action = str(row["操作"] or "").strip().upper()
            if action == "ADD":
                product = Product.all_objects.get(
                    owner=owner, code=str(row["货主商品编码"]).strip()
                )
                package = None
                if row["包装单位编码"] not in (None, ""):
                    package = ProductPackage.all_objects.get(
                        product=product, uom__code=str(row["包装单位编码"]).strip()
                    )
                add_product_barcode(
                    product=product,
                    barcode=row["条码"],
                    barcode_type=row["条码类型"],
                    package=package,
                    is_primary=_bool(row["是否主码"]),
                    valid_from=_datetime(row["生效时间"]),
                    valid_to=_datetime(row["失效时间"]),
                    is_active=_bool(row["启用状态"], True),
                )
            else:
                record = ProductBarcode.all_objects.get(pk=row["记录ID"], owner=owner)
                if action == "SET_PRIMARY":
                    set_barcode_primary(record)
                elif action == "RETIRE":
                    set_identifier_active(record, False)
                elif action == "REACTIVATE":
                    set_identifier_active(record, True)
                else:
                    raise IdentifierExcelError(
                        "操作必须为 ADD、SET_PRIMARY、RETIRE 或 REACTIVATE。"
                    )
            changed += 1
        for row_number, row in _rows(workbook[EXTERNAL_SHEET], EXTERNAL_HEADERS):
            action = str(row["操作"] or "").strip().upper()
            if action == "ADD":
                product = Product.all_objects.get(
                    owner=owner, code=str(row["货主商品编码"]).strip()
                )
                add_external_identifier(
                    product=product,
                    source_system=row["来源系统"],
                    external_code=row["外部系统商品编码"],
                    is_primary=_bool(row["是否主标识"]),
                    valid_from=_datetime(row["生效时间"]),
                    valid_to=_datetime(row["失效时间"]),
                    is_active=_bool(row["启用状态"], True),
                )
            else:
                record = ProductExternalIdentifier.all_objects.get(
                    pk=row["记录ID"], owner=owner
                )
                if action == "SET_PRIMARY":
                    set_external_primary(record)
                elif action == "RETIRE":
                    set_identifier_active(record, False)
                elif action == "REACTIVATE":
                    set_identifier_active(record, True)
                else:
                    raise IdentifierExcelError(
                        "操作必须为 ADD、SET_PRIMARY、RETIRE 或 REACTIVATE。"
                    )
            changed += 1
    except Exception as exc:
        if isinstance(exc, IdentifierExcelError):
            raise
        messages = getattr(exc, "messages", None)
        detail = "；".join(messages) if messages else str(exc)
        if isinstance(exc, IdentifierConcurrencyError):
            raise IdentifierExcelConflictError(
                f"第 {row_number} 行并发冲突：{detail}"
            ) from exc
        raise IdentifierExcelError(f"第 {row_number} 行处理失败：{detail}") from exc
    return {"changed_count": changed}
