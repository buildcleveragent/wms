from __future__ import annotations

import re
import hashlib
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, Optional, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction

from openpyxl import load_workbook

from allapp.baseinfo.models import Owner
from allapp.products.models import Product, ProductUom


def _norm_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _header_norm(h: Any) -> str:
    s = _norm_str(h)
    s = re.sub(r"\s+", "", s)
    return s


def _as_decimal(x: Any, default: Decimal) -> Decimal:
    if x is None or str(x).strip() == "":
        return default
    try:
        d = Decimal(str(x).strip())
        return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return default


def _as_int(x: Any) -> Optional[int]:
    if x is None or str(x).strip() == "":
        return None
    try:
        # 兼容 excel 里可能是 12.0
        v = int(float(str(x).strip()))
        return v
    except Exception:
        return None


def _as_bool(x: Any) -> Optional[bool]:
    """
    支持：是/否、true/false、1/0、Y/N
    返回 None 表示“未提供”
    """
    if x is None:
        return None
    s = str(x).strip().lower()
    if s == "":
        return None
    if s in ("1", "true", "t", "y", "yes", "是", "启用", "开启"):
        return True
    if s in ("0", "false", "f", "n", "no", "否", "不启用", "关闭"):
        return False
    return None


def _pick_first(row: Dict[str, Any], keys: Tuple[str, ...]) -> Any:
    for k in keys:
        if k in row and row[k] is not None and str(row[k]).strip() != "":
            return row[k]
    return None


# 常见中文单位 -> UOM code（code 必须是字母数字/_/-，不能是中文）
UOM_CODE_MAP = {
    "个": ("EA", "COUNT", 0),
    "件": ("PCS", "COUNT", 0),
    "包": ("PK", "COUNT", 0),
    "袋": ("BAG", "COUNT", 0),
    "箱": ("CTN", "COUNT", 0),
    "托": ("PLT", "COUNT", 0),
    "卷": ("JUAN", "COUNT", 0),
    "瓶": ("BTL", "COUNT", 0),
    "公斤": ("KG", "WEIGHT", 3),
    "千克": ("KG", "WEIGHT", 3),
    "克": ("G", "WEIGHT", 0),
    "升": ("L", "VOLUME", 3),
    "毫升": ("ML", "VOLUME", 0),
}


def _safe_uom_code_from_name(name: str) -> str:
    h = hashlib.md5(name.encode("utf-8")).hexdigest()[:6].upper()
    return f"UOM_{h}"


def _norm_expiry_basis(x: Any) -> Optional[str]:
    """
    支持：MFG/INBOUND、生产日期/入库日期
    """
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s_up = s.upper()
    if s_up in ("MFG", "INBOUND"):
        return s_up
    if s in ("生产日期", "生产", "按生产日期"):
        return "MFG"
    if s in ("入库日期", "入库", "按入库日期"):
        return "INBOUND"
    return None


class Command(BaseCommand):
    help = "从 Excel 批量导入商品到 Product 表（支持软删恢复、可选更新已存在记录）"

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Excel 文件路径，例如 goods2.xlsx 或 /mnt/data/xxx.xlsx")
        parser.add_argument("--sheet", default="Sheet1", help="工作表名称，默认 Sheet1")
        parser.add_argument("--default-price", default="0", help="当表里没有价格列时的默认价格，默认 0")
        parser.add_argument(
            "--update",
            action="store_true",
            help="若 Product 已存在（按 owner+code 匹配），是否更新 name/spec/base_uom/price（不改 code/sku）",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="只校验与打印统计，不写入数据库（但会 full_clean 以发现校验错误）",
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        sheet_name = options["sheet"]
        default_price = _as_decimal(options.get("default_price"), Decimal("0.00"))
        do_update = bool(options["update"])
        dry_run = bool(options["dry_run"])

        try:
            wb = load_workbook(filename=file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"无法打开 Excel：{file_path}，错误：{e}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Excel 中找不到工作表：{sheet_name}，实际有：{wb.sheetnames}")

        ws = wb[sheet_name]

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise CommandError("第一行没有表头")

        headers = [_header_norm(h) for h in header_row]
        if not any(headers):
            raise CommandError("表头为空，无法识别列")

        data_rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None:
                continue
            if all(v is None or str(v).strip() == "" for v in r):
                continue
            row_dict = {}
            for idx, h in enumerate(headers):
                if not h:
                    continue
                if idx < len(r):
                    row_dict[h] = r[idx]
            data_rows.append(row_dict)

        if not data_rows:
            self.stdout.write(self.style.WARNING("没有数据行可导入。"))
            return

        # 你表里的关键列（并支持别名）
        OWNER_KEYS = ("货主", "owner", "货主名称", "货主代码")
        NAME_KEYS = ("商品名称", "name", "品名")
        SPEC_KEYS = ("规格", "spec", "规格型号")
        UOM_KEYS = ("单位", "基本单位", "base_uom", "基础单位")
        CODE_KEYS = ("code", "商品编号", "商品编码")
        SKU_KEYS = ("sku", "SKU", "SKU编码", "sku编码")
        PRICE_KEYS = ("价格", "单价", "售价", "price")

        # 效期相关（可选列；如果都没提供，则自动关闭 expiry_control）
        EXPIRY_CONTROL_KEYS = ("保质期管理", "效期管理", "expiry_control")
        EXPIRY_BASIS_KEYS = ("效期基准", "expiry_basis")
        SHELF_LIFE_KEYS = ("保质期天数", "shelf_life_days", "保质期")
        INBOUND_VALID_KEYS = ("入库有效天数", "inbound_valid_days")
        EXPIRY_WARNING_KEYS = ("预警天数", "expiry_warning_days")
        FEFO_KEYS = ("FEFO", "fefo_required", "先到期先出")

        owner_cache: Dict[str, Owner] = {}
        uom_cache: Dict[str, ProductUom] = {}

        def get_owner(owner_val: str) -> Optional[Owner]:
            key = owner_val.strip()
            if not key:
                return None
            if key in owner_cache:
                return owner_cache[key]

            o = Owner.all_objects.filter(code__iexact=key).first()
            if not o:
                o = Owner.all_objects.filter(name__iexact=key).first()
            if not o:
                o = Owner.all_objects.filter(name__icontains=key).first()

            if o:
                if getattr(o, "is_deleted", False):
                    if not dry_run:
                        o.restore()
                        o.is_active = True
                        o.save(update_fields=["is_active"])
                owner_cache[key] = o
            return o

        def get_uom(uom_val: str) -> Optional[ProductUom]:
            key = uom_val.strip()
            if not key:
                return None
            if key in uom_cache:
                return uom_cache[key]

            u = ProductUom.all_objects.filter(code__iexact=key).first()
            if not u:
                u = ProductUom.all_objects.filter(name__iexact=key).first()

            if u:
                if getattr(u, "is_deleted", False):
                    if not dry_run:
                        u.restore()
                        u.is_active = True
                        u.save(update_fields=["is_active"])
                uom_cache[key] = u
                return u

            # 不存在则创建
            code, kind, dp = None, "COUNT", 0
            if key in UOM_CODE_MAP:
                code, kind, dp = UOM_CODE_MAP[key]
            else:
                code = _safe_uom_code_from_name(key)

            if dry_run:
                return None

            u = ProductUom(code=code, name=key, kind=kind, decimal_places=dp)
            u.full_clean()
            u.save()
            uom_cache[key] = u
            return u

        created = 0
        updated = 0
        restored = 0
        skipped = 0
        errors = 0

        for idx, row in enumerate(data_rows, start=2):
            try:
                owner_val = _norm_str(_pick_first(row, OWNER_KEYS))
                name_val = _norm_str(_pick_first(row, NAME_KEYS))
                spec_val = _norm_str(_pick_first(row, SPEC_KEYS)) or None
                uom_val = _norm_str(_pick_first(row, UOM_KEYS))
                code_val = _norm_str(_pick_first(row, CODE_KEYS))
                sku_val = _norm_str(_pick_first(row, SKU_KEYS)) or code_val

                price_raw = _pick_first(row, PRICE_KEYS)
                price_val = _as_decimal(price_raw, default_price)

                if not owner_val or not name_val or not uom_val or not code_val:
                    raise ValueError(
                        f"必填缺失：货主/商品名称/单位/code 必须有值（货主={owner_val} 名称={name_val} 单位={uom_val} code={code_val}）"
                    )

                owner = get_owner(owner_val)
                if not owner:
                    raise ValueError(f"找不到货主：{owner_val}（请先在后台创建 Owner，或确认 code/name）")

                base_uom = get_uom(uom_val)
                if not base_uom:
                    raise ValueError(f"找不到单位：{uom_val}（dry-run 不创建单位；或请先建 ProductUom）")

                # ====== 关键：效期字段自动决策 ======
                expiry_control_in = _as_bool(_pick_first(row, EXPIRY_CONTROL_KEYS))
                expiry_basis_in = _norm_expiry_basis(_pick_first(row, EXPIRY_BASIS_KEYS))
                shelf_life_days_in = _as_int(_pick_first(row, SHELF_LIFE_KEYS))
                inbound_valid_days_in = _as_int(_pick_first(row, INBOUND_VALID_KEYS))
                expiry_warning_days_in = _as_int(_pick_first(row, EXPIRY_WARNING_KEYS))
                fefo_in = _as_bool(_pick_first(row, FEFO_KEYS))

                # 1) 若明确给了“保质期管理”列：以它为准
                # 2) 否则：如果提供了任何效期天数/基准，则启用；否则自动关闭（满足你“这些商品无保质期”的需求）
                if expiry_control_in is not None:
                    expiry_control = expiry_control_in
                else:
                    expiry_control = bool(
                        (shelf_life_days_in and shelf_life_days_in > 0) or
                        (inbound_valid_days_in and inbound_valid_days_in > 0) or
                        (expiry_basis_in is not None)
                    )

                if not expiry_control:
                    expiry_basis = None
                    shelf_life_days = None
                    inbound_valid_days = None
                    expiry_warning_days = None
                    fefo_required = False
                else:
                    # 没填基准则根据天数推断
                    if expiry_basis_in:
                        expiry_basis = expiry_basis_in
                    elif shelf_life_days_in and shelf_life_days_in > 0:
                        expiry_basis = "MFG"
                    elif inbound_valid_days_in and inbound_valid_days_in > 0:
                        expiry_basis = "INBOUND"
                    else:
                        expiry_basis = "MFG"  # 兜底

                    shelf_life_days = shelf_life_days_in
                    inbound_valid_days = inbound_valid_days_in
                    expiry_warning_days = expiry_warning_days_in
                    fefo_required = True if fefo_in is None else bool(fefo_in)

                # 先用 all_objects 防软删撞唯一
                existing = Product.all_objects.filter(owner=owner, code=code_val).first()

                # sku 冲突检查（同货主下唯一）
                sku_conflict = (
                    Product.all_objects.filter(owner=owner, sku=sku_val)
                    .exclude(code=code_val)
                    .first()
                )
                if sku_conflict and not existing:
                    raise ValueError(f"SKU 冲突：同货主下 sku={sku_val} 已被 code={sku_conflict.code} 使用")

                with transaction.atomic():
                    if existing:
                        if getattr(existing, "is_deleted", False):
                            restored += 1
                            if not dry_run:
                                existing.restore()
                                existing.is_active = True
                                existing.save(update_fields=["is_active"])

                        if do_update:
                            existing.name = name_val
                            existing.spec = spec_val
                            existing.base_uom = base_uom
                            existing.price = price_val

                            # ✅ 如果 Excel 没提供效期信息，则不覆盖现有；如果提供了（或明确给了保质期管理），才更新
                            provided_expiry = (
                                (_pick_first(row, EXPIRY_CONTROL_KEYS) is not None) or
                                (_pick_first(row, EXPIRY_BASIS_KEYS) is not None) or
                                (_pick_first(row, SHELF_LIFE_KEYS) is not None) or
                                (_pick_first(row, INBOUND_VALID_KEYS) is not None) or
                                (_pick_first(row, EXPIRY_WARNING_KEYS) is not None)
                            )
                            if provided_expiry:
                                existing.expiry_control = expiry_control
                                existing.expiry_basis = expiry_basis
                                existing.shelf_life_days = shelf_life_days
                                existing.inbound_valid_days = inbound_valid_days
                                existing.expiry_warning_days = expiry_warning_days
                                existing.fefo_required = fefo_required

                            # ✅ dry-run 也 full_clean，确保提前发现校验问题
                            existing.full_clean()
                            if not dry_run:
                                existing.save()
                            updated += 1
                        else:
                            skipped += 1
                    else:
                        p = Product(
                            owner=owner,
                            code=code_val,
                            name=name_val,
                            spec=spec_val,
                            sku=sku_val,
                            base_uom=base_uom,
                            price=price_val,

                            # ✅ 关键：无保质期商品自动关闭
                            expiry_control=expiry_control,
                            expiry_basis=expiry_basis,
                            shelf_life_days=shelf_life_days,
                            inbound_valid_days=inbound_valid_days,
                            expiry_warning_days=expiry_warning_days,
                            fefo_required=fefo_required,
                        )

                        p.full_clean()   # ✅ dry-run 也校验
                        if not dry_run:
                            p.save()
                        created += 1

            except (ValidationError, IntegrityError, ValueError) as e:
                errors += 1
                self.stdout.write(self.style.ERROR(f"[行 {idx}] 导入失败：{e}"))
            except Exception as e:
                errors += 1
                self.stdout.write(self.style.ERROR(f"[行 {idx}] 未知错误：{e}"))

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("==== 导入完成 ===="))
        self.stdout.write(f"created : {created}")
        self.stdout.write(f"updated : {updated} (需要 --update)")
        self.stdout.write(f"restored: {restored} (软删恢复)")
        self.stdout.write(f"skipped : {skipped}")
        self.stdout.write(f"errors  : {errors}")
        if dry_run:
            self.stdout.write(self.style.WARNING("这是 dry-run：未写入数据库（但已 full_clean 校验）。"))
