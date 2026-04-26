from __future__ import annotations

import ast
import hashlib
import json
import re
from typing import Any, Dict, Optional, Tuple

from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError, transaction

from openpyxl import load_workbook

from allapp.baseinfo.models import Owner
from allapp.products.models import Product, ProductUom


UOM_CODE_MAP = {
    "个": ("GE", "COUNT", 0),
    "件": ("JIAN", "COUNT", 0),
    "包": ("BAO", "COUNT", 0),
    "袋": ("DAI", "COUNT", 0),
    "箱": ("XIANG", "COUNT", 0),
    "托": ("TUO", "COUNT", 0),
    "提": ("TI", "COUNT", 0),
    "盒": ("HE", "COUNT", 0),
    "套": ("TAO", "COUNT", 0),
    "罐": ("GUAN", "COUNT", 0),
    "瓶": ("PING", "COUNT", 0),
    "卷": ("JUAN", "COUNT", 0),
    "公斤": ("KG", "WEIGHT", 3),
    "千克": ("KG", "WEIGHT", 3),
    "克": ("G", "WEIGHT", 0),
    "升": ("L", "VOLUME", 3),
    "毫升": ("ML", "VOLUME", 0),
}


def _norm_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _header_norm(value: Any) -> str:
    s = _norm_str(value)
    s = re.sub(r"\s+", "", s)
    return s.lower()


def _pick_first(row: Dict[str, Any], keys: Tuple[str, ...]) -> Any:
    for k in keys:
        key = _header_norm(k)
        if key in row and row[key] is not None and _norm_str(row[key]) != "":
            return row[key]
    return None


def _safe_uom_code_from_name(name: str) -> str:
    h = hashlib.md5(name.encode("utf-8")).hexdigest()[:8].upper()
    return f"UOM_{h}"


def _as_int(value: Any, field_name: str = "整数") -> Optional[int]:
    if value is None or _norm_str(value) == "":
        return None

    try:
        return int(float(_norm_str(value)))
    except Exception:
        raise ValueError(f"{field_name} 必须是整数，当前值：{value}")


def _as_bool(value: Any, field_name: str = "布尔值") -> Optional[bool]:
    if value is None or _norm_str(value) == "":
        return None

    s = _norm_str(value).lower()

    if s in ("1", "true", "t", "yes", "y", "是", "启用", "开启", "有", "需要"):
        return True

    if s in ("0", "false", "f", "no", "n", "否", "不启用", "关闭", "无", "不需要"):
        return False

    raise ValueError(f"{field_name} 只能填 是/否、true/false、1/0，当前值：{value}")


def _as_gtin(value: Any) -> Optional[str]:
    s = _norm_str(value)

    if not s:
        return None

    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    if not re.fullmatch(r"\d{8}|\d{12}|\d{13}|\d{14}", s):
        raise ValueError(f"gtin 必须是 8/12/13/14 位纯数字，当前值：{value}")

    return s


def _as_json_dict(value: Any) -> dict:
    if value is None or _norm_str(value) == "":
        return {}

    if isinstance(value, dict):
        return value

    s = _norm_str(value)

    try:
        data = json.loads(s)
        if isinstance(data, dict):
            return data
        return {"value": data}
    except Exception:
        pass

    try:
        data = ast.literal_eval(s)
        if isinstance(data, dict):
            return data
        return {"value": data}
    except Exception:
        return {"remark": s}


def _norm_code(value: Any, field_name: str) -> str:
    s = _norm_str(value)
    if not s:
        raise ValueError(f"{field_name} 不能为空")
    return s.upper()


class Command(BaseCommand):
    help = "从 Excel Sheet1 批量导入商品档案到 Product 表"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Excel 文件路径，例如 /root/newinventoryofapril.xlsx",
        )
        parser.add_argument(
            "--sheet",
            default="sheet1",
            help="工作表名称，默认 sheet1，大小写不敏感",
        )
        parser.add_argument(
            "--update",
            action="store_true",
            help="若 owner+code 已存在，是否更新 name/spec/base_uom/extra/expiry_control/shelf_life_days 等非标识字段",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="只校验，不提交数据库。会在事务里模拟创建，再整体回滚。",
        )

    def pick_sheet(self, wb, sheet_name: str):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]

        lower_map = {s.lower(): s for s in wb.sheetnames}
        real_name = lower_map.get(sheet_name.lower())

        if real_name:
            return wb[real_name]

        raise CommandError(
            f"Excel 中找不到工作表：{sheet_name}，实际工作表：{wb.sheetnames}"
        )

    def get_owner(self, value: Any) -> Owner:
        key = _norm_str(value)

        if not key:
            raise ValueError("owner 不能为空")

        owner = None

        if re.fullmatch(r"\d+", key):
            owner = Owner.all_objects.filter(id=int(key)).first()

        if owner is None:
            owner = Owner.all_objects.filter(code__iexact=key).first()

        if owner is None:
            owner = Owner.all_objects.filter(name__iexact=key).first()

        if owner is None:
            owner = Owner.all_objects.filter(name__icontains=key).first()

        if owner is None:
            raise ValueError(f"找不到 owner：{key}。请先在后台创建 Owner，或确认 owner 填的是 id/code/name。")

        if getattr(owner, "is_deleted", False):
            owner.restore()
            owner.is_active = True
            owner.save(update_fields=["is_active"])

        return owner

    def get_or_create_uom(self, value: Any) -> ProductUom:
        name = _norm_str(value)

        if not name:
            raise ValueError("base_uom 不能为空")

        uom = ProductUom.all_objects.filter(code__iexact=name).first()

        if uom is None:
            uom = ProductUom.all_objects.filter(name__iexact=name).first()

        if uom:
            if getattr(uom, "is_deleted", False):
                uom.restore()
                uom.is_active = True
                uom.save(update_fields=["is_active"])

            return uom

        if name in UOM_CODE_MAP:
            code, kind, decimal_places = UOM_CODE_MAP[name]
        else:
            code, kind, decimal_places = _safe_uom_code_from_name(name), "COUNT", 0

        code_exists = ProductUom.all_objects.filter(code__iexact=code).first()
        if code_exists and code_exists.name != name:
            code = _safe_uom_code_from_name(name)

        uom = ProductUom(
            code=code,
            name=name,
            kind=kind,
            decimal_places=decimal_places,
        )
        uom.full_clean()
        uom.save()

        return uom

    def parse_headers(self, ws) -> Dict[str, int]:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)

        if not header_row:
            raise CommandError("第 1 行没有表头")

        headers: Dict[str, int] = {}

        for idx, header in enumerate(header_row):
            key = _header_norm(header)
            if key:
                headers[key] = idx

        if not headers:
            raise CommandError("表头为空，无法识别列")

        return headers

    def row_to_dict(self, headers: Dict[str, int], row_values: Tuple[Any, ...]) -> Dict[str, Any]:
        row: Dict[str, Any] = {}

        for key, idx in headers.items():
            if idx < len(row_values):
                row[key] = row_values[idx]

        return row

    def handle(self, *args, **options):
        file_path = options["file"]
        sheet_name = options["sheet"]
        do_update = bool(options["update"])
        dry_run = bool(options["dry_run"])

        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"无法打开 Excel：{file_path}，错误：{e}")

        ws = self.pick_sheet(wb, sheet_name)
        headers = self.parse_headers(ws)

        OWNER_KEYS = ("owner", "owner_id", "货主id", "货主", "货主名称", "商家")
        GTIN_KEYS = ("gtin", "商品条码", "条码", "barcode", "国际条码")
        NAME_KEYS = ("name", "商品名称", "品名")
        SPEC_KEYS = ("spec", "规格", "箱规", "规格型号")
        BASE_UOM_KEYS = ("base_uom", "基本单位", "基础单位", "单位")
        EXTRA_KEYS = ("extra", "扩展属性", "扩展信息", "备注")
        CODE_KEYS = ("code", "商品编号", "商品编码")
        SKU_KEYS = ("sku", "sku编码", "SKU", "SKU编码")
        EXPIRY_CONTROL_KEYS = ("expiry_control", "保质期管理", "效期管理")
        SHELF_LIFE_DAYS_KEYS = ("shelf_life_days", "保质期天数", "保质期")

        required_any = {
            "owner": OWNER_KEYS,
            "name": NAME_KEYS,
            "base_uom": BASE_UOM_KEYS,
            "code": CODE_KEYS,
            "sku": SKU_KEYS,
        }

        missing = []
        header_keys = set(headers.keys())

        for label, keys in required_any.items():
            if not any(_header_norm(k) in header_keys for k in keys):
                missing.append(label)

        if missing:
            raise CommandError(
                f"缺少必要列：{missing}。当前识别到的表头：{list(headers.keys())}"
            )

        created = 0
        updated = 0
        restored = 0
        skipped = 0
        errors = 0

        with transaction.atomic():
            for row_no, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if values is None:
                    continue

                if all(v is None or _norm_str(v) == "" for v in values):
                    continue

                row = self.row_to_dict(headers, values)

                try:
                    with transaction.atomic():
                        owner_raw = _pick_first(row, OWNER_KEYS)
                        gtin_raw = _pick_first(row, GTIN_KEYS)
                        name_raw = _pick_first(row, NAME_KEYS)
                        spec_raw = _pick_first(row, SPEC_KEYS)
                        base_uom_raw = _pick_first(row, BASE_UOM_KEYS)
                        extra_raw = _pick_first(row, EXTRA_KEYS)
                        code_raw = _pick_first(row, CODE_KEYS)
                        sku_raw = _pick_first(row, SKU_KEYS)
                        expiry_control_raw = _pick_first(row, EXPIRY_CONTROL_KEYS)
                        shelf_life_days_raw = _pick_first(row, SHELF_LIFE_DAYS_KEYS)

                        owner = self.get_owner(owner_raw)
                        gtin_val = _as_gtin(gtin_raw)
                        name_val = _norm_str(name_raw)
                        spec_val = _norm_str(spec_raw) or None
                        base_uom = self.get_or_create_uom(base_uom_raw)
                        extra_val = _as_json_dict(extra_raw)
                        code_val = _norm_code(code_raw, "code")
                        sku_val = _norm_code(sku_raw, "sku")

                        if not name_val:
                            raise ValueError("name 不能为空")

                        shelf_life_days_val = _as_int(
                            shelf_life_days_raw,
                            field_name="shelf_life_days",
                        )

                        expiry_control_in = _as_bool(
                            expiry_control_raw,
                            field_name="expiry_control",
                        )

                        if expiry_control_in is not None:
                            expiry_control_val = expiry_control_in
                        else:
                            expiry_control_val = bool(
                                shelf_life_days_val and shelf_life_days_val > 0
                            )

                        if expiry_control_val:
                            if not shelf_life_days_val or shelf_life_days_val <= 0:
                                raise ValueError(
                                    "expiry_control=True 时，shelf_life_days 必须大于 0"
                                )

                            expiry_basis_val = "MFG"
                            fefo_required_val = True
                        else:
                            expiry_basis_val = None
                            shelf_life_days_val = None
                            fefo_required_val = False

                        existing = Product.all_objects.filter(
                            owner=owner,
                            code=code_val,
                        ).first()

                        sku_conflict_qs = Product.all_objects.filter(
                            owner=owner,
                            sku=sku_val,
                        )

                        if existing:
                            sku_conflict_qs = sku_conflict_qs.exclude(pk=existing.pk)

                        sku_conflict = sku_conflict_qs.first()

                        if sku_conflict:
                            raise ValueError(
                                f"SKU 冲突：同一货主下 sku={sku_val} "
                                f"已被商品 code={sku_conflict.code}, name={sku_conflict.name} 使用"
                            )

                        if gtin_val:
                            gtin_conflict_qs = Product.all_objects.filter(
                                owner=owner,
                                gtin=gtin_val,
                            )

                            if existing:
                                gtin_conflict_qs = gtin_conflict_qs.exclude(pk=existing.pk)

                            gtin_conflict = gtin_conflict_qs.first()

                            if gtin_conflict:
                                raise ValueError(
                                    f"GTIN 冲突：同一货主下 gtin={gtin_val} "
                                    f"已被商品 code={gtin_conflict.code}, name={gtin_conflict.name} 使用"
                                )

                        if existing:
                            if getattr(existing, "is_deleted", False):
                                existing.restore()
                                existing.is_active = True
                                existing.save(update_fields=["is_active"])
                                restored += 1

                            if not do_update:
                                skipped += 1
                                continue

                            if existing.sku and existing.sku != sku_val:
                                raise ValueError(
                                    f"已有商品 code={code_val} 的 sku 是 {existing.sku}，"
                                    f"Excel 中是 {sku_val}。当前模型禁止修改 sku，请新建商品或先处理旧商品。"
                                )

                            if gtin_val and existing.gtin and existing.gtin != gtin_val:
                                raise ValueError(
                                    f"已有商品 code={code_val} 的 gtin 是 {existing.gtin}，"
                                    f"Excel 中是 {gtin_val}。当前模型禁止修改 gtin，请新建商品或先处理旧商品。"
                                )

                            if gtin_val and not existing.gtin:
                                raise ValueError(
                                    f"已有商品 code={code_val} 当前 gtin 为空，但 Excel 提供了 gtin={gtin_val}。"
                                    f"你当前 Product.clean() 禁止修改已有商品的 gtin。"
                                    f"如果确实要补 gtin，需要先调整模型策略或用专门的数据修复脚本。"
                                )

                            existing.name = name_val
                            existing.spec = spec_val
                            existing.base_uom = base_uom
                            existing.extra = extra_val
                            existing.expiry_control = expiry_control_val
                            existing.expiry_basis = expiry_basis_val
                            existing.shelf_life_days = shelf_life_days_val
                            existing.fefo_required = fefo_required_val

                            existing.full_clean()
                            existing.save()

                            updated += 1

                        else:
                            product = Product(
                                owner=owner,
                                gtin=gtin_val,
                                name=name_val,
                                spec=spec_val,
                                base_uom=base_uom,
                                extra=extra_val,
                                code=code_val,
                                sku=sku_val,
                                expiry_control=expiry_control_val,
                                expiry_basis=expiry_basis_val,
                                shelf_life_days=shelf_life_days_val,
                                fefo_required=fefo_required_val,
                            )

                            product.full_clean()
                            product.save()

                            created += 1

                except (ValidationError, IntegrityError, ValueError) as e:
                    errors += 1
                    self.stdout.write(self.style.ERROR(f"[第 {row_no} 行] 导入失败：{e}"))

                except Exception as e:
                    errors += 1
                    self.stdout.write(self.style.ERROR(f"[第 {row_no} 行] 未知错误：{e}"))

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("==== 商品档案导入完成 ===="))
        self.stdout.write(f"created : {created}")
        self.stdout.write(f"updated : {updated}")
        self.stdout.write(f"restored: {restored}")
        self.stdout.write(f"skipped : {skipped}")
        self.stdout.write(f"errors  : {errors}")

        if dry_run:
            self.stdout.write(self.style.WARNING("dry-run 模式：已模拟校验，但没有写入数据库。"))