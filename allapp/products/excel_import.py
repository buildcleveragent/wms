from __future__ import annotations

import io
import re
import uuid
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import IntegrityError, transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from pypinyin import lazy_pinyin
from rest_framework.exceptions import PermissionDenied

from allapp.accounts.access import AccessScope
from allapp.accounts.audit import record_audit_event
from allapp.baseinfo.models import Owner
from allapp.baseinfo.owner_warehouse_access import owner_can_use_warehouse
from allapp.inbound.services import receive_goods_without_order
from allapp.locations.models import Warehouse

from .identifier_services import add_product_barcode
from .models import (
    Brand,
    Product,
    ProductBarcode,
    ProductCategory,
    ProductPackage,
    ProductUom,
    normalize_product_identifier,
)
from .permissions import can_manage_all_owner_products

MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
MAX_XLSX_UNCOMPRESSED_SIZE = 50 * 1024 * 1024
MAX_XLSX_ENTRIES = 300
IMPORT_SHEET_NAME = "商品导入"
PACKAGE_SHEET_NAME = "商品包装"
TEMPLATE_VERSION = "6"


BASE_PRODUCT_HEADERS = (
    "货主编码",
    "货主商品编码",
    "标准贸易条码",
    "商品名称",
    "分类",
    "基本单位",
    "基本单位类型",
    "单位小数位数",
    "基本单位数量",
    "规格",
    "品牌编码",
    "零码",
    "箱码",
    "箱码对应包装单位编码",
    "外部系统商品编码",
    "默认价格",
    "最低价格",
    "最高折扣%",
    "重量kg",
    "体积m³",
    "净含量g",
    "厂家",
    "材质",
    "描述",
    "最低库存",
    "最高库存",
    "序列号管理",
    "批次管理",
    "启用",
    "保质期管理",
    "效期基准",
    "保质期天数",
    "入库有效天数",
    "效期预警天数",
    "FEFO",
)

LEGACY_PACKAGE_HEADERS = (
    "包装单位编码",
    "包装换算数量",
    "包装条码",
    "采购默认",
    "销售默认",
)

# The v5 template exposes the legacy inline package fields again so products whose
# code is generated during persistence can still import one package level.
PRODUCT_HEADERS = BASE_PRODUCT_HEADERS + LEGACY_PACKAGE_HEADERS

# Backwards-compatible public constant used by legacy v2 import callers/tests.
HEADERS = PRODUCT_HEADERS

PACKAGE_HEADERS = (
    "货主编码",
    "货主商品编码",
    "包装单位编码",
    "包装换算数量",
    "包装条码",
    "长cm",
    "宽cm",
    "高cm",
    "毛重kg",
    "体积m³",
    "体积自动计算",
    "可直接拣配",
    "采购默认",
    "销售默认",
    "启用",
    "排序",
)

REQUIRED_PACKAGE_HEADERS = frozenset(
    {"货主编码", "货主商品编码", "包装单位编码", "包装换算数量"}
)

REQUIRED_HEADERS = frozenset(
    {"货主编码", "货主商品编码", "商品名称", "分类", "基本单位"}
)
REQUIRED_VALUE_HEADERS = frozenset({"货主编码", "商品名称", "分类", "基本单位"})
CONDITIONAL_HEADERS = frozenset({"基本单位类型", "单位小数位数"})
TEXT_HEADERS = frozenset(
    {
        "货主编码",
        "货主商品编码",
        "分类",
        "品牌编码",
        "基本单位",
        "标准贸易条码",
        "零码",
        "箱码",
        "箱码对应包装单位编码",
        "外部系统商品编码",
        "包装单位编码",
        "包装条码",
    }
)

PACKAGE_TEXT_HEADERS = frozenset(
    {"货主编码", "货主商品编码", "包装单位编码", "包装条码"}
)

MODEL_FIELD_LABELS = {
    "owner": "货主编码",
    "code": "货主商品编码",
    "sku": "仓库SKU编码",
    "name": "商品名称",
    "spec": "规格",
    "category": "分类",
    "brand": "品牌编码",
    "base_uom": "基本单位",
    "gtin": "标准贸易条码",
    "unit_barcode": "零码",
    "carton_barcode": "箱码",
    "carton_package": "箱码对应包装单位编码",
    "external_code": "外部系统商品编码",
    "price": "默认价格",
    "min_price": "最低价格",
    "max_discount": "最高折扣%",
    "weight": "重量kg",
    "volume": "体积m³",
    "net_content": "净含量g",
    "vender": "厂家",
    "material_quality": "材质",
    "description": "描述",
    "min_stock": "最低库存",
    "max_stock": "最高库存",
    "serial_control": "序列号管理",
    "batch_control": "批次管理",
    "is_active": "启用",
    "expiry_control": "保质期管理",
    "expiry_basis": "效期基准",
    "shelf_life_days": "保质期天数",
    "inbound_valid_days": "入库有效天数",
    "expiry_warning_days": "效期预警天数",
    "fefo_required": "FEFO",
    "uom": "包装单位编码",
    "qty_in_base": "包装换算数量",
    "barcode": "包装条码",
    "is_purchase_default": "采购默认",
    "is_sales_default": "销售默认",
    "length_cm": "长cm",
    "width_cm": "宽cm",
    "height_cm": "高cm",
    "gross_weight_kg": "毛重kg",
    "volume_m3": "体积m³",
    "volume_auto": "体积自动计算",
    "is_pickable": "可直接拣配",
    "sort_order": "排序",
    "__all__": "整行",
}


class ProductImportFileError(ValueError):
    pass


class ProductImportConflictError(ProductImportFileError):
    pass


@dataclass(frozen=True)
class ProductImportAccess:
    allowed_owner_ids: frozenset[int] | None

    def allows_owner(self, owner_id: int) -> bool:
        return self.allowed_owner_ids is None or owner_id in self.allowed_owner_ids


@dataclass
class ParsedProductRow:
    row_number: int
    product: Product
    package_data: dict[str, Any] | None
    carton_barcode: str | None = None
    carton_package_uom_code: str | None = None
    receipt_qty: Decimal = Decimal("0")


@dataclass
class ParsedPackageRow:
    row_number: int
    product_key: tuple[str, str]
    package_data: dict[str, Any]


def resolve_product_import_access(user) -> ProductImportAccess:
    if not user or not getattr(user, "is_authenticated", False):
        raise PermissionDenied("请先登录。")
    if not (
        getattr(user, "is_superuser", False) or user.has_perm("products.add_product")
    ):
        raise PermissionDenied("当前账号没有新增商品权限。")

    scope = AccessScope.for_user(user)
    if getattr(user, "is_superuser", False):
        return ProductImportAccess(allowed_owner_ids=None)
    if not scope.is_valid:
        raise PermissionDenied("当前账号没有有效的角色数据范围。")
    if can_manage_all_owner_products(user):
        return ProductImportAccess(allowed_owner_ids=None)
    if len(scope.owner_ids) == 1:
        owner_id = next(iter(scope.owner_ids))
        return ProductImportAccess(allowed_owner_ids=frozenset({owner_id}))
    raise PermissionDenied("当前账号没有单一货主范围或跨货主管理权限。")


def can_import_products(user) -> bool:
    try:
        resolve_product_import_access(user)
    except PermissionDenied:
        return False
    return True


def product_import_warehouse_queryset(user):
    """Warehouses the current product importer may select for automatic receipt."""

    access = resolve_product_import_access(user)
    queryset = Warehouse.objects.filter(is_active=True).order_by("code", "id")
    if getattr(user, "is_superuser", False):
        return queryset
    scope = AccessScope.for_user(user)
    if scope.warehouse_ids:
        return queryset.filter(pk__in=scope.warehouse_ids)
    if access.allowed_owner_ids is not None:
        return queryset.filter(
            owner_bindings__owner_id__in=access.allowed_owner_ids,
            owner_bindings__is_active=True,
            owner_bindings__is_deleted=False,
        ).distinct()
    return queryset


def can_receive_product_import(user) -> bool:
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if not (
        getattr(user, "is_superuser", False)
        or user.has_perm("accounts.receive_without_order")
    ):
        return False
    try:
        return product_import_warehouse_queryset(user).exists()
    except PermissionDenied:
        return False


def _owner_queryset_for_access(access: ProductImportAccess):
    qs = Owner.objects.filter(is_active=True).order_by("code")
    if access.allowed_owner_ids is not None:
        qs = qs.filter(pk__in=access.allowed_owner_ids)
    return qs


def build_product_import_template(user) -> bytes:
    access = resolve_product_import_access(user)
    owners = list(_owner_queryset_for_access(access).values_list("code", "name"))
    uoms = list(
        ProductUom.objects.filter(is_active=True)
        .order_by("code")
        .values_list("code", "name", "kind", "decimal_places")
    )
    categories = [
        (category.code, category.full_path)
        for category in ProductCategory.objects.filter(is_active=True)
        .select_related("parent__parent")
        .order_by("code")
        if category.has_active_path()
    ]
    brands = list(
        Brand.objects.filter(is_active=True)
        .order_by("code")
        .values_list("code", "name")
    )

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "填写说明"
    data_sheet = workbook.create_sheet(IMPORT_SHEET_NAME)
    package_sheet = workbook.create_sheet(PACKAGE_SHEET_NAME)
    refs = workbook.create_sheet("基础资料")
    meta = workbook.create_sheet("_meta")

    _write_instruction_sheet(instructions)
    _write_import_sheet(data_sheet)
    _write_package_sheet(package_sheet)
    _write_reference_sheet(refs, owners, uoms, categories, brands)
    _write_meta_sheet(meta)
    _add_template_validations(
        workbook, data_sheet, package_sheet, owners, uoms, categories, brands
    )
    meta.sheet_state = "hidden"
    workbook.active = workbook.sheetnames.index(IMPORT_SHEET_NAME)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_instruction_sheet(sheet) -> None:
    sheet.append(["商品批量导入模板", f"模板版本：{TEMPLATE_VERSION}"])
    sheet.append(
        [
            "使用步骤",
            _joined_text(
                "下载模板 → 填写“商品导入”和“商品包装” → ",
                "在 wmspda 上传导入。",
            ),
        ]
    )
    sheet.append(["必填字段", "货主编码、商品名称、分类、基本单位。"])
    sheet.append(
        [
            "分类规则",
            "按完整分类路径或唯一分类名称匹配；不存在的简单名称自动创建为一级分类。",
        ]
    )
    sheet.append(
        [
            "基本单位规则",
            _joined_text(
                "按单位名称匹配；不存在时自动创建，并必须填写基本单位类型和单位小数位数。",
                "新分类和单位代码使用名称的汉语拼音全拼，重码自动追加序号。",
            ),
        ]
    )
    sheet.append(
        [
            "自动收货",
            _joined_text(
                "基本单位数量留空或填 0 时仅建档；填写正数时，上传页必须选择仓库，",
                "系统使用该仓库的默认收货库位立即完成无订单收货。",
            ),
        ]
    )
    sheet.append(
        [
            "货主商品编码规则",
            _joined_text(
                "优先使用填写的货主商品编码；留空时使用标准贸易条码（GTIN）；",
                "两者均留空时，使用系统自动生成的仓库SKU编码。仅含空格视为留空。",
            ),
        ]
    )
    sheet.append(
        [
            "包装必填字段",
            "货主编码、货主商品编码、包装单位编码、包装换算数量；无包装可不填写包装表。",
        ]
    )
    sheet.append(
        ["货主规则", "货主编码必须填写，且只能填写“基础资料”中当前账号有权使用的编码。"]
    )
    sheet.append(
        [
            "仓库SKU编码规则",
            "模板不提供仓库SKU编码列，系统按“货主编码-货主下一个仓库SKU序号”自动生成。",
        ]
    )
    sheet.append(
        [
            "布尔值",
            _joined_text(
                "填写 是/否、1/0、true/false；批次、序列号和保质期管理默认否，",
                "启用默认是。",
            ),
        ]
    )
    sheet.append(
        [
            "效期规则",
            _joined_text(
                "保质期管理留空默认否；", "启用时效期基准可填 MFG/INBOUND。"
            ),  # noqa: E501
        ]
    )
    sheet.append(
        [
            "包装规则",
            _joined_text(
                "多层包装在“商品包装”中每层单独占一行；主表同行包装列可填写一个包装层级；",
                "同一商品不得同时使用两处。无码且无货主商品编码的商品只能使用主表同行包装列。",
            ),
        ]
    )
    sheet.append(
        [
            "箱码绑定",
            _joined_text(
                "填写箱码时必须同时填写“箱码对应包装单位编码”；",
                "该包装单位必须在“商品包装”中属于同一商品且处于启用状态。",
            ),
        ]
    )
    sheet.append(
        [
            "重复规则",
            _joined_text(
                "货主商品编码、条码或外部系统商品编码与已有商品冲突，",
                "或 Excel 内部重复，都会使整批不写入。",
            ),
        ]
    )
    sheet.append(
        [
            "注意",
            _joined_text(
                "编码和条码请按文本填写以保留前导零；",
                "不要在业务单元格使用公式。",
            ),
        ]
    )
    sheet.append([])
    sheet.append(["示例（仅说明，不会导入）"])
    sheet.append(
        [
            "货主编码",
            "货主商品编码",
            "标准贸易条码",
            "商品名称",
            "基本单位",
            "基本单位数量",
            "批次管理",
            "保质期管理",
        ]
    )
    sheet.append(
        ["OWNER-001", "ITEM-001", "06901234567890", "示例商品", "件", 0, "否", "否"]
    )
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="2563EB")
    sheet["B1"].fill = PatternFill("solid", fgColor="2563EB")
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 88
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_import_sheet(sheet) -> None:
    _write_business_sheet(
        sheet,
        PRODUCT_HEADERS,
        required_headers=REQUIRED_VALUE_HEADERS,
        conditional_headers=CONDITIONAL_HEADERS,
        text_headers=TEXT_HEADERS,
    )


def _write_package_sheet(sheet) -> None:
    _write_business_sheet(
        sheet,
        PACKAGE_HEADERS,
        required_headers=REQUIRED_PACKAGE_HEADERS,
        conditional_headers=frozenset(),
        text_headers=PACKAGE_TEXT_HEADERS,
    )


def _write_business_sheet(
    sheet, headers, *, required_headers, conditional_headers, text_headers
) -> None:
    sheet.append(list(headers))
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    required_fill = PatternFill("solid", fgColor="DC2626")
    conditional_fill = PatternFill("solid", fgColor="D97706")
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = (
            required_fill
            if header in required_headers
            else conditional_fill if header in conditional_headers else header_fill
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        width = (
            18
            if header in {"商品名称", "描述"}
            else max(12, min(len(header) * 2 + 4, 18))
        )
        sheet.column_dimensions[get_column_letter(index)].width = width
        if header in text_headers:
            sheet.column_dimensions[get_column_letter(index)].number_format = "@"
            for row_number in range(2, MAX_IMPORT_ROWS + 2):
                sheet.cell(row=row_number, column=index).number_format = "@"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    sheet.row_dimensions[1].height = 32


def _write_reference_sheet(sheet, owners, uoms, categories, brands) -> None:
    simple_uoms = [(code, name) for code, name, _kind, _places in uoms]
    blocks = (
        (1, "货主编码", "货主名称", owners),
        (4, "单位编码", "单位名称", simple_uoms),
        (9, "分类编码", "分类完整路径", categories),
        (12, "品牌编码", "品牌名称", brands),
    )
    for start_col, code_header, name_header, rows in blocks:
        sheet.cell(row=1, column=start_col, value=code_header)
        sheet.cell(row=1, column=start_col + 1, value=name_header)
        for column in (start_col, start_col + 1):
            sheet.cell(row=1, column=column).font = Font(color="FFFFFF", bold=True)
            sheet.cell(row=1, column=column).fill = PatternFill(
                "solid", fgColor="475569"
            )
            sheet.column_dimensions[get_column_letter(column)].width = 22
        for row_number, (code, name) in enumerate(rows, start=2):
            sheet.cell(row=row_number, column=start_col, value=code)
            sheet.cell(row=row_number, column=start_col + 1, value=name)
            sheet.cell(row=row_number, column=start_col).number_format = "@"
    sheet.cell(row=1, column=6, value="单位类型")
    sheet.cell(row=1, column=7, value="小数位数")
    for row_number, (_code, _name, kind, decimal_places) in enumerate(uoms, start=2):
        sheet.cell(
            row=row_number,
            column=6,
            value=ProductUom.Kind(kind).label,
        )
        sheet.cell(row=row_number, column=7, value=decimal_places)
    unique_uom_names = sorted({name for _code, name, _kind, _places in uoms})
    category_paths = [path for _code, path in categories]
    sheet.cell(row=1, column=15, value="基本单位下拉")
    sheet.cell(row=1, column=16, value="分类下拉")
    for row_number, name in enumerate(unique_uom_names, start=2):
        sheet.cell(row=row_number, column=15, value=name)
    for row_number, path in enumerate(category_paths, start=2):
        sheet.cell(row=row_number, column=16, value=path)
    for column in (6, 7, 15, 16):
        sheet.cell(row=1, column=column).font = Font(color="FFFFFF", bold=True)
        sheet.cell(row=1, column=column).fill = PatternFill("solid", fgColor="475569")
        sheet.column_dimensions[get_column_letter(column)].width = 24
    sheet.freeze_panes = "A2"


def _write_meta_sheet(sheet) -> None:
    sheet.append(["schema", "product_import"])
    sheet.append(["version", TEMPLATE_VERSION])
    sheet.append(["max_rows", MAX_IMPORT_ROWS])


def _add_template_validations(
    workbook, data_sheet, package_sheet, owners, uoms, categories, brands
) -> None:
    end_row = MAX_IMPORT_ROWS + 1

    def add_list(sheet, headers, header: str, formula: str) -> None:
        column = get_column_letter(headers.index(header) + 1)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "请选择下拉列表中的有效值。"
        validation.errorTitle = "无效值"
        sheet.add_data_validation(validation)
        validation.add(f"{column}2:{column}{end_row}")

    def add_named_range(name: str, column: str, count: int) -> None:
        workbook.defined_names.add(
            DefinedName(name, attr_text=f"'基础资料'!${column}$2:${column}${count + 1}")
        )

    if owners:
        add_named_range("ProductImportOwnerCodes", "A", len(owners))
        add_list(data_sheet, PRODUCT_HEADERS, "货主编码", "ProductImportOwnerCodes")
        add_list(package_sheet, PACKAGE_HEADERS, "货主编码", "ProductImportOwnerCodes")
    if uoms:
        add_named_range("ProductImportUomCodes", "D", len(uoms))
        unique_uom_count = len({name for _code, name, _kind, _places in uoms})
        add_named_range("ProductImportUomNames", "O", unique_uom_count)
        add_list(data_sheet, PRODUCT_HEADERS, "基本单位", "ProductImportUomNames")
        add_list(data_sheet, PRODUCT_HEADERS, "包装单位编码", "ProductImportUomCodes")
        add_list(
            package_sheet, PACKAGE_HEADERS, "包装单位编码", "ProductImportUomCodes"
        )
    if categories:
        add_named_range("ProductImportCategoryCodes", "I", len(categories))
        add_named_range("ProductImportCategoryNames", "P", len(categories))
        add_list(data_sheet, PRODUCT_HEADERS, "分类", "ProductImportCategoryNames")
    if brands:
        add_named_range("ProductImportBrandCodes", "L", len(brands))
        add_list(data_sheet, PRODUCT_HEADERS, "品牌编码", "ProductImportBrandCodes")
    add_list(
        data_sheet, PRODUCT_HEADERS, "基本单位类型", '"计数,重量,体积,长度,面积,其他"'
    )
    for header in (
        "序列号管理",
        "批次管理",
        "启用",
        "保质期管理",
        "FEFO",
    ):
        add_list(data_sheet, PRODUCT_HEADERS, header, '"是,否"')
    add_list(data_sheet, PRODUCT_HEADERS, "效期基准", '"MFG,INBOUND"')
    for header in ("采购默认", "销售默认"):
        add_list(data_sheet, PRODUCT_HEADERS, header, '"是,否"')
    for header in (
        "体积自动计算",
        "可直接拣配",
        "采购默认",
        "销售默认",
        "启用",
    ):
        add_list(package_sheet, PACKAGE_HEADERS, header, '"是,否"')


class ProductExcelImporter:
    def __init__(self, *, user, request=None, warehouse_id=None):
        self.user = user
        self.request = request
        self.warehouse_id = warehouse_id
        self.warehouse: Warehouse | None = None
        self.access = resolve_product_import_access(user)
        self.errors: list[dict[str, Any]] = []
        self.skipped: list[dict[str, Any]] = []
        self.total_rows = 0
        self._owner_cache: dict[str, Owner | None] = {}
        self._uom_cache: dict[str, ProductUom | None] = {}
        self._category_cache: dict[str, ProductCategory | None] = {}
        self._brand_cache: dict[str, Brand | None] = {}
        self._created_uom_specs: dict[str, tuple[str, int]] = {}
        self.created_categories: list[ProductCategory] = []
        self.created_uoms: list[ProductUom] = []
        self.receipts: list[dict[str, Any]] = []
        self.received_product_count = 0
        self._active_sheet = IMPORT_SHEET_NAME

    def import_file(self, uploaded_file) -> dict[str, Any]:
        workbook = self._load_workbook(uploaded_file)
        try:
            with transaction.atomic():
                self.warehouse = self._resolve_selected_warehouse()
                rows, package_rows = self._parse_workbook(workbook)
                self._validate_receipt_context(rows)
                if self.errors:
                    transaction.set_rollback(True)
                    return self._result(created=[])
                created = self._persist(
                    rows, package_rows, getattr(uploaded_file, "name", "")
                )
        except IntegrityError as exc:
            message = _joined_text(
                "导入期间发生唯一性冲突，整批已回滚；",
                "请重新下载数据或检查并发导入。",
            )
            raise ProductImportConflictError(message) from exc
        except DjangoValidationError as exc:
            raise ProductImportFileError(
                f"导入校验失败，整批已回滚：{_validation_text(exc)}"
            ) from exc
        return self._result(created=created)

    def _resolve_selected_warehouse(self) -> Warehouse | None:
        if self.warehouse_id in (None, ""):
            return None
        try:
            warehouse_id = int(self.warehouse_id)
        except (TypeError, ValueError) as exc:
            raise ProductImportFileError("请选择有效的收货仓库。") from exc
        warehouse = (
            product_import_warehouse_queryset(self.user)
            .select_related("default_receive_location")
            .filter(pk=warehouse_id)
            .first()
        )
        if warehouse is None:
            raise PermissionDenied("当前账号无权使用所选仓库进行商品导入收货。")
        return warehouse

    def _validate_receipt_context(self, rows: list[ParsedProductRow]) -> None:
        receipt_rows = [row for row in rows if row.receipt_qty > 0]
        if not receipt_rows:
            return
        if self.warehouse is None:
            for row in receipt_rows:
                self._error(
                    row.row_number,
                    "基本单位数量",
                    "填写正数时必须在上传页选择收货仓库。",
                )
            return
        if not (
            getattr(self.user, "is_superuser", False)
            or self.user.has_perm("accounts.receive_without_order")
        ):
            for row in receipt_rows:
                self._error(
                    row.row_number, "基本单位数量", "当前账号没有无订单收货权限。"
                )
            return
        location = self.warehouse.default_receive_location
        if location is None:
            for row in receipt_rows:
                self._error(
                    row.row_number, "基本单位数量", "所选仓库尚未配置默认收货库位。"
                )
            return
        if (
            location.warehouse_id != self.warehouse.pk
            or not location.is_active
            or location.is_deleted
            or location.is_disabled
            or location.is_frozen
        ):
            for row in receipt_rows:
                self._error(
                    row.row_number, "基本单位数量", "所选仓库的默认收货库位当前不可用。"
                )
            return
        for row in receipt_rows:
            if not owner_can_use_warehouse(row.product.owner_id, self.warehouse.pk):
                self._error(
                    row.row_number,
                    "基本单位数量",
                    f"货主 {row.product.owner.code} 未授权使用所选仓库。",
                )

    def _load_workbook(self, uploaded_file):
        name = Path(getattr(uploaded_file, "name", "") or "").name
        if Path(name).suffix.lower() != ".xlsx":
            raise ProductImportFileError("仅支持 .xlsx 格式的 Excel 文件。")
        size = getattr(uploaded_file, "size", None)
        if size is not None and size > MAX_IMPORT_FILE_SIZE:
            raise ProductImportFileError("Excel 文件不能超过 5 MB。")
        data = uploaded_file.read(MAX_IMPORT_FILE_SIZE + 1)
        if len(data) > MAX_IMPORT_FILE_SIZE:
            raise ProductImportFileError("Excel 文件不能超过 5 MB。")
        if not data:
            raise ProductImportFileError("Excel 文件为空。")
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                entries = archive.infolist()
                if len(entries) > MAX_XLSX_ENTRIES:
                    raise ProductImportFileError("Excel 文件结构异常。")
                if (
                    sum(entry.file_size for entry in entries)
                    > MAX_XLSX_UNCOMPRESSED_SIZE
                ):
                    raise ProductImportFileError("Excel 解压后内容过大。")
                if archive.testzip() is not None:
                    raise ProductImportFileError("Excel 文件已损坏。")
        except ProductImportFileError:
            raise
        except (zipfile.BadZipFile, OSError) as exc:
            message = "无法解析 Excel 文件，请使用系统模板重新保存。"
            raise ProductImportFileError(message) from exc
        try:
            return load_workbook(io.BytesIO(data), data_only=False, read_only=False)
        except Exception as exc:
            raise ProductImportFileError(f"无法打开 Excel：{exc}") from exc

    def _parse_workbook(
        self, workbook
    ) -> tuple[list[ParsedProductRow], list[ParsedPackageRow]]:
        if IMPORT_SHEET_NAME not in workbook.sheetnames:
            raise ProductImportFileError(f"Excel 中缺少“{IMPORT_SHEET_NAME}”工作表。")
        sheet = workbook[IMPORT_SHEET_NAME]
        header_cells = list(sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        raw_headers = [_text(cell.value) for cell in header_cells]
        nonempty_headers = [header for header in raw_headers if header]
        duplicates = sorted(
            {
                header
                for header in nonempty_headers
                if nonempty_headers.count(header) > 1
            }
        )
        if duplicates:
            raise ProductImportFileError(f"Excel 表头重复：{', '.join(duplicates)}。")
        missing = [
            header for header in REQUIRED_HEADERS if header not in nonempty_headers
        ]
        if missing:
            message = (
                f"Excel 缺少 v6 必要表头：{', '.join(sorted(missing))}。"
                "请重新下载最新商品批量导入模板。"
            )
            raise ProductImportFileError(message)
        unknown = [header for header in nonempty_headers if header not in HEADERS]
        if unknown:
            message = f"Excel 包含不支持的表头：{', '.join(unknown)}。"
            raise ProductImportFileError(message)

        column_by_header = {
            header: index + 1 for index, header in enumerate(raw_headers) if header
        }
        raw_rows: list[tuple[int, dict[str, Any]]] = []
        for row_number in range(2, sheet.max_row + 1):
            values = {
                header: sheet.cell(row_number, column).value
                for header, column in column_by_header.items()
            }
            if not any(_text(value) for value in values.values()):
                continue
            self.total_rows += 1
            if self.total_rows > MAX_IMPORT_ROWS:
                raise ProductImportFileError(f"一次最多导入 {MAX_IMPORT_ROWS} 条商品。")
            formula_header = next(
                (
                    header
                    for header, column in column_by_header.items()
                    if sheet.cell(row_number, column).data_type == "f"
                ),
                None,
            )
            if formula_header:
                self._error(row_number, formula_header, "业务单元格不能使用公式。")
                continue
            raw_rows.append((row_number, values))
        if self.total_rows == 0:
            raise ProductImportFileError("“商品导入”工作表没有数据行。")

        parsed: list[ParsedProductRow] = []
        seen_identifiers: dict[tuple[object, str], dict[str, Any]] = {}
        for row_number, values in raw_rows:
            row = self._parse_row(row_number, values)
            self._check_file_duplicates(row_number, values, seen_identifiers)
            if row is not None:
                parsed.append(row)
        package_rows: list[ParsedPackageRow] = []
        if PACKAGE_SHEET_NAME in workbook.sheetnames:
            package_rows = self._parse_package_sheet(
                workbook[PACKAGE_SHEET_NAME], parsed, seen_identifiers
            )
        package_uoms = {
            (*row.product_key, row.package_data["uom"].code.upper())
            for row in package_rows
            if row.package_data.get("uom") is not None
            and row.package_data.get("is_active", True)
        }
        for row in parsed:
            if (
                row.package_data
                and row.package_data.get("uom") is not None
                and row.package_data.get("is_active", True)
            ):
                package_uoms.add(
                    (
                        row.product.owner.code.upper(),
                        row.product.code.upper(),
                        row.package_data["uom"].code.upper(),
                    )
                )
            if (
                row.carton_barcode
                and (
                    row.product.owner.code.upper(),
                    row.product.code.upper(),
                    row.carton_package_uom_code or "",
                )
                not in package_uoms
            ):
                self._error(
                    row.row_number,
                    "箱码对应包装单位编码",
                    "找不到该商品对应的启用包装层级，请在商品包装表中提供相同包装单位编码。",
                )
        return parsed, package_rows

    def _parse_package_sheet(
        self,
        sheet,
        product_rows: list[ParsedProductRow],
        seen_identifiers: dict[tuple[object, str], dict[str, Any]],
    ) -> list[ParsedPackageRow]:
        self._active_sheet = PACKAGE_SHEET_NAME
        header_cells = list(sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        raw_headers = [_text(cell.value) for cell in header_cells]
        nonempty_headers = [header for header in raw_headers if header]
        duplicates = sorted(
            {
                header
                for header in nonempty_headers
                if nonempty_headers.count(header) > 1
            }
        )
        if duplicates:
            raise ProductImportFileError(
                f"“{PACKAGE_SHEET_NAME}”表头重复：{', '.join(duplicates)}。"
            )
        unknown = [
            header for header in nonempty_headers if header not in PACKAGE_HEADERS
        ]
        if unknown:
            raise ProductImportFileError(
                f"“{PACKAGE_SHEET_NAME}”包含不支持的表头：{', '.join(unknown)}。"
            )
        missing = [
            header
            for header in REQUIRED_PACKAGE_HEADERS
            if header not in nonempty_headers
        ]
        if missing:
            raise ProductImportFileError(
                f"“{PACKAGE_SHEET_NAME}”缺少必要表头：{', '.join(sorted(missing))}。"
            )

        column_by_header = {
            header: index + 1 for index, header in enumerate(raw_headers) if header
        }
        product_by_key = {
            (row.product.owner.code.upper(), row.product.code.upper()): row
            for row in product_rows
            if row.product.code
        }
        generated_code_owners = {
            row.product.owner.code.upper()
            for row in product_rows
            if getattr(row.product, "_derive_code_from_sku_on_create", False)
        }
        parsed: list[ParsedPackageRow] = []
        seen_uoms: dict[tuple[str, str, str], int] = {}
        default_rows: dict[tuple[str, str, str], int] = {}
        for row_number in range(2, sheet.max_row + 1):
            values = {
                header: sheet.cell(row_number, column).value
                for header, column in column_by_header.items()
            }
            if not any(_text(value) for value in values.values()):
                continue
            formula_header = next(
                (
                    header
                    for header, column in column_by_header.items()
                    if sheet.cell(row_number, column).data_type == "f"
                ),
                None,
            )
            if formula_header:
                self._error(
                    row_number,
                    formula_header,
                    "业务单元格不能使用公式。",
                    sheet=PACKAGE_SHEET_NAME,
                )
                continue
            parsed_row = self._parse_package_row(
                row_number,
                values,
                product_by_key,
                seen_uoms,
                seen_identifiers,
                default_rows,
                generated_code_owners,
            )
            if parsed_row is not None:
                parsed.append(parsed_row)
        self._active_sheet = IMPORT_SHEET_NAME
        return parsed

    def _parse_package_row(
        self,
        row_number,
        values,
        product_by_key,
        seen_uoms,
        seen_identifiers,
        default_rows,
        generated_code_owners,
    ) -> ParsedPackageRow | None:
        before_error_count = len(self.errors)
        owner_code = _text(values.get("货主编码")).upper()
        product_code = _text(values.get("货主商品编码")).upper()
        uom_code = _text(values.get("包装单位编码")).upper()
        qty_text = _text(values.get("包装换算数量"))
        for field, value in (
            ("货主编码", owner_code),
            ("货主商品编码", product_code),
            ("包装单位编码", uom_code),
            ("包装换算数量", qty_text),
        ):
            if not value:
                message = "不能为空。"
                if field == "货主商品编码" and owner_code in generated_code_owners:
                    message = (
                        "该货主存在由系统生成货主商品编码的商品，商品包装表无法预先引用；"
                        "请改填商品导入主表同行包装列。"
                    )
                self._error(row_number, field, message, sheet=PACKAGE_SHEET_NAME)
        key = (owner_code, product_code)
        product_row = product_by_key.get(key)
        product = product_row.product if product_row is not None else None
        if owner_code and product_code and product is None:
            self._error(
                row_number,
                "货主商品编码",
                "包装引用的商品必须同时存在于“商品导入”工作表。",
                sheet=PACKAGE_SHEET_NAME,
            )
        elif product_row is not None and product_row.package_data is not None:
            self._error(
                row_number,
                "货主商品编码",
                "该商品已在商品导入主表填写同行包装，不得同时在商品包装表填写。",
                sheet=PACKAGE_SHEET_NAME,
            )
        uom = self._resolve_uom(
            row_number, "包装单位编码", uom_code, required=bool(uom_code)
        )
        qty = self._integer(row_number, "包装换算数量", qty_text, minimum=1)
        barcode = self._optional_text(
            row_number, "包装条码", values.get("包装条码"), max_length=50
        )
        dimensions = {
            field: self._decimal(
                row_number, label, values.get(label), minimum=Decimal("0")
            )
            for field, label in (
                ("length_cm", "长cm"),
                ("width_cm", "宽cm"),
                ("height_cm", "高cm"),
                ("gross_weight_kg", "毛重kg"),
                ("volume_m3", "体积m³"),
            )
        }
        sort_order = self._integer(row_number, "排序", values.get("排序"), minimum=0)
        package_data = {
            "uom": uom,
            "qty_in_base": qty,
            "barcode": barcode,
            **dimensions,
            "volume_auto": self._boolean(
                row_number, "体积自动计算", values.get("体积自动计算"), True
            ),
            "is_pickable": self._boolean(
                row_number, "可直接拣配", values.get("可直接拣配"), False
            ),
            "is_purchase_default": self._nullable_boolean(
                row_number, "采购默认", values.get("采购默认")
            ),
            "is_sales_default": self._nullable_boolean(
                row_number, "销售默认", values.get("销售默认")
            ),
            "is_active": self._boolean(row_number, "启用", values.get("启用"), True),
            "sort_order": 0 if sort_order is None else sort_order,
        }
        dims = [dimensions[name] for name in ("length_cm", "width_cm", "height_cm")]
        if any(value is not None for value in dims) and not all(
            value is not None and value > 0 for value in dims
        ):
            self._error(
                row_number,
                "长cm/宽cm/高cm",
                "长、宽、高必须同时填写且均大于 0。",
                sheet=PACKAGE_SHEET_NAME,
            )
        if product is not None and uom is not None and qty is not None:
            if uom.pk == product.base_uom_id and qty != 1:
                self._error(
                    row_number,
                    "包装换算数量",
                    "包装单位与基本单位相同时，换算数量必须为 1。",
                    sheet=PACKAGE_SHEET_NAME,
                )
            uom_key = (*key, uom.code.upper())
            if uom_key in seen_uoms:
                self._error(
                    row_number,
                    "包装单位编码",
                    f"与第 {seen_uoms[uom_key]} 行重复。",
                    sheet=PACKAGE_SHEET_NAME,
                )
            else:
                seen_uoms[uom_key] = row_number
            if barcode:
                normalized = normalize_product_identifier(barcode)
                owner = self._owner_cache.get(owner_code)
                owner_key: object = owner.pk if owner is not None else owner_code
                barcode_key = (owner_key, normalized)
                previous = seen_identifiers.get(barcode_key)
                if previous:
                    self._error(
                        row_number,
                        "包装条码",
                        f"标识“{normalized}”与“{previous['sheet']}”第 "
                        f"{previous['row']} 行的{previous['field']}冲突。",
                        sheet=PACKAGE_SHEET_NAME,
                    )
                else:
                    seen_identifiers[barcode_key] = {
                        "sheet": PACKAGE_SHEET_NAME,
                        "row": row_number,
                        "field": "包装条码",
                        "exclusive": True,
                    }
            for field, flag in (
                ("采购默认", package_data["is_purchase_default"]),
                ("销售默认", package_data["is_sales_default"]),
            ):
                default_key = (*key, field)
                if flag is True and default_key in default_rows:
                    self._error(
                        row_number,
                        field,
                        f"同一商品只能设置一个默认包装；与第 {default_rows[default_key]} 行冲突。",
                        sheet=PACKAGE_SHEET_NAME,
                    )
                elif flag is True:
                    default_rows[default_key] = row_number
        if len(self.errors) == before_error_count and product is not None:
            package = ProductPackage(product=product, **package_data)
            try:
                package.full_clean(exclude={"product"}, validate_constraints=False)
            except DjangoValidationError as exc:
                self._add_validation_errors(row_number, exc, sheet=PACKAGE_SHEET_NAME)
        if len(self.errors) != before_error_count or product is None:
            return None
        return ParsedPackageRow(
            row_number=row_number, product_key=key, package_data=package_data
        )

    def _parse_row(
        self, row_number: int, values: dict[str, Any]
    ) -> ParsedProductRow | None:
        before_error_count = len(self.errors)
        owner = self._resolve_owner(row_number, values.get("货主编码"))
        supplied_code = self._optional_code(
            row_number,
            "货主商品编码",
            values.get("货主商品编码"),
            max_length=50,
        )
        gtin = self._optional_text(
            row_number, "标准贸易条码", values.get("标准贸易条码"), max_length=20
        )
        code = supplied_code or gtin or ""
        derive_code_from_sku = not supplied_code and not _text(
            values.get("标准贸易条码")
        )
        if owner is None:
            return None

        existing = (
            Product.all_objects.filter(owner=owner, code__iexact=code).first()
            if code
            else None
        )
        if existing:
            message = (
                "货主商品编码命中已软删除商品，请恢复旧商品或更换编号；整批不会写入。"
                if existing.is_deleted
                else "货主商品编码已存在；整批不会写入。"
            )
            self._error(row_number, "货主商品编码", message)
            return None

        name = self._required_text(
            row_number,
            "商品名称",
            values.get("商品名称"),
            max_length=200,
        )
        base_uom = self._resolve_base_uom(
            row_number,
            values.get("基本单位"),
            values.get("基本单位类型"),
            values.get("单位小数位数"),
        )
        if not name or base_uom is None:
            return None

        category = self._resolve_category(row_number, values.get("分类"))
        brand = self._resolve_brand(row_number, values.get("品牌编码"))
        package_uom_value = values.get("包装单位编码")
        package_uom = self._resolve_uom(
            row_number,
            "包装单位编码",
            package_uom_value,
            required=False,
        )

        carton_barcode = self._optional_text(
            row_number, "箱码", values.get("箱码"), max_length=50
        )
        carton_package_uom_code = (
            _text(values.get("箱码对应包装单位编码")).upper() or None
        )
        if bool(carton_barcode) != bool(carton_package_uom_code):
            self._error(
                row_number,
                "箱码对应包装单位编码",
                "箱码和箱码对应包装单位编码必须同时填写。",
            )

        product = Product(
            owner=owner,
            code=code,
            sku="",
            name=name,
            spec=self._optional_text(
                row_number, "规格", values.get("规格"), max_length=200
            ),
            category=category,
            brand=brand,
            base_uom=base_uom,
            gtin=gtin,
            unit_barcode=self._optional_text(
                row_number,
                "零码",
                values.get("零码"),
                max_length=50,
            ),
            carton_barcode=None,
            external_code=self._optional_text(
                row_number,
                "外部系统商品编码",
                values.get("外部系统商品编码"),
                max_length=50,
            ),
            price=self._decimal(
                row_number, "默认价格", values.get("默认价格"), minimum=Decimal("0")
            ),
            min_price=self._decimal(
                row_number, "最低价格", values.get("最低价格"), minimum=Decimal("0")
            ),
            max_discount=self._decimal(
                row_number,
                "最高折扣%",
                values.get("最高折扣%"),
                minimum=Decimal("0"),
                maximum=Decimal("100"),
            ),
            weight=self._decimal(
                row_number,
                "重量kg",
                values.get("重量kg"),
                minimum=Decimal("0"),
            ),
            volume=self._decimal(
                row_number,
                "体积m³",
                values.get("体积m³"),
                minimum=Decimal("0"),
            ),
            net_content=self._decimal(
                row_number, "净含量g", values.get("净含量g"), minimum=Decimal("0")
            ),
            vender=self._optional_text(
                row_number, "厂家", values.get("厂家"), max_length=50
            ),
            material_quality=self._optional_text(
                row_number, "材质", values.get("材质"), max_length=20
            ),
            description=self._optional_text(row_number, "描述", values.get("描述")),
            min_stock=self._decimal(
                row_number, "最低库存", values.get("最低库存"), minimum=Decimal("0")
            ),
            max_stock=self._decimal(
                row_number, "最高库存", values.get("最高库存"), minimum=Decimal("0")
            ),
            serial_control=self._boolean(
                row_number,
                "序列号管理",
                values.get("序列号管理"),
                False,
            ),
            batch_control=self._boolean(
                row_number,
                "批次管理",
                values.get("批次管理"),
                False,
            ),
            is_active=self._boolean(row_number, "启用", values.get("启用"), True),
            created_by=self.user,
            updated_by=self.user,
        )
        if derive_code_from_sku:
            product._derive_code_from_sku_on_create = True
        expiry_control = self._boolean(
            row_number,
            "保质期管理",
            values.get("保质期管理"),
            False,
        )
        product.expiry_control = expiry_control
        if expiry_control:
            basis = (_text(values.get("效期基准")) or "MFG").upper()
            if basis not in {"MFG", "INBOUND"}:
                self._error(row_number, "效期基准", "只能填写 MFG 或 INBOUND。")
            product.expiry_basis = basis
            product.shelf_life_days = self._integer(
                row_number, "保质期天数", values.get("保质期天数"), minimum=1
            )
            product.inbound_valid_days = self._integer(
                row_number, "入库有效天数", values.get("入库有效天数"), minimum=1
            )
            product.expiry_warning_days = self._integer(
                row_number, "效期预警天数", values.get("效期预警天数"), minimum=1
            )
            product.fefo_required = self._boolean(
                row_number, "FEFO", values.get("FEFO"), True
            )
        else:
            product.expiry_basis = None
            product.shelf_life_days = None
            product.inbound_valid_days = None
            product.expiry_warning_days = None
            product.fefo_required = False

        receipt_qty = self._receipt_quantity(
            row_number, values.get("基本单位数量"), base_uom
        )
        if receipt_qty > 0:
            if not product.is_active:
                self._error(row_number, "基本单位数量", "自动收货的商品必须启用。")
            if (
                product.serial_control
                or product.batch_control
                or product.expiry_control
            ):
                self._error(
                    row_number,
                    "基本单位数量",
                    "序列号、批次或保质期管理商品不能在商品导入中自动收货，请使用专用收货流程。",
                )

        package_qty_value = values.get("包装换算数量")
        package_qty = self._integer(
            row_number,
            "包装换算数量",
            package_qty_value,
            minimum=1,
        )
        package_barcode = self._optional_text(
            row_number, "包装条码", values.get("包装条码"), max_length=50
        )
        package_data = None
        package_requested = (
            bool(_text(package_uom_value))
            or bool(_text(package_qty_value))
            or package_barcode is not None
            or bool(_text(values.get("采购默认")))
            or bool(_text(values.get("销售默认")))
        )
        if package_requested:
            if package_uom is None and not _text(package_uom_value):
                self._error(
                    row_number,
                    "包装单位编码",
                    "填写包装信息时包装单位不能为空。",
                )
            if package_qty is None and not _text(package_qty_value):
                self._error(
                    row_number,
                    "包装换算数量",
                    "填写包装信息时换算数量不能为空。",
                )
            if package_uom is not None and package_qty is not None:
                if package_uom.pk == base_uom.pk and package_qty != 1:
                    self._error(
                        row_number,
                        "包装换算数量",
                        "包装单位与基本单位相同时，换算数量必须为 1。",
                    )
                package_data = {
                    "uom": package_uom,
                    "qty_in_base": package_qty,
                    "barcode": package_barcode,
                    "is_pickable": True,
                    "is_purchase_default": self._nullable_boolean(
                        row_number, "采购默认", values.get("采购默认")
                    ),
                    "is_sales_default": self._nullable_boolean(
                        row_number, "销售默认", values.get("销售默认")
                    ),
                }

        if len(self.errors) == before_error_count:
            try:
                product.full_clean(
                    exclude={"code", "sku"} if derive_code_from_sku else None
                )
            except DjangoValidationError as exc:
                self._add_validation_errors(row_number, exc)
        if len(self.errors) == before_error_count and package_data:
            package = ProductPackage(product=product, **package_data)
            try:
                package.full_clean(
                    exclude={"product"},
                )
            except DjangoValidationError as exc:
                self._add_validation_errors(row_number, exc)
        if len(self.errors) != before_error_count:
            return None
        return ParsedProductRow(
            row_number=row_number,
            product=product,
            package_data=package_data,
            carton_barcode=carton_barcode,
            carton_package_uom_code=carton_package_uom_code,
            receipt_qty=receipt_qty,
        )

    def _resolve_owner(self, row_number: int, value) -> Owner | None:
        code = _text(value).upper()
        if not code:
            self._error(row_number, "货主编码", "不能为空。")
            return None
        if code not in self._owner_cache:
            self._owner_cache[code] = Owner.objects.filter(
                code__iexact=code, is_active=True
            ).first()
        owner = self._owner_cache[code]
        if owner is None:
            self._error(row_number, "货主编码", f"找不到启用的货主：{code}。")
            return None
        if not self.access.allows_owner(owner.pk):
            self._error(row_number, "货主编码", "当前账号无权为该货主导入商品。")
            return None
        return owner

    def _resolve_uom(self, row_number, field, value, *, required):
        code = _text(value).upper()
        if not code:
            if required:
                self._error(row_number, field, "不能为空。")
            return None
        if code not in self._uom_cache:
            self._uom_cache[code] = ProductUom.objects.filter(
                code__iexact=code, is_active=True
            ).first()
        uom = self._uom_cache[code]
        if uom is None:
            self._error(row_number, field, f"找不到启用的单位：{code}。")
        return uom

    def _resolve_base_uom(self, row_number, value, kind_value, decimal_places_value):
        label = _text(value)
        if not label:
            self._error(row_number, "基本单位", "不能为空。")
            return None
        cache_key = label.casefold()
        if cache_key in self._uom_cache:
            uom = self._uom_cache[cache_key]
            if uom is not None and cache_key in self._created_uom_specs:
                spec = self._parse_new_uom_spec(
                    row_number, kind_value, decimal_places_value
                )
                if spec and spec != self._created_uom_specs[cache_key]:
                    self._error(
                        row_number,
                        "基本单位",
                        "同一批次中新建的同名单位，其类型和小数位数必须一致。",
                    )
            return uom

        active = list(
            ProductUom.objects.filter(name__iexact=label, is_active=True).order_by("id")
        )
        if not active:
            by_code = ProductUom.objects.filter(
                code__iexact=label, is_active=True
            ).first()
            if by_code is not None:
                active = [by_code]
        if len(active) > 1:
            self._error(
                row_number,
                "基本单位",
                f"单位名称“{label}”对应多个单位，请改填单位编码。",
            )
            self._uom_cache[cache_key] = None
            return None
        if active:
            self._uom_cache[cache_key] = active[0]
            return active[0]

        occupied = ProductUom.all_objects.filter(name__iexact=label).first()
        if occupied is not None:
            state = "已删除" if occupied.is_deleted else "已停用"
            self._error(
                row_number,
                "基本单位",
                f"同名单位“{occupied.name}”（编码 {occupied.code}）已{state}，请先恢复或启用。",
            )
            self._uom_cache[cache_key] = None
            return None
        spec = self._parse_new_uom_spec(row_number, kind_value, decimal_places_value)
        if spec is None:
            self._uom_cache[cache_key] = None
            return None
        kind, decimal_places = spec
        uom, created = self._create_uom(label, kind, decimal_places)
        self._uom_cache[cache_key] = uom
        self._uom_cache[uom.code.upper()] = uom
        self._created_uom_specs[cache_key] = spec
        if created:
            self.created_uoms.append(uom)
        return uom

    def _parse_new_uom_spec(self, row_number, kind_value, decimal_places_value):
        kind_text = _text(kind_value)
        kind_map = {choice.label.casefold(): choice.value for choice in ProductUom.Kind}
        kind_map.update(
            {choice.value.casefold(): choice.value for choice in ProductUom.Kind}
        )
        kind = kind_map.get(kind_text.casefold()) if kind_text else None
        if kind is None:
            self._error(
                row_number,
                "基本单位类型",
                "新单位必须填写：计数、重量、体积、长度、面积或其他。",
            )
        decimal_places = self._integer(
            row_number, "单位小数位数", decimal_places_value, minimum=0
        )
        if not _text(decimal_places_value):
            self._error(row_number, "单位小数位数", "新单位必须填写 0–6。")
        elif decimal_places is not None and decimal_places > 6:
            self._error(row_number, "单位小数位数", "不能大于 6。")
        if kind is None or decimal_places is None or decimal_places > 6:
            return None
        return kind, decimal_places

    def _create_uom(
        self, name: str, kind: str, decimal_places: int
    ) -> tuple[ProductUom, bool]:
        base_code = _pinyin_code(name)
        if not base_code:
            raise DjangoValidationError({"code": "单位名称无法生成有效拼音代码。"})
        for candidate in _code_candidates(base_code):
            if ProductUom.all_objects.filter(code__iexact=candidate).exists():
                continue
            try:
                with transaction.atomic():
                    return (
                        ProductUom.objects.create(
                            code=candidate,
                            name=name,
                            kind=kind,
                            decimal_places=decimal_places,
                            created_by=self.user,
                            updated_by=self.user,
                        ),
                        True,
                    )
            except IntegrityError:
                concurrent = ProductUom.objects.filter(
                    name__iexact=name, is_active=True
                ).first()
                if concurrent is not None:
                    if (
                        concurrent.kind != kind
                        or concurrent.decimal_places != decimal_places
                    ):
                        raise ProductImportConflictError(
                            f"单位“{name}”已被并发创建，但类型或小数位数不一致。"
                        )
                    return concurrent, False
                continue
        raise ProductImportConflictError("无法为基本单位生成唯一拼音代码。")

    def _resolve_category(self, row_number, value):
        label = _text(value)
        if not label:
            self._error(row_number, "分类", "不能为空；新商品至少需要选择一个大类。")
            return None
        cache_key = label.casefold()
        if cache_key in self._category_cache:
            return self._category_cache[cache_key]
        categories = list(
            ProductCategory.objects.filter(is_active=True)
            .select_related("parent__parent")
            .order_by("id")
        )
        path_matches = [
            category
            for category in categories
            if category.has_active_path() and category.full_path.casefold() == cache_key
        ]
        if len(path_matches) == 1:
            self._category_cache[cache_key] = path_matches[0]
            return path_matches[0]
        code_match = next(
            (
                category
                for category in categories
                if category.has_active_path() and category.code.casefold() == cache_key
            ),
            None,
        )
        if code_match is not None:
            self._category_cache[cache_key] = code_match
            return code_match
        name_matches = [
            category
            for category in categories
            if category.has_active_path() and category.name.casefold() == cache_key
        ]
        if len(name_matches) == 1:
            self._category_cache[cache_key] = name_matches[0]
            return name_matches[0]
        if len(name_matches) > 1:
            self._error(
                row_number, "分类", f"分类名称“{label}”不唯一，请填写完整分类路径。"
            )
            self._category_cache[cache_key] = None
            return None
        if ">" in label:
            self._error(
                row_number, "分类", f"找不到分类路径“{label}”，不会自动创建多级分类。"
            )
            self._category_cache[cache_key] = None
            return None
        occupied = ProductCategory.all_objects.filter(name__iexact=label).first()
        if occupied is not None:
            state = "已删除" if occupied.is_deleted else "已停用"
            self._error(
                row_number,
                "分类",
                f"同名分类“{occupied.name}”（编码 {occupied.code}）已{state}，请先恢复或启用。",
            )
            self._category_cache[cache_key] = None
            return None
        category, created = self._create_category(label)
        self._category_cache[cache_key] = category
        self._category_cache[category.code.casefold()] = category
        if created:
            self.created_categories.append(category)
        return category

    def _create_category(self, name: str) -> tuple[ProductCategory, bool]:
        base_code = _pinyin_code(name)
        if not base_code:
            raise DjangoValidationError({"code": "分类名称无法生成有效拼音代码。"})
        for candidate in _code_candidates(base_code):
            if ProductCategory.all_objects.filter(code__iexact=candidate).exists():
                continue
            try:
                with transaction.atomic():
                    return (
                        ProductCategory.objects.create(
                            code=candidate,
                            name=name,
                            parent=None,
                            sort_order=0,
                            created_by=self.user,
                            updated_by=self.user,
                        ),
                        True,
                    )
            except IntegrityError:
                concurrent = ProductCategory.objects.filter(
                    name__iexact=name, parent__isnull=True, is_active=True
                ).first()
                if concurrent is not None:
                    return concurrent, False
                continue
        raise ProductImportConflictError("无法为分类生成唯一拼音代码。")

    def _receipt_quantity(self, row_number, value, base_uom) -> Decimal:
        quantity = self._decimal(
            row_number, "基本单位数量", value, minimum=Decimal("0")
        )
        if quantity is None:
            return Decimal("0")
        sign, digits, exponent = quantity.as_tuple()
        decimal_places = max(0, -exponent)
        integer_digits = max(0, len(digits) + exponent)
        if decimal_places > 3 or integer_digits + decimal_places > 18:
            self._error(row_number, "基本单位数量", "最多允许 18 位数字和 3 位小数。")
        if base_uom is not None and decimal_places > base_uom.decimal_places:
            self._error(
                row_number,
                "基本单位数量",
                f"不能超过基本单位允许的 {base_uom.decimal_places} 位小数。",
            )
        return quantity

    def _resolve_brand(self, row_number, value):
        code = _text(value).upper()
        if not code:
            return None
        if code not in self._brand_cache:
            self._brand_cache[code] = Brand.objects.filter(
                code__iexact=code, is_active=True
            ).first()
        brand = self._brand_cache[code]
        if brand is None:
            self._error(row_number, "品牌编码", f"找不到启用的品牌：{code}。")
        return brand

    def _required_code(self, row, field, value, *, max_length):
        code = self._optional_code(row, field, value, max_length=max_length)
        if not code:
            self._error(row, field, "不能为空。")
        return code

    def _optional_code(self, row, field, value, *, max_length):
        result = _text(value).upper()
        if len(result) > max_length:
            self._error(row, field, f"长度不能超过 {max_length} 个字符。")
            return None
        return result or None

    def _required_text(self, row, field, value, *, max_length):
        result = self._optional_text(row, field, value, max_length=max_length)
        if not result:
            self._error(row, field, "不能为空。")
        return result

    def _optional_text(self, row, field, value, *, max_length=None):
        result = _text(value)
        if max_length is not None and len(result) > max_length:
            self._error(row, field, f"长度不能超过 {max_length} 个字符。")
            return None
        return result or None

    def _decimal(self, row, field, value, *, minimum=None, maximum=None):
        text = _text(value)
        if not text:
            return None
        try:
            result = Decimal(text)
        except (InvalidOperation, ValueError):
            self._error(row, field, "必须是有效数字。")
            return None
        if not result.is_finite():
            self._error(row, field, "必须是有限数字。")
            return None
        if minimum is not None and result < minimum:
            self._error(row, field, f"不能小于 {minimum}。")
        if maximum is not None and result > maximum:
            self._error(row, field, f"不能大于 {maximum}。")
        return result

    def _integer(self, row, field, value, *, minimum=None):
        text = _text(value)
        if not text:
            return None
        try:
            decimal_value = Decimal(text)
            if decimal_value != decimal_value.to_integral_value():
                raise InvalidOperation
            result = int(decimal_value)
        except (InvalidOperation, ValueError, OverflowError):
            self._error(row, field, "必须是整数。")
            return None
        if minimum is not None and result < minimum:
            self._error(row, field, f"不能小于 {minimum}。")
        return result

    def _boolean(self, row, field, value, default):
        result = self._nullable_boolean(row, field, value)
        return default if result is None else result

    def _nullable_boolean(self, row, field, value):
        text = _text(value).lower()
        if not text:
            return None
        if text in {"是", "1", "true", "t", "yes", "y", "启用", "开启"}:
            return True
        if text in {"否", "0", "false", "f", "no", "n", "停用", "关闭"}:
            return False
        self._error(row, field, "只能填写 是/否、1/0 或 true/false。")
        return None

    def _check_file_duplicates(self, row_number: int, row_values, seen) -> None:
        owner_code = _text(row_values.get("货主编码")).upper()
        owner = self._owner_cache.get(owner_code) if owner_code else None
        owner_key: object = (
            owner.pk if owner is not None else owner_code or "missing-owner"
        )
        code = _text(row_values.get("货主商品编码")).upper()
        identifiers = (
            ("货主商品编码", code, False),
            ("标准贸易条码", _text(row_values.get("标准贸易条码")), False),
            ("零码", _text(row_values.get("零码")), False),
            ("箱码", _text(row_values.get("箱码")), False),
            (
                "外部系统商品编码",
                _text(row_values.get("外部系统商品编码")),
                False,
            ),
            ("包装条码", _text(row_values.get("包装条码")), True),
        )
        for field, value, exclusive in identifiers:
            if not value:
                continue
            normalized = normalize_product_identifier(value)
            key = (owner_key, normalized)
            previous = seen.get(key)
            conflicts = previous and (
                exclusive
                or previous["exclusive"]
                or previous["row"] != row_number
                or previous["sheet"] != IMPORT_SHEET_NAME
            )
            if conflicts:
                self._error(
                    row_number,
                    field,
                    f"标识“{normalized}”与“{previous['sheet']}”第 "
                    f"{previous['row']} 行的{previous['field']}冲突。",
                )
            elif previous is None:
                seen[key] = {
                    "sheet": IMPORT_SHEET_NAME,
                    "row": row_number,
                    "field": field,
                    "exclusive": exclusive,
                }

    def _add_validation_errors(
        self, row_number, exc: DjangoValidationError, *, sheet=None
    ) -> None:
        if hasattr(exc, "message_dict"):
            for field, messages in exc.message_dict.items():
                label = MODEL_FIELD_LABELS.get(field, field)
                for message in messages:
                    self._error(row_number, label, str(message), sheet=sheet)
            return
        for message in exc.messages:
            self._error(row_number, "整行", str(message), sheet=sheet)

    def _error(self, row, field, message, *, sheet=None):
        self.errors.append(
            {
                "sheet": sheet or self._active_sheet,
                "row": row,
                "field": field,
                "message": message,
            }
        )

    def _persist(
        self,
        rows: list[ParsedProductRow],
        package_rows: list[ParsedPackageRow],
        filename: str,
    ) -> list[dict[str, Any]]:
        created = []
        with transaction.atomic():
            product_by_key = {}
            package_by_key = {}
            for row in rows:
                product = row.product
                product.full_clean(
                    exclude=(
                        {"code", "sku"}
                        if getattr(product, "_derive_code_from_sku_on_create", False)
                        else None
                    )
                )
                product.save()
                if row.package_data:
                    package = ProductPackage(
                        product=product,
                        created_by=self.user,
                        updated_by=self.user,
                        **row.package_data,
                    )
                    package.full_clean()
                    package.save()
                    package_by_key[
                        (
                            product.owner.code.upper(),
                            product.code.upper(),
                            package.uom.code.upper(),
                        )
                    ] = package
                product_by_key[(product.owner.code.upper(), product.code.upper())] = (
                    product
                )
                created.append(
                    {
                        "row": row.row_number,
                        "product_id": product.pk,
                        "owner_code": product.owner.code,
                        "code": product.code,
                        "name": product.name,
                        "received_qty": str(row.receipt_qty),
                    }
                )
            for row in package_rows:
                package = ProductPackage(
                    product=product_by_key[row.product_key],
                    created_by=self.user,
                    updated_by=self.user,
                    **row.package_data,
                )
                package.full_clean()
                package.save()
                package_by_key[(*row.product_key, package.uom.code.upper())] = package
            for row in rows:
                if not row.carton_barcode:
                    continue
                product = row.product
                package = package_by_key[
                    (
                        product.owner.code.upper(),
                        product.code.upper(),
                        row.carton_package_uom_code,
                    )
                ]
                add_product_barcode(
                    product=product,
                    barcode=row.carton_barcode,
                    barcode_type=ProductBarcode.BarcodeType.CARTON,
                    package=package,
                    is_primary=True,
                )
            receipt_rows_by_owner: dict[int, list[ParsedProductRow]] = defaultdict(list)
            for row in rows:
                if row.receipt_qty > 0:
                    receipt_rows_by_owner[row.product.owner_id].append(row)
            if receipt_rows_by_owner:
                assert self.warehouse is not None
                assert self.warehouse.default_receive_location_id is not None
                import_request_id = uuid.uuid4().hex
                for owner_id, owner_rows in sorted(receipt_rows_by_owner.items()):
                    receipt = receive_goods_without_order(
                        owner_id=owner_id,
                        warehouse_id=self.warehouse.pk,
                        location_id=self.warehouse.default_receive_location_id,
                        items=[
                            {"product_id": row.product.pk, "qty": row.receipt_qty}
                            for row in owner_rows
                        ],
                        request_id=f"product-import:{import_request_id}:{owner_id}",
                        by_user=self.user,
                        remark="商品批量导入自动收货",
                        request=self.request,
                        source="product_excel_import",
                    )
                    owner = owner_rows[0].product.owner
                    self.receipts.append(
                        {
                            "owner_id": owner_id,
                            "owner_code": owner.code,
                            "warehouse_id": self.warehouse.pk,
                            "warehouse_code": self.warehouse.code,
                            "location_id": self.warehouse.default_receive_location_id,
                            "location_code": self.warehouse.default_receive_location.code,
                            "task_id": receipt.get("task_id"),
                            "task_no": receipt.get("task_no"),
                            "item_count": len(owner_rows),
                            "posted": bool(receipt.get("posted")),
                        }
                    )
                    self.received_product_count += len(owner_rows)
            owner_ids = sorted({row.product.owner_id for row in rows})
            record_audit_event(
                action="products.import_excel",
                module="products",
                request=self.request,
                user=self.user,
                owner_id=owner_ids[0] if len(owner_ids) == 1 else None,
                succeeded=True,
                after={"product_ids": [item["product_id"] for item in created]},
                metadata={
                    "filename": Path(filename).name,
                    "total_rows": self.total_rows,
                    "created_count": len(created),
                    "package_count": len(package_rows)
                    + sum(bool(row.package_data) for row in rows),
                    "skipped_count": len(self.skipped),
                    "created_categories": [
                        {"id": item.pk, "code": item.code, "name": item.name}
                        for item in self.created_categories
                    ],
                    "created_uoms": [
                        {"id": item.pk, "code": item.code, "name": item.name}
                        for item in self.created_uoms
                    ],
                    "warehouse_id": self.warehouse.pk if self.warehouse else None,
                    "default_receive_location_id": (
                        self.warehouse.default_receive_location_id
                        if self.warehouse
                        else None
                    ),
                    "received_product_count": self.received_product_count,
                    "receipts": self.receipts,
                },
            )
        return created

    def _result(self, *, created):
        committed = bool(created) and not self.errors
        return {
            "total_rows": self.total_rows,
            "created_count": len(created),
            "skipped_count": len(self.skipped),
            "error_count": len(self.errors),
            "created": created,
            "created_categories": (
                [
                    {"id": item.pk, "code": item.code, "name": item.name}
                    for item in self.created_categories
                ]
                if committed
                else []
            ),
            "created_uoms": (
                [
                    {
                        "id": item.pk,
                        "code": item.code,
                        "name": item.name,
                        "kind": item.kind,
                        "decimal_places": item.decimal_places,
                    }
                    for item in self.created_uoms
                ]
                if committed
                else []
            ),
            "received_product_count": self.received_product_count if committed else 0,
            "receipts": self.receipts if committed else [],
            "skipped": self.skipped,
            "errors": self.errors,
        }


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _joined_text(*parts: str) -> str:
    return "".join(parts)


def _validation_text(exc: DjangoValidationError) -> str:
    if hasattr(exc, "message_dict"):
        return "; ".join(
            f"{MODEL_FIELD_LABELS.get(field, field)}: "
            f"{', '.join(str(message) for message in messages)}"
            for field, messages in exc.message_dict.items()
        )
    return "; ".join(str(message) for message in exc.messages)


def _pinyin_code(name: str) -> str:
    pinyin = "".join(lazy_pinyin(name, errors=lambda chars: list(chars)))
    return re.sub(r"[^A-Za-z0-9]+", "_", pinyin).strip("_").upper()


def _code_candidates(base_code: str):
    yield base_code
    sequence = 2
    while True:
        yield f"{base_code}_{sequence}"
        sequence += 1
