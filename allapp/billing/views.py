from __future__ import annotations

import datetime
import io
from decimal import Decimal

from django.db import transaction
from django.db.models import Count, Q, QuerySet, Sum
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from rest_framework import filters, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework.views import APIView

from allapp.accounts.access import AccessScope
from allapp.accounts.audit import object_snapshot, record_audit_event
from allapp.baseinfo.models import Owner, OwnerWarehouseBinding
from allapp.locations.models import Warehouse
from allapp.reports.boss_contract import build_meta, warning

from .enums import AccrualStatus, BillStatus, PeriodStatus
from .models import (
    Bill,
    BillingAccrual,
    BillingEvent,
    BillingMetricDaily,
    BillingPeriod,
    BillingRule,
    BillingRuleTier,
    PaymentReceipt,
    ReceivableCollectionCase,
)
from .serializers import (
    BillDetailSerializer,
    BillingAccrualDetailSerializer,
    BillingAccrualSerializer,
    BillingEventSerializer,
    BillingMetricDailySerializer,
    BillingMetricGenerateSerializer,
    BillingPeriodInvoiceSerializer,
    BillingPeriodSerializer,
    BillingRuleSerializer,
    BillingRuleTierSerializer,
    BillListSerializer,
    CollectionActivitySerializer,
    PaymentReceiptSerializer,
    ReceivableCollectionCaseSerializer,
    RepriceUnpricedSerializer,
    UnlockPeriodSerializer,
)
from .services import (
    BillingCloseBlocked,
    accrue_metrics_for_date,
    accrue_order_processing_from_posted,
    accrue_storage_for_date,
    build_close_readiness,
    generate_invoice_for_period,
    generate_metrics_for_date,
    generate_metrics_for_range,
    lock_period,
    preview_lock_period,
    reprice_unpriced_events,
    unlock_period,
)
from .services.dashboard import build_warehouse_overview_payload
from .services.ledger import financial_ledger_accruals
from .services.receivables import post_receipt, reverse_receipt


def _require_billing_perm(request, perm_codename: str):
    """检查用户是否有指定的 billing 权限，无则抛 PermissionDenied。"""
    if not request.user.has_perm(f"billing.{perm_codename}"):
        raise PermissionDenied(f"需要 billing.{perm_codename} 权限。")


def _require_financial_export_perm(request):
    if request.user.is_superuser:
        return
    if not (
        request.user.has_perm("reports.export_operations")
        or request.user.has_perm("accounts.export_operational_reports")
    ):
        raise PermissionDenied("需要运营报表导出权限。")


class BillingModelPermissions(permissions.DjangoModelPermissions):
    perms_map = {
        **permissions.DjangoModelPermissions.perms_map,
        "GET": ["%(app_label)s.view_%(model_name)s"],
        "HEAD": ["%(app_label)s.view_%(model_name)s"],
        "OPTIONS": ["%(app_label)s.view_%(model_name)s"],
    }


def _explicit_billing_scope(user) -> AccessScope:
    """Return the tenant scope accepted for a state-changing billing request.

    Billing is a financial write surface.  In particular, a legacy ``owner``
    or ``warehouse`` field on ``User`` is only a migration aid and must never
    decide where a rule, metric or period is written.  Require a current
    ``UserRoleScope`` row for every non-superuser mutation.
    """

    scope = AccessScope.for_user(user)
    if not scope.is_valid or scope.source != "user_role_scope":
        raise PermissionDenied("计费写入必须使用有效的显式角色范围授权。")
    return scope


class PaymentReceiptViewSet(viewsets.ModelViewSet):
    serializer_class = PaymentReceiptSerializer
    permission_classes = [BillingModelPermissions]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["owner", "warehouse", "currency", "status", "receipt_date"]
    ordering_fields = ["receipt_date", "receipt_no", "created_at"]
    ordering = ["-receipt_date", "-id"]

    def get_queryset(self):
        qs = PaymentReceipt.objects.select_related("owner", "warehouse").prefetch_related(
            "allocations"
        )
        return AccessScope.for_user(self.request.user).filter_queryset(
            qs, owner_field="owner_id", warehouse_field="warehouse_id"
        )

    def perform_create(self, serializer):
        if not self.request.user.has_perm("billing.add_paymentreceipt"):
            raise PermissionDenied("需要收款登记权限。")
        scope = _explicit_billing_scope(self.request.user)
        owner = serializer.validated_data["owner"]
        warehouse = serializer.validated_data["warehouse"]
        if not scope.allows(owner_id=owner.pk, warehouse_id=warehouse.pk):
            raise PermissionDenied("收款单超出授权货主或仓库范围。")
        serializer.save()

    def perform_update(self, serializer):
        if not self.request.user.has_perm("billing.change_paymentreceipt"):
            raise PermissionDenied("需要收款修改权限。")
        serializer.save()

    def perform_destroy(self, instance):
        if instance.status != "DRAFT":
            raise PermissionDenied("已过账或已冲销收款单禁止删除。")
        return super().perform_destroy(instance)

    @action(detail=True, methods=["post"])
    def post(self, request, pk=None):
        if not request.user.has_perm("billing.change_paymentreceipt"):
            raise PermissionDenied("需要收款过账权限。")
        try:
            receipt = post_receipt(self.get_object().pk, by_user=request.user)
        except ValueError as exc:
            return Response({"code": "PAYMENT_POST_BLOCKED", "detail": str(exc)}, status=409)
        return Response(self.get_serializer(receipt).data)

    @action(detail=True, methods=["post"])
    def reverse(self, request, pk=None):
        if not request.user.has_perm("billing.change_paymentreceipt"):
            raise PermissionDenied("需要收款冲销权限。")
        receipt_no = str(request.data.get("receipt_no") or "").strip()
        if not receipt_no:
            return Response({"detail": "receipt_no is required."}, status=400)
        raw_date = request.data.get("receipt_date")
        reversal_date = parse_date(str(raw_date)) if raw_date else None
        if raw_date and reversal_date is None:
            return Response({"detail": "receipt_date must be YYYY-MM-DD."}, status=400)
        try:
            reversal = reverse_receipt(
                self.get_object().pk,
                receipt_no=receipt_no,
                reversal_date=reversal_date,
                memo=str(request.data.get("memo") or ""),
                by_user=request.user,
            )
        except ValueError as exc:
            return Response({"code": "PAYMENT_REVERSE_BLOCKED", "detail": str(exc)}, status=409)
        return Response(self.get_serializer(reversal).data, status=201)


class ReceivableCollectionCaseViewSet(viewsets.ModelViewSet):
    serializer_class = ReceivableCollectionCaseSerializer
    permission_classes = [BillingModelPermissions]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["bill", "status", "assignee"]
    ordering = ["-updated_at", "-id"]

    def get_queryset(self):
        qs = ReceivableCollectionCase.objects.select_related(
            "bill", "bill__owner", "bill__warehouse", "assignee"
        ).prefetch_related("activities")
        return AccessScope.for_user(self.request.user).filter_queryset(
            qs, owner_field="bill__owner_id", warehouse_field="bill__warehouse_id"
        )

    def _require_manage(self):
        if not self.request.user.has_perm("billing.manage_collections"):
            raise PermissionDenied("需要催收记录维护权限。")

    def perform_create(self, serializer):
        self._require_manage()
        bill = serializer.validated_data["bill"]
        if not AccessScope.for_user(self.request.user).allows(
            owner_id=bill.owner_id, warehouse_id=bill.warehouse_id
        ):
            raise PermissionDenied("催收案件超出授权范围。")
        serializer.save()

    def perform_update(self, serializer):
        self._require_manage()
        bill = serializer.validated_data.get("bill", serializer.instance.bill)
        if not AccessScope.for_user(self.request.user).allows(
            owner_id=bill.owner_id, warehouse_id=bill.warehouse_id
        ):
            raise PermissionDenied("催收案件超出授权范围。")
        assignee = serializer.validated_data.get("assignee")
        if assignee and not AccessScope.for_user(assignee).allows(
            warehouse_id=serializer.instance.bill.warehouse_id
        ):
            raise PermissionDenied("责任人没有账单对应仓库范围。")
        serializer.save()

    def perform_destroy(self, instance):
        raise PermissionDenied("催收案件和历史禁止删除。")

    @action(detail=True, methods=["post"])
    def activities(self, request, pk=None):
        self._require_manage()
        serializer = CollectionActivitySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        activity = serializer.save(case=self.get_object(), created_by=request.user)
        if activity.next_follow_up_at:
            ReceivableCollectionCase.objects.filter(pk=activity.case_id).update(
                next_follow_up_at=activity.next_follow_up_at
            )
        return Response(CollectionActivitySerializer(activity).data, status=201)


def _scope_id(value):
    """Return a model instance's primary key without trusting raw input."""

    if value is None:
        return None
    return getattr(value, "pk", value)


class BillingDataPermission(permissions.DjangoModelPermissions):
    """Role-aware read access plus Django model permissions for mutations."""

    perms_map = {
        **permissions.DjangoModelPermissions.perms_map,
        "GET": [],
        "HEAD": [],
        "OPTIONS": [],
    }

    # DRF maps every custom POST action to ``add_<view-model>`` by default.
    # Billing actions are semantically different: lock/unlock changes a
    # period, while invoice creation creates a bill.  Map them explicitly so
    # a correctly authorized scoped finance user is neither blocked nor given
    # a broader create permission than the operation needs.
    action_permissions = {
        "activate": "billing.change_billingrule",
        "deactivate": "billing.change_billingrule",
        "generate_metrics": "billing.change_billingperiod",
        "accrue_storage": "billing.change_billingperiod",
        "accrue_orders_posted": "billing.change_billingperiod",
        "lock": "billing.change_billingperiod",
        "unlock": "billing.change_billingperiod",
        "invoice": "billing.add_bill",
        "generate": "billing.change_billingperiod",
        "preview_lock": "billing.change_billingperiod",
        "reprice_unpriced": "billing.change_billingperiod",
    }

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.user.is_superuser:
            return True
        scope = AccessScope.for_user(request.user)
        if request.method in permissions.SAFE_METHODS:
            return scope.is_valid and (
                request.user.has_perm("accounts.view_owner_financials")
                or request.user.has_perm("reports.view_warehouse_finance")
            )
        required = self.action_permissions.get(getattr(view, "action", ""))
        if required:
            return (
                scope.is_valid
                and scope.source == "user_role_scope"
                and request.user.has_perm(required)
            )
        return super().has_permission(request, view)


class OwnerWarehouseScopedQuerysetMixin:
    permission_classes = [BillingDataPermission]

    def scope_queryset(self, qs: QuerySet):
        request = getattr(self, "request", None)
        user = getattr(request, "user", None)
        return AccessScope.for_user(user).filter_queryset(
            qs,
            owner_field=getattr(self, "scope_owner_field", "owner_id"),
            warehouse_field=getattr(self, "scope_warehouse_field", "warehouse_id"),
        )

    def get_queryset(self):  # type: ignore[override]
        qs = super().get_queryset()  # type: ignore[misc]
        return self.scope_queryset(qs)


class OwnerWarehouseSaveMixin:
    """Validate the *final* serializer tenant target before persisting it.

    ``serializer.validated_data`` is authoritative here.  Mutating a payload
    with the legacy bindings stored on ``User`` both hides the requested
    target and permits unbound users to choose a tenant.  This mixin is used
    only by models that carry direct ``owner`` and ``warehouse`` fields.
    """

    def _final_serializer_scope_ids(self, serializer):
        instance = getattr(serializer, "instance", None)
        data = serializer.validated_data
        owner = data.get("owner", getattr(instance, "owner", None))
        warehouse = data.get("warehouse", getattr(instance, "warehouse", None))
        return _scope_id(owner), _scope_id(warehouse)

    def _validate_serializer_write_scope(self, serializer):
        owner_id, warehouse_id = self._final_serializer_scope_ids(serializer)
        self._validate_owner_warehouse_write_scope(owner_id, warehouse_id)

    def _validate_owner_warehouse_write_scope(self, owner_id, warehouse_id):
        user = self.request.user
        if getattr(user, "is_superuser", False):
            return
        scope = _explicit_billing_scope(user)
        if (
            not owner_id
            or not warehouse_id
            or not scope.allows(
                owner_id=owner_id,
                warehouse_id=warehouse_id,
            )
        ):
            raise PermissionDenied("无权写入目标货主或仓库的计费数据。")

    def perform_create(self, serializer):
        self._validate_serializer_write_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_serializer_write_scope(serializer)
        serializer.save()

    def perform_destroy(self, instance):
        self._validate_owner_warehouse_write_scope(
            getattr(instance, "owner_id", None),
            getattr(instance, "warehouse_id", None),
        )
        instance.delete()


class BillingWarehouseOverviewApi(OwnerWarehouseScopedQuerysetMixin, APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _scope_dashboard_queryset(self, qs: QuerySet, *, scope: AccessScope):
        return scope.filter_queryset(
            qs,
            owner_field="owner_id",
            warehouse_field="warehouse_id",
        )

    def _resolve_scope_label(self, model_cls, pk):
        if not pk:
            return ""
        obj = model_cls.objects.filter(pk=pk).only("name").first()
        if obj is None:
            return ""
        return getattr(obj, "name", "")

    def get(self, request):
        owner_raw = (request.query_params.get("owner") or "").strip()
        warehouse_raw = (request.query_params.get("warehouse") or "").strip()
        charge_type = (request.query_params.get("charge_type") or "").strip()
        accrual_status = (request.query_params.get("status") or "").strip()
        date_from_raw = (request.query_params.get("date_from") or "").strip()
        date_to_raw = (request.query_params.get("date_to") or "").strip()
        recent_limit_raw = (request.query_params.get("recent_limit") or "").strip()
        # The client may choose presentation filters, never the authorization
        # mode.  The server resolves owner-vs-warehouse access from role scope.
        scope = AccessScope.for_user(request.user)
        if not scope.is_valid:
            raise PermissionDenied("No valid billing data scope.")
        if not request.user.is_superuser:
            if scope.warehouse_ids and not request.user.has_perm("reports.view_warehouse_finance"):
                raise PermissionDenied("No permission to view warehouse financial data.")
            if scope.owner_ids and not request.user.has_perm("accounts.view_owner_financials"):
                raise PermissionDenied("No permission to view owner financial data.")

        if owner_raw and not owner_raw.isdigit():
            return Response(
                {"detail": "owner must be an integer id."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if warehouse_raw and not warehouse_raw.isdigit():
            return Response(
                {"detail": "warehouse must be an integer id."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        owner_id = int(owner_raw) if owner_raw else None
        warehouse_id = int(warehouse_raw) if warehouse_raw else None

        if scope.owner_ids and owner_id and not scope.allows(owner_id=owner_id):
            exc = PermissionDenied("No access to other owners in billing dashboard.")
            exc.detail = {
                "code": "SCOPE_FORBIDDEN",
                "detail": "No access to other owners in billing dashboard.",
            }
            raise exc
        if warehouse_id and not scope.allows(warehouse_id=warehouse_id):
            exc = PermissionDenied("No access to other warehouses in billing dashboard.")
            exc.detail = {
                "code": "SCOPE_FORBIDDEN",
                "detail": "No access to other warehouses in billing dashboard.",
            }
            raise exc

        current = timezone.now()
        today = timezone.localtime(current).date() if timezone.is_aware(current) else current.date()
        date_to = parse_date(date_to_raw) if date_to_raw else today
        date_from = parse_date(date_from_raw) if date_from_raw else date_to.replace(day=1)
        if date_from_raw and date_from is None:
            return Response(
                {"detail": "date_from must be YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if date_to_raw and date_to is None:
            return Response(
                {"detail": "date_to must be YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if date_from and date_to and date_from > date_to:
            return Response(
                {"detail": "date_from cannot be after date_to."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if date_to > today:
            return Response(
                {"detail": "date_to cannot be after today."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if date_from and date_to and (date_to - date_from).days > 366:
            return Response(
                {"detail": "date range cannot exceed 367 days."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        recent_limit = int(recent_limit_raw) if recent_limit_raw.isdigit() else 10
        recent_limit = max(1, min(recent_limit, 50))

        scoped_accrual_qs = self._scope_dashboard_queryset(
            BillingAccrual.objects.select_related(
                "owner", "warehouse", "period", "rule", "event", "created_by"
            ),
            scope=scope,
        )
        base_accrual_qs = scoped_accrual_qs.filter(is_reversal=False).exclude(
            status=AccrualStatus.VOID
        )
        ledger_accrual_qs = financial_ledger_accruals(scoped_accrual_qs)
        base_bill_qs = self._scope_dashboard_queryset(
            Bill.objects.select_related("owner", "warehouse", "period")
            .prefetch_related("lines")
            .exclude(status=BillStatus.VOID),
            scope=scope,
        )

        if warehouse_id:
            base_accrual_qs = base_accrual_qs.filter(warehouse_id=warehouse_id)
            ledger_accrual_qs = ledger_accrual_qs.filter(warehouse_id=warehouse_id)
            base_bill_qs = base_bill_qs.filter(warehouse_id=warehouse_id)

        binding_qs = OwnerWarehouseBinding.objects.select_related("owner").filter(
            is_active=True,
            is_deleted=False,
            owner__is_active=True,
            owner__is_deleted=False,
        )
        if warehouse_id:
            binding_qs = binding_qs.filter(warehouse_id=warehouse_id)
        elif scope.warehouse_ids:
            binding_qs = binding_qs.filter(warehouse_id__in=scope.warehouse_ids)
        owner_options_map = {
            binding.owner_id: {
                "id": binding.owner_id,
                "name": binding.owner.name,
            }
            for binding in binding_qs.order_by("owner__name", "owner_id")
        }
        owner_options = sorted(
            owner_options_map.values(), key=lambda item: (item["name"], item["id"])
        )

        if owner_id:
            base_accrual_qs = base_accrual_qs.filter(owner_id=owner_id)
            ledger_accrual_qs = ledger_accrual_qs.filter(owner_id=owner_id)
            base_bill_qs = base_bill_qs.filter(owner_id=owner_id)

        if charge_type:
            base_accrual_qs = base_accrual_qs.filter(charge_type=charge_type)
            ledger_accrual_qs = ledger_accrual_qs.filter(charge_type=charge_type)

        if accrual_status:
            base_accrual_qs = base_accrual_qs.filter(status=accrual_status)
            ledger_accrual_qs = ledger_accrual_qs.filter(status=accrual_status)

        if date_from:
            base_accrual_qs = base_accrual_qs.filter(service_date__gte=date_from)
            ledger_accrual_qs = ledger_accrual_qs.filter(service_date__gte=date_from)
            base_bill_qs = base_bill_qs.filter(issue_date__gte=date_from)

        if date_to:
            base_accrual_qs = base_accrual_qs.filter(service_date__lte=date_to)
            ledger_accrual_qs = ledger_accrual_qs.filter(service_date__lte=date_to)
            base_bill_qs = base_bill_qs.filter(issue_date__lte=date_to)

        scope_owner_id = owner_id
        scope_owner_name = ""
        if scope_owner_id:
            scope_owner_name = owner_options_map.get(scope_owner_id, {}).get("name", "")
            if not scope_owner_name and (scope.is_global or scope_owner_id in scope.owner_ids):
                scope_owner_name = (
                    Owner.objects.filter(pk=scope_owner_id).values_list("name", flat=True).first()
                    or ""
                )

        payload = build_warehouse_overview_payload(
            accrual_qs=base_accrual_qs,
            ledger_accrual_qs=ledger_accrual_qs,
            bill_qs=base_bill_qs,
            recent_limit=recent_limit,
        )
        payload["scope"] = {
            "mode": "WAREHOUSE" if warehouse_id else "ALL_AUTHORIZED",
            "owner": scope_owner_id,
            "owner_name": scope_owner_name,
            "warehouse": warehouse_id,
            "warehouse_name": self._resolve_scope_label(
                Warehouse,
                warehouse_id,
            )
            or "全部授权仓库",
            "date_from": date_from,
            "date_to": date_to,
            "charge_type": charge_type,
            "status": accrual_status,
        }
        payload["owner_options"] = owner_options
        unknown_currency_count = (
            ledger_accrual_qs.filter(Q(currency__isnull=True) | Q(currency="")).count()
            + base_bill_qs.filter(Q(currency__isnull=True) | Q(currency="")).count()
        )
        warnings = []
        if unknown_currency_count:
            warnings.append(warning("UNKNOWN_CURRENCY", unknown_currency_count))
        payload["meta"] = build_meta(scope=payload["scope"], warnings=warnings)
        return Response(payload)


class BillingRuleViewSet(
    OwnerWarehouseScopedQuerysetMixin, OwnerWarehouseSaveMixin, viewsets.ModelViewSet
):
    queryset = (
        BillingRule.objects.select_related("owner", "warehouse").prefetch_related("tiers").all()
    )
    serializer_class = BillingRuleSerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = {
        "owner": ["exact"],
        "warehouse": ["exact"],
        "charge_type": ["exact", "in"],
        "calc_method": ["exact", "in"],
        "active": ["exact"],
        "bundle_key": ["exact", "icontains"],
        "effective_from": ["exact", "gte", "lte"],
        "effective_to": ["exact", "gte", "lte"],
    }
    search_fields = ["note", "bundle_key"]
    ordering_fields = ["id", "priority", "effective_from", "effective_to"]
    ordering = ["owner_id", "priority", "id"]

    def scope_queryset(self, qs: QuerySet):
        request = getattr(self, "request", None)
        user = getattr(request, "user", None)
        scope = AccessScope.for_user(user)
        scoped = scope.filter_queryset(qs, owner_field="owner_id", warehouse_field="warehouse_id")
        # Global rate cards carry no tenant data. They are intentionally
        # readable so a scoped finance user can understand which fallback
        # rule applies, but `_validate_rule_write_scope` still rejects every
        # mutation of them.
        if request and request.method in permissions.SAFE_METHODS and scope.is_valid:
            return qs.filter(
                Q(owner__isnull=True, warehouse__isnull=True) | Q(pk__in=scoped.values("pk"))
            )
        return scoped

    def _validate_rule_write_scope(self, rule: BillingRule):
        user = self.request.user
        if getattr(user, "is_superuser", False):
            return
        scope = _explicit_billing_scope(user)
        if not scope.allows(owner_id=rule.owner_id, warehouse_id=rule.warehouse_id):
            raise PermissionDenied("无权修改通用规则或范围外规则。")

    def perform_update(self, serializer):
        # The inherited mixin validates the post-patch owner/warehouse pair,
        # rather than only the row that happened to be retrieved.
        super().perform_update(serializer)

    def perform_destroy(self, instance):
        self._validate_rule_write_scope(instance)
        instance.delete()

    @action(detail=True, methods=["post"], url_path="activate")
    def activate(self, request, pk=None):
        _require_billing_perm(request, "change_billingrule")
        rule = self.get_object()
        self._validate_rule_write_scope(rule)
        rule.active = True
        rule.save(update_fields=["active"])
        return Response(self.get_serializer(rule).data)

    @action(detail=True, methods=["post"], url_path="deactivate")
    def deactivate(self, request, pk=None):
        _require_billing_perm(request, "change_billingrule")
        rule = self.get_object()
        self._validate_rule_write_scope(rule)
        rule.active = False
        rule.save(update_fields=["active"])
        return Response(self.get_serializer(rule).data)


class BillingRuleTierViewSet(OwnerWarehouseScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = BillingRuleTier.objects.select_related(
        "rule", "rule__owner", "rule__warehouse"
    ).all()
    serializer_class = BillingRuleTierSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = {
        "rule": ["exact"],
        "rule__owner": ["exact"],
        "rule__warehouse": ["exact"],
    }
    ordering_fields = ["id", "threshold_from", "threshold_to"]
    ordering = ["rule_id", "threshold_from", "id"]

    def scope_queryset(self, qs: QuerySet):
        user = getattr(self.request, "user", None)
        scope = AccessScope.for_user(user)
        scoped = scope.filter_queryset(
            qs,
            owner_field="rule__owner_id",
            warehouse_field="rule__warehouse_id",
        )
        if self.request.method in permissions.SAFE_METHODS and scope.is_valid:
            return qs.filter(
                Q(rule__owner__isnull=True, rule__warehouse__isnull=True)
                | Q(pk__in=scoped.values("pk"))
            )
        return scoped

    def _validate_rule_scope(self, rule):
        user = self.request.user
        if user.is_superuser:
            return
        scope = _explicit_billing_scope(user)
        if not scope.allows(owner_id=rule.owner_id, warehouse_id=rule.warehouse_id):
            raise PermissionDenied("无权操作通用规则或范围外规则的阶梯。")

    def perform_create(self, serializer):
        _require_billing_perm(self.request, "change_billingrule")
        rule = serializer.validated_data["rule"]
        self._validate_rule_scope(rule)
        serializer.save()

    def perform_update(self, serializer):
        _require_billing_perm(self.request, "change_billingrule")
        rule = serializer.validated_data.get("rule", serializer.instance.rule)
        self._validate_rule_scope(rule)
        serializer.save()

    def perform_destroy(self, instance):
        _require_billing_perm(self.request, "change_billingrule")
        self._validate_rule_scope(instance.rule)
        instance.delete()


class BillingMetricDailyViewSet(
    OwnerWarehouseScopedQuerysetMixin, OwnerWarehouseSaveMixin, viewsets.ModelViewSet
):
    queryset = BillingMetricDaily.objects.select_related("owner", "warehouse").all()
    serializer_class = BillingMetricDailySerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = {
        "owner": ["exact"],
        "warehouse": ["exact"],
        "service_date": ["exact", "gte", "lte"],
        "metric_type": ["exact", "in"],
    }
    search_fields = ["source", "note"]
    ordering_fields = ["id", "service_date", "metric_type", "created_at"]
    ordering = ["-service_date", "metric_type", "-id"]

    @action(detail=False, methods=["post"], url_path="generate")
    @transaction.atomic
    def generate(self, request):
        payload = BillingMetricGenerateSerializer(data=request.data)
        payload.is_valid(raise_exception=True)
        data = payload.validated_data

        if request.user.is_superuser:
            scope = AccessScope.for_user(request.user)
        else:
            scope = _explicit_billing_scope(request.user)

        # Defaults may only come from an unambiguous explicit role scope.
        # Never fall back to User.owner/User.warehouse, which may be stale or
        # deliberately populated for another application workflow.
        owner_id = data.get("owner") or (
            next(iter(scope.owner_ids)) if len(scope.owner_ids) == 1 else None
        )
        warehouse_id = data.get("warehouse") or (
            next(iter(scope.warehouse_ids)) if len(scope.warehouse_ids) == 1 else None
        )
        if not owner_id or not warehouse_id:
            return Response(
                {"detail": "必须提供 owner 和 warehouse，或由显式角色范围唯一确定。"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not scope.allows(owner_id=owner_id, warehouse_id=warehouse_id):
            raise PermissionDenied("无权为范围外货主或仓库生成计费指标。")

        summary = generate_metrics_for_range(
            owner_id,
            warehouse_id,
            data["start_date"],
            data["end_date"],
            metric_types=data.get("metric_types"),
            overwrite=data.get("overwrite", False),
            allow_area_fallback=data.get("allow_area_fallback", False),
        )
        return Response(summary)


class BillingEventViewSet(OwnerWarehouseScopedQuerysetMixin, viewsets.ReadOnlyModelViewSet):
    queryset = BillingEvent.objects.select_related(
        "owner",
        "warehouse",
        "task",
        "task_line",
        "scan_log",
        "posting_journal",
        "metric",
        "pricing_rule",
    ).all()
    serializer_class = BillingEventSerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = {
        "owner": ["exact"],
        "warehouse": ["exact"],
        "charge_type": ["exact", "in"],
        "calc_method": ["exact", "in"],
        "pricing_status": ["exact", "in"],
        "pricing_reason": ["exact", "in", "icontains"],
        "pricing_rule": ["exact", "isnull"],
        "priced_at": ["exact", "gte", "lte", "isnull"],
        "metric": ["exact", "isnull"],
        "bundle_key": ["exact", "icontains"],
        "service_date": ["exact", "gte", "lte"],
        "task": ["exact"],
        "task_line": ["exact"],
        "scan_log": ["exact"],
    }
    search_fields = ["event_fp"]
    ordering_fields = ["id", "service_date", "created_at"]
    ordering = ["-service_date", "-id"]

    @action(detail=False, methods=["post"], url_path="reprice-unpriced")
    def reprice_unpriced(self, request):
        payload = RepriceUnpricedSerializer(data=request.data)
        payload.is_valid(raise_exception=True)
        data = payload.validated_data
        if not request.user.is_superuser:
            scope = _explicit_billing_scope(request.user)
            if not scope.allows(owner_id=data["owner"], warehouse_id=data["warehouse"]):
                raise PermissionDenied("无权重算范围外货主或仓库的计费事件。")
        return Response(
            reprice_unpriced_events(
                owner_id=data["owner"],
                warehouse_id=data["warehouse"],
                date_from=data["date_from"],
                date_to=data["date_to"],
                dry_run=data["dry_run"],
                by_user=request.user,
            )
        )


class BillingAccrualViewSet(OwnerWarehouseScopedQuerysetMixin, viewsets.ReadOnlyModelViewSet):
    queryset = (
        BillingAccrual.objects.select_related(
            "owner",
            "warehouse",
            "period",
            "rule",
            "event",
            "event__task",
            "created_by",
        )
        .prefetch_related("billline_set__bill")
        .all()
    )
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = {
        "owner": ["exact"],
        "warehouse": ["exact"],
        "period": ["exact", "isnull"],
        "charge_type": ["exact", "in"],
        "rule": ["exact"],
        "service_date": ["exact", "gte", "lte"],
        "status": ["exact", "in"],
        "event": ["exact"],
        "is_reversal": ["exact"],
        "reversal_of": ["exact"],
        "bundle_key": ["exact", "icontains"],
    }
    search_fields = ["acc_fingerprint", "bundle_key", "event__event_fp"]
    ordering_fields = ["id", "service_date", "amount", "tax_amount", "created_at"]
    ordering = ["-service_date", "-id"]

    def get_serializer_class(self):  # type: ignore[override]
        if self.action == "retrieve":
            return BillingAccrualDetailSerializer
        return BillingAccrualSerializer


class BillingPeriodViewSet(
    OwnerWarehouseScopedQuerysetMixin, OwnerWarehouseSaveMixin, viewsets.ModelViewSet
):
    queryset = BillingPeriod.objects.select_related("owner", "warehouse").all()
    serializer_class = BillingPeriodSerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = {
        "owner": ["exact"],
        "warehouse": ["exact"],
        "status": ["exact", "in"],
        "start_date": ["exact", "gte", "lte"],
        "end_date": ["exact", "gte", "lte"],
    }
    search_fields = ["label"]
    ordering_fields = ["id", "label", "start_date", "end_date"]
    ordering = ["-start_date", "-id"]

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        scoped = self.get_object()
        period = BillingPeriod.objects.select_for_update().get(pk=scoped.pk)
        immutable_scope = set()
        if "owner" in request.data and str(request.data.get("owner")) != str(period.owner_id):
            immutable_scope.add("owner")
        if "warehouse" in request.data and str(request.data.get("warehouse")) != str(
            period.warehouse_id
        ):
            immutable_scope.add("warehouse")
        protected_fields = {"label", "start_date", "end_date", "currency"} & set(
            request.data.keys()
        )
        if immutable_scope:
            record_audit_event(
                action="UPDATE_REJECTED",
                module="billing.period",
                request=request,
                obj=period,
                succeeded=False,
                before=object_snapshot(period),
                metadata={"fields": sorted(immutable_scope), "reason": "scope_change"},
            )
            return Response(
                {
                    "code": "SCOPE_FORBIDDEN",
                    "detail": "无权修改账期的货主或仓库范围。",
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        blocked = (
            protected_fields
            if period.status in {PeriodStatus.CLOSED, PeriodStatus.INVOICED}
            else set()
        )
        if blocked:
            record_audit_event(
                action="UPDATE_REJECTED",
                module="billing.period",
                request=request,
                obj=period,
                succeeded=False,
                before=object_snapshot(period),
                metadata={"fields": sorted(blocked), "reason": "immutable_period"},
            )
            return Response(
                {
                    "code": "PERIOD_IMMUTABLE",
                    "period_status": period.status,
                    "detail": "账期范围创建后不可变，关闭或开票后业务字段不可修改。",
                },
                status=status.HTTP_409_CONFLICT,
            )
        before = object_snapshot(period)
        serializer = self.get_serializer(
            period, data=request.data, partial=kwargs.pop("partial", False)
        )
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        record_audit_event(
            action="UPDATE",
            module="billing.period",
            request=request,
            obj=period,
            before=before,
            after=object_snapshot(period),
        )
        return Response(serializer.data)

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        scoped = self.get_object()
        period = BillingPeriod.objects.select_for_update().get(pk=scoped.pk)
        blocked = (
            period.status != PeriodStatus.OPEN
            or period.billingaccrual_set.exists()
            or period.bill_set.exists()
        )
        if blocked:
            record_audit_event(
                action="DELETE_REJECTED",
                module="billing.period",
                request=request,
                obj=period,
                succeeded=False,
                before=object_snapshot(period),
            )
            return Response(
                {
                    "code": "PERIOD_DELETE_BLOCKED",
                    "period_status": period.status,
                    "detail": "仅无应计、无账单关联的 OPEN 账期可删除。",
                },
                status=status.HTTP_409_CONFLICT,
            )
        before = object_snapshot(period)
        record_audit_event(
            action="DELETE",
            module="billing.period",
            request=request,
            obj=period,
            before=before,
        )
        period.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=["get"], url_path="close-readiness")
    def close_readiness(self, request, pk=None):
        period = self.get_object()
        return Response(
            build_close_readiness(
                owner_id=period.owner_id,
                warehouse_id=period.warehouse_id,
                start_date=period.start_date,
                end_date=period.end_date,
                for_invoice=period.status == PeriodStatus.CLOSED,
            )
        )

    def _guard_status(self, period: BillingPeriod, allowed_statuses):
        if period.status in allowed_statuses:
            return None
        allowed = ", ".join(allowed_statuses)
        return Response(
            {"detail": f"该操作仅允许在账期状态 {allowed} 时执行，当前为 {period.status}。"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _preview_queryset(self, period: BillingPeriod):
        base_qs = BillingAccrual.objects.select_related("rule", "event").filter(
            owner_id=period.owner_id,
            warehouse_id=period.warehouse_id,
        )
        if period.status == PeriodStatus.OPEN:
            return base_qs.filter(
                period__isnull=True,
                status=AccrualStatus.OPEN,
                service_date__gte=period.start_date,
                service_date__lte=period.end_date,
            )
        return base_qs.filter(period=period)

    @staticmethod
    def _preview_group_totals(qs, *group_fields):
        """Return one consistent subtotal/tax/total contract for preview groups."""

        rows = list(
            qs.values(*group_fields)
            .annotate(
                accrual_count=Count("id"),
                subtotal=Sum("amount"),
                tax_total=Sum("tax_amount"),
            )
            .order_by(*group_fields)
        )
        for row in rows:
            row["subtotal"] = row["subtotal"] or Decimal("0.00")
            row["tax_total"] = row["tax_total"] or Decimal("0.00")
            row["total"] = row["subtotal"] + row["tax_total"]
        return rows

    @action(detail=True, methods=["get"], url_path="preview")
    def preview(self, request, pk=None):
        period = self.get_object()
        qs = self._preview_queryset(period)
        accruals = list(qs)

        subtotal = sum((Decimal(a.amount) for a in accruals), Decimal("0.00"))
        tax_total = sum((Decimal(a.tax_amount) for a in accruals), Decimal("0.00"))

        data = {
            "period": self.get_serializer(period).data,
            "scope": ("open_unlocked" if period.status == PeriodStatus.OPEN else "period_locked"),
            "accrual_count": len(accruals),
            "quantity_total": sum((Decimal(a.quantity) for a in accruals), Decimal("0.0000")),
            "subtotal": subtotal,
            "tax_total": tax_total,
            "total": subtotal + tax_total,
            "by_charge_type": self._preview_group_totals(qs, "charge_type"),
            "by_status": self._preview_group_totals(qs, "status"),
            "by_service_date": self._preview_group_totals(qs, "service_date"),
        }
        return Response(data)

    @action(detail=True, methods=["post"], url_path="generate-metrics")
    @transaction.atomic
    def generate_metrics(self, request, pk=None):
        _require_billing_perm(request, "change_billingperiod")
        period = self.get_object()
        summary = generate_metrics_for_range(
            period.owner_id,
            period.warehouse_id,
            period.start_date,
            period.end_date,
            metric_types=request.data.get("metric_types"),
            overwrite=bool(request.data.get("overwrite", False)),
            allow_area_fallback=bool(request.data.get("allow_area_fallback", False)),
        )
        return Response({"period": self.get_serializer(period).data, "summary": summary})

    @action(detail=True, methods=["post"], url_path="accrue-storage")
    @transaction.atomic
    def accrue_storage(self, request, pk=None):
        _require_billing_perm(request, "change_billingperiod")
        period = self.get_object()
        blocked = self._guard_status(period, [PeriodStatus.OPEN])
        if blocked is not None:
            return blocked

        service_date = period.start_date
        total_events = 0
        total_accruals = 0
        total_metrics_created = 0
        total_metrics_updated = 0
        while service_date <= period.end_date:
            metric_summary = generate_metrics_for_date(
                period.owner_id,
                period.warehouse_id,
                service_date,
                allow_area_fallback=bool(request.data.get("allow_area_fallback", False)),
            )
            total_metrics_created += metric_summary["created"]
            total_metrics_updated += metric_summary["updated"]
            ev1, acc1 = accrue_storage_for_date(
                period.owner_id,
                period.warehouse_id,
                service_date,
                by_user=request.user,
            )
            ev2, acc2 = accrue_metrics_for_date(
                period.owner_id,
                period.warehouse_id,
                service_date,
                by_user=request.user,
            )
            total_events += ev1 + ev2
            total_accruals += acc1 + acc2
            service_date += datetime.timedelta(days=1)

        return Response(
            {
                "period": self.get_serializer(period).data,
                "events_created": total_events,
                "accruals_created": total_accruals,
                "metrics_created": total_metrics_created,
                "metrics_updated": total_metrics_updated,
            }
        )

    @action(detail=True, methods=["post"], url_path="accrue-orders-posted")
    @transaction.atomic
    def accrue_orders_posted(self, request, pk=None):
        _require_billing_perm(request, "change_billingperiod")
        period = self.get_object()
        blocked = self._guard_status(period, [PeriodStatus.OPEN])
        if blocked is not None:
            return blocked

        events_created, accruals_created = accrue_order_processing_from_posted(
            period.owner_id,
            period.warehouse_id,
            period.start_date,
            period.end_date,
            by_user=request.user,
        )
        return Response(
            {
                "period": self.get_serializer(period).data,
                "events_created": events_created,
                "accruals_created": accruals_created,
            }
        )

    @action(detail=True, methods=["post"], url_path="lock")
    @transaction.atomic
    def lock(self, request, pk=None):
        _require_billing_perm(request, "change_billingperiod")
        period = self.get_object()
        blocked = self._guard_status(period, [PeriodStatus.OPEN])
        if blocked is not None:
            return blocked

        try:
            locked_period = lock_period(
                period.owner_id,
                period.warehouse_id,
                period.label,
                period.start_date,
                period.end_date,
                by_user=request.user,
            )
        except BillingCloseBlocked as exc:
            return Response(
                {"code": "BILLING_CLOSE_BLOCKED", "readiness": exc.readiness},
                status=status.HTTP_409_CONFLICT,
            )
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(locked_period).data)

    @action(detail=True, methods=["post"], url_path="invoice")
    @transaction.atomic
    def invoice(self, request, pk=None):
        _require_billing_perm(request, "add_bill")
        period = self.get_object()
        blocked = self._guard_status(period, [PeriodStatus.CLOSED])
        if blocked is not None:
            return blocked

        payload = BillingPeriodInvoiceSerializer(data=request.data)
        payload.is_valid(raise_exception=True)

        invoice_no = payload.validated_data.get("invoice_no")
        if not invoice_no:
            seq = (
                Bill.objects.filter(
                    period__owner=period.owner, period__warehouse=period.warehouse
                ).count()
                + 1
            )
            invoice_no = f"INV-{period.label}-{period.owner_id}-{period.warehouse_id}-{seq:04d}"

        try:
            bill = generate_invoice_for_period(
                period,
                invoice_no=invoice_no,
                issue_date=payload.validated_data.get("issue_date"),
                due_date=payload.validated_data.get("due_date"),
                by_user=request.user,
            )
        except BillingCloseBlocked as exc:
            return Response(
                {"code": "BILLING_CLOSE_BLOCKED", "readiness": exc.readiness},
                status=status.HTTP_409_CONFLICT,
            )
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(
            BillDetailSerializer(bill, context={"request": request}).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="preview-lock")
    def preview_lock(self, request, pk=None):
        period = self.get_object()
        blocked = self._guard_status(period, [PeriodStatus.OPEN])
        if blocked is not None:
            return blocked

        result = preview_lock_period(
            period.owner_id,
            period.warehouse_id,
            period.label,
            period.start_date,
            period.end_date,
        )
        return Response(result)

    @action(detail=True, methods=["post"], url_path="unlock")
    @transaction.atomic
    def unlock(self, request, pk=None):
        _require_billing_perm(request, "change_billingperiod")
        period = self.get_object()
        blocked = self._guard_status(period, [PeriodStatus.CLOSED, PeriodStatus.INVOICED])
        if blocked is not None:
            return blocked

        payload = UnlockPeriodSerializer(data=request.data)
        payload.is_valid(raise_exception=True)

        try:
            result = unlock_period(
                period,
                by_user=request.user,
                reason=payload.validated_data.get("reason", ""),
            )
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(result)


class BillViewSet(OwnerWarehouseScopedQuerysetMixin, viewsets.ReadOnlyModelViewSet):
    queryset = (
        Bill.objects.select_related("owner", "warehouse", "period")
        .prefetch_related("lines__accrual", "payment_allocations__receipt")
        .all()
    )
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = {
        "owner": ["exact"],
        "warehouse": ["exact"],
        "period": ["exact"],
        "status": ["exact", "in"],
        "document_status": ["exact", "in"],
        "payment_status": ["exact", "in"],
        "issue_date": ["exact", "gte", "lte"],
        "due_date": ["exact", "gte", "lte"],
    }
    search_fields = ["invoice_no", "memo", "period__label"]
    ordering_fields = ["id", "issue_date", "due_date", "subtotal", "tax_total", "total"]
    ordering = ["-issue_date", "-id"]

    def get_serializer_class(self):  # type: ignore[override]
        if self.action == "retrieve":
            return BillDetailSerializer
        return BillListSerializer

    def _xlsx_response(self, workbook: Workbook, filename: str) -> HttpResponse:
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["get"], url_path="export")
    def export_list(self, request):
        _require_financial_export_perm(request)
        qs = self.filter_queryset(self.get_queryset())

        workbook = Workbook()
        info = workbook.active
        info.title = "报告信息"
        info.append(["字段", "值"])
        info.append(["报告类型", "账单列表"])
        info.append(["仓库", request.query_params.get("warehouse") or "全部授权仓库"])
        info.append(["货主", request.query_params.get("owner") or "全部货主"])
        info.append(
            [
                "日期范围",
                (
                    f"{request.query_params.get('issue_date__gte', '')} ~ "
                    f"{request.query_params.get('issue_date__lte', '')}"
                ),
            ]
        )
        info.append(["数据截至时间", timezone.now().isoformat()])
        info.append(["数据口径", "原币种；单据状态与回款状态分离"])
        sheet = workbook.create_sheet("Bills")
        sheet.append(
            [
                "Invoice No",
                "Status",
                "Document Status",
                "Payment Status",
                "Owner",
                "Warehouse",
                "Period",
                "Issue Date",
                "Due Date",
                "Currency",
                "Subtotal",
                "Tax Total",
                "Total",
                "Line Count",
                "Memo",
            ]
        )

        for bill in qs:
            lines = list(bill.lines.all())
            sheet.append(
                [
                    bill.invoice_no,
                    bill.status,
                    bill.document_status,
                    bill.payment_status,
                    getattr(bill.owner, "name", "") if bill.owner_id else "",
                    getattr(bill.warehouse, "name", "") if bill.warehouse_id else "",
                    getattr(bill.period, "label", "") if bill.period_id else "",
                    bill.issue_date.isoformat() if bill.issue_date else "",
                    bill.due_date.isoformat() if bill.due_date else "",
                    bill.currency,
                    Decimal(bill.subtotal),
                    Decimal(bill.tax_total),
                    Decimal(bill.total),
                    len(lines),
                    bill.memo or "",
                ]
            )

        response = self._xlsx_response(
            workbook,
            f"billing-bills-{datetime.date.today().isoformat()}.xlsx",
        )
        record_audit_event(
            action="EXPORT",
            module="billing.bill.list",
            request=request,
            metadata={"rows": qs.count()},
        )
        return response

    @action(detail=True, methods=["get"], url_path="export")
    def export_detail(self, request, pk=None):
        _require_financial_export_perm(request)
        bill = self.get_object()
        workbook = Workbook()

        info = workbook.active
        info.title = "报告信息"
        info.append(["字段", "值"])
        info.append(["报告类型", "单张账单明细"])
        info.append(["仓库", getattr(bill.warehouse, "name", bill.warehouse_id)])
        info.append(["货主", getattr(bill.owner, "name", bill.owner_id)])
        info.append(["日期范围", f"{bill.period.start_date} ~ {bill.period.end_date}"])
        info.append(["币种", bill.currency])
        info.append(["数据截至时间", timezone.now().isoformat()])
        summary_sheet = workbook.create_sheet("Bill")
        summary_sheet.append(["Field", "Value"])
        summary_sheet.append(["Invoice No", bill.invoice_no])
        summary_sheet.append(["Status", bill.status])
        summary_sheet.append(["Document Status", bill.document_status])
        summary_sheet.append(["Payment Status", bill.payment_status])
        summary_sheet.append(["Paid Amount", bill.paid_amount])
        summary_sheet.append(["Outstanding Amount", bill.outstanding_amount])
        summary_sheet.append(["Owner", getattr(bill.owner, "name", "") if bill.owner_id else ""])
        summary_sheet.append(
            [
                "Warehouse",
                getattr(bill.warehouse, "name", "") if bill.warehouse_id else "",
            ]
        )
        summary_sheet.append(
            ["Period", getattr(bill.period, "label", "") if bill.period_id else ""]
        )
        summary_sheet.append(["Issue Date", bill.issue_date.isoformat() if bill.issue_date else ""])
        summary_sheet.append(["Due Date", bill.due_date.isoformat() if bill.due_date else ""])
        summary_sheet.append(["Currency", bill.currency])
        summary_sheet.append(["Subtotal", Decimal(bill.subtotal)])
        summary_sheet.append(["Tax Total", Decimal(bill.tax_total)])
        summary_sheet.append(["Total", Decimal(bill.total)])
        summary_sheet.append(["Memo", bill.memo or ""])

        lines_sheet = workbook.create_sheet("Lines")
        lines_sheet.append(
            [
                "Service Date",
                "Charge Type",
                "Quantity",
                "Unit Price",
                "Amount",
                "Tax Amount",
                "Description",
                "Accrual Fingerprint",
            ]
        )
        for line in bill.lines.all():
            lines_sheet.append(
                [
                    line.service_date.isoformat() if line.service_date else "",
                    line.charge_type,
                    Decimal(line.quantity),
                    Decimal(line.unit_price),
                    Decimal(line.amount),
                    Decimal(line.tax_amount),
                    line.description or "",
                    getattr(line.accrual, "acc_fingerprint", ""),
                ]
            )

        invoice_token = "".join(
            ch if ch.isalnum() or ch in ("-", "_") else "-"
            for ch in bill.invoice_no or f"bill-{bill.id}"
        )
        return self._xlsx_response(workbook, f"{invoice_token or f'bill-{bill.id}'}.xlsx")
