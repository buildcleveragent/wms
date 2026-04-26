import datetime
import csv
import io
import threading
import tempfile
from decimal import Decimal
from unittest import mock

from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import close_old_connections
from django.db.utils import IntegrityError
from django.db import models
from django.test import TestCase, TransactionTestCase
from django.utils import timezone
from django.urls import reverse
from rest_framework.test import APIClient, APIRequestFactory, force_authenticate

from allapp.baseinfo.models import Customer, Owner
from allapp.core.choices import InvTxType
from allapp.billing.enums import (
    AccrualStatus,
    BillStatus,
    BundleScope,
    BundleType,
    CalcMethod,
    CapMode,
    ChargeType,
    MetricType,
    PeriodStatus,
)
from allapp.billing.models import (
    Bill,
    BillLine,
    BillingAccrual,
    BillingEvent,
    BillingJobRun,
    BillingMetricDaily,
    BillingPeriod,
    BillingRule,
    BillingRuleTier,
)
from allapp.billing.services import (
    accrue_order_processing_for_task,
    accrue_order_processing_from_posted,
    accrue_storage_for_date,
    generate_invoice_for_period,
    generate_metrics_for_date,
    lock_period,
    run_scheduled_metric_generation_for_date,
    unlock_period,
)
from allapp.billing.services._common import _compute_fee_with_rule
from allapp.inventory.models import InventoryDetail, InventorySnapshotDaily, InventoryTransaction, PostingJournal
from allapp.inventory.snapshot_services import generate_inventory_snapshot_for_date
from allapp.locations.models import Location, Subwarehouse, Warehouse
from allapp.products.models import Product, ProductUom
from allapp.tasking.models import TaskScanLog, WmsTask, WmsTaskLine
from allapp.tasking.plugins.handlers import DefaultPostingHandler
from wmsmaster.views import profile_view
from allapp.tasking.plugins.handlers import DefaultPostingHandler


class BillingServiceTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(name="Owner A", code="OWNA")
        self.other_owner = Owner.objects.create(name="Owner B", code="OWNB")
        self.warehouse = Warehouse.objects.create(code="WH1", name="Warehouse 1")
        self.user = get_user_model().objects.create_user(
            username="billing-user",
            password="x",
            warehouse=self.warehouse,
        )

    def _create_rule(
        self,
        *,
        owner=None,
        charge_type=ChargeType.DISPATCH,
        calc_method=CalcMethod.PER_ORDER,
        unit_price="10.00",
        effective_from=None,
        effective_to=None,
        bundle_key="",
        bundle_scope=BundleScope.NONE,
        bundle_type=BundleType.CAP,
        bundle_price=None,
    ):
        return BillingRule.objects.create(
            owner=owner or self.owner,
            warehouse=self.warehouse,
            charge_type=charge_type,
            calc_method=calc_method,
            unit_price=Decimal(unit_price),
            effective_from=effective_from,
            effective_to=effective_to,
            bundle_key=bundle_key,
            bundle_scope=bundle_scope,
            bundle_type=bundle_type,
            bundle_price=Decimal(bundle_price) if bundle_price is not None else None,
        )

    def _create_task(self, task_no: str, task_type=WmsTask.TaskType.DISPATCH):
        return WmsTask.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            task_no=task_no,
            task_type=task_type,
        )

    def _create_scan_log(self, task, task_line, service_date: datetime.date, *, fp: str, label_key: str):
        posted_at = datetime.datetime.combine(service_date, datetime.time(10, 0))
        return TaskScanLog.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            task=task,
            task_line=task_line,
            status=TaskScanLog.ScanStatus.OK,
            qty_base_delta=Decimal("1"),
            fp=fp,
            label_key=label_key,
            scan_snapshot_rev=0,
            posted_at=posted_at,
        )

    def _create_outbound_order_line(
        self,
        service_date: datetime.date,
        *,
        suffix: str,
        qty="1.000",
        price="10.0000",
        final_line_amount="0.00",
    ):
        from allapp.outbound.models import OutboundOrder, OutboundOrderLine

        uom = ProductUom.objects.create(code=f"PCS-{suffix}", name=f"件-{suffix}", is_active=True)
        product = Product.objects.create(
            owner=self.owner,
            code=f"SKU-{suffix}",
            name=f"Product {suffix}",
            sku=f"SKU-{suffix}",
            base_uom=uom,
            price=Decimal(price),
        )
        customer = Customer.objects.create(
            owner=self.owner,
            salesperson=self.user,
            code=f"CUST-{suffix}",
            name=f"Customer {suffix}",
        )
        order = OutboundOrder.objects.create(
            owner=self.owner,
            customer=customer,
            warehouse=self.warehouse,
            order_no=f"OUT-{suffix}",
            biz_date=service_date,
            submit_status="SUBMITTED",
            approval_status="OWNER_APPROVED",
            created_by=self.user,
        )
        order_line = OutboundOrderLine.objects.create(
            order=order,
            product=product,
            base_qty=Decimal(qty),
            base_price=Decimal(price),
            base_uom=uom,
            line_no=10,
            final_line_amount=Decimal(final_line_amount),
        )
        return order, order_line, product

    def _create_accrual(
        self,
        *,
        owner=None,
        rule=None,
        amount="10.00",
        service_date=None,
        bundle_key="",
        period=None,
        status=AccrualStatus.OPEN,
        fingerprint: str,
    ):
        owner = owner or self.owner
        rule = rule or self._create_rule(owner=owner)
        return BillingAccrual.objects.create(
            owner=owner,
            warehouse=self.warehouse,
            period=period,
            charge_type=rule.charge_type,
            rule=rule,
            service_date=service_date or datetime.date(2026, 3, 1),
            currency="CNY",
            quantity=Decimal("1"),
            unit_price=Decimal(amount),
            amount=Decimal(amount),
            tax_amount=Decimal("0.00"),
            status=status,
            bundle_key=bundle_key,
            acc_fingerprint=fingerprint,
            created_by=self.user,
        )

    def test_accrue_order_processing_creates_one_accrual_per_order(self):
        service_date = datetime.date(2026, 3, 1)
        self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="10.00")

        task = self._create_task("TASK-ORDER-1")
        line1 = WmsTaskLine.objects.create(task=task)
        line2 = WmsTaskLine.objects.create(task=task)
        self._create_scan_log(task, line1, service_date, fp="fp-order-1", label_key="LBL-ORDER-1")
        self._create_scan_log(task, line2, service_date, fp="fp-order-2", label_key="LBL-ORDER-2")

        def resolver(task_line):
            return {"order_ids": {1000 + task_line.id}}

        with mock.patch("allapp.billing.services.accrual._load_taskline_order_resolver", return_value=resolver):
            events, accruals = accrue_order_processing_from_posted(
                self.owner.id,
                self.warehouse.id,
                service_date,
                service_date,
                by_user=self.user,
            )

        self.assertEqual(events, 2)
        self.assertEqual(accruals, 2)
        self.assertEqual(BillingAccrual.objects.count(), 2)
        self.assertEqual(
            BillingAccrual.objects.aggregate(total=models.Sum("amount"))["total"],
            Decimal("20.00"),
        )

    def test_accrue_order_processing_selects_rule_by_service_date(self):
        day1 = datetime.date(2026, 3, 1)
        day2 = datetime.date(2026, 3, 2)
        self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="10.00", effective_to=day1)
        self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="20.00", effective_from=day2)

        task = self._create_task("TASK-ORDER-2")
        line1 = WmsTaskLine.objects.create(task=task)
        line2 = WmsTaskLine.objects.create(task=task)
        self._create_scan_log(task, line1, day1, fp="fp-date-1", label_key="LBL-DATE-1")
        self._create_scan_log(task, line2, day2, fp="fp-date-2", label_key="LBL-DATE-2")

        def resolver(task_line):
            return {"order_ids": {2000 + task_line.id}}

        with mock.patch("allapp.billing.services.accrual._load_taskline_order_resolver", return_value=resolver):
            accrue_order_processing_from_posted(
                self.owner.id,
                self.warehouse.id,
                day1,
                day2,
                by_user=self.user,
            )

        amounts = {
            accrual.service_date: accrual.amount
            for accrual in BillingAccrual.objects.order_by("service_date")
        }
        self.assertEqual(amounts[day1], Decimal("10.00"))
        self.assertEqual(amounts[day2], Decimal("20.00"))

    def test_accrue_order_processing_resolves_direct_outbound_order_line_binding(self):
        service_date = datetime.date(2026, 3, 3)
        self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="10.00")
        _order, order_line, product = self._create_outbound_order_line(
            service_date,
            suffix="DIRECT",
        )

        task = self._create_task("TASK-ORDER-DIRECT")
        task_line = WmsTaskLine.objects.create(
            task=task,
            product=product,
            src_model="OutboundOrderLine",
            src_id=order_line.id,
        )
        self._create_scan_log(task, task_line, service_date, fp="fp-direct-order", label_key="LBL-DIRECT")

        events, accruals = accrue_order_processing_from_posted(
            self.owner.id,
            self.warehouse.id,
            service_date,
            service_date,
            by_user=self.user,
        )

        self.assertEqual(events, 1)
        self.assertEqual(accruals, 1)
        accrual = BillingAccrual.objects.get()
        self.assertEqual(accrual.amount, Decimal("10.00"))

    def test_accrue_order_processing_resolves_nested_task_line_binding(self):
        service_date = datetime.date(2026, 3, 4)
        self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="12.00")
        _order, order_line, product = self._create_outbound_order_line(
            service_date,
            suffix="CHAIN",
        )

        pick_task = self._create_task("TASK-PICK-CHAIN", task_type=WmsTask.TaskType.PICK)
        pick_line = WmsTaskLine.objects.create(
            task=pick_task,
            product=product,
            src_model="OutboundOrderLine",
            src_id=order_line.id,
        )
        review_task = self._create_task("TASK-REVIEW-CHAIN", task_type=WmsTask.TaskType.REVIEW)
        review_line = WmsTaskLine.objects.create(
            task=review_task,
            product=product,
            src_model="wmstaskline",
            src_id=pick_line.id,
        )
        dispatch_task = self._create_task("TASK-DISPATCH-CHAIN")
        dispatch_line = WmsTaskLine.objects.create(
            task=dispatch_task,
            product=product,
            src_model="WmsTaskLine",
            src_id=review_line.id,
        )
        self._create_scan_log(dispatch_task, dispatch_line, service_date, fp="fp-chain-order", label_key="LBL-CHAIN")

        events, accruals = accrue_order_processing_from_posted(
            self.owner.id,
            self.warehouse.id,
            service_date,
            service_date,
            by_user=self.user,
        )

        self.assertEqual(events, 1)
        self.assertEqual(accruals, 1)
        accrual = BillingAccrual.objects.get()
        self.assertEqual(accrual.amount, Decimal("12.00"))

    def test_accrue_order_processing_uses_base_amount_when_final_line_amount_is_zero(self):
        service_date = datetime.date(2026, 3, 5)
        self._create_rule(
            calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            unit_price="0.1000",
        )
        _order, order_line, product = self._create_outbound_order_line(
            service_date,
            suffix="PCTZERO",
            qty="2.000",
            price="15.0000",
            final_line_amount="0.00",
        )

        task = self._create_task("TASK-ORDER-PCTZERO")
        task_line = WmsTaskLine.objects.create(
            task=task,
            product=product,
            src_model="OutboundOrderLine",
            src_id=order_line.id,
        )
        self._create_scan_log(task, task_line, service_date, fp="fp-pct-order", label_key="LBL-PCTZERO")

        events, accruals = accrue_order_processing_from_posted(
            self.owner.id,
            self.warehouse.id,
            service_date,
            service_date,
            by_user=self.user,
        )

        self.assertEqual(events, 1)
        self.assertEqual(accruals, 1)
        accrual = BillingAccrual.objects.get()
        self.assertEqual(accrual.quantity, Decimal("30.00"))
        self.assertEqual(accrual.amount, Decimal("3.00"))

    def test_lock_period_uses_bundle_rule_scoped_to_owner_and_warehouse(self):
        service_date = datetime.date(2026, 3, 1)
        owner_rule = self._create_rule(
            owner=self.owner,
            bundle_key="BUNDLE-A",
            bundle_scope=BundleScope.PER_PERIOD,
            bundle_type=BundleType.CAP,
            bundle_price="100.00",
        )
        self._create_rule(
            owner=self.other_owner,
            bundle_key="BUNDLE-A",
            bundle_scope=BundleScope.PER_PERIOD,
            bundle_type=BundleType.CAP,
            bundle_price="1.00",
        )

        self._create_accrual(
            owner=self.owner,
            rule=owner_rule,
            amount="60.00",
            service_date=service_date,
            bundle_key="BUNDLE-A",
            fingerprint="acc-bundle-scope-1",
        )
        self._create_accrual(
            owner=self.owner,
            rule=owner_rule,
            amount="70.00",
            service_date=service_date + datetime.timedelta(days=1),
            bundle_key="BUNDLE-A",
            fingerprint="acc-bundle-scope-2",
        )

        period = lock_period(
            self.owner.id,
            self.warehouse.id,
            "2026-03-A",
            service_date,
            service_date + datetime.timedelta(days=1),
        )

        total = BillingAccrual.objects.filter(period=period).aggregate(total=models.Sum("amount"))["total"]
        self.assertEqual(period.status, PeriodStatus.CLOSED)
        self.assertEqual(total, Decimal("100.00"))

    def test_lock_period_fixed_bundle_hits_target_total(self):
        start_date = datetime.date(2026, 3, 10)
        fixed_rule = self._create_rule(
            bundle_key="FIXED-A",
            bundle_scope=BundleScope.PER_PERIOD,
            bundle_type=BundleType.FIXED,
            bundle_price="30.00",
        )

        self._create_accrual(
            rule=fixed_rule,
            amount="50.00",
            service_date=start_date,
            bundle_key="FIXED-A",
            fingerprint="acc-fixed-1",
        )
        self._create_accrual(
            rule=fixed_rule,
            amount="40.00",
            service_date=start_date + datetime.timedelta(days=1),
            bundle_key="FIXED-A",
            fingerprint="acc-fixed-2",
        )
        self._create_accrual(
            rule=fixed_rule,
            amount="10.00",
            service_date=start_date + datetime.timedelta(days=2),
            bundle_key="FIXED-A",
            fingerprint="acc-fixed-3",
        )

        period = lock_period(
            self.owner.id,
            self.warehouse.id,
            "2026-03-FIXED",
            start_date,
            start_date + datetime.timedelta(days=2),
        )

        accruals = list(BillingAccrual.objects.filter(period=period).order_by("service_date", "id"))
        self.assertEqual(sum((a.amount for a in accruals), Decimal("0.00")), Decimal("30.00"))
        self.assertTrue(all(a.amount >= 0 for a in accruals))

    def test_lock_period_rejects_non_open_period(self):
        BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-03-CLOSED",
            start_date=datetime.date(2026, 3, 1),
            end_date=datetime.date(2026, 3, 31),
            status=PeriodStatus.CLOSED,
        )

        with self.assertRaises(ValueError):
            lock_period(
                self.owner.id,
                self.warehouse.id,
                "2026-03-CLOSED",
                datetime.date(2026, 3, 1),
                datetime.date(2026, 3, 31),
            )

    def test_lock_period_rejects_existing_open_period_with_different_dates(self):
        BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-03-RANGE",
            start_date=datetime.date(2026, 3, 1),
            end_date=datetime.date(2026, 3, 31),
            status=PeriodStatus.OPEN,
        )

        with self.assertRaises(ValueError):
            lock_period(
                self.owner.id,
                self.warehouse.id,
                "2026-03-RANGE",
                datetime.date(2026, 3, 5),
                datetime.date(2026, 3, 20),
            )

    def test_lock_period_rejects_when_reconciliation_gate_fails(self):
        rule = self._create_rule()
        accrual = self._create_accrual(
            rule=rule,
            amount="15.00",
            service_date=datetime.date(2026, 3, 6),
            fingerprint="acc-lock-gate-fail",
        )
        BillingAccrual.objects.filter(pk=accrual.pk).update(charge_type=ChargeType.STORAGE)

        with self.assertRaises(ValueError) as exc:
            lock_period(
                self.owner.id,
                self.warehouse.id,
                "2026-03-GATE",
                datetime.date(2026, 3, 6),
                datetime.date(2026, 3, 6),
            )

        self.assertIn("数据对账未通过", str(exc.exception))

    def test_generate_invoice_requires_closed_period_and_is_single_use(self):
        open_period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-03-OPEN",
            start_date=datetime.date(2026, 3, 1),
            end_date=datetime.date(2026, 3, 31),
            status=PeriodStatus.OPEN,
        )
        rule = self._create_rule()
        self._create_accrual(
            rule=rule,
            amount="10.00",
            service_date=datetime.date(2026, 3, 1),
            period=open_period,
            status=AccrualStatus.LOCKED,
            fingerprint="acc-open-invoice",
        )

        with self.assertRaises(ValueError):
            generate_invoice_for_period(open_period, invoice_no="INV-OPEN")

        closed_period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-03-CINV",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.CLOSED,
        )
        self._create_accrual(
            rule=rule,
            amount="15.00",
            service_date=datetime.date(2026, 4, 2),
            period=closed_period,
            status=AccrualStatus.LOCKED,
            fingerprint="acc-closed-invoice",
        )

        bill = generate_invoice_for_period(closed_period, invoice_no="INV-CLOSED")

        self.assertEqual(Bill.objects.filter(period=closed_period).count(), 1)
        self.assertEqual(bill.status, "ISSUED")
        with self.assertRaises(ValueError):
            generate_invoice_for_period(closed_period, invoice_no="INV-CLOSED-2")

    def test_generate_invoice_rejects_when_reconciliation_gate_fails(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-GATE",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.CLOSED,
        )
        rule = self._create_rule()
        accrual = self._create_accrual(
            rule=rule,
            amount="15.00",
            service_date=datetime.date(2026, 4, 2),
            period=period,
            status=AccrualStatus.LOCKED,
            fingerprint="acc-invoice-gate-fail",
        )
        BillingAccrual.objects.filter(pk=accrual.pk).update(charge_type=ChargeType.STORAGE)

        with self.assertRaises(ValueError) as exc:
            generate_invoice_for_period(period, invoice_no="INV-GATE-FAIL")

        self.assertIn("数据对账未通过", str(exc.exception))


    def test_accrue_order_processing_can_filter_allowed_methods(self):
        service_date = datetime.date(2026, 3, 6)
        self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="10.00")
        self._create_rule(
            calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            unit_price="0.1000",
        )
        _order, order_line, product = self._create_outbound_order_line(
            service_date,
            suffix="FILTER",
            qty="2.000",
            price="15.0000",
            final_line_amount="0.00",
        )

        task = self._create_task("TASK-ORDER-FILTER")
        task_line = WmsTaskLine.objects.create(
            task=task,
            product=product,
            src_model="OutboundOrderLine",
            src_id=order_line.id,
        )
        self._create_scan_log(task, task_line, service_date, fp="fp-filter-order", label_key="LBL-FILTER")

        events, accruals = accrue_order_processing_from_posted(
            self.owner.id,
            self.warehouse.id,
            service_date,
            service_date,
            by_user=self.user,
            allowed_methods={CalcMethod.PER_ORDER},
        )

        self.assertEqual(events, 1)
        self.assertEqual(accruals, 1)
        self.assertEqual(BillingAccrual.objects.count(), 1)
        accrual = BillingAccrual.objects.get()
        self.assertEqual(accrual.rule.calc_method, CalcMethod.PER_ORDER)
        self.assertEqual(accrual.amount, Decimal("10.00"))

    def test_handler_triggers_review_order_processing_with_allowed_methods(self):
        review_task = self._create_task("TASK-REVIEW-AUTO", task_type=WmsTask.TaskType.REVIEW)
        handler = DefaultPostingHandler()
        now_ts = timezone.make_aware(datetime.datetime(2026, 3, 7, 10, 0, 0))

        with mock.patch.object(DefaultPostingHandler, "_handle_atomic", return_value=1), \
             mock.patch("allapp.billing.services.accrue_for_posting") as accrue_for_posting_mock, \
             mock.patch("allapp.billing.services.accrue_order_processing_for_task") as accrue_for_task_mock:
            affected = handler.handle(task=review_task, now=now_ts, note="AUTO", by_user=self.user)

        self.assertEqual(affected, 1)
        accrue_for_posting_mock.assert_called_once()
        accrue_for_task_mock.assert_called_once_with(
            review_task,
            mock.ANY,
            by_user=self.user,
            allowed_methods={
                CalcMethod.PER_ORDER,
                CalcMethod.PER_ORDER_LINE,
                CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            },
        )

    def test_handler_does_not_trigger_order_processing_for_non_review_tasks(self):
        pick_task = self._create_task("TASK-PICK-AUTO", task_type=WmsTask.TaskType.PICK)
        handler = DefaultPostingHandler()
        now_ts = timezone.make_aware(datetime.datetime(2026, 3, 8, 10, 0, 0))

        with mock.patch.object(DefaultPostingHandler, "_handle_atomic", return_value=1), \
             mock.patch("allapp.billing.services.accrue_for_posting") as accrue_for_posting_mock, \
             mock.patch("allapp.billing.services.accrue_order_processing_for_task") as accrue_for_task_mock:
            affected = handler.handle(task=pick_task, now=now_ts, note="AUTO", by_user=self.user)

        self.assertEqual(affected, 1)
        accrue_for_posting_mock.assert_called_once()
        accrue_for_task_mock.assert_not_called()

    def test_accrue_order_processing_for_task_uses_precise_order_line_amount_not_daily_metric(self):
        service_date = datetime.date(2026, 3, 9)
        self._create_rule(
            calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            unit_price="0.1000",
        )
        order, order_line, product = self._create_outbound_order_line(
            service_date,
            suffix="TASK-AMT",
            qty="2.000",
            price="15.0000",
            final_line_amount="0.00",
        )

        task = self._create_task("TASK-REVIEW-AMT", task_type=WmsTask.TaskType.REVIEW)
        task_line = WmsTaskLine.objects.create(
            task=task,
            product=product,
            src_model="OutboundOrderLine",
            src_id=order_line.id,
        )
        posting_journal = PostingJournal.objects.create(
            src_model="WmsTask",
            src_id=task.id,
            tx_type="POST",
            status="PENDING",
        )
        TaskScanLog.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            task=task,
            task_line=task_line,
            posting_journal=posting_journal,
            status=TaskScanLog.ScanStatus.OK,
            qty_base_delta=Decimal("1"),
            fp="fp-review-amt-task",
            label_key="LBL-REVIEW-AMT-TASK",
            scan_snapshot_rev=0,
            posted_at=datetime.datetime.combine(service_date, datetime.time(10, 0)),
        )

        # 故意写一个夸张的日指标，证明新函数不会再用它
        BillingMetricDaily.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type=MetricType.ORDER_AMT,
            value=Decimal("9999.0000"),
            source="TEST",
        )

        def resolver(_task_line):
            return {
                "order_ids": {order.id},
                "order_line_ids": {(order.id, order_line.id)},
            }

        with mock.patch(
            "allapp.billing.services.accrual._load_taskline_order_resolver",
            return_value=resolver,
        ):
            events, accruals = accrue_order_processing_for_task(
                task,
                posting_journal,
                by_user=self.user,
                allowed_methods={CalcMethod.PERCENT_OF_ORDER_AMOUNT},
            )

        self.assertEqual(events, 1)
        self.assertEqual(accruals, 1)
        accrual = BillingAccrual.objects.get(rule__calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT)

        # 当前 task 真实金额 = 2 * 15 = 30，不应使用 9999 的日指标
        self.assertEqual(accrual.quantity, Decimal("30.00"))
        self.assertEqual(accrual.amount, Decimal("3.00"))

    def test_handler_triggers_review_task_level_order_processing(self):
        review_task = self._create_task("TASK-REVIEW-TASK-LEVEL", task_type=WmsTask.TaskType.REVIEW)
        handler = DefaultPostingHandler()
        now_ts = timezone.make_aware(datetime.datetime(2026, 3, 10, 10, 0, 0))

        with mock.patch.object(DefaultPostingHandler, "_handle_atomic", return_value=1), \
             mock.patch("allapp.billing.services.accrue_for_posting") as accrue_for_posting_mock, \
             mock.patch("allapp.billing.services.accrue_order_processing_for_task") as accrue_for_task_mock:
            affected = handler.handle(task=review_task, now=now_ts, note="AUTO", by_user=self.user)

        self.assertEqual(affected, 1)
        accrue_for_posting_mock.assert_called_once()
        accrue_for_task_mock.assert_called_once_with(
            review_task,
            mock.ANY,
            by_user=self.user,
            allowed_methods={
                CalcMethod.PER_ORDER,
                CalcMethod.PER_ORDER_LINE,
                CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            },
        )

    # ── Bugfix regression tests ──────────────────────────────────────────

    def test_storage_fingerprint_unique_per_owner(self):
        """A-2: 两个 owner 同量同日应各自生成独立的 BillingEvent。"""
        from allapp.billing.enums import LadderMode
        service_date = datetime.date(2026, 3, 15)
        other_owner = Owner.objects.create(name="Owner B", code="OWNB")

        for ow in (self.owner, other_owner):
            BillingRule.objects.create(
                owner=ow, warehouse=self.warehouse,
                charge_type=ChargeType.STORAGE, calc_method=CalcMethod.PER_DAY_ONHAND_BASE,
                unit_price=Decimal("1.00"),
            )

        from allapp.inventory.models import InventoryDetail
        for ow in (self.owner, other_owner):
            InventoryDetail.objects.create(
                owner=ow, warehouse=self.warehouse,
                product=Product.objects.create(
                    owner=ow, code=f"SKU-STORE-{ow.code}", name=f"P-{ow.code}",
                    sku=f"SKU-STORE-{ow.code}",
                    base_uom=ProductUom.objects.create(code=f"EA-{ow.code}", name=f"EA-{ow.code}", is_active=True),
                ),
                location=Location.objects.create(
                    warehouse=self.warehouse, code=f"LOC-{ow.code}", name=f"LOC-{ow.code}",
                ),
                onhand_qty=Decimal("100"),
                is_active=True,
            )

        accrue_storage_for_date(self.owner.id, self.warehouse.id, service_date, by_user=self.user)
        accrue_storage_for_date(other_owner.id, self.warehouse.id, service_date, by_user=self.user)

        self.assertEqual(BillingEvent.objects.count(), 2)
        self.assertEqual(BillingAccrual.objects.count(), 2)
        fps = list(BillingEvent.objects.values_list("event_fp", flat=True))
        self.assertNotEqual(fps[0], fps[1])

    def test_task_level_skips_percent_when_batch_exists(self):
        """A-1: batch 已存在时 task-level 跳过 PERCENT_OF_ORDER_AMOUNT。"""
        service_date = datetime.date(2026, 3, 16)
        self._create_rule(
            calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            unit_price="0.1000",
        )
        order, order_line, product = self._create_outbound_order_line(
            service_date, suffix="BATCH-VS-TASK", qty="5.000", price="20.0000",
            final_line_amount="0.00",
        )
        task = self._create_task("TASK-BATCH-CHECK", task_type=WmsTask.TaskType.REVIEW)
        task_line = WmsTaskLine.objects.create(
            task=task, product=product,
            src_model="OutboundOrderLine", src_id=order_line.id,
        )
        posting_journal = PostingJournal.objects.create(
            src_model="WmsTask", src_id=task.id, tx_type="POST", status="PENDING",
        )
        TaskScanLog.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            task=task, task_line=task_line, posting_journal=posting_journal,
            status=TaskScanLog.ScanStatus.OK, qty_base_delta=Decimal("1"),
            fp="fp-batch-vs-task", label_key="LBL-BVT", scan_snapshot_rev=0,
            posted_at=datetime.datetime.combine(service_date, datetime.time(10, 0)),
        )

        def resolver(_tl):
            return {"order_ids": {order.id}, "order_line_ids": {(order.id, order_line.id)}}

        # 先跑 batch 函数
        with mock.patch("allapp.billing.services.accrual._load_taskline_order_resolver", return_value=resolver):
            accrue_order_processing_from_posted(
                self.owner.id, self.warehouse.id, service_date, service_date,
                by_user=self.user,
                allowed_methods={CalcMethod.PERCENT_OF_ORDER_AMOUNT},
            )
        self.assertEqual(
            BillingAccrual.objects.filter(rule__calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT).count(), 1,
        )

        # 再跑 task-level 函数，应该跳过
        with mock.patch("allapp.billing.services.accrual._load_taskline_order_resolver", return_value=resolver):
            events, accruals = accrue_order_processing_for_task(
                task, posting_journal, by_user=self.user,
                allowed_methods={CalcMethod.PERCENT_OF_ORDER_AMOUNT},
            )
        self.assertEqual(events, 0)
        self.assertEqual(accruals, 0)
        # 仍然只有 batch 那一条
        self.assertEqual(
            BillingAccrual.objects.filter(rule__calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT).count(), 1,
        )

    def test_reinvoice_after_unlock_with_void_bill(self):
        """A-3: unlock 后 VOID bill 不应阻止重新开票。"""
        service_date = datetime.date(2026, 3, 17)
        rule = self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="10.00")

        # 手工创建一条 OPEN accrual
        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("10.0000"),
            amount=Decimal("10.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.OPEN,
            acc_fingerprint="fp-reinvoice-test-1",
        )

        period = lock_period(self.owner.id, self.warehouse.id, "T-REINV",
                             service_date, service_date)
        bill = generate_invoice_for_period(period, invoice_no="INV-REINV-1",
                                           issue_date=service_date, due_date=service_date)
        self.assertEqual(bill.status, "ISSUED")

        # unlock（红冲）
        unlock_period(period, by_user=self.user, reason="test")
        period.refresh_from_db()

        # 重建 OPEN accrual 供重新 lock
        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("10.0000"),
            amount=Decimal("10.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.OPEN,
            acc_fingerprint="fp-reinvoice-test-2",
        )

        # 重新 lock 和开票，不应报错
        period2 = lock_period(self.owner.id, self.warehouse.id, "T-REINV-2",
                              service_date, service_date)
        bill2 = generate_invoice_for_period(period2, invoice_no="INV-REINV-2",
                                            issue_date=service_date, due_date=service_date)
        self.assertEqual(bill2.status, "ISSUED")

    def test_unlock_cleans_void_dedup_accruals(self):
        """B-2: unlock 应清理 dedup 产生的 VOID accrual。"""
        service_date = datetime.date(2026, 3, 18)
        rule = self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="10.00")

        # 同一个 event 两条不同价的 accrual（模拟 repricing）
        event = BillingEvent.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, service_date=service_date,
            quantity=Decimal("1"), quantity_uom="ORDER",
            event_fp="fp-dedup-test-event",
        )
        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("10.0000"),
            amount=Decimal("10.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.OPEN, event=event,
            acc_fingerprint="fp-dedup-old",
        )
        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("12.0000"),
            amount=Decimal("12.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.OPEN, event=event,
            acc_fingerprint="fp-dedup-new",
        )

        period = lock_period(self.owner.id, self.warehouse.id, "T-DEDUP",
                             service_date, service_date)

        # lock 后应有 1 条 LOCKED + 1 条 VOID
        self.assertEqual(BillingAccrual.objects.filter(period=period, status=AccrualStatus.LOCKED).count(), 1)
        self.assertEqual(BillingAccrual.objects.filter(period=period, status=AccrualStatus.VOID).count(), 1)

        # unlock
        unlock_period(period, by_user=self.user, reason="test")

        # VOID accrual 应已与 period 解绑
        self.assertEqual(BillingAccrual.objects.filter(period=period).count(), 0)
        self.assertEqual(BillingAccrual.objects.filter(status=AccrualStatus.VOID, period__isnull=True).count(), 1)

    def test_incremental_tier_gap_logs_warning(self):
        """B-3: INCREMENTAL 模式 tier 有空档时应 log warning。"""
        from allapp.billing.enums import LadderMode
        rule = self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="1.00")
        rule.ladder_mode = LadderMode.INCREMENTAL
        rule.save(update_fields=["ladder_mode"])

        # 创建有间隙的 tier: [0,10) 和 [20,∞)，间隙 [10,20)
        BillingRuleTier.objects.create(rule=rule, threshold_from=0, threshold_to=10, unit_price=Decimal("1.00"))
        BillingRuleTier.objects.create(rule=rule, threshold_from=20, threshold_to=None, unit_price=Decimal("2.00"))

        with self.assertLogs("allapp.billing", level="WARNING") as cm:
            amt, eff = _compute_fee_with_rule(rule, Decimal("15"))

        self.assertTrue(any("tier gap" in msg for msg in cm.output))
        # 只有 [0,10) 被定价: 10 * 1.00 = 10.00, [10,15) 未被定价
        self.assertEqual(amt, Decimal("10.00"))

    def test_resolver_max_depth(self):
        """B-6: 超深 task_line 链不应 RecursionError。"""
        from allapp.billing.resolvers import taskline_to_order_mapping

        # 创建 60 层深的 task_line chain（超过 _MAX_RECURSION_DEPTH=50）
        task = self._create_task("TASK-DEEP-CHAIN")
        uom = ProductUom.objects.create(code="EA-DEEP", name="EA-DEEP", is_active=True)
        product = Product.objects.create(
            owner=self.owner, code="SKU-DEEP", name="P-DEEP", sku="SKU-DEEP", base_uom=uom,
        )
        prev_line = None
        for i in range(60):
            line = WmsTaskLine.objects.create(
                task=task, product=product,
                src_model="WmsTaskLine" if prev_line else "",
                src_id=prev_line.id if prev_line else None,
            )
            prev_line = line

        # 从最后一层开始解析，不应爆栈
        with self.assertLogs("allapp.billing", level="WARNING") as cm:
            result = taskline_to_order_mapping(prev_line)

        self.assertTrue(any("max recursion depth" in msg for msg in cm.output))
        self.assertIsNotNone(result)

    def test_void_accrual_not_counted_in_daily_cap(self):
        """日口径 cap 不应把 VOID/reversal accrual 算进已用额度。"""
        from allapp.billing.enums import CapMode
        service_date = datetime.date(2026, 3, 20)
        rule = BillingRule.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("10.00"), cap_mode=CapMode.PER_DAY, cap_amount=Decimal("20.00"),
        )

        # 创建一条 VOID accrual（占 10 额度）
        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("10.0000"),
            amount=Decimal("10.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.VOID, is_reversal=True,
            acc_fingerprint="fp-void-cap-test",
        )

        # 日封顶是 20，VOID 的 10 不应占用额度，所以还能用 20
        from allapp.billing.services._common import _apply_caps_bundles_day
        result = _apply_caps_bundles_day(rule, self.owner.id, self.warehouse.id, service_date, Decimal("15.00"))
        self.assertEqual(result, Decimal("15.00"))

    def test_void_accrual_not_counted_in_period_cap(self):
        """账期口径 cap 不应把 VOID accrual 算进封顶额度。"""
        from allapp.billing.enums import CapMode
        service_date = datetime.date(2026, 3, 21)
        rule = BillingRule.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("10.00"), cap_mode=CapMode.PER_PERIOD, cap_amount=Decimal("25.00"),
        )

        # 创建 3 条 accrual：2 条 OPEN + 1 条 VOID（dedup 残留）
        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("10.0000"),
            amount=Decimal("10.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.OPEN,
            acc_fingerprint="fp-period-cap-1",
        )
        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("10.0000"),
            amount=Decimal("10.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.OPEN,
            acc_fingerprint="fp-period-cap-2",
        )

        period = lock_period(self.owner.id, self.warehouse.id, "T-PCAP",
                             service_date, service_date)

        # 两条有效 accrual 各 10 = 20，cap = 25 → 不应被截断
        locked = BillingAccrual.objects.filter(
            period=period, status=AccrualStatus.LOCKED, is_reversal=False,
        )
        total = sum(a.amount for a in locked)
        self.assertEqual(total, Decimal("20.00"))

    def test_invoiced_period_unlock_reopens_for_rebilling(self):
        """红冲后 period 应回到 OPEN，可以重新 lock 和 invoice。"""
        service_date = datetime.date(2026, 3, 22)
        rule = self._create_rule(calc_method=CalcMethod.PER_ORDER, unit_price="10.00")

        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("10.0000"),
            amount=Decimal("10.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.OPEN,
            acc_fingerprint="fp-rebill-1",
        )

        # lock → invoice → unlock (红冲)
        period = lock_period(self.owner.id, self.warehouse.id, "T-REBILL",
                             service_date, service_date)
        generate_invoice_for_period(period, invoice_no="INV-RB-1",
                                    issue_date=service_date, due_date=service_date)
        result = unlock_period(period, by_user=self.user, reason="test rebill")

        period.refresh_from_db()
        self.assertEqual(result["action"], "red_reversal")
        self.assertEqual(period.status, "OPEN")

        # period 上不应有任何 accrual 残留
        self.assertEqual(BillingAccrual.objects.filter(period=period).count(), 0)

        # 重新创建 accrual → lock → invoice，应成功
        BillingAccrual.objects.create(
            owner=self.owner, warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH, rule=rule,
            service_date=service_date, currency="CNY",
            quantity=Decimal("1"), unit_price=Decimal("12.0000"),
            amount=Decimal("12.00"), tax_amount=Decimal("0.00"),
            status=AccrualStatus.OPEN,
            acc_fingerprint="fp-rebill-2",
        )

        period2 = lock_period(self.owner.id, self.warehouse.id, "T-REBILL-2",
                              service_date, service_date)
        bill2 = generate_invoice_for_period(period2, invoice_no="INV-RB-2",
                                            issue_date=service_date, due_date=service_date)
        self.assertEqual(bill2.status, "ISSUED")
        self.assertEqual(bill2.subtotal, Decimal("12.00"))
    def test_accrue_metrics_for_date_skips_order_amount_metric(self):
        service_date = datetime.date(2026, 3, 17)
        self._create_rule(
            calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            unit_price="0.1000",
        )
        BillingMetricDaily.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type=MetricType.ORDER_AMT,
            value=Decimal("1620.0000"),
            source="TEST",
        )

        events, accruals = accrue_metrics_for_date(
            self.owner.id,
            self.warehouse.id,
            service_date,
            by_user=self.user,
        )

        self.assertEqual(events, 0)
        self.assertEqual(accruals, 0)
        self.assertFalse(
            BillingAccrual.objects.filter(rule__calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT).exists()
        )

    def test_batch_percent_backfills_missing_task_without_rebilling_existing_task(self):
        service_date = datetime.date(2026, 3, 18)
        self._create_rule(
            calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            unit_price="0.1000",
        )
        old_order, old_line, old_product = self._create_outbound_order_line(
            service_date,
            suffix="PCT-OLD",
            final_line_amount="1320.00",
        )
        new_order, new_line, new_product = self._create_outbound_order_line(
            service_date,
            suffix="PCT-NEW",
            final_line_amount="300.00",
        )

        old_task = self._create_task("TASK-PCT-OLD", task_type=WmsTask.TaskType.REVIEW)
        old_task_line = WmsTaskLine.objects.create(
            task=old_task,
            product=old_product,
            src_model="OutboundOrderLine",
            src_id=old_line.id,
        )
        self._create_scan_log(
            old_task,
            old_task_line,
            service_date,
            fp="fp-pct-old",
            label_key="LBL-PCT-OLD",
        )

        new_task = self._create_task("TASK-PCT-NEW", task_type=WmsTask.TaskType.REVIEW)
        new_task_line = WmsTaskLine.objects.create(
            task=new_task,
            product=new_product,
            src_model="OutboundOrderLine",
            src_id=new_line.id,
        )
        posting_journal = PostingJournal.objects.create(
            src_model="WmsTask",
            src_id=new_task.id,
            tx_type="POST",
            status="PENDING",
        )
        TaskScanLog.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            task=new_task,
            task_line=new_task_line,
            posting_journal=posting_journal,
            status=TaskScanLog.ScanStatus.OK,
            qty_base_delta=Decimal("1"),
            fp="fp-pct-new",
            label_key="LBL-PCT-NEW",
            scan_snapshot_rev=0,
            posted_at=datetime.datetime.combine(service_date, datetime.time(11, 0)),
        )

        resolver_map = {
            old_task_line.id: {
                "order_ids": {old_order.id},
                "order_line_ids": {(old_order.id, old_line.id)},
            },
            new_task_line.id: {
                "order_ids": {new_order.id},
                "order_line_ids": {(new_order.id, new_line.id)},
            },
        }

        def resolver(task_line):
            return resolver_map[task_line.id]

        with mock.patch(
            "allapp.billing.services.accrual._load_taskline_order_resolver",
            return_value=resolver,
        ):
            task_events, task_accruals = accrue_order_processing_for_task(
                new_task,
                posting_journal,
                by_user=self.user,
                allowed_methods={CalcMethod.PERCENT_OF_ORDER_AMOUNT},
            )
            batch_events, batch_accruals = accrue_order_processing_from_posted(
                self.owner.id,
                self.warehouse.id,
                service_date,
                service_date,
                by_user=self.user,
                allowed_methods={CalcMethod.PERCENT_OF_ORDER_AMOUNT},
            )

        self.assertEqual(task_events, 1)
        self.assertEqual(task_accruals, 1)
        self.assertEqual(batch_events, 1)
        self.assertEqual(batch_accruals, 1)

        percent_accruals = BillingAccrual.objects.filter(
            rule__calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT
        ).order_by("amount")
        self.assertEqual(percent_accruals.count(), 2)
        self.assertEqual(
            list(percent_accruals.values_list("amount", flat=True)),
            [Decimal("30.00"), Decimal("132.00")],
        )
        self.assertEqual(
            list(percent_accruals.values_list("quantity", flat=True)),
            [Decimal("300.00"), Decimal("1320.00")],
        )
        self.assertFalse(
            BillingAccrual.objects.filter(
                rule__calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
                quantity=Decimal("1620.00"),
            ).exists()
        )


class BillingModelGuardrailTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(name="Owner Guardrail", code="OWGRD")
        self.other_owner = Owner.objects.create(name="Owner Guardrail Other", code="OWGRX")
        self.warehouse = Warehouse.objects.create(code="WHGR1", name="Warehouse Guardrail 1")
        self.other_warehouse = Warehouse.objects.create(code="WHGR2", name="Warehouse Guardrail 2")
        self.user = get_user_model().objects.create_user(
            username="billing-guardrail-user",
            password="x",
            warehouse=self.warehouse,
        )

    def _create_rule(
        self,
        *,
        owner=None,
        warehouse=None,
        charge_type=ChargeType.DISPATCH,
        calc_method=CalcMethod.PER_ORDER,
        unit_price="10.00",
    ):
        return BillingRule.objects.create(
            owner=owner or self.owner,
            warehouse=warehouse or self.warehouse,
            charge_type=charge_type,
            calc_method=calc_method,
            unit_price=Decimal(unit_price),
        )

    def _create_period(self, *, label="2026-05-A", start_date=None, end_date=None):
        start_date = start_date or datetime.date(2026, 5, 1)
        end_date = end_date or datetime.date(2026, 5, 31)
        return BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label=label,
            start_date=start_date,
            end_date=end_date,
            status=PeriodStatus.OPEN,
        )

    def _create_accrual(self, *, rule=None, period=None, fingerprint="guardrail-acc-1"):
        rule = rule or self._create_rule()
        period = period or self._create_period()
        return BillingAccrual.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=period,
            charge_type=rule.charge_type,
            rule=rule,
            service_date=period.start_date,
            currency="CNY",
            quantity=Decimal("1.0000"),
            unit_price=Decimal("10.0000"),
            amount=Decimal("10.00"),
            tax_amount=Decimal("0.00"),
            status=AccrualStatus.LOCKED,
            acc_fingerprint=fingerprint,
            created_by=self.user,
        )

    def test_billing_period_rejects_overlapping_ranges(self):
        self._create_period(label="2026-05-A")

        with self.assertRaises(ValidationError):
            BillingPeriod.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                label="2026-05-B",
                start_date=datetime.date(2026, 5, 15),
                end_date=datetime.date(2026, 6, 14),
                status=PeriodStatus.OPEN,
            )

    def test_billing_rule_tier_rejects_overlapping_ranges(self):
        rule = self._create_rule(charge_type=ChargeType.STORAGE, calc_method=CalcMethod.PER_CBM_DAY)
        BillingRuleTier.objects.create(
            rule=rule,
            threshold_from=Decimal("0.0000"),
            threshold_to=Decimal("100.0000"),
            unit_price=Decimal("1.0000"),
        )

        with self.assertRaises(ValidationError):
            BillingRuleTier.objects.create(
                rule=rule,
                threshold_from=Decimal("50.0000"),
                threshold_to=Decimal("200.0000"),
                unit_price=Decimal("2.0000"),
            )

    def test_billing_rule_rejects_invalid_cap_bundle_and_percent_config(self):
        with self.assertRaises(ValidationError):
            BillingRule.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                charge_type=ChargeType.DISPATCH,
                calc_method=CalcMethod.PER_ORDER,
                unit_price=Decimal("10.0000"),
                cap_mode=CapMode.NONE,
                cap_amount=Decimal("1.00"),
            )

        with self.assertRaises(ValidationError):
            BillingRule.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                charge_type=ChargeType.DISPATCH,
                calc_method=CalcMethod.PER_ORDER,
                unit_price=Decimal("10.0000"),
                bundle_scope=BundleScope.NONE,
                bundle_key="BUNDLE-X",
            )

        with self.assertRaises(ValidationError):
            BillingRule.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                charge_type=ChargeType.DISPATCH,
                calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
                unit_price=Decimal("1.5000"),
            )

    def test_billing_rule_tier_rejects_invalid_numeric_ranges(self):
        rule = self._create_rule(
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            unit_price="0.1000",
        )

        with self.assertRaises(ValidationError):
            BillingRuleTier.objects.create(
                rule=rule,
                threshold_from=Decimal("-1.0000"),
                threshold_to=Decimal("100.0000"),
                percent_rate=Decimal("0.100000"),
            )

        with self.assertRaises(ValidationError):
            BillingRuleTier.objects.create(
                rule=rule,
                threshold_from=Decimal("0.0000"),
                threshold_to=Decimal("100.0000"),
                percent_rate=Decimal("1.500000"),
            )

        with self.assertRaises(ValidationError):
            BillingRuleTier.objects.create(
                rule=self._create_rule(charge_type=ChargeType.STORAGE, calc_method=CalcMethod.PER_CBM_DAY),
                threshold_from=Decimal("0.0000"),
                threshold_to=Decimal("100.0000"),
                unit_price=Decimal("-1.0000"),
            )

    def test_billing_rule_tier_requires_price_field_to_match_calc_method(self):
        percent_rule = self._create_rule(
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PERCENT_OF_ORDER_AMOUNT,
            unit_price="0.1000",
        )
        qty_rule = self._create_rule(
            charge_type=ChargeType.STORAGE,
            calc_method=CalcMethod.PER_CBM_DAY,
            unit_price="10.0000",
        )

        with self.assertRaises(ValidationError):
            BillingRuleTier.objects.create(
                rule=percent_rule,
                threshold_from=Decimal("0.0000"),
                threshold_to=Decimal("100.0000"),
                unit_price=Decimal("0.1000"),
            )

        with self.assertRaises(ValidationError):
            BillingRuleTier.objects.create(
                rule=qty_rule,
                threshold_from=Decimal("0.0000"),
                threshold_to=Decimal("100.0000"),
                percent_rate=Decimal("0.100000"),
            )

    def test_billing_accrual_rejects_rule_scope_mismatch(self):
        rule = self._create_rule(owner=self.owner, warehouse=self.warehouse)

        with self.assertRaises(ValidationError):
            BillingAccrual.objects.create(
                owner=self.other_owner,
                warehouse=self.warehouse,
                charge_type=rule.charge_type,
                rule=rule,
                service_date=datetime.date(2026, 5, 1),
                currency="CNY",
                quantity=Decimal("1.0000"),
                unit_price=Decimal("10.0000"),
                amount=Decimal("10.00"),
                tax_amount=Decimal("0.00"),
                status=AccrualStatus.OPEN,
                acc_fingerprint="guardrail-bad-accrual",
                created_by=self.user,
            )

    def test_billing_accrual_rejects_currency_mismatch_with_rule_and_period(self):
        rule = self._create_rule(owner=self.owner, warehouse=self.warehouse)
        period = self._create_period(label="2026-05-CURRENCY")

        with self.assertRaises(ValidationError):
            BillingAccrual.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                period=period,
                charge_type=rule.charge_type,
                rule=rule,
                service_date=period.start_date,
                currency="USD",
                quantity=Decimal("1.0000"),
                unit_price=Decimal("10.0000"),
                amount=Decimal("10.00"),
                tax_amount=Decimal("0.00"),
                status=AccrualStatus.OPEN,
                acc_fingerprint="guardrail-acc-currency-mismatch",
                created_by=self.user,
            )

    def test_billing_event_rejects_task_scope_mismatch(self):
        task = WmsTask.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            task_no="TASK-GUARDRAIL-1",
            task_type=WmsTask.TaskType.DISPATCH,
        )

        with self.assertRaises(ValidationError):
            BillingEvent.objects.create(
                owner=self.other_owner,
                warehouse=self.warehouse,
                charge_type=ChargeType.DISPATCH,
                service_date=datetime.date(2026, 5, 1),
                task=task,
                quantity=Decimal("1.0000"),
                quantity_uom="BASE",
                event_fp="event-guardrail-1",
            )

    def test_billing_event_rejects_posting_journal_task_mismatch(self):
        task = WmsTask.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            task_no="TASK-GUARDRAIL-2",
            task_type=WmsTask.TaskType.DISPATCH,
        )
        other_task = WmsTask.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            task_no="TASK-GUARDRAIL-3",
            task_type=WmsTask.TaskType.DISPATCH,
        )
        posting_journal = PostingJournal.objects.create(
            src_model="WmsTask",
            src_id=other_task.id,
            tx_type="POST",
        )

        with self.assertRaises(ValidationError):
            BillingEvent.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                charge_type=ChargeType.DISPATCH,
                service_date=datetime.date(2026, 5, 1),
                task=task,
                posting_journal=posting_journal,
                quantity=Decimal("1.0000"),
                quantity_uom="BASE",
                event_fp="event-guardrail-posting-mismatch",
            )

    def test_billing_event_rejects_posting_journal_without_task_context(self):
        posting_journal = PostingJournal.objects.create(
            src_model="WmsTask",
            src_id=999999,
            tx_type="POST",
        )

        with self.assertRaises(ValidationError):
            BillingEvent.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                charge_type=ChargeType.DISPATCH,
                service_date=datetime.date(2026, 5, 1),
                posting_journal=posting_journal,
                quantity=Decimal("1.0000"),
                quantity_uom="BASE",
                event_fp="event-guardrail-posting-without-task",
            )

    def test_billing_metric_daily_rejects_negative_value(self):
        with self.assertRaises(ValidationError):
            BillingMetricDaily.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=datetime.date(2026, 5, 1),
                metric_type="CBM",
                value=Decimal("-1.0000"),
                source="AUTO:TEST",
            )

    def test_billing_job_run_rejects_invalid_attempts_and_time_order(self):
        with self.assertRaises(ValidationError):
            BillingJobRun.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
                service_date=datetime.date(2026, 5, 1),
                attempts=0,
                started_at=timezone.now(),
            )

        started_at = timezone.now()
        with self.assertRaises(ValidationError):
            BillingJobRun.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
                service_date=datetime.date(2026, 5, 2),
                attempts=1,
                started_at=started_at,
                finished_at=started_at - datetime.timedelta(minutes=1),
            )

        with self.assertRaises(ValidationError):
            BillingJobRun.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
                service_date=datetime.date(2026, 5, 3),
                status=BillingJobRun.Status.SUCCESS,
                started_at=started_at,
            )

        with self.assertRaises(ValidationError):
            BillingJobRun.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
                service_date=datetime.date(2026, 5, 4),
                status=BillingJobRun.Status.RUNNING,
                started_at=started_at,
                finished_at=started_at,
            )

    def test_bill_and_billline_reject_duplicate_billing(self):
        period = self._create_period()
        accrual = self._create_accrual(period=period, fingerprint="guardrail-acc-dup")
        bill = Bill.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=period,
            invoice_no="INV-GUARDRAIL-1",
            currency=period.currency,
        )
        BillLine.objects.create(
            bill=bill,
            accrual=accrual,
            charge_type=accrual.charge_type,
            service_date=accrual.service_date,
            quantity=accrual.quantity,
            unit_price=accrual.unit_price,
            amount=accrual.amount,
            tax_amount=accrual.tax_amount,
        )

        with self.assertRaises(ValidationError):
            Bill.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                period=period,
                invoice_no="INV-GUARDRAIL-2",
                currency=period.currency,
            )

        with self.assertRaises(ValidationError):
            BillLine.objects.create(
                bill=bill,
                accrual=accrual,
                charge_type=accrual.charge_type,
                service_date=accrual.service_date,
                quantity=accrual.quantity,
                unit_price=accrual.unit_price,
                amount=accrual.amount,
                tax_amount=accrual.tax_amount,
            )

    def test_billing_accrual_rejects_negative_unit_price_and_tax_amount(self):
        rule = self._create_rule(owner=self.owner, warehouse=self.warehouse)

        with self.assertRaises(ValidationError):
            BillingAccrual.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                charge_type=rule.charge_type,
                rule=rule,
                service_date=datetime.date(2026, 5, 3),
                currency="CNY",
                quantity=Decimal("1.0000"),
                unit_price=Decimal("-1.0000"),
                amount=Decimal("1.00"),
                tax_amount=Decimal("0.00"),
                status=AccrualStatus.OPEN,
                acc_fingerprint="guardrail-acc-negative-unit-price",
                created_by=self.user,
            )

        with self.assertRaises(ValidationError):
            BillingAccrual.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                charge_type=rule.charge_type,
                rule=rule,
                service_date=datetime.date(2026, 5, 4),
                currency="CNY",
                quantity=Decimal("1.0000"),
                unit_price=Decimal("1.0000"),
                amount=Decimal("1.00"),
                tax_amount=Decimal("-0.01"),
                status=AccrualStatus.OPEN,
                acc_fingerprint="guardrail-acc-negative-tax",
                created_by=self.user,
            )

    def test_bill_rejects_invalid_due_date_and_total(self):
        period = self._create_period(label="2026-05-BILL-CHECK")

        with self.assertRaises(ValidationError):
            Bill.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                period=period,
                invoice_no="INV-GUARDRAIL-BAD-DUE",
                issue_date=datetime.date(2026, 5, 10),
                due_date=datetime.date(2026, 5, 9),
                currency=period.currency,
            )

        with self.assertRaises(ValidationError):
            Bill.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                period=period,
                invoice_no="INV-GUARDRAIL-BAD-TOTAL",
                issue_date=datetime.date(2026, 5, 10),
                due_date=datetime.date(2026, 5, 12),
                currency=period.currency,
                subtotal=Decimal("10.00"),
                tax_total=Decimal("1.00"),
                total=Decimal("12.00"),
            )

    def test_bill_rejects_negative_amount_fields(self):
        period = self._create_period(label="2026-05-BILL-NONNEG")

        with self.assertRaises(ValidationError):
            Bill.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                period=period,
                invoice_no="INV-GUARDRAIL-BAD-NEG",
                currency=period.currency,
                subtotal=Decimal("-1.00"),
                tax_total=Decimal("0.00"),
                total=Decimal("-1.00"),
            )

    def test_bill_defaults_issue_date_to_localdate(self):
        period = self._create_period(label="2026-05-BILL-DATE")

        bill = Bill.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=period,
            invoice_no="INV-GUARDRAIL-LOCALDATE",
            currency=period.currency,
        )

        now = timezone.now()
        expected_issue_date = timezone.localtime(now).date() if timezone.is_aware(now) else now.date()
        self.assertEqual(bill.issue_date, expected_issue_date)

    def test_billline_rejects_amount_fields_that_do_not_match_accrual(self):
        period = self._create_period(label="2026-05-BILLLINE")
        accrual = self._create_accrual(period=period, fingerprint="guardrail-acc-billline")
        bill = Bill.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=period,
            invoice_no="INV-GUARDRAIL-BILLLINE",
            currency=period.currency,
        )

        with self.assertRaises(ValidationError):
            BillLine.objects.create(
                bill=bill,
                accrual=accrual,
                charge_type=accrual.charge_type,
                service_date=accrual.service_date,
                quantity=Decimal("2.0000"),
                unit_price=accrual.unit_price,
                amount=accrual.amount,
                tax_amount=accrual.tax_amount,
            )

        with self.assertRaises(ValidationError):
            BillLine.objects.create(
                bill=bill,
                accrual=accrual,
                charge_type=accrual.charge_type,
                service_date=accrual.service_date,
                quantity=accrual.quantity,
                unit_price=Decimal("9.9999"),
                amount=accrual.amount,
                tax_amount=accrual.tax_amount,
            )

        with self.assertRaises(ValidationError):
            BillLine.objects.create(
                bill=bill,
                accrual=accrual,
                charge_type=accrual.charge_type,
                service_date=accrual.service_date,
                quantity=accrual.quantity,
                unit_price=accrual.unit_price,
                amount=Decimal("9.99"),
                tax_amount=accrual.tax_amount,
            )

        with self.assertRaises(ValidationError):
            BillLine.objects.create(
                bill=bill,
                accrual=accrual,
                charge_type=accrual.charge_type,
                service_date=accrual.service_date,
                quantity=accrual.quantity,
                unit_price=accrual.unit_price,
                amount=accrual.amount,
                tax_amount=Decimal("0.01"),
            )

    def test_billline_rejects_negative_values(self):
        period = self._create_period(label="2026-05-BLLNNEG")
        accrual = self._create_accrual(period=period, fingerprint="guardrail-acc-billline-nonneg")
        bill = Bill.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=period,
            invoice_no="INV-GUARDRAIL-BILLLINE-NONNEG",
            currency=period.currency,
        )

        with self.assertRaises(ValidationError):
            BillLine.objects.create(
                bill=bill,
                accrual=accrual,
                charge_type=accrual.charge_type,
                service_date=accrual.service_date,
                quantity=Decimal("-1.0000"),
                unit_price=Decimal("-10.0000"),
                amount=Decimal("-10.00"),
                tax_amount=Decimal("-0.01"),
            )

    def test_billing_import_rules_from_csv_is_scoped_by_warehouse(self):
        fieldnames = [
            "owner_id",
            "warehouse_id",
            "charge_type",
            "calc_method",
            "unit_price",
            "currency",
            "taxable",
            "tax_rate",
            "min_charge",
            "priority",
            "effective_from",
            "effective_to",
            "note",
        ]
        with tempfile.NamedTemporaryFile("w+", newline="", suffix=".csv") as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerow(
                {
                    "owner_id": self.owner.id,
                    "warehouse_id": self.warehouse.id,
                    "charge_type": ChargeType.DISPATCH,
                    "calc_method": CalcMethod.PER_ORDER,
                    "unit_price": "10.00",
                    "currency": "CNY",
                    "taxable": "1",
                    "tax_rate": "0.00",
                    "min_charge": "0.00",
                    "priority": "100",
                    "effective_from": "",
                    "effective_to": "",
                    "note": "warehouse-1",
                }
            )
            writer.writerow(
                {
                    "owner_id": self.owner.id,
                    "warehouse_id": self.other_warehouse.id,
                    "charge_type": ChargeType.DISPATCH,
                    "calc_method": CalcMethod.PER_ORDER,
                    "unit_price": "20.00",
                    "currency": "CNY",
                    "taxable": "1",
                    "tax_rate": "0.00",
                    "min_charge": "0.00",
                    "priority": "100",
                    "effective_from": "",
                    "effective_to": "",
                    "note": "warehouse-2",
                }
            )
            csv_file.flush()

            call_command("billing_import_rules_from_csv", csv_file.name)

        rules = list(
            BillingRule.objects
            .filter(owner=self.owner, charge_type=ChargeType.DISPATCH, calc_method=CalcMethod.PER_ORDER)
            .order_by("warehouse_id")
        )
        self.assertEqual(len(rules), 2)
        self.assertEqual([rule.warehouse_id for rule in rules], [self.warehouse.id, self.other_warehouse.id])
        self.assertEqual([Decimal(rule.unit_price) for rule in rules], [Decimal("10.0000"), Decimal("20.0000")])


class BillingMenuTests(TestCase):
    def test_profile_menu_points_billing_users_to_admin(self):
        warehouse = Warehouse.objects.create(code="WHM", name="Warehouse Menu")
        user = get_user_model().objects.create_user(
            username="billing-menu",
            password="x",
            warehouse=warehouse,
        )
        permission = Permission.objects.get(codename="view_bill")
        user.user_permissions.add(permission)

        request = APIRequestFactory().get("/api/auth/profile/")
        force_authenticate(request, user=user)
        response = profile_view(request)

        menus = response.data["menus"]
        self.assertIn(
            {"path": "/admin/billing/", "title": "计费", "icon": "el-icon-credit-card"},
            menus,
        )

    def test_console_ribbon_exposes_billing_overview_for_billing_users(self):
        warehouse = Warehouse.objects.create(code="WHM2", name="Warehouse Menu 2")
        user = get_user_model().objects.create_user(
            username="billing-menu-console",
            password="x",
            warehouse=warehouse,
        )
        permission = Permission.objects.get(codename="view_bill")
        user.user_permissions.add(permission)

        self.client.force_login(user)
        response = self.client.get("/console/billing/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("console:billing_overview"))


class BillingApiTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(name="Owner API", code="OWAPI")
        self.other_owner = Owner.objects.create(name="Owner Other", code="OWOTH")
        self.warehouse = Warehouse.objects.create(code="WHAPI", name="Warehouse API")
        self.user = get_user_model().objects.create_user(
            username="billing-api-user",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _create_rule(self, *, owner=None, calc_method=CalcMethod.PER_ORDER, unit_price="10.00"):
        return BillingRule.objects.create(
            owner=owner or self.owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH,
            calc_method=calc_method,
            unit_price=Decimal(unit_price),
        )

    def _create_accrual(
        self,
        *,
        rule,
        amount,
        service_date,
        period=None,
        status=AccrualStatus.OPEN,
        fingerprint,
        owner=None,
        warehouse=None,
    ):
        return BillingAccrual.objects.create(
            owner=owner or self.owner,
            warehouse=warehouse or self.warehouse,
            period=period,
            charge_type=rule.charge_type,
            rule=rule,
            service_date=service_date,
            currency="CNY",
            quantity=Decimal("1"),
            unit_price=Decimal(amount),
            amount=Decimal(amount),
            tax_amount=Decimal("0.00"),
            status=status,
            bundle_key="",
            acc_fingerprint=fingerprint,
            created_by=self.user,
        )

    def test_rules_list_includes_scoped_and_generic_rules(self):
        own_rule = self._create_rule(unit_price="10.00")
        generic_rule = BillingRule.objects.create(
            owner=None,
            warehouse=None,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("8.00"),
        )
        self._create_rule(owner=self.other_owner, unit_price="20.00")
        BillingRule.objects.create(
            owner=self.owner,
            warehouse=Warehouse.objects.create(code="WHAPIX", name="Warehouse API X"),
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("30.00"),
        )

        response = self.client.get("/api/billing/rules/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            sorted(item["id"] for item in response.data),
            sorted([own_rule.id, generic_rule.id]),
        )

    def test_rule_tiers_list_includes_scoped_and_generic_rules(self):
        own_rule = self._create_rule(unit_price="10.00")
        generic_rule = BillingRule.objects.create(
            owner=None,
            warehouse=None,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("8.00"),
        )
        other_rule = self._create_rule(owner=self.other_owner, unit_price="20.00")
        own_tier = BillingRuleTier.objects.create(
            rule=own_rule,
            threshold_from=Decimal("0.0000"),
            threshold_to=Decimal("100.0000"),
            unit_price=Decimal("10.0000"),
        )
        generic_tier = BillingRuleTier.objects.create(
            rule=generic_rule,
            threshold_from=Decimal("0.0000"),
            threshold_to=Decimal("100.0000"),
            unit_price=Decimal("8.0000"),
        )
        BillingRuleTier.objects.create(
            rule=other_rule,
            threshold_from=Decimal("0.0000"),
            threshold_to=Decimal("100.0000"),
            unit_price=Decimal("20.0000"),
        )

        response = self.client.get("/api/billing/rule-tiers/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            sorted(item["id"] for item in response.data),
            sorted([own_tier.id, generic_tier.id]),
        )

    def test_generic_rule_is_readable_but_not_mutable_for_scoped_user(self):
        generic_rule = BillingRule.objects.create(
            owner=None,
            warehouse=None,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("8.00"),
        )

        get_response = self.client.get(f"/api/billing/rules/{generic_rule.id}/")
        patch_response = self.client.patch(
            f"/api/billing/rules/{generic_rule.id}/",
            {"note": "should-fail"},
            format="json",
        )

        self.assertEqual(get_response.status_code, 200)
        self.assertEqual(patch_response.status_code, 403)

    def test_period_preview_returns_open_accrual_summary(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-OPEN",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.OPEN,
        )
        rule = self._create_rule(unit_price="12.50")
        self._create_accrual(
            rule=rule,
            amount="12.50",
            service_date=datetime.date(2026, 4, 2),
            fingerprint="acc-preview-1",
        )

        response = self.client.get(f"/api/billing/periods/{period.id}/preview/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["accrual_count"], 1)
        self.assertEqual(Decimal(response.data["subtotal"]), Decimal("12.50"))
        self.assertEqual(response.data["scope"], "open_unlocked")

    def test_period_lock_and_invoice_actions_work_end_to_end(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-LOCK",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.OPEN,
        )
        rule = self._create_rule(unit_price="15.00")
        self._create_accrual(
            rule=rule,
            amount="15.00",
            service_date=datetime.date(2026, 4, 3),
            fingerprint="acc-lock-1",
        )

        lock_response = self.client.post(f"/api/billing/periods/{period.id}/lock/")
        period.refresh_from_db()

        self.assertEqual(lock_response.status_code, 200)
        self.assertEqual(lock_response.data["status"], PeriodStatus.CLOSED)
        self.assertEqual(period.status, PeriodStatus.CLOSED)
        self.assertEqual(BillingAccrual.objects.get(acc_fingerprint="acc-lock-1").status, AccrualStatus.LOCKED)

        invoice_response = self.client.post(
            f"/api/billing/periods/{period.id}/invoice/",
            {"invoice_no": "INV-API-0001"},
            format="json",
        )
        period.refresh_from_db()

        self.assertEqual(invoice_response.status_code, 201)
        self.assertEqual(invoice_response.data["invoice_no"], "INV-API-0001")
        self.assertEqual(period.status, PeriodStatus.INVOICED)
        self.assertEqual(Bill.objects.filter(period=period).count(), 1)
        self.assertEqual(len(invoice_response.data["lines"]), 1)

    def test_period_lock_action_returns_400_when_reconciliation_gate_fails(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-LOCK-GATE",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.OPEN,
        )
        rule = self._create_rule(unit_price="15.00")
        accrual = self._create_accrual(
            rule=rule,
            amount="15.00",
            service_date=datetime.date(2026, 4, 3),
            fingerprint="acc-lock-api-gate",
        )
        BillingAccrual.objects.filter(pk=accrual.pk).update(charge_type=ChargeType.STORAGE)

        response = self.client.post(f"/api/billing/periods/{period.id}/lock/")

        self.assertEqual(response.status_code, 400)
        self.assertIn("数据对账未通过", response.data["detail"])

    def test_period_invoice_action_returns_400_when_reconciliation_gate_fails(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-INVOICE-GATE",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.CLOSED,
        )
        rule = self._create_rule(unit_price="15.00")
        accrual = self._create_accrual(
            rule=rule,
            amount="15.00",
            service_date=datetime.date(2026, 4, 3),
            period=period,
            status=AccrualStatus.LOCKED,
            fingerprint="acc-invoice-api-gate",
        )
        BillingAccrual.objects.filter(pk=accrual.pk).update(charge_type=ChargeType.STORAGE)

        response = self.client.post(
            f"/api/billing/periods/{period.id}/invoice/",
            {"invoice_no": "INV-API-GATE"},
            format="json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("数据对账未通过", response.data["detail"])

    def test_bill_detail_endpoint_includes_lines(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-BILL",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.CLOSED,
        )
        rule = self._create_rule(unit_price="18.00")
        self._create_accrual(
            rule=rule,
            amount="18.00",
            service_date=datetime.date(2026, 4, 4),
            period=period,
            status=AccrualStatus.LOCKED,
            fingerprint="acc-bill-detail-1",
        )
        bill = generate_invoice_for_period(period, invoice_no="INV-API-DETAIL")

        response = self.client.get(f"/api/billing/bills/{bill.id}/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["invoice_no"], "INV-API-DETAIL")
        self.assertEqual(len(response.data["lines"]), 1)

    def test_bill_list_export_endpoint_returns_workbook(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-EXPORT-LIST",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.CLOSED,
        )
        rule = self._create_rule(unit_price="20.00")
        self._create_accrual(
            rule=rule,
            amount="20.00",
            service_date=datetime.date(2026, 4, 5),
            period=period,
            status=AccrualStatus.LOCKED,
            fingerprint="acc-bill-export-list-1",
        )
        bill = generate_invoice_for_period(period, invoice_no="INV-API-EXPORT-LIST")

        response = self.client.get("/api/billing/bills/export/", {"period": period.id})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook["Bills"]
        self.assertEqual(sheet["A2"].value, bill.invoice_no)
        self.assertEqual(sheet["E2"].value, period.label)

    def test_bill_detail_export_endpoint_returns_line_workbook(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-EXP-DET",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.CLOSED,
        )
        rule = self._create_rule(unit_price="21.00")
        self._create_accrual(
            rule=rule,
            amount="21.00",
            service_date=datetime.date(2026, 4, 6),
            period=period,
            status=AccrualStatus.LOCKED,
            fingerprint="acc-bill-export-detail-1",
        )
        bill = generate_invoice_for_period(period, invoice_no="INV-API-EXPORT-DETAIL")

        response = self.client.get(f"/api/billing/bills/{bill.id}/export/")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook["Bill"]["B2"].value, bill.invoice_no)
        self.assertEqual(workbook["Lines"]["A2"].value, "2026-04-06")
        self.assertEqual(workbook["Lines"]["B2"].value, ChargeType.DISPATCH)

    def test_superuser_can_create_period_for_explicit_warehouse(self):
        other_warehouse = Warehouse.objects.create(code="WHAPI2", name="Warehouse API 2")
        superuser = get_user_model().objects.create_superuser(
            username="billing-api-superuser",
            password="x",
            warehouse=self.warehouse,
        )
        client = APIClient()
        client.force_authenticate(superuser)

        response = client.post(
            "/api/billing/periods/",
            {
                "owner": self.owner.id,
                "warehouse": other_warehouse.id,
                "label": "2026-04-SUPER",
                "start_date": "2026-04-01",
                "end_date": "2026-04-30",
                "currency": "CNY",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        period = BillingPeriod.objects.get(id=response.data["id"])
        self.assertEqual(period.warehouse_id, other_warehouse.id)
        self.assertEqual(response.data["warehouse"], other_warehouse.id)

    def test_warehouse_dashboard_overview_aggregates_multiple_owners_in_same_warehouse(self):
        boss = get_user_model().objects.create_user(
            username="billing-warehouse-boss",
            password="x",
            warehouse=self.warehouse,
        )
        client = APIClient()
        client.force_authenticate(boss)

        own_rule = self._create_rule(unit_price="20.00")
        other_rule = BillingRule.objects.create(
            owner=self.other_owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("35.00"),
        )
        other_warehouse = Warehouse.objects.create(code="WHAPI3", name="Warehouse API 3")
        ignored_rule = BillingRule.objects.create(
            owner=self.other_owner,
            warehouse=other_warehouse,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("99.00"),
        )

        self._create_accrual(
            rule=own_rule,
            amount="20.00",
            service_date=datetime.date(2026, 4, 10),
            fingerprint="acc-dash-own",
            owner=self.owner,
        )
        self._create_accrual(
            rule=other_rule,
            amount="35.00",
            service_date=datetime.date(2026, 4, 11),
            fingerprint="acc-dash-other",
            owner=self.other_owner,
        )
        self._create_accrual(
            rule=ignored_rule,
            amount="99.00",
            service_date=datetime.date(2026, 4, 12),
            fingerprint="acc-dash-ignored",
            owner=self.other_owner,
            warehouse=other_warehouse,
        )

        response = client.get(
            "/api/billing/dashboard/warehouse-overview/",
            {"date_from": "2026-04-01", "date_to": "2026-04-30"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["scope"]["warehouse"], self.warehouse.id)
        self.assertEqual(response.data["summary"]["owner_count"], 2)
        self.assertEqual(response.data["summary"]["accrual_count"], 2)
        self.assertEqual(Decimal(response.data["summary"]["subtotal"]), Decimal("55.00"))
        self.assertEqual(len(response.data["by_owner"]), 2)
        self.assertEqual(
            {row["owner"] for row in response.data["by_owner"]},
            {self.owner.id, self.other_owner.id},
        )
        self.assertEqual(
            {row["id"] for row in response.data["owner_options"]},
            {self.owner.id, self.other_owner.id},
        )

    def test_warehouse_dashboard_overview_owner_filter_narrows_results(self):
        boss = get_user_model().objects.create_user(
            username="billing-warehouse-boss-filter",
            password="x",
            warehouse=self.warehouse,
        )
        client = APIClient()
        client.force_authenticate(boss)

        own_rule = self._create_rule(unit_price="20.00")
        other_rule = BillingRule.objects.create(
            owner=self.other_owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("35.00"),
        )

        self._create_accrual(
            rule=own_rule,
            amount="20.00",
            service_date=datetime.date(2026, 4, 10),
            fingerprint="acc-dash-filter-own",
            owner=self.owner,
        )
        self._create_accrual(
            rule=other_rule,
            amount="35.00",
            service_date=datetime.date(2026, 4, 11),
            fingerprint="acc-dash-filter-other",
            owner=self.other_owner,
        )

        response = client.get(
            "/api/billing/dashboard/warehouse-overview/",
            {
                "owner": self.owner.id,
                "date_from": "2026-04-01",
                "date_to": "2026-04-30",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["summary"]["owner_count"], 1)
        self.assertEqual(response.data["summary"]["accrual_count"], 1)
        self.assertEqual(Decimal(response.data["summary"]["subtotal"]), Decimal("20.00"))
        self.assertEqual(len(response.data["by_owner"]), 1)
        self.assertEqual(response.data["by_owner"][0]["owner"], self.owner.id)

    def test_owner_scoped_dashboard_rejects_other_owner_query(self):
        boss = get_user_model().objects.create_user(
            username="billing-owner-boss-filter",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        client = APIClient()
        client.force_authenticate(boss)

        response = client.get(
            "/api/billing/dashboard/warehouse-overview/",
            {"owner": self.other_owner.id},
        )

        self.assertEqual(response.status_code, 403)

    def test_warehouse_boss_scope_mode_keeps_multi_owner_view_for_owner_bound_user(self):
        boss = get_user_model().objects.create_user(
            username="billing-owner-bound-warehouse-boss",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        client = APIClient()
        client.force_authenticate(boss)

        own_rule = self._create_rule(unit_price="20.00")
        other_rule = BillingRule.objects.create(
            owner=self.other_owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.DISPATCH,
            calc_method=CalcMethod.PER_ORDER,
            unit_price=Decimal("35.00"),
        )

        self._create_accrual(
            rule=own_rule,
            amount="20.00",
            service_date=datetime.date(2026, 4, 10),
            fingerprint="acc-dash-scope-own",
            owner=self.owner,
        )
        self._create_accrual(
            rule=other_rule,
            amount="35.00",
            service_date=datetime.date(2026, 4, 11),
            fingerprint="acc-dash-scope-other",
            owner=self.other_owner,
        )

        response = client.get(
            "/api/billing/dashboard/warehouse-overview/",
            {
                "scope_mode": "warehouse_boss",
                "date_from": "2026-04-01",
                "date_to": "2026-04-30",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["summary"]["owner_count"], 2)
        self.assertEqual(len(response.data["by_owner"]), 2)
        self.assertEqual(
            {row["id"] for row in response.data["owner_options"]},
            {self.owner.id, self.other_owner.id},
        )

    def test_warehouse_dashboard_scope_owner_name_does_not_leak_unscoped_owner(self):
        boss = get_user_model().objects.create_user(
            username="billing-warehouse-boss-scope",
            password="x",
            warehouse=self.warehouse,
        )
        client = APIClient()
        client.force_authenticate(boss)

        response = client.get(
            "/api/billing/dashboard/warehouse-overview/",
            {"owner": self.other_owner.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["scope"]["owner"], self.other_owner.id)
        self.assertEqual(response.data["scope"]["owner_name"], "")


class BillingMetricGenerationTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(name="Owner Metrics", code="OWMET")
        self.warehouse = Warehouse.objects.create(code="WHMT", name="Warehouse Metrics")
        self.user = get_user_model().objects.create_user(
            username="billing-metric-user",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        self.uom = ProductUom.objects.create(code="PCS-MT", name="件", is_active=True)
        self.subwarehouse = Subwarehouse.objects.create(
            warehouse=self.warehouse,
            code="SWMT",
            name="Subwarehouse Metrics",
        )
        self.location1 = Location.objects.create(
            warehouse=self.warehouse,
            code="SWMT-01-01-01",
            name="L1",
        )
        self.location2 = Location.objects.create(
            warehouse=self.warehouse,
            code="SWMT-01-01-02",
            name="L2",
        )
        self.product = Product.objects.create(
            owner=self.owner,
            code="SKU-MT",
            name="Metric Product",
            sku="SKU-MT",
            base_uom=self.uom,
            batch_control=False,
            expiry_control=False,
            volume=Decimal("0.500000"),
            price=Decimal("10.00"),
        )
        self.customer = Customer.objects.create(
            owner=self.owner,
            salesperson=self.user,
            code="CUST-MT",
            name="Metric Customer",
        )

    def _create_inventory(self, *, location, qty):
        from allapp.inventory.models import InventoryDetail

        return InventoryDetail.objects.create(
            owner=self.owner,
            product=self.product,
            warehouse=self.warehouse,
            location=location,
            onhand_qty=Decimal(qty),
            allocated_qty=Decimal("0"),
            locked_qty=Decimal("0"),
            damaged_qty=Decimal("0"),
        )

    def _create_outbound_order(self, service_date: datetime.date, *, qty="3.000", price="10.0000"):
        from allapp.outbound.models import OutboundOrder, OutboundOrderLine

        order = OutboundOrder.objects.create(
            owner=self.owner,
            customer=self.customer,
            warehouse=self.warehouse,
            order_no=f"OUT-{service_date:%Y%m%d}",
            biz_date=service_date,
            submit_status="SUBMITTED",
            approval_status="OWNER_APPROVED",
            created_by=self.user,
        )
        OutboundOrderLine.objects.create(
            order=order,
            product=self.product,
            base_qty=Decimal(qty),
            base_price=Decimal(price),
            base_uom=self.uom,
            line_no=10,
        )
        return order

    def _create_inventory_transaction(self, service_date: datetime.date, *, qty_delta, location=None):
        qty_delta = Decimal(qty_delta)
        tx_type = InvTxType.RECEIVE if qty_delta > 0 else InvTxType.ISSUE
        next_src_id = InventoryTransaction.objects.count() + 1
        return InventoryTransaction.objects.create(
            tx_type=tx_type,
            owner=self.owner,
            product=self.product,
            warehouse=self.warehouse,
            subwarehouse=self.subwarehouse,
            location=location or self.location1,
            qty_delta=qty_delta,
            src_model="billing_metric_test",
            src_id=next_src_id,
            src_line_id=next_src_id,
            src_no=f"TX-{next_src_id}",
            posted_at=datetime.datetime.combine(service_date, datetime.time(10, 0)),
        )

    def test_generate_metrics_for_date_builds_inventory_and_order_metrics(self):
        service_date = datetime.date(2026, 4, 5)
        self._create_inventory(location=self.location1, qty="3.0000")
        self._create_inventory(location=self.location2, qty="1.0000")
        self._create_outbound_order(service_date, qty="3.000", price="10.0000")

        summary = generate_metrics_for_date(self.owner.id, self.warehouse.id, service_date)

        self.assertEqual(summary["created"], 3)
        self.assertEqual(summary["unsupported"], 1)
        metrics = {
            metric.metric_type: metric
            for metric in BillingMetricDaily.objects.filter(
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=service_date,
            )
        }
        self.assertEqual(Decimal(metrics["PALLET"].value), Decimal("2.0000"))
        self.assertEqual(Decimal(metrics["CBM"].value), Decimal("2.0000"))
        self.assertEqual(Decimal(metrics["ORDER_AMT"].value), Decimal("30.0000"))

    def test_generate_metrics_for_date_does_not_overwrite_manual_rows(self):
        service_date = datetime.date(2026, 4, 6)
        self._create_outbound_order(service_date, qty="5.000", price="12.0000")
        BillingMetricDaily.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type="ORDER_AMT",
            value=Decimal("999.0000"),
            source="MANUAL",
            note="manual override",
        )

        summary = generate_metrics_for_date(
            self.owner.id,
            self.warehouse.id,
            service_date,
            metric_types=["ORDER_AMT"],
        )

        self.assertEqual(summary["skipped_manual"], 1)
        metric = BillingMetricDaily.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type="ORDER_AMT",
        )
        self.assertEqual(Decimal(metric.value), Decimal("999.0000"))
        self.assertEqual(metric.source, "MANUAL")

    def test_generate_metrics_for_date_second_run_is_idempotent(self):
        service_date = datetime.date(2026, 4, 6)
        self._create_inventory(location=self.location1, qty="3.0000")
        self._create_outbound_order(service_date, qty="5.000", price="12.0000")

        first = generate_metrics_for_date(self.owner.id, self.warehouse.id, service_date)
        second = generate_metrics_for_date(self.owner.id, self.warehouse.id, service_date)

        self.assertEqual(first["created"], 3)
        self.assertEqual(second["created"], 0)
        self.assertEqual(second["noop"], 3)
        self.assertEqual(
            BillingMetricDaily.objects.filter(
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=service_date,
            ).count(),
            3,
        )

    def test_period_generate_metrics_action_creates_rows(self):
        service_date = datetime.date(2026, 4, 7)
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-MET",
            start_date=service_date,
            end_date=service_date,
            status=PeriodStatus.OPEN,
        )
        self._create_inventory(location=self.location1, qty="2.0000")
        self._create_outbound_order(service_date, qty="2.000", price="8.0000")

        response = self.client.post(
            f"/api/billing/periods/{period.id}/generate-metrics/",
            {"metric_types": ["PALLET", "CBM", "ORDER_AMT"]},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["summary"]["created"], 3)
        self.assertEqual(
            BillingMetricDaily.objects.filter(
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=service_date,
            ).count(),
            3,
        )

    def test_billing_run_scheduler_once_records_success_and_skips_repeat_date(self):
        service_date = datetime.date(2026, 4, 8)
        self._create_inventory(location=self.location1, qty="4.0000")
        self._create_inventory_transaction(service_date, qty_delta="4.0000", location=self.location1)
        self._create_outbound_order(service_date, qty="2.000", price="15.0000")

        call_command(
            "billing_run_scheduler",
            "--once",
            "--date",
            service_date.isoformat(),
            "--owner",
            str(self.owner.id),
            "--warehouse",
            str(self.warehouse.id),
        )

        job_run = BillingJobRun.objects.get(
            job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
        )
        self.assertEqual(job_run.status, BillingJobRun.Status.SUCCESS)
        self.assertEqual(job_run.attempts, 1)
        self.assertEqual(job_run.summary["created"], 3)

        call_command(
            "billing_run_scheduler",
            "--once",
            "--date",
            service_date.isoformat(),
            "--owner",
            str(self.owner.id),
            "--warehouse",
            str(self.warehouse.id),
        )

        job_run.refresh_from_db()
        self.assertEqual(job_run.status, BillingJobRun.Status.SUCCESS)
        self.assertEqual(job_run.attempts, 1)
        self.assertEqual(
            BillingJobRun.objects.filter(
                job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=service_date,
            ).count(),
            1,
        )

    def test_billing_run_scheduler_once_fails_when_reconciliation_gate_fails(self):
        service_date = datetime.date(2026, 4, 9)
        self._create_inventory(location=self.location1, qty="4.0000")

        with self.assertRaises(CommandError):
            call_command(
                "billing_run_scheduler",
                "--once",
                "--date",
                service_date.isoformat(),
                "--owner",
                str(self.owner.id),
                "--warehouse",
                str(self.warehouse.id),
            )

        job_run = BillingJobRun.objects.get(
            job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
        )
        self.assertEqual(job_run.status, BillingJobRun.Status.FAILED)
        self.assertIn("数据对账未通过", job_run.message)
        self.assertEqual(
            BillingMetricDaily.objects.filter(
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=service_date,
            ).count(),
            0,
        )

    def test_generate_metrics_for_past_date_uses_snapshot_not_current_inventory(self):
        service_date = timezone.now().date() - datetime.timedelta(days=1)
        bootstrap_date = service_date - datetime.timedelta(days=1)
        detail = self._create_inventory(location=self.location1, qty="5.0000")

        generate_inventory_snapshot_for_date(
            bootstrap_date,
            owner_id=self.owner.id,
            warehouse_id=self.warehouse.id,
            bootstrap=True,
        )
        self._create_inventory_transaction(service_date, qty_delta="-2.0000", location=self.location1)

        detail.onhand_qty = Decimal("5.0000")
        detail.save()

        summary = generate_metrics_for_date(
            self.owner.id,
            self.warehouse.id,
            service_date,
            metric_types=["CBM", "PALLET"],
        )

        self.assertEqual(summary["created"], 2)
        snapshot_row = InventorySnapshotDaily.objects.get(
            snapshot_date=service_date,
            owner=self.owner,
            warehouse=self.warehouse,
            location=self.location1,
            product=self.product,
        )
        self.assertEqual(Decimal(snapshot_row.onhand_qty), Decimal("3.0000"))

        cbm_metric = BillingMetricDaily.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type="CBM",
        )
        self.assertEqual(Decimal(cbm_metric.value), Decimal("1.5000"))
        self.assertEqual(cbm_metric.source, "AUTO:INVENTORY_SNAPSHOT_ONHAND_VOLUME")

    def test_billing_run_scheduler_generates_inventory_snapshot_for_past_date(self):
        service_date = timezone.now().date() - datetime.timedelta(days=1)
        bootstrap_date = service_date - datetime.timedelta(days=1)
        detail = self._create_inventory(location=self.location1, qty="4.0000")
        self._create_inventory_transaction(bootstrap_date, qty_delta="4.0000", location=self.location1)

        generate_inventory_snapshot_for_date(
            bootstrap_date,
            owner_id=self.owner.id,
            warehouse_id=self.warehouse.id,
            bootstrap=True,
        )

        detail.onhand_qty = Decimal("3.0000")
        detail.save()
        self._create_inventory_transaction(service_date, qty_delta="-1.0000", location=self.location1)

        call_command(
            "billing_run_scheduler",
            "--once",
            "--date",
            service_date.isoformat(),
            "--owner",
            str(self.owner.id),
            "--warehouse",
            str(self.warehouse.id),
        )

        snapshot_row = InventorySnapshotDaily.objects.get(
            snapshot_date=service_date,
            owner=self.owner,
            warehouse=self.warehouse,
            location=self.location1,
            product=self.product,
        )
        self.assertEqual(Decimal(snapshot_row.onhand_qty), Decimal("3.0000"))

        cbm_metric = BillingMetricDaily.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type="CBM",
        )
        self.assertEqual(Decimal(cbm_metric.value), Decimal("1.5000"))


class BillingSchedulerConcurrencyTests(TransactionTestCase):
    def setUp(self):
        self.owner = Owner.objects.create(name="Owner Billing Concurrent", code="OWN-BILL-C")
        self.warehouse = Warehouse.objects.create(code="WH-BILL-C", name="Warehouse Billing Concurrent")

    def test_scheduler_claims_single_job_under_concurrency(self):
        service_date = datetime.date(2026, 4, 10)
        generation_entered = threading.Event()
        release_generation = threading.Event()
        generation_calls = 0
        generation_lock = threading.Lock()
        results = [None, None]
        errors = []

        def fake_generate_metrics_for_date(owner_id, warehouse_id, requested_date, **kwargs):
            nonlocal generation_calls
            with generation_lock:
                generation_calls += 1
            generation_entered.set()
            if not release_generation.wait(timeout=5):
                raise AssertionError("timed out waiting to release billing scheduler concurrent test")
            return {
                "service_date": requested_date,
                "created": 1,
                "updated": 0,
                "deleted_zero": 0,
                "skipped_zero": 0,
                "skipped_manual": 0,
                "unsupported": 0,
                "noop": 0,
            }

        def invoke(index):
            close_old_connections()
            try:
                results[index] = run_scheduled_metric_generation_for_date(
                    service_date,
                    owner_id=self.owner.id,
                    warehouse_id=self.warehouse.id,
                    metric_types=["ORDER_AMT"],
                )
            except BaseException as exc:
                errors.append(exc)
            finally:
                close_old_connections()

        with mock.patch(
            "allapp.billing.services.generate_metrics_for_date",
            side_effect=fake_generate_metrics_for_date,
        ), mock.patch(
            "allapp.billing.services._billing_accuracy_gate_enabled",
            return_value=False,
        ):
            thread1 = threading.Thread(target=invoke, args=(0,))
            thread1.start()
            self.assertTrue(generation_entered.wait(timeout=5))

            thread2 = threading.Thread(target=invoke, args=(1,))
            thread2.start()
            thread2.join(timeout=5)

            release_generation.set()
            thread1.join(timeout=5)

        if thread1.is_alive() or thread2.is_alive():
            self.fail("concurrent billing scheduler threads did not finish")
        if errors:
            raise errors[0]

        self.assertEqual(generation_calls, 1)
        self.assertEqual(
            sorted(result["runs"][0]["status"] for result in results),
            ["skipped_running", "success"],
        )

        job_run = BillingJobRun.objects.get(
            job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
        )
        self.assertEqual(job_run.status, BillingJobRun.Status.SUCCESS)
        self.assertEqual(job_run.attempts, 1)
        self.assertEqual(job_run.summary["created"], 1)
        self.assertEqual(
            BillingJobRun.objects.filter(
                job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=service_date,
            ).count(),
            1,
        )

    def test_generate_metrics_for_date_survives_concurrent_create_race(self):
        service_date = datetime.date(2026, 4, 11)
        user = get_user_model().objects.create_user(
            username="billing-metric-race-user",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        uom = ProductUom.objects.create(code="PCS-BMR", name="件", is_active=True)
        product = Product.objects.create(
            owner=self.owner,
            code="SKU-BMR",
            name="Metric Race Product",
            sku="SKU-BMR",
            base_uom=uom,
            volume=Decimal("0.500000"),
            price=Decimal("10.00"),
        )
        customer = Customer.objects.create(
            owner=self.owner,
            salesperson=user,
            code="CUST-BMR",
            name="Metric Race Customer",
        )
        from allapp.outbound.models import OutboundOrder, OutboundOrderLine

        order = OutboundOrder.objects.create(
            owner=self.owner,
            customer=customer,
            warehouse=self.warehouse,
            order_no=f"OUT-{service_date:%Y%m%d}-RACE",
            biz_date=service_date,
            submit_status="SUBMITTED",
            approval_status="OWNER_APPROVED",
            created_by=user,
        )
        OutboundOrderLine.objects.create(
            order=order,
            product=product,
            base_qty=Decimal("3.000"),
            base_price=Decimal("10.0000"),
            base_uom=uom,
            line_no=10,
        )

        create_entered = threading.Event()
        release_create = threading.Event()
        create_calls = 0
        create_lock = threading.Lock()
        results = [None, None]
        errors = []
        real_metric_create = BillingMetricDaily.objects.create

        def fake_metric_create(*args, **kwargs):
            nonlocal create_calls
            with create_lock:
                create_calls += 1
                current_call = create_calls
            if current_call == 1:
                create_entered.set()
                if not release_create.wait(timeout=5):
                    raise AssertionError("timed out waiting to release metric concurrent test")
            return real_metric_create(*args, **kwargs)

        def invoke(index):
            close_old_connections()
            try:
                results[index] = generate_metrics_for_date(
                    self.owner.id,
                    self.warehouse.id,
                    service_date,
                    metric_types=["ORDER_AMT"],
                )
            except BaseException as exc:
                errors.append(exc)
            finally:
                close_old_connections()

        with mock.patch("allapp.billing.services.BillingMetricDaily.objects.create", side_effect=fake_metric_create):
            thread1 = threading.Thread(target=invoke, args=(0,))
            thread1.start()
            self.assertTrue(create_entered.wait(timeout=5))

            thread2 = threading.Thread(target=invoke, args=(1,))
            thread2.start()

            release_create.set()
            thread1.join(timeout=5)
            thread2.join(timeout=5)

        if thread1.is_alive() or thread2.is_alive():
            self.fail("concurrent metric generation threads did not finish")
        if errors:
            raise errors[0]

        self.assertEqual(
            BillingMetricDaily.objects.filter(
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=service_date,
                metric_type="ORDER_AMT",
            ).count(),
            1,
        )
        metric = BillingMetricDaily.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type="ORDER_AMT",
        )
        self.assertEqual(metric.value, Decimal("30.0000"))
        self.assertGreaterEqual(create_calls, 1)
        self.assertEqual(sum(result["created"] for result in results), 1)
        self.assertEqual(sum(result["noop"] for result in results), 1)

    def test_generate_metrics_for_date_recovers_from_integrity_error_after_concurrent_insert(self):
        service_date = datetime.date(2026, 4, 12)
        user = get_user_model().objects.create_user(
            username="billing-metric-integrity-user",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        uom = ProductUom.objects.create(code="PCS-BMI", name="件", is_active=True)
        product = Product.objects.create(
            owner=self.owner,
            code="SKU-BMI",
            name="Metric Integrity Product",
            sku="SKU-BMI",
            base_uom=uom,
            volume=Decimal("0.500000"),
            price=Decimal("10.00"),
        )
        customer = Customer.objects.create(
            owner=self.owner,
            salesperson=user,
            code="CUST-BMI",
            name="Metric Integrity Customer",
        )
        from allapp.outbound.models import OutboundOrder, OutboundOrderLine

        order = OutboundOrder.objects.create(
            owner=self.owner,
            customer=customer,
            warehouse=self.warehouse,
            order_no=f"OUT-{service_date:%Y%m%d}-INT",
            biz_date=service_date,
            submit_status="SUBMITTED",
            approval_status="OWNER_APPROVED",
            created_by=user,
        )
        OutboundOrderLine.objects.create(
            order=order,
            product=product,
            base_qty=Decimal("2.000"),
            base_price=Decimal("10.0000"),
            base_uom=uom,
            line_no=10,
        )

        def fake_metric_create(*args, **kwargs):
            BillingMetricDaily(
                owner_id=kwargs["owner_id"],
                warehouse_id=kwargs["warehouse_id"],
                service_date=kwargs["service_date"],
                metric_type=kwargs["metric_type"],
                value=kwargs["value"],
                source=kwargs["source"],
                note=kwargs["note"],
            ).save(force_insert=True)
            raise IntegrityError("duplicate metric row")

        with mock.patch("allapp.billing.services.BillingMetricDaily.objects.create", side_effect=fake_metric_create):
            summary = generate_metrics_for_date(
                self.owner.id,
                self.warehouse.id,
                service_date,
                metric_types=["ORDER_AMT"],
            )

        self.assertEqual(summary["created"], 0)
        self.assertEqual(summary["noop"], 1)
        metric = BillingMetricDaily.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type="ORDER_AMT",
        )
        self.assertEqual(metric.value, Decimal("20.0000"))


class BillingSettlementConcurrencyTests(TransactionTestCase):
    def setUp(self):
        self.owner = Owner.objects.create(name="Owner Billing Settle C", code="OWBSC")
        self.warehouse = Warehouse.objects.create(code="WHBSC", name="Warehouse Billing Settle C")
        self.user = get_user_model().objects.create_user(
            username="billing-settle-concurrent-user",
            password="x",
            warehouse=self.warehouse,
        )

    def _create_rule(
        self,
        *,
        charge_type=ChargeType.DISPATCH,
        calc_method=CalcMethod.PER_ORDER,
        unit_price="10.00",
        cap_mode=CapMode.NONE,
        cap_amount=None,
    ):
        return BillingRule.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            charge_type=charge_type,
            calc_method=calc_method,
            unit_price=Decimal(unit_price),
            cap_mode=cap_mode,
            cap_amount=Decimal(cap_amount) if cap_amount is not None else None,
        )

    def _create_accrual(
        self,
        *,
        rule,
        amount,
        service_date,
        period=None,
        status=AccrualStatus.OPEN,
        fingerprint,
    ):
        return BillingAccrual.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=period,
            charge_type=rule.charge_type,
            rule=rule,
            service_date=service_date,
            currency="CNY",
            quantity=Decimal("1.0000"),
            unit_price=Decimal(amount),
            amount=Decimal(amount),
            tax_amount=Decimal("0.00"),
            status=status,
            acc_fingerprint=fingerprint,
            created_by=self.user,
        )

    def test_lock_period_allows_only_one_success_under_concurrency(self):
        start_date = datetime.date(2026, 4, 11)
        end_date = datetime.date(2026, 4, 12)
        rule = self._create_rule(cap_mode=CapMode.PER_PERIOD, cap_amount="100.00")
        self._create_accrual(
            rule=rule,
            amount="60.00",
            service_date=start_date,
            fingerprint="acc-lock-concurrent-1",
        )
        self._create_accrual(
            rule=rule,
            amount="70.00",
            service_date=end_date,
            fingerprint="acc-lock-concurrent-2",
        )

        adjustment_entered = threading.Event()
        release_adjustment = threading.Event()
        results = [None, None]
        errors = []
        real_save_adjusted_accrual = __import__("allapp.billing.services", fromlist=["_save_adjusted_accrual"])._save_adjusted_accrual

        def fake_save_adjusted_accrual(accrual, new_amount):
            if not adjustment_entered.is_set():
                adjustment_entered.set()
                if not release_adjustment.wait(timeout=5):
                    raise AssertionError("timed out waiting to release concurrent lock_period test")
            return real_save_adjusted_accrual(accrual, new_amount)

        def invoke(index):
            close_old_connections()
            try:
                results[index] = lock_period(
                    self.owner.id,
                    self.warehouse.id,
                    "2026-04-CONC-LOCK",
                    start_date,
                    end_date,
                )
            except BaseException as exc:
                errors.append(exc)
            finally:
                close_old_connections()

        with mock.patch("allapp.billing.services._save_adjusted_accrual", side_effect=fake_save_adjusted_accrual):
            thread1 = threading.Thread(target=invoke, args=(0,))
            thread1.start()
            self.assertTrue(adjustment_entered.wait(timeout=5))

            thread2 = threading.Thread(target=invoke, args=(1,))
            thread2.start()

            release_adjustment.set()
            thread1.join(timeout=5)
            thread2.join(timeout=5)

        if thread1.is_alive() or thread2.is_alive():
            self.fail("concurrent lock_period threads did not finish")

        period = BillingPeriod.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-CONC-LOCK",
        )
        accruals = list(BillingAccrual.objects.filter(period=period).order_by("service_date", "id"))

        self.assertEqual(sum(result is not None for result in results), 1)
        self.assertEqual(len(errors), 1)
        self.assertIsInstance(errors[0], ValueError)
        self.assertIn("already", str(errors[0]))
        self.assertEqual(period.status, PeriodStatus.CLOSED)
        self.assertEqual(sum((a.amount for a in accruals), Decimal("0.00")), Decimal("100.00"))
        self.assertTrue(all(a.status == AccrualStatus.LOCKED for a in accruals))
        self.assertEqual(
            BillingPeriod.objects.filter(
                owner=self.owner,
                warehouse=self.warehouse,
                label="2026-04-CONC-LOCK",
            ).count(),
            1,
        )

    def test_generate_invoice_allows_only_one_success_under_concurrency(self):
        period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-04-CONC-INV",
            start_date=datetime.date(2026, 4, 1),
            end_date=datetime.date(2026, 4, 30),
            status=PeriodStatus.CLOSED,
        )
        rule = self._create_rule()
        accrual = self._create_accrual(
            rule=rule,
            amount="15.00",
            service_date=datetime.date(2026, 4, 2),
            period=period,
            status=AccrualStatus.LOCKED,
            fingerprint="acc-invoice-concurrent-1",
        )

        billline_entered = threading.Event()
        release_billline = threading.Event()
        results = [None, None]
        errors = []
        real_billline_create = BillLine.objects.create

        def fake_billline_create(*args, **kwargs):
            if not billline_entered.is_set():
                billline_entered.set()
                if not release_billline.wait(timeout=5):
                    raise AssertionError("timed out waiting to release concurrent invoice test")
            return real_billline_create(*args, **kwargs)

        def invoke(index, invoice_no):
            close_old_connections()
            try:
                thread_period = BillingPeriod.objects.get(pk=period.pk)
                results[index] = generate_invoice_for_period(
                    thread_period,
                    invoice_no=invoice_no,
                )
            except BaseException as exc:
                errors.append(exc)
            finally:
                close_old_connections()

        with mock.patch("allapp.billing.services.BillLine.objects.create", side_effect=fake_billline_create):
            thread1 = threading.Thread(target=invoke, args=(0, "INV-CONC-1"))
            thread1.start()
            self.assertTrue(billline_entered.wait(timeout=5))

            thread2 = threading.Thread(target=invoke, args=(1, "INV-CONC-2"))
            thread2.start()

            release_billline.set()
            thread1.join(timeout=5)
            thread2.join(timeout=5)

        if thread1.is_alive() or thread2.is_alive():
            self.fail("concurrent generate_invoice threads did not finish")

        period.refresh_from_db()
        accrual.refresh_from_db()

        self.assertEqual(sum(result is not None for result in results), 1)
        self.assertEqual(len(errors), 1)
        self.assertIsInstance(errors[0], ValueError)
        self.assertTrue(
            "Only closed periods can be invoiced." in str(errors[0])
            or "Invoice already exists for this period." in str(errors[0])
        )
        self.assertEqual(period.status, PeriodStatus.INVOICED)
        self.assertEqual(accrual.status, AccrualStatus.INVOICED)
        self.assertEqual(Bill.objects.filter(period=period).count(), 1)
        self.assertEqual(BillLine.objects.filter(bill__period=period).count(), 1)


class BillingConsolePageTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(name="Owner Console", code="OWCON")
        self.warehouse = Warehouse.objects.create(code="WHCON", name="Warehouse Console")
        self.user = get_user_model().objects.create_user(
            username="billing-console-user",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        self.user.user_permissions.add(
            Permission.objects.get(codename="view_bill"),
            Permission.objects.get(codename="view_billingperiod"),
            Permission.objects.get(codename="view_billingaccrual"),
        )
        self.client.force_login(self.user)

        self.period = BillingPeriod.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            label="2026-03-CONSOLE",
            start_date=datetime.date(2026, 3, 1),
            end_date=datetime.date(2026, 3, 31),
            status=PeriodStatus.CLOSED,
        )
        self.rule = BillingRule.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.STORAGE,
            calc_method=CalcMethod.PER_CBM_DAY,
            unit_price=Decimal("3.5000"),
        )
        self.accrual_a = BillingAccrual.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=self.period,
            charge_type=ChargeType.STORAGE,
            rule=self.rule,
            service_date=datetime.date(2026, 3, 1),
            currency="CNY",
            quantity=Decimal("12.0000"),
            unit_price=Decimal("3.5000"),
            amount=Decimal("42.00"),
            tax_amount=Decimal("2.52"),
            status=AccrualStatus.LOCKED,
            acc_fingerprint="console-acc-1",
            created_by=self.user,
        )
        self.accrual_b = BillingAccrual.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=self.period,
            charge_type=ChargeType.DISPATCH,
            rule=BillingRule.objects.create(
                owner=self.owner,
                warehouse=self.warehouse,
                charge_type=ChargeType.DISPATCH,
                calc_method=CalcMethod.PER_ORDER,
                unit_price=Decimal("10.0000"),
            ),
            service_date=datetime.date(2026, 3, 2),
            currency="CNY",
            quantity=Decimal("5.0000"),
            unit_price=Decimal("10.0000"),
            amount=Decimal("50.00"),
            tax_amount=Decimal("3.00"),
            status=AccrualStatus.LOCKED,
            acc_fingerprint="console-acc-2",
            created_by=self.user,
        )
        self.bill = Bill.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            period=self.period,
            invoice_no="INV-CONSOLE-0001",
            issue_date=datetime.date(2026, 4, 1),
            due_date=datetime.date(2026, 4, 10),
            currency="CNY",
            subtotal=Decimal("92.00"),
            tax_total=Decimal("5.52"),
            total=Decimal("97.52"),
            status=BillStatus.ISSUED,
        )
        BillLine.objects.create(
            bill=self.bill,
            accrual=self.accrual_a,
            charge_type=self.accrual_a.charge_type,
            service_date=self.accrual_a.service_date,
            quantity=self.accrual_a.quantity,
            unit_price=self.accrual_a.unit_price,
            amount=self.accrual_a.amount,
            tax_amount=self.accrual_a.tax_amount,
            description="按日仓储费",
        )
        BillLine.objects.create(
            bill=self.bill,
            accrual=self.accrual_b,
            charge_type=self.accrual_b.charge_type,
            service_date=self.accrual_b.service_date,
            quantity=self.accrual_b.quantity,
            unit_price=self.accrual_b.unit_price,
            amount=self.accrual_b.amount,
            tax_amount=self.accrual_b.tax_amount,
            description="订单处理费",
        )

    def test_billing_overview_page_renders_period_summary_and_bill_link(self):
        response = self.client.get(
            reverse("console:billing_overview"),
            {"period": self.period.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "计费总览")
        self.assertContains(response, self.period.label)
        self.assertContains(response, self.bill.invoice_no)
        self.assertContains(response, "¥92.00")
        self.assertContains(response, reverse("console:billing_bill_detail", args=[self.bill.id]))

    def test_bill_detail_page_filters_lines_by_charge_type(self):
        response = self.client.get(
            reverse("console:billing_bill_detail", args=[self.bill.id]),
            {
                "period": self.period.id,
                "owner": self.owner.id,
                "warehouse": self.warehouse.id,
                "charge_type": ChargeType.STORAGE,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "账单详情")
        self.assertContains(response, "按日仓储费")
        self.assertNotContains(response, "订单处理费")
