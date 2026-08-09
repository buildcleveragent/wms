from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from allapp.accounts.models import UserRoleScope
from allapp.baseinfo.models import Owner
from allapp.inventory.models import InventoryDetail
from allapp.locations.models import Location, Subwarehouse, Warehouse
from allapp.products.models import Product, ProductUom


class InventoryDetailAdminExportTests(TestCase):
    def setUp(self):
        self.owner = Owner.objects.create(code="ADMEXP", name="Admin Export Owner")
        self.uom = ProductUom.objects.create(code="ADEA", name="件", is_active=True)
        self.first_warehouse, self.first_location = self._create_location("ADEXP1", "ADEXA")
        self.second_warehouse, self.second_location = self._create_location("ADEXP2", "ADEXB")
        self.first_product = self._create_product("PROD-KEEP", "SKU-KEEP", "保留商品")
        self.second_product = self._create_product("PROD-HIDE", "SKU-HIDE", "排除商品")
        self._create_inventory(
            self.first_product,
            self.first_location,
            Decimal("12.0000"),
        )
        self._create_inventory(
            self.second_product,
            self.second_location,
            Decimal("25.0000"),
        )
        self.export_url = reverse(
            "admin:inventory_inventorydetail_export_filtered_excel"
        )
        self.changelist_url = reverse("admin:inventory_inventorydetail_changelist")

    def _create_location(self, warehouse_code, subwarehouse_code):
        warehouse = Warehouse.objects.create(
            code=warehouse_code,
            name=f"Warehouse {warehouse_code}",
        )
        subwarehouse = Subwarehouse.objects.create(
            warehouse=warehouse,
            code=subwarehouse_code,
            name=f"Subwarehouse {subwarehouse_code}",
        )
        location = Location.objects.create(
            warehouse=warehouse,
            subwarehouse=subwarehouse,
            code=f"{subwarehouse_code}-01-01-01",
            name=f"Location {subwarehouse_code}",
        )
        return warehouse, location

    def _create_product(self, code, sku, name):
        return Product.objects.create(
            owner=self.owner,
            code=code,
            sku=sku,
            name=name,
            base_uom=self.uom,
            volume=Decimal("0.100000"),
            price=Decimal("10.00"),
            batch_control=False,
            expiry_control=False,
        )

    def _create_inventory(self, product, location, quantity):
        return InventoryDetail.objects.create(
            owner=self.owner,
            product=product,
            warehouse=location.warehouse,
            subwarehouse=location.subwarehouse,
            location=location,
            base_unit=self.uom.code,
            onhand_qty=quantity,
            allocated_qty=Decimal("0"),
            locked_qty=Decimal("0"),
            damaged_qty=Decimal("0"),
        )

    def _rows(self, response):
        content = b"".join(response.streaming_content)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        return list(workbook["库存现存量"].iter_rows(values_only=True))

    def test_export_uses_current_admin_search(self):
        admin_user = get_user_model().objects.create_superuser(
            username="inventory-export-admin",
            password="x",
        )
        self.client.force_login(admin_user)

        changelist_response = self.client.get(
            self.changelist_url,
            {"q": self.first_product.sku},
        )
        self.assertContains(changelist_response, "导出当前筛选结果（Excel）")
        self.assertContains(
            changelist_response,
            f"{self.export_url}?q={self.first_product.sku}",
        )

        response = self.client.get(self.export_url, {"q": self.first_product.sku})

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response["Content-Disposition"])
        rows = self._rows(response)
        self.assertEqual(len(rows), 2)
        header, exported = rows
        self.assertIn("仓库SKU编码", header)
        self.assertEqual(
            exported[header.index("仓库SKU编码")], self.first_product.sku
        )
        self.assertEqual(exported[header.index("账面库存")], 12)

    def test_export_keeps_warehouse_access_scope(self):
        staff_user = get_user_model().objects.create_user(
            username="inventory-export-staff",
            password="x",
            is_staff=True,
        )
        staff_user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="inventory",
                codename="view_inventorydetail",
            )
        )
        UserRoleScope.objects.create(
            user=staff_user,
            role=UserRoleScope.Role.WAREHOUSE_OPERATOR,
            warehouse=self.first_warehouse,
        )
        self.client.force_login(staff_user)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 200)
        rows = self._rows(response)
        self.assertEqual(len(rows), 2)
        header, exported = rows
        self.assertEqual(exported[header.index("仓库编号")], "ADEXP1")
        self.assertEqual(
            exported[header.index("仓库SKU编码")], self.first_product.sku
        )
