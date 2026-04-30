# allapp/outbound/views.py  或  allapp/outbound/api_views.py
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.permissions import AllowAny
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import quote
from django.conf import settings
from django.http import FileResponse, Http404
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import load_workbook
from django.apps import apps
from datetime import datetime
from django.db.models import Q, Sum, Prefetch
import logging
from ..products.models import ProductPackage
from django.db.models import F
from rest_framework.exceptions import ValidationError
logger = logging.getLogger(__name__)
from rest_framework import viewsets, mixins, status
from rest_framework.pagination import PageNumberPagination
from .models import OutboundOrder
from rest_framework import serializers
from rest_framework.permissions import IsAuthenticated
from allapp.tasking.models import WmsTask, WmsTaskLine
from allapp.tasking import services as task_services
from allapp.tasking.services import _run_posting_handler, adjust_pick_line_qty
from django.db import transaction
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from allapp.billing.enums import PeriodStatus
from allapp.billing.models import BillingPeriod
from allapp.outbound.enums import PricingStatus
from allapp.outbound.models import OutboundOrderLine
from allapp.outbound.serializers import OutboundOrderCreateSerializer,ConfirmPricingSerializer, OutboundOrderReadSerializer
from allapp.core.utils.log_context import build_log_payload


# 放到 OutboundOrderViewSet 类中
@action(detail=True, methods=["post"], url_path="confirm-pricing")
@transaction.atomic
def confirm_pricing(self, request, pk=None):
    order = self.get_object()

    locked_period_exists = BillingPeriod.objects.filter(
        owner_id=order.owner_id,
        warehouse_id=order.warehouse_id,
        start_date__lte=order.biz_date,
        end_date__gte=order.biz_date,
        status__in=[PeriodStatus.CLOSED, PeriodStatus.INVOICED],
    ).exists()
    if locked_period_exists:
        return Response(
            {"detail": "该订单所属账期已关账或已开票，禁止确认/修改价格。"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    serializer = ConfirmPricingSerializer(
        data=request.data,
        context={"order": order, "request": request},
    )
    serializer.is_valid(raise_exception=True)

    lines_map = {
        line.id: line
        for line in OutboundOrderLine.objects.select_for_update().filter(
            order=order,
            is_deleted=False,
        )
    }

    total_amount = Decimal("0.00")
    for item in serializer.validated_data["lines"]:
        line = lines_map[item["line_id"]]
        line.base_price = item["base_price"]

        line_amount = (
            Decimal(line.base_qty or 0) * Decimal(line.base_price or 0)
        ).quantize(Decimal("0.01"))
        line.final_line_amount = line_amount
        line.save(update_fields=["base_price", "final_line_amount", "updated_at"])
        total_amount += line_amount

    order.final_order_amount = total_amount.quantize(Decimal("0.01"))
    order.pricing_status = PricingStatus.CONFIRMED
    order.priced_at = timezone.now()
    order.priced_by = request.user
    order.save(
        update_fields=[
            "final_order_amount",
            "pricing_status",
            "priced_at",
            "priced_by",
            "updated_at",
        ]
    )

    return Response(
        OutboundOrderReadSerializer(order, context={"request": request}).data,
        status=status.HTTP_200_OK,
    )



def _q3(x) -> Decimal:
    # 你项目里已有的数量标准化函数
    return _q3_impl(x)

class DefaultPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class ReceiveProductPagination(PageNumberPagination):
    page_size = 300
    page_size_query_param = "page_size"
    max_page_size = 500

class ProductViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = DefaultPagination

    def list(self, request, *args, **kwargs):
        Product   = apps.get_model("products", "Product")
        InvDetail = apps.get_model("inventory", "InventoryDetail")

        owner_id = getattr(request.user, "owner_id", None)
        if not owner_id:
            owner_id = request.query_params.get("owner")
            if not owner_id:
                ctx, ctx_text = build_log_payload(
                    user=request.user,
                    warehouse_id=request.query_params.get("warehouse_id"),
                )
                logger.warning("outbound.product_list.owner_missing %s", ctx_text, extra=ctx)
                return Response([])
        ctx, ctx_text = build_log_payload(
            user=request.user,
            owner_id=owner_id,
            warehouse_id=request.query_params.get("warehouse_id"),
        )

        # ✅ 只预取“当前查询命中的产品”的包装，并连带 uom，限制字段，避免 N+1
        pkg_qs = (ProductPackage.objects
                  .select_related("uom")
                  .only("id", "product_id", "uom_id", "qty_in_base", "barcode",
                        "length_cm", "width_cm", "height_cm",
                        "gross_weight_kg", "volume_m3",
                        "is_purchase_default", "sort_order",
                        "uom__name", "uom__code"))

        qs = (Product.objects
              .filter(owner_id=owner_id)
              .select_related("base_uom", "replenish_uom",)
              .only("id", "owner_id", "code", "name", "sku", "spec","product_image","gtin","min_price","max_discount",
                    "base_uom__code","base_uom__name", "replenish_uom__code", "replenish_uom__name","replenish_uom_id")
              .order_by("id"))

        q = request.query_params.get("search")
        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(code__icontains=q) |
                Q(sku__icontains=q) |
                Q(gtin__icontains=q) |
                Q(unit_barcode__icontains=q) |
                Q(carton_barcode__icontains=q)
            )

        page = self.paginate_queryset(qs)
        if not page:
            return self.get_paginated_response([])

        pid_list = [p.id for p in page]
        wh_id = request.query_params.get("warehouse_id")
        inv_filter = {"owner_id": owner_id, "product_id__in": pid_list}
        if wh_id:
            inv_filter["warehouse_id"] = int(wh_id)

        rows = (InvDetail.objects
                .filter(**inv_filter)
                .values("product_id")
                .annotate(avail=Sum("available_qty")))
        avail_map = {r["product_id"]: r["avail"] for r in rows}

        def carton_info(p):
            unit_code = getattr(getattr(p, "replenish_uom", None), "code", None)
            conv = None
            if unit_code and hasattr(p, "packages"):
                pkg = p.packages.filter(uom_id=p.replenish_uom_id).only("id", "base_qty", "multiplier").first()
                if pkg is not None:
                    conv = getattr(pkg, "base_qty", None) or getattr(pkg, "multiplier", None)
            return unit_code, conv

        def default_sales_uom(p):
            # 查找 product 的所有 package，返回 is_sales_default 为 True 的 UOM 的名称
            default_package = p.packages.filter(is_sales_default=True).first()
            if default_package:
                return default_package.uom.name,default_package.qty_in_base
            return None


        def product_packaging(p):
            """获取商品所有包装信息，使用 ProductPackage 来表示包装"""
            packaging = []
            for pkg in p.packages.all():  # 使用 p.packages 获取商品的所有包装信息
                packaging.append({
                    'id': pkg.id,  # 包装的 ID
                    'uom_type': pkg.uom.name,  # 获取包装单位名称
                    'quantity_in_base': pkg.qty_in_base,  # 获取换算数量
                    'barcode': pkg.barcode,  # 获取条码
                    'length_cm': pkg.length_cm,  # 获取包装尺寸
                    'width_cm': pkg.width_cm,
                    'height_cm': pkg.height_cm,
                    'gross_weight_kg': pkg.gross_weight_kg,  # 获取毛重
                    'volume_m3': pkg.volume_m3,  # 获取体积
                })
            return packaging

        # === 在服务端拼出 _unitOptions / _selectedUnitIndex（最小新增逻辑） ===
        def build_unit_options(p, packaging):
            base_name = (getattr(getattr(p, "base_uom", None), "name", None) or
                         getattr(getattr(p, "base_uom", None), "code", None))
            opts = []
            if base_name:
                opts.append({
                    "key": "BASE",
                    "kind": "base",
                    "label": base_name,
                    "multiplier": 1,
                    "package_id": None,
                    "barcode": None,
                })
            for row in packaging:
                # 跳过与基本单位 1:1 的冗余项
                if row["quantity_in_base"] == 1 and row["uom_type"] == base_name:
                    continue
                opts.append({
                    "key": row["id"],
                    "kind": "package",
                    "label": row["uom_type"],
                    "multiplier": row["quantity_in_base"],
                    "package_id": row["id"],
                    "barcode": row["barcode"],
                })
            return opts

        def default_selected_index(packaging, unit_opts):
            # 优先选择 is_sales_default 的包装；否则 0（通常为基本单位）
            sales_pkg_id = next((r["id"] for r in packaging if r.get("is_sales_default")), None)
            if sales_pkg_id is not None:
                for i, o in enumerate(unit_opts):
                    if o["package_id"] == sales_pkg_id:
                        return i
            return 0


        data = []
        for p in page:
            packaging = product_packaging(p)
            unit_opts = build_unit_options(p, packaging)  # ← 基本单位 + 包装
            sel_idx = default_selected_index(packaging, unit_opts)
            sel_unit = unit_opts[sel_idx] if unit_opts else None


            carton_unit, carton_conv = carton_info(p)

            if default_sales_uom(p):
              aux_uom_name,aux_qty_in_base=default_sales_uom(p)
            else:
              aux_uom_name=None
              aux_qty_in_base=None

            product_image_url = None
            if p.product_image:
                product_image_url = request.build_absolute_uri(p.product_image.url)
                # product_image_url = "http://192.168.1.6:8001"+p.product_image.url  # 获取图片的 URL 地址

            if avail_map.get(p.id, 0)>0:
                data.append({
                    "id": p.id,
                    "sku": p.sku or p.code or "",
                    "name": p.name or "",
                    "spec": p.spec,
                    "base_unit": getattr(getattr(p, "base_uom", None), "code", None),
                    "base_unit_name": getattr(getattr(p, "base_uom", None), "name", None),
                    "carton_unit": carton_unit,
                    "carton_conv": carton_conv,
                    "available": avail_map.get(p.id, 0),
                    "price": getattr(p, "price", None) or getattr(p, "sale_price", None) or 0,
                    "product_image_url":product_image_url,
                    "gtin":p.gtin,
                    "aux_uom_name":aux_uom_name,
                    "aux_qty_in_base":aux_qty_in_base,
                    "max_discount": p.max_discount ,
                    "product_min_price": p.min_price,
                    "unitOptions": unit_opts,
                    "selectedUnitIndex": sel_idx,
                    "base_quantity": 0,
                })
        logger.debug("outbound.product_list.response %s count=%s", ctx_text, len(data), extra=ctx)
        return self.get_paginated_response(data)

class CustomerViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = DefaultPagination     # 你项目里已有的分页类
    filter_backends = []
    queryset = apps.get_model("baseinfo", "Customer").objects.none()

    def get_queryset(self):
        Customer = apps.get_model("baseinfo", "Customer")
        user = self.request.user

        qs = Customer.objects.filter(salesperson=user)
        if getattr(user, "owner_id", None):
            qs = qs.filter(owner_id=user.owner_id)

        q = self.request.query_params.get("search")
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q))

        return qs.order_by("id")

    def list(self, request, *args, **kwargs):
        page = self.paginate_queryset(self.get_queryset())
        data = [{"id": c.id, "code": c.code, "name": c.name} for c in page]
        return self.get_paginated_response(data)

class OwnerViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = DefaultPagination     # 你项目里已有的分页类
    filter_backends = []
    queryset = apps.get_model("baseinfo", "Owner").objects.none()

    def get_queryset(self):
        Owner = apps.get_model("baseinfo", "Owner")
        user = self.request.user

        qs = Owner.objects.all()
        # if getattr(user, "owner_id", None):
        #     qs = qs.filter(owner_id=user.owner_id)

        q = self.request.query_params.get("search")
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q))

        return qs.order_by("id")

    def list(self, request, *args, **kwargs):
        page = self.paginate_queryset(self.get_queryset())
        data = [{"id": c.id, "code": c.code, "name": c.name} for c in page]
        return self.get_paginated_response(data)

class SupplierViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = DefaultPagination     # 你项目里已有的分页类
    filter_backends = []
    queryset = apps.get_model("baseinfo", "Supplier").objects.none()

    def get_queryset(self):
        Supplier = apps.get_model("baseinfo", "Supplier")
        user = self.request.user

        qs = Supplier.objects.all()
        # if getattr(user, "owner_id", None):
        #     qs = qs.filter(owner_id=user.owner_id)

        q = self.request.query_params.get("search")
        o = self.request.query_params.get("owner")
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q))

        if o:
            qs = qs.filter(Q(owner_id=o))
        else:

            raise ValidationError({"detail": "owner 参数是必需的"})

        return qs.order_by("id")

    def list(self, request, *args, **kwargs):
        page = self.paginate_queryset(self.get_queryset())
        data = [{"id": c.id, "code": c.code, "name": c.name} for c in page]
        return self.get_paginated_response(data)

class ReceiveProductViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]
    pagination_class = ReceiveProductPagination

    def list(self, request, *args, **kwargs):
        Product   = apps.get_model("products", "Product")
        ProductPackage = apps.get_model("products", "ProductPackage")

        owner_id = request.query_params.get("owner")
        if not owner_id:
            ctx, ctx_text = build_log_payload(user=request.user)
            logger.warning("outbound.receive_product_list.owner_missing %s", ctx_text, extra=ctx)
            return Response([])
        ctx, ctx_text = build_log_payload(user=request.user, owner_id=owner_id)

        # 预取 packages + uom，减少 N+1
        # pkg_qs = (ProductPackage.objects
        #           .select_related("uom")
        #           .only("id","product_id","uom_id","qty_in_base","barcode",
        #                 "is_sales_default","sort_order",
        #                 "uom__code","uom__name"))

        # ✅ 只预取“当前查询命中的产品”的包装，并连带 uom，限制字段，避免 N+1
        pkg_qs = (ProductPackage.objects
                  .select_related("uom")
                  .only("id", "product_id", "uom_id", "qty_in_base", "barcode",
                        "length_cm", "width_cm", "height_cm",
                        "gross_weight_kg", "volume_m3",
                        "is_purchase_default", "sort_order",
                        "uom__name", "uom__code"))


        # qs = (Product.objects
        #       .filter(owner_id=owner_id)
        #       .select_related("base_uom", "replenish_uom",)
        #       .only("id", "owner_id", "code", "name", "sku", "spec","product_image","gtin","min_price","max_discount",
        #             "base_uom__code","base_uom__name", "replenish_uom__code", "replenish_uom__name","replenish_uom_id")
        #       .order_by("id"))

        qs = (Product.objects
              .filter(owner_id=owner_id)
              .select_related("base_uom")
              .prefetch_related(Prefetch("packages", queryset=pkg_qs))
              .only("id", "owner_id", "code", "name", "sku", "spec", "product_image", "gtin",
                    "min_price", "max_discount", "base_uom__name", "base_uom__code")
              .order_by("id"))


        q = request.query_params.get("search")
        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(code__icontains=q) |
                Q(sku__icontains=q) |
                Q(gtin__icontains=q) |
                Q(unit_barcode__icontains=q) |
                Q(carton_barcode__icontains=q)
            ).distinct()



        page = self.paginate_queryset(qs)
        if not page:
            return self.get_paginated_response([])

        def carton_info(p):
            unit_code = getattr(getattr(p, "replenish_uom", None), "code", None)
            conv = None
            if unit_code and hasattr(p, "packages"):
                pkg = p.packages.filter(uom_id=p.replenish_uom_id).only("id", "base_qty", "multiplier").first()
                if pkg is not None:
                    conv = getattr(pkg, "base_qty", None) or getattr(pkg, "multiplier", None)
            return unit_code, conv

        def default_sales_uom(p):
            # 查找 product 的所有 package，返回 is_sales_default 为 True 的 UOM 的名称
            default_package = p.packages.filter(is_sales_default=True).first()
            if default_package:
                return default_package.uom.name,default_package.qty_in_base
            return None

        def product_packaging(p):
            """获取商品所有包装信息，使用 ProductPackage 来表示包装"""
            packaging = []
            for pkg in p.packages.all():  # 使用 p.packages 获取商品的所有包装信息
                packaging.append({
                    'id': pkg.id,  # 包装的 ID
                    'uom_type': pkg.uom.name,  # 获取包装单位名称
                    'quantity_in_base': pkg.qty_in_base,  # 获取换算数量
                    'barcode': pkg.barcode,  # 获取条码
                    'length_cm': pkg.length_cm,  # 获取包装尺寸
                    'width_cm': pkg.width_cm,
                    'height_cm': pkg.height_cm,
                    'gross_weight_kg': pkg.gross_weight_kg,  # 获取毛重
                    'volume_m3': pkg.volume_m3,  # 获取体积
                })
            return packaging

        # === 在服务端拼出 _unitOptions / _selectedUnitIndex（最小新增逻辑） ===
        def build_unit_options(p, packaging):
            base_name = (getattr(getattr(p, "base_uom", None), "name", None) or
                         getattr(getattr(p, "base_uom", None), "code", None))
            opts = []
            if base_name:
                opts.append({
                    "key": "BASE",
                    "kind": "base",
                    "label": base_name,
                    "multiplier": 1,
                    "package_id": None,
                    "barcode": None,
                })
            for row in packaging:
                # 跳过与基本单位 1:1 的冗余项
                if row["quantity_in_base"] == 1 and row["uom_type"] == base_name:
                    continue
                opts.append({
                    "key": row["id"],
                    "kind": "package",
                    "label": row["uom_type"],
                    "multiplier": row["quantity_in_base"],
                    "package_id": row["id"],
                    "barcode": row["barcode"],
                })
            return opts

        def default_selected_index(packaging, unit_opts):
            # 优先选择 is_sales_default 的包装；否则 0（通常为基本单位）
            sales_pkg_id = next((r["id"] for r in packaging if r.get("is_sales_default")), None)
            if sales_pkg_id is not None:
                for i, o in enumerate(unit_opts):
                    if o["package_id"] == sales_pkg_id:
                        return i
            return 0


        data = []
        for p in page:
            # product_packaging_info = product_packaging(p)
            packaging = product_packaging(p)
            unit_opts = build_unit_options(p, packaging)  # ← 基本单位 + 包装
            sel_idx = default_selected_index(packaging, unit_opts)
            sel_unit = unit_opts[sel_idx] if unit_opts else None

            carton_unit, carton_conv = carton_info(p)

            if default_sales_uom(p):
              aux_uom_name,aux_qty_in_base=default_sales_uom(p)
            else:
              aux_uom_name=None
              aux_qty_in_base=None

            product_image_url = None
            if p.product_image:
                product_image_url = request.build_absolute_uri(p.product_image.url)
                # product_image_url = "http://192.168.1.6:8001"+p.product_image.url  # 获取图片的 URL 地址


            data.append({
                "id": p.id,
                "sku": p.sku or p.code or "",
                "name": p.name or "",
                "spec": p.spec,
                "base_unit": getattr(getattr(p, "base_uom", None), "code", None),
                "base_unit_name": getattr(getattr(p, "base_uom", None), "name", None),
                "carton_unit": carton_unit,
                "carton_conv": carton_conv,
                "price": getattr(p, "price", None) or getattr(p, "sale_price", None) or 0,
                "product_image_url":product_image_url,
                "gtin":p.gtin,
                "aux_uom_name":aux_uom_name,
                "aux_qty_in_base":aux_qty_in_base,
                "max_discount": p.max_discount ,
                "product_min_price": p.min_price,
                "packaging": packaging,  # 包装信息返回，包括 id

                # === 新增：用于前端单选/回填到购物车
                "unitOptions": unit_opts,
                "selectedUnitIndex": sel_idx,
                "base_quantity":0,
                # "selectedUnit": sel_unit,  # 若不想冗余，可不下发这个，前端用 index 取
            })
        logger.debug("outbound.receive_product_list.response %s count=%s", ctx_text, len(data), extra=ctx)
        return self.get_paginated_response(data)

class OutboundOrderViewSet(
    mixins.CreateModelMixin, mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet
):
    queryset = OutboundOrder.objects.all().order_by("-biz_date", "-id")
    permission_classes = [IsAuthenticated]
    pagination_class = DefaultPagination

    def get_serializer_class(self):
        if self.action == "create":
            return OutboundOrderCreateSerializer
        return OutboundOrderReadSerializer

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        q = request.query_params.get("search")
        owner_id = request.query_params.get("owner_id")
        warehouse_id = request.query_params.get("warehouse_id")
        submit_status = request.query_params.get("submit_status")
        approval_status = request.query_params.get("approval_status")

        if q:
            qs = qs.filter(
                Q(order_no__icontains=q) |
                Q(customer__name__icontains=q) | Q(customer__code__icontains=q) |
                Q(supplier__name__icontains=q)
            )
        if owner_id:      qs = qs.filter(owner_id=owner_id)
        if warehouse_id:  qs = qs.filter(warehouse_id=warehouse_id)
        if submit_status: qs = qs.filter(submit_status=submit_status)
        if approval_status: qs = qs.filter(approval_status=approval_status)

        page = self.paginate_queryset(qs)
        ser = OutboundOrderReadSerializer(page, many=True)
        return self.get_paginated_response(ser.data)

    def create(self, request, *args, **kwargs):
        ser = OutboundOrderCreateSerializer(data=request.data, context={"request": request})
        ser.is_valid(raise_exception=True)
        order = ser.save()
        return Response(OutboundOrderReadSerializer(order).data, status=status.HTTP_201_CREATED)

    # 提交：DRAFT -> SUBMITTED
    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        order = self.get_object()
        if getattr(order, "submit_status", None) != "DRAFT":
            return Response({"detail": "仅 DRAFT 可提交"}, status=400)
        order.submit_status = "SUBMITTED"
        order.save(update_fields=["submit_status"])
        return Response(OutboundOrderReadSerializer(order).data)

    @action(detail=True, methods=["post"], url_path="owner-approve")
    def owner_approve(self, request, pk=None):
        order = self.get_object()

        if getattr(order, "approval_status", None) not in ("OWNER_PENDING", "OWNER_REJECTED"):
            return Response(
                {"detail": "当前状态不可进行货主管理员审核"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            order.owner_approve(by_user=request.user, allow_backorder=True)
        except DjangoValidationError as e:
            return Response(
                {"detail": e.message_dict if hasattr(e, "message_dict") else e.messages},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            return Response(
                {"detail": f"订单确认失败：{e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(OutboundOrderReadSerializer(order, context={"request": request}).data)

    @action(detail=True, methods=["post"], url_path="owner-reject")
    def owner_reject(self, request, pk=None):
        order = self.get_object()
        print("owner_reject is tabbed")
        if getattr(order, "approval_status", None) != "OWNER_PENDING":
            return Response(
                {"detail": "仅待审核订单可退回修改"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            with transaction.atomic():
                order = type(order).objects.select_for_update().get(pk=order.pk)
                order.approval_status = "OWNER_REJECTED"
                if hasattr(order, "approved_by_ownermanager"):
                    order.approved_by_ownermanager = request.user
                if hasattr(order, "approved_at_ownermanager"):
                    order.approved_at_ownermanager = timezone.now()
                order.save(update_fields=[
                    "approval_status",
                    "approved_by_ownermanager",
                    "approved_at_ownermanager",
                    "updated_at",
                ])
        except DjangoValidationError as e:
            return Response(
                {"detail": e.message_dict if hasattr(e, "message_dict") else e.messages},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            return Response(
                {"detail": f"退回修改失败：{e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            OutboundOrderReadSerializer(order, context={"request": request}).data,
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel(self, request, pk=None):
        from allapp.outbound import services as ob_services

        order = self.get_object()
        print("1 cancel is tabbed")

        if getattr(order, "approval_status", None) == "CANCELLED":
            return Response(
                {"detail": "订单已取消，请勿重复操作"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if getattr(order, "approval_status", None) == "WHS_APPROVED":
            return Response(
                {"detail": "订单已进入仓库处理流程，不能直接取消"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        print("2 cancel is tabbed")
        try:
            with transaction.atomic():
                order = type(order).objects.select_for_update().get(pk=order.pk)

                if getattr(order, "approval_status", None) in ("OWNER_APPROVED", "WHS_PENDING"):
                    print("2.1 cancel is tabbed before ob_services.unallocate_for_order")
                    ob_services.unallocate_for_order(order)
                print("3 cancel is tabbed")
                order.approval_status = "CANCELLED"
                # order.is_closed = True
                if not order.close_reason:
                    order.close_reason = "货主管理员取消订单"

                order.save(update_fields=[
                    "approval_status",
                    "close_reason",
                    "updated_at",
                ])
        except DjangoValidationError as e:
            print("4 cancel is tabbed")
            return Response(
                {"detail": e.message_dict if hasattr(e, "message_dict") else e.messages},

                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            print("5 cancel is tabbed")
            return Response(
                {"detail": f"取消订单失败：{e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        print("6 cancel is tabbed")
        return Response(
            OutboundOrderReadSerializer(order, context={"request": request}).data,
            status=status.HTTP_200_OK,
        )

    def _excel_str(self, v):
        return "" if v is None else str(v).strip()

    def _excel_decimal(self, v):
        if v in (None, ""):
            raise ValueError("数量不能为空")
        try:
            d = Decimal(str(v).strip())
        except (InvalidOperation, ValueError):
            raise ValueError("数量格式不正确")
        if d <= 0:
            raise ValueError("数量必须大于 0")
        return d

    def _build_ship_to(self, row_dict):
        parts = [
            self._excel_str(row_dict.get("收件人省")),
            self._excel_str(row_dict.get("收件人市")),
            self._excel_str(row_dict.get("收件人区")),
            self._excel_str(row_dict.get("收件人详细地址")),
        ]
        return "".join([p for p in parts if p])

    def _build_remark(self, row_dict):
        parts = []

        remark = self._excel_str(row_dict.get("备注"))
        if remark:
            parts.append(f"备注:{remark}")

        express_no = self._excel_str(row_dict.get("物流单号"))
        if express_no:
            parts.append(f"物流单号:{express_no}")

        sale_attr = self._excel_str(row_dict.get("销售属性"))
        if sale_attr:
            parts.append(f"销售属性:{sale_attr}")

        goods_name = self._excel_str(row_dict.get("商品名称"))
        if goods_name:
            parts.append(f"商品名称:{goods_name}")

        sender_name = self._excel_str(row_dict.get("发货人姓名"))
        sender_phone = self._excel_str(row_dict.get("发货人手机/电话"))
        sender_addr = "".join([
            self._excel_str(row_dict.get("发货人省")),
            self._excel_str(row_dict.get("发货人市")),
            self._excel_str(row_dict.get("发货人区")),
            self._excel_str(row_dict.get("发货人详细地址")),
        ])
        if sender_name or sender_phone or sender_addr:
            parts.append(
                f"发货人:{sender_name} {sender_phone} {sender_addr}".strip()
            )

        return " | ".join(parts)

    def _get_default_price(self, product):
        """
        这份模板没有价格列，先沿用你系统里的商品默认售价逻辑。
        如果取不到，就按 0 处理。
        """
        for attr in ("price", "sale_price", "base_price"):
            val = getattr(product, attr, None)
            if val not in (None, ""):
                try:
                    return Decimal(str(val))
                except Exception:
                    pass
        return Decimal("0")

    def _find_product_for_import(self, owner_id, row_dict):
        Product = apps.get_model("products", "Product")

        sku = self._excel_str(row_dict.get("商家编码"))
        goods_name = self._excel_str(row_dict.get("商品名称"))

        if sku:
            p = Product.objects.filter(owner_id=owner_id, sku=sku).order_by("id").first()
            if p:
                return p
            raise ValueError(f"商家编码[{sku}]匹配不到商品")

        if goods_name:
            qs = Product.objects.filter(owner_id=owner_id, name=goods_name).order_by("id")
            cnt = qs.count()
            if cnt == 1:
                return qs.first()
            if cnt > 1:
                raise ValueError(f"商品名称[{goods_name}]匹配到多个商品，请改填商家编码")
            raise ValueError(f"商品名称[{goods_name}]匹配不到商品")

        raise ValueError("商家编码和商品名称不能同时为空")

    @action(
        detail=False,
        methods=["post"],
        url_path="import-drop-ship-excel",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_drop_ship_excel(self, request):
        """
        上传一件代发 Excel，按行生成 OutboundOrder。
        每行 1 单、每单 1 条明细。
        """
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "请上传 Excel 文件，字段名 file"}, status=status.HTTP_400_BAD_REQUEST)

        user = request.user
        owner_id = getattr(user, "owner_id", None)
        warehouse_id = getattr(user, "warehouse_id", None)

        if not owner_id:
            return Response({"detail": "当前用户未绑定货主(owner)"}, status=status.HTTP_400_BAD_REQUEST)
        if not warehouse_id:
            return Response({"detail": "当前用户未绑定仓库(warehouse)"}, status=status.HTTP_400_BAD_REQUEST)

        Customer = apps.get_model("baseinfo", "Customer")
        OutboundOrder = apps.get_model("outbound", "OutboundOrder")

        cash_customer = Customer.objects.filter(owner_id=owner_id, code="CASH").order_by("id").first()
        if not cash_customer:
            return Response(
                {"detail": f"当前货主[{owner_id}]下不存在 code=CASH 的散客客户，请先创建"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({"detail": f"Excel 解析失败: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Excel 为空"}, status=status.HTTP_400_BAD_REQUEST)

        headers = [self._excel_str(x) for x in rows[0]]
        required_headers = [
            "收件人姓名",
            "收件人手机/电话",
            "收件人详细地址",
            "数量",
            "订单编号",
        ]
        missing = [h for h in required_headers if h not in headers]
        if missing:
            return Response(
                {"detail": f"模板缺少必要列: {missing}", "headers": headers},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = {
            "total_rows": 0,
            "success_count": 0,
            "skip_count": 0,
            "fail_count": 0,
            "successes": [],
            "skips": [],
            "errors": [],
        }

        serializer_cls = self.get_serializer_class()
        if serializer_cls is OutboundOrderReadSerializer:
            # create 动作外手工拿创建 serializer
            from .serializers import OutboundOrderCreateSerializer
            create_serializer_cls = OutboundOrderCreateSerializer
        else:
            create_serializer_cls = serializer_cls

        for excel_row_no, raw_row in enumerate(rows[1:], start=2):
            if not raw_row or all(self._excel_str(v) == "" for v in raw_row):
                continue

            result["total_rows"] += 1
            row_dict = dict(zip(headers, raw_row))

            try:
                src_bill_no = self._excel_str(row_dict.get("订单编号"))
                contact = self._excel_str(row_dict.get("收件人姓名"))
                contact_phone = self._excel_str(row_dict.get("收件人手机/电话"))
                ship_to = self._build_ship_to(row_dict)
                qty = self._excel_decimal(row_dict.get("数量"))

                if not src_bill_no:
                    raise ValueError("订单编号不能为空")
                if not contact:
                    raise ValueError("收件人姓名不能为空")
                if not contact_phone:
                    raise ValueError("收件人手机/电话不能为空")
                if not ship_to:
                    raise ValueError("收货地址不能为空")

                # 幂等：同 owner + src_bill_no 已存在则跳过
                existing = OutboundOrder.objects.filter(
                    owner_id=owner_id,
                    src_bill_no=src_bill_no,
                ).order_by("id").first()
                if existing:
                    result["skip_count"] += 1
                    result["skips"].append({
                        "row": excel_row_no,
                        "src_bill_no": src_bill_no,
                        "reason": f"订单已存在，order_id={existing.id}, order_no={existing.order_no}",
                    })
                    continue

                product = self._find_product_for_import(owner_id, row_dict)
                price = self._get_default_price(product)
                remark = self._build_remark(row_dict)

                payload = {
                    "customer_id": cash_customer.id,
                    "remark": remark,
                    "src_bill_no": src_bill_no,
                    "contact": contact,
                    "contact_phone": contact_phone,
                    "ship_to": ship_to,
                    "items": [
                        {
                            "product_id": product.id,
                            "qty": qty,
                            "price": price,
                        }
                    ],
                }

                ser = create_serializer_cls(data=payload, context={"request": request})
                ser.is_valid(raise_exception=True)

                with transaction.atomic():
                    order = ser.save()

                result["success_count"] += 1
                result["successes"].append({
                    "row": excel_row_no,
                    "src_bill_no": src_bill_no,
                    "order_id": order.id,
                    "order_no": getattr(order, "order_no", ""),
                })

            except Exception as e:
                result["fail_count"] += 1
                result["errors"].append({
                    "row": excel_row_no,
                    "src_bill_no": self._excel_str(row_dict.get("订单编号")) if "row_dict" in locals() else "",
                    "reason": str(e),
                })

        return Response(result, status=status.HTTP_200_OK)


    @action(detail=False, methods=["get"], url_path="import-drop-ship-template")
    def import_drop_ship_template(self, request):
        """
        下载一件代发 Excel 模板
        """
        template_path = Path(settings.BASE_DIR) / "allapp" / "outbound" / "resources" / "一件代发.xlsx"

        if not template_path.exists():
            raise Http404("模板文件不存在，请联系管理员。")

        filename = "一件代发.xlsx"
        response = FileResponse(
            open(template_path, "rb"),
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # 兼容中文文件名下载
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{quote(filename)}"
        )
        return response

    @action(
        detail=False,
        methods=["get"],
        url_path="import-drop-ship-template",
        permission_classes=[AllowAny],
        authentication_classes=[],
    )
    def import_drop_ship_template(self, request):
        template_path = Path(settings.BASE_DIR) / "allapp" / "outbound" / "resources" / "yi-jian-dai-fa-mo-ban.xlsx"

        if not template_path.exists():
            raise Http404("模板文件不存在，请联系管理员。")

        filename = "一件代发模板.xlsx"
        response = FileResponse(
            open(template_path, "rb"),
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return response

class PickTaskSerializer(serializers.ModelSerializer):
    owner_name = serializers.CharField(source="owner.name", read_only=True)
    warehouse_name = serializers.CharField(source="warehouse.name", read_only=True)

    class Meta:
        model = WmsTask
        fields = [
            "id",
            "task_no",
            "task_type",
            "status",
            "owner_id",
            "owner_name",
            "warehouse_id",
            "warehouse_name",
            "remark",
            "review_status",
        ]

class PickTaskLineSerializer(serializers.ModelSerializer):
    product_sku = serializers.CharField(source="product.sku", read_only=True)
    product_name = serializers.CharField(source="product.name", read_only=True)
    from_loc_code = serializers.CharField(source="from_location.code", read_only=True)
    to_loc_code = serializers.CharField(source="to_location.code", read_only=True)

    class Meta:
        model = WmsTaskLine
        fields = [
            "id",
            "task_id",
            "product_id",
            "product_sku",
            "product_name",
            "from_location_id",
            "from_loc_code",
            "to_location_id",
            "to_loc_code",
            "qty_plan",
            "qty_done",
            "status",
        ]


class PickTaskViewSet(viewsets.ReadOnlyModelViewSet):
    """
    PDA 拣货任务接口：
      - GET  /api/outbound/pda/pick-tasks/             列表
      - GET  /api/outbound/pda/pick-tasks/<id>/        任务头
      - GET  /api/outbound/pda/pick-tasks/<id>/lines/  行
      - POST /api/outbound/pda/pick-tasks/<id>/scan/   扫码拣货
      - POST /api/outbound/pda/pick-tasks/<id>/create-review-task/ complete拣货
      - POST /api/outbound/pda/pick-tasks/<id>/post/   完成并过账
    """
    permission_classes = [IsAuthenticated]
    serializer_class = PickTaskSerializer

    def get_queryset(self):
        qs = WmsTask.objects.filter(
            task_type=WmsTask.TaskType.PICK,
        ).order_by("-id")

        action = getattr(self, "action", None)

        # 1) status 过滤
        status_list = self.request.query_params.getlist("status")
        if status_list:
            qs = qs.filter(status__in=status_list)
        else:
            # 只在 list 时应用“默认进行中状态”的过滤
            if action == "list":
                qs = qs.filter(
                    status__in=[
                        WmsTask.Status.RELEASED,
                        WmsTask.Status.IN_PROGRESS,
                        WmsTask.Status.RESERVED,
                    ]
                )

        # 2) review_status 过滤（你列表已经传了 ?review_status=PENDING）
        review_status_list = self.request.query_params.getlist("review_status")
        if review_status_list:
            qs = qs.filter(review_status__in=review_status_list)

        # 3) for_review=1 时，进一步强制“已完成 + 待审核”
        #    （可选，但很直观）
        for_review = self.request.query_params.get("for_review")
        if for_review in ("1", "true", "True"):
            qs = qs.filter(
                status=WmsTask.Status.COMPLETED,
                review_status=WmsTask.ReviewStatus.PENDING,
            ).exclude(picked_by=self.request.user)

        return qs

    @action(methods=["get"], detail=True)
    def lines(self, request, pk=None):
        ctx, ctx_text = build_log_payload(task_id=pk, user=request.user)
        logger.info("outbound.pick.lines.request %s", ctx_text, extra=ctx)
        lines = (
            WmsTaskLine.objects
            .filter(task_id=pk)
            .select_related("product", "from_location", "to_location")
            .order_by("id")
        )
        data = PickTaskLineSerializer(lines, many=True).data
        logger.debug("outbound.pick.lines.response %s count=%s", ctx_text, len(data), extra=ctx)
        return Response(data)

    @action(methods=["post"], detail=True)
    def scan(self, request, pk=None):
        """
        请求体：
          { "barcode": "...", "qty": 1 }

        内部调用统一的 scan_task()，根据任务类型 = PICK 处理。
        """
        payload = request.data or {}
        barcode = (payload.get("barcode") or "").strip()
        qty = payload.get("qty") or 1
        location_id = payload.get("location_id") or None

        if not barcode:
            return Response({"detail": "缺少条码"}, status=400)

        res = task_services.scan_task(
            task_id=int(pk),
            barcode=barcode,
            qty=qty,
            location_id=location_id,
            by_user=request.user,
            client_seq=payload.get("client_seq"),
        )

        # scan_task 返回里至少有 line_id / qty_done
        line_id = res.get("line_id")
        if line_id:
            line = (
                WmsTaskLine.objects
                .filter(id=line_id)
                .only("id", "qty_plan", "qty_done", "status")
                .first()
            )
            if line:
                res["line"] = {
                    "id": line.id,
                    "qty_plan": line.qty_plan,
                    "qty_done": line.qty_done,
                    "status": line.status,
                }

        return Response(res)

# === 新增：拣货完成 → 提交复核 =========================================
    @action(methods=["post"], detail=True, url_path="create-review-task")
    @transaction.atomic
    def create_review_task(self, request, pk=None):
        """
        拣货完成，不直接过账：
          - 校验所有明细 qty_done >= qty_plan
          - 状态流转：
              status         → COMPLETED
              review_status  → PENDING
              posting_status → NOT_READY
          - 记录 picked_by = 当前用户（拣货人）
        """
        task = (
            WmsTask.objects
            .select_for_update()
            .get(id=pk, task_type=WmsTask.TaskType.PICK)
        )
        ctx, ctx_text = build_log_payload(task=task, user=request.user)
        logger.info("outbound.pick.create_review.begin %s", ctx_text, extra=ctx)
        # 1）必须是进行中或已发布等“可完成”的状态
        if task.status not in (
            WmsTask.Status.RELEASED,
            WmsTask.Status.IN_PROGRESS,
            WmsTask.Status.RESERVED,
        ):
            raise ValidationError(f"当前任务状态为 {task.status}，不能提交复核。")
        # 2）检查是否还有未拣完的明细
        has_undone = WmsTaskLine.objects.filter(
            task_id=task.id,
            qty_done__lt=F("qty_plan"),
        ).exists()
        if has_undone:
            logger.warning("outbound.pick.create_review.pending_lines %s", ctx_text, extra=ctx)
            raise ValidationError("还有未拣完的明细，不能提交复核。")
        # 3）流转到“已完成，待审核”
        task.status = WmsTask.Status.COMPLETED
        task.review_status = WmsTask.ReviewStatus.PENDING
        task.updated_at=datetime.now()
        # 过账还没到，不要改 posting_status（保持 NOT_READY）
        # 记录拣货人（如果你加了 picked_by 字段）
        if task.picked_by_id is None:
            task.picked_by = request.user

        task.finished_at = task.finished_at or timezone.now()
        task.save(update_fields=[
            "status",
            "review_status",
            "finished_at",
            "picked_by",
            "updated_at",
        ])
        logger.info(
            "outbound.pick.create_review.completed %s status=%s review_status=%s posting_status=%s",
            ctx_text,
            task.status,
            task.review_status,
            task.posting_status,
            extra=ctx,
        )
        return Response({
            "task_id": task.id,
            "status": task.status,
            "review_status": task.review_status,
            "posting_status": task.posting_status,
            "message": "拣货已完成，已提交复核。",
        })

    # === 修改：复核通过 + 过账 ==============================================
    @action(methods=["post"], detail=True)
    @transaction.atomic
    def post(self, request, pk=None):
        """
        复核通过并过账：
          - 仅允许在：
              status=COMPLETED & review_status=PENDING
          - 当前登录用户 != 拣货人（picked_by）
          - 设置：
              review_status = APPROVED
              approved_by   = 当前用户
              approved_at   = now
              posting_status = PENDING
          - 然后调用 _run_posting_handler 真正过账
        """
        task = (
            WmsTask.objects
            .select_for_update()
            .get(id=pk, task_type=WmsTask.TaskType.PICK)
        )
        ctx, ctx_text = build_log_payload(task=task, user=request.user)
        logger.info("outbound.pick.post.begin %s", ctx_text, extra=ctx)
        # 1）必须是“已完成，待审核”
        if task.status != WmsTask.Status.COMPLETED:
            raise ValidationError("任务未处于已完成状态，不能过账。")
        if task.review_status != WmsTask.ReviewStatus.PENDING:
            raise ValidationError("当前任务不在待审核状态，不能过账。")
        # 2）禁止拣货人自己复核自己
        picker = getattr(task, "picked_by", None)
        if picker and picker.id == request.user.id:
            logger.warning("outbound.pick.post.self_review_blocked %s", ctx_text, extra=ctx)
            raise ValidationError("拣货人不能作为本任务的复核人。")
        # 3）设置审核通过 & 准备过账
        task.review_status = WmsTask.ReviewStatus.APPROVED
        task.approved_by = request.user
        task.approved_at = timezone.now()
        task.approval_note = (task.approval_note or "") + " PDA拣货复核通过"

        task.posting_status = WmsTask.PostingStatus.PENDING
        task.save(update_fields=[
            "review_status",
            "approved_by",
            "approved_at",
            "approval_note",
            "posting_status",
        ])

        # 4）调用统一过账逻辑
        result = _run_posting_handler(
            task_id=int(pk),
            by_user=request.user,
            note="PDA拣货复核通过并过账",
        )
        logger.info("outbound.pick.post.completed %s", ctx_text, extra=ctx)
        # 这里假设 _run_posting_handler 内部会把 posting_status / posted_by / posted_at 更新好

        return Response({
            "task_id": int(pk),
            "posted": True,
            **(result or {}),
        })

    @action(methods=["post"], detail=True, url_path="adjust-line-qty")
    def adjust_line_qty(self, request, pk=None):
        """
        PDA 手工调整拣货行数量：
          - 请求体：{ line_id, final_qty_done, client_seq? }
          - 调用 tasking.services.adjust_pick_line_qty
        """
        ctx, ctx_text = build_log_payload(task_id=pk, user=request.user)
        logger.info("outbound.pick.adjust_line_qty.begin %s", ctx_text, extra=ctx)
        task_id = int(pk)
        line_id = request.data.get("line_id")
        final_qty_done = request.data.get("final_qty_done")
        client_seq = request.data.get("client_seq")

        if not line_id:
            return Response(
                {"detail": "缺少 line_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if final_qty_done is None:
            return Response(
                {"detail": "缺少 final_qty_done"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            res = adjust_pick_line_qty(
                task_id=task_id,
                line_id=int(line_id),
                final_qty=final_qty_done,
                by_user=request.user,
                client_seq=client_seq,
            )
        except ValidationError as e:
            # 和其它接口一样，抛 Django/DRF 的 ValidationError
            raise e

        return Response(res)
