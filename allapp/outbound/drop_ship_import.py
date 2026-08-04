"""Bounded, tenant-scoped one-piece drop-shipping Excel import."""

from __future__ import annotations

import io
import logging
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.apps import apps
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import IntegrityError, transaction
from django.db.models import Q
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError as DRFValidationError

from allapp.outbound import services as outbound_services
from allapp.outbound.serializers import OutboundOrderCreateSerializer

logger = logging.getLogger(__name__)

MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
MAX_XLSX_UNCOMPRESSED_SIZE = 50 * 1024 * 1024
MAX_XLSX_ENTRIES = 300

REQUIRED_HEADERS = frozenset(
    {
        "收件人姓名",
        "收件人手机/电话",
        "收件人详细地址",
        "数量",
        "订单编号",
    }
)

GENERIC_ROW_ERROR = "系统处理失败，请联系管理员。"


class DropShipImportFileError(Exception):
    """A safe, client-visible file-level validation error."""


class DropShipRowError(Exception):
    """A safe, client-visible business-row validation error."""


@dataclass(frozen=True)
class ParsedRow:
    row_number: int
    values: dict[str, object]
    has_formula: bool


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _quantity(value) -> Decimal:
    if value in (None, ""):
        raise DropShipRowError("数量不能为空")
    try:
        quantity = Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise DropShipRowError("数量格式不正确") from exc
    if not quantity.is_finite() or quantity <= 0:
        raise DropShipRowError("数量必须大于 0")
    return quantity


def _build_ship_to(row: dict[str, object]) -> str:
    return "".join(
        _text(row.get(header))
        for header in ("收件人省", "收件人市", "收件人区", "收件人详细地址")
        if _text(row.get(header))
    )


def _build_remark(row: dict[str, object]) -> str:
    parts: list[str] = []
    for header, label in (
        ("备注", "备注"),
        ("物流单号", "物流单号"),
        ("销售属性", "销售属性"),
        ("商品名称", "商品名称"),
    ):
        value = _text(row.get(header))
        if value:
            parts.append(f"{label}:{value}")

    sender_name = _text(row.get("发货人姓名"))
    sender_phone = _text(row.get("发货人手机/电话"))
    sender_address = "".join(
        _text(row.get(header))
        for header in ("发货人省", "发货人市", "发货人区", "发货人详细地址")
        if _text(row.get(header))
    )
    if sender_name or sender_phone or sender_address:
        parts.append(f"发货人:{sender_name} {sender_phone} {sender_address}".strip())
    return " | ".join(parts)


def _safe_validation_message(detail) -> str:
    """Flatten DRF's known business validation detail without exposing internals."""

    messages: list[str] = []

    def collect(value):
        if isinstance(value, dict):
            for child in value.values():
                collect(child)
        elif isinstance(value, (list, tuple)):
            for child in value:
                collect(child)
        elif value not in (None, ""):
            messages.append(str(value))

    collect(detail)
    return "；".join(messages[:3])[:300] or "导入数据校验失败。"


class DropShipImportService:
    def __init__(self, *, request, owner_id: int, warehouse_id: int, cash_customer):
        self.request = request
        self.owner_id = owner_id
        self.warehouse_id = warehouse_id
        self.cash_customer = cash_customer
        self.user_id = getattr(request.user, "id", None)

    def import_file(self, uploaded_file) -> dict[str, object]:
        workbook = self._load_workbook(uploaded_file)
        try:
            rows = self._parse_rows(workbook)
        except DropShipImportFileError:
            raise
        except Exception as exc:
            logger.exception(
                "drop_ship_import_parse_failed owner_id=%s user_id=%s row=%s",
                self.owner_id,
                self.user_id,
                None,
            )
            raise DropShipImportFileError(
                "无法解析 Excel 文件，请使用系统模板重新保存。"
            ) from exc
        finally:
            workbook.close()

        return self._persist_rows(rows)

    def _load_workbook(self, uploaded_file):
        name = Path(getattr(uploaded_file, "name", "") or "").name
        if Path(name).suffix.lower() != ".xlsx":
            raise DropShipImportFileError("仅支持 .xlsx 格式的 Excel 文件。")

        size = getattr(uploaded_file, "size", None)
        if size is not None and size > MAX_IMPORT_FILE_SIZE:
            raise DropShipImportFileError("Excel 文件不能超过 5 MB。")
        try:
            uploaded_file.seek(0)
            data = uploaded_file.read(MAX_IMPORT_FILE_SIZE + 1)
        except (OSError, ValueError) as exc:
            raise DropShipImportFileError(
                "无法读取 Excel 文件，请重新选择文件。"
            ) from exc
        if len(data) > MAX_IMPORT_FILE_SIZE:
            raise DropShipImportFileError("Excel 文件不能超过 5 MB。")
        if not data:
            raise DropShipImportFileError("Excel 文件为空。")

        try:
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                entries = archive.infolist()
                if len(entries) > MAX_XLSX_ENTRIES:
                    raise DropShipImportFileError("Excel 压缩条目不能超过 300 个。")
                if (
                    sum(entry.file_size for entry in entries)
                    > MAX_XLSX_UNCOMPRESSED_SIZE
                ):
                    raise DropShipImportFileError("Excel 解压后内容不能超过 50 MB。")
                if archive.testzip() is not None:
                    raise DropShipImportFileError("Excel 文件已损坏。")
        except DropShipImportFileError:
            raise
        except (OSError, RuntimeError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
            raise DropShipImportFileError(
                "无法解析 Excel 文件，请使用系统模板重新保存。"
            ) from exc

        try:
            return load_workbook(io.BytesIO(data), data_only=False, read_only=True)
        except Exception as exc:
            logger.exception(
                "drop_ship_import_open_failed owner_id=%s user_id=%s row=%s",
                self.owner_id,
                self.user_id,
                None,
            )
            raise DropShipImportFileError(
                "无法解析 Excel 文件，请使用系统模板重新保存。"
            ) from exc

    def _parse_rows(self, workbook) -> list[ParsedRow]:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=False)
        try:
            header_cells = next(iterator)
        except StopIteration as exc:
            raise DropShipImportFileError("Excel 文件为空。") from exc

        if any(cell.data_type == "f" for cell in header_cells):
            raise DropShipImportFileError("Excel 表头不允许使用公式。")
        headers = [_text(cell.value) for cell in header_cells]
        nonempty_headers = [header for header in headers if header]
        duplicates = sorted(
            header
            for header in set(nonempty_headers)
            if nonempty_headers.count(header) > 1
        )
        if duplicates:
            raise DropShipImportFileError(f"Excel 表头重复：{', '.join(duplicates)}。")
        missing = sorted(REQUIRED_HEADERS.difference(nonempty_headers))
        if missing:
            raise DropShipImportFileError(f"模板缺少必要列：{', '.join(missing)}。")

        parsed: list[ParsedRow] = []
        for row_number, cells in enumerate(iterator, start=2):
            values = [cell.value for cell in cells]
            if not any(_text(value) for value in values):
                continue
            if len(parsed) >= MAX_IMPORT_ROWS:
                raise DropShipImportFileError(
                    "Excel 非空业务行不能超过 1000 行，请拆分文件。"
                )
            parsed.append(
                ParsedRow(
                    row_number=row_number,
                    values=dict(zip(headers, values)),
                    has_formula=any(
                        cell.data_type == "f" and bool(headers[index])
                        for index, cell in enumerate(cells[: len(headers)])
                    ),
                )
            )
        return parsed

    def _product_maps(self, rows: list[ParsedRow]):
        Product = apps.get_model("products", "Product")
        skus = {_text(row.values.get("商家编码")) for row in rows}
        names = {_text(row.values.get("商品名称")) for row in rows}
        skus.discard("")
        names.discard("")
        products = (
            Product.objects.filter(owner_id=self.owner_id)
            .filter(Q(sku__in=skus) | Q(name__in=names))
            .order_by("id")
        )
        by_sku: dict[str, object] = {}
        by_name: dict[str, list[object]] = {}
        for product in products:
            if product.sku and product.sku not in by_sku:
                by_sku[product.sku] = product
            by_name.setdefault(product.name, []).append(product)
        return by_sku, by_name

    def _find_product(self, row, by_sku, by_name):
        sku = _text(row.get("商家编码"))
        product_name = _text(row.get("商品名称"))
        if sku:
            product = by_sku.get(sku)
            if product is None:
                raise DropShipRowError(f"商家编码[{sku}]匹配不到商品")
            return product
        if product_name:
            products = by_name.get(product_name, [])
            if len(products) == 1:
                return products[0]
            if len(products) > 1:
                raise DropShipRowError(
                    f"商品名称[{product_name}]匹配到多个商品，请改填商家编码"
                )
            raise DropShipRowError(f"商品名称[{product_name}]匹配不到商品")
        raise DropShipRowError("商家编码和商品名称不能同时为空")

    def _persist_rows(self, rows: list[ParsedRow]) -> dict[str, object]:
        OutboundOrder = apps.get_model("outbound", "OutboundOrder")
        source_numbers = {
            _text(row.values.get("订单编号"))
            for row in rows
            if _text(row.values.get("订单编号"))
        }
        existing_orders = {
            order.src_bill_no: order
            for order in OutboundOrder.all_objects.filter(
                owner_id=self.owner_id,
                src_bill_no__in=source_numbers,
            ).order_by("id")
        }
        by_sku, by_name = self._product_maps(rows)
        result = {
            "total_rows": len(rows),
            "success_count": 0,
            "skip_count": 0,
            "fail_count": 0,
            "successes": [],
            "skips": [],
            "errors": [],
        }

        for parsed in rows:
            row = parsed.values
            source_number = _text(row.get("订单编号"))
            try:
                if parsed.has_formula:
                    raise DropShipRowError("业务单元格不允许使用公式")
                contact = _text(row.get("收件人姓名"))
                contact_phone = _text(row.get("收件人手机/电话"))
                ship_to = _build_ship_to(row)
                quantity = _quantity(row.get("数量"))
                if not source_number:
                    raise DropShipRowError("订单编号不能为空")
                if not contact:
                    raise DropShipRowError("收件人姓名不能为空")
                if not contact_phone:
                    raise DropShipRowError("收件人手机/电话不能为空")
                if not ship_to:
                    raise DropShipRowError("收货地址不能为空")

                existing = existing_orders.get(source_number)
                if existing:
                    self._append_skip(
                        result, parsed.row_number, source_number, existing
                    )
                    continue

                product = self._find_product(row, by_sku, by_name)
                payload = {
                    "warehouse_id": self.warehouse_id,
                    "customer_id": self.cash_customer.id,
                    "remark": _build_remark(row),
                    "src_bill_no": source_number,
                    "contact": contact,
                    "contact_phone": contact_phone,
                    "ship_to": ship_to,
                    "items": [
                        {
                            "product_id": product.id,
                            "qty": quantity,
                            "price": outbound_services.get_default_product_price(
                                product
                            ),
                        }
                    ],
                }
                serializer = OutboundOrderCreateSerializer(
                    data=payload,
                    context={"request": self.request},
                )
                serializer.is_valid(raise_exception=True)
                with transaction.atomic():
                    order = serializer.save()
                existing_orders[source_number] = order
                result["success_count"] += 1
                result["successes"].append(
                    {
                        "row": parsed.row_number,
                        "src_bill_no": source_number,
                        "order_id": order.id,
                        "order_no": order.order_no or "",
                    }
                )
            except DropShipRowError as exc:
                self._append_error(result, parsed.row_number, source_number, str(exc))
            except DRFValidationError as exc:
                self._append_error(
                    result,
                    parsed.row_number,
                    source_number,
                    _safe_validation_message(exc.detail),
                )
            except (IntegrityError, DjangoValidationError):
                existing = (
                    OutboundOrder.all_objects.filter(
                        owner_id=self.owner_id,
                        src_bill_no__iexact=source_number,
                    )
                    .order_by("id")
                    .first()
                    if source_number
                    else None
                )
                if existing:
                    existing_orders[source_number] = existing
                    self._append_skip(
                        result, parsed.row_number, source_number, existing
                    )
                else:
                    logger.exception(
                        "drop_ship_import_save_failed owner_id=%s user_id=%s row=%s",
                        self.owner_id,
                        self.user_id,
                        parsed.row_number,
                    )
                    self._append_error(
                        result, parsed.row_number, source_number, GENERIC_ROW_ERROR
                    )
            except Exception:
                logger.exception(
                    "drop_ship_import_unexpected owner_id=%s user_id=%s row=%s",
                    self.owner_id,
                    self.user_id,
                    parsed.row_number,
                )
                self._append_error(
                    result, parsed.row_number, source_number, GENERIC_ROW_ERROR
                )
        return result

    @staticmethod
    def _append_skip(result, row_number, source_number, order):
        result["skip_count"] += 1
        result["skips"].append(
            {
                "row": row_number,
                "src_bill_no": source_number,
                "reason": (
                    f"订单已存在，order_id={order.id}, order_no={order.order_no or ''}"
                ),
            }
        )

    @staticmethod
    def _append_error(result, row_number, source_number, reason):
        result["fail_count"] += 1
        result["errors"].append(
            {"row": row_number, "src_bill_no": source_number, "reason": reason}
        )


def import_drop_ship_workbook(
    *, uploaded_file, request, owner_id, warehouse_id, cash_customer
):
    return DropShipImportService(
        request=request,
        owner_id=owner_id,
        warehouse_id=warehouse_id,
        cash_customer=cash_customer,
    ).import_file(uploaded_file)
