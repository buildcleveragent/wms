import datetime
import io
from decimal import Decimal

from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from django.db.models import Sum
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from rest_framework.test import APIClient

from allapp.baseinfo.models import Customer, Owner
from allapp.inbound.models import InboundOrder, InboundOrderLine
from allapp.inbound.services import create_receive_task_draft
from allapp.billing.enums import AccrualStatus, CalcMethod, ChargeType, MetricType
from allapp.billing.models import Bill, BillingAccrual, BillingJobRun, BillingMetricDaily, BillingRule
from allapp.billing.services import (
    accrue_metrics_for_date,
    generate_invoice_for_period,
    generate_metrics_for_date,
    lock_period,
    run_scheduled_metric_generation_for_date,
)
from allapp.inventory.models import InventoryDetail, InventorySummary, InventoryTransaction
from allapp.inventory.services_quick_adjust import QuickAdjustInput, quick_adjust_via_post_task
from allapp.inventory.snapshot_services import generate_inventory_snapshot_for_date
from allapp.locations.models import Location, Subwarehouse, Warehouse
from allapp.outbound.models import OutboundOrder, OutboundOrderLine
from allapp.outbound.services import allocate_inventory, promote_reserved_pick, unallocate_for_order
from allapp.products.models import Product, ProductUom
from allapp.tasking.services import _run_posting_handler, approve_task, scan_task
from allapp.tasking.models import PutawayLineExtra, TaskAssignment, WmsTask, WmsTaskLine
from allapp.baseinfo.models import Supplier


def business_flow_barcode_resolver(owner_id, barcode):
    if not barcode.startswith("SKU:"):
        return {}

    code = barcode.split(":", 1)[1].strip().upper()
    product = Product.objects.get(owner_id=owner_id, code=code)
    return {
        "product_id": product.id,
        "pack_qty": Decimal("1"),
        "label_key": barcode,
        "code_type": "SKU",
        "uom_code": product.base_uom.code,
    }


@override_settings(TASKING_BARCODE_RESOLVER="allapp.test_business_flows.business_flow_barcode_resolver")
class BusinessFlowTests(TestCase):
    def setUp(self):
        user_model = get_user_model()

        self.owner = Owner.objects.create(name="Flow Owner", code="FLOWOWN")
        self.warehouse = Warehouse.objects.create(code="FLOWWH", name="Flow Warehouse")
        self.subwarehouse = Subwarehouse.objects.create(
            warehouse=self.warehouse,
            code="FLOWSW",
            name="Flow Subwarehouse",
        )
        self.receive_location = Location.objects.create(
            warehouse=self.warehouse,
            code="FLOWSW-01-01-01",
            name="Receive Location",
        )
        self.pick_location = Location.objects.create(
            warehouse=self.warehouse,
            code="FLOWSW-01-01-02",
            name="Pick Location",
        )
        self.adjust_location = Location.objects.create(
            warehouse=self.warehouse,
            code="FLOWSW-01-01-03",
            name="Adjust Location",
        )

        self.owner_user = user_model.objects.create_user(
            username="flow-owner",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        self.picker_user = user_model.objects.create_user(
            username="flow-picker",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        self.reviewer_user = user_model.objects.create_user(
            username="flow-reviewer",
            password="x",
            owner=self.owner,
            warehouse=self.warehouse,
        )
        self.superuser = user_model.objects.create_superuser(
            username="flow-admin",
            email="flow-admin@example.com",
            password="x",
        )

        self.customer = Customer.objects.create(
            owner=self.owner,
            salesperson=self.owner_user,
            code="FLOWCUS",
            name="Flow Customer",
        )
        self.supplier = Supplier.objects.create(
            owner=self.owner,
            code="FLOWSUP",
            name="Flow Supplier",
        )
        self.base_uom = ProductUom.objects.create(code="PCS", name="Piece", decimal_places=0)

    def api_client_for(self, user):
        client = APIClient()
        client.force_authenticate(user=user)
        return client

    def web_client_for(self, user):
        client = Client()
        client.force_login(user)
        return client

    def response_rows(self, response):
        payload = response.json()
        if isinstance(payload, dict) and "results" in payload:
            return payload["results"]
        return payload

    def create_product(self, code, *, volume="0.500000", price="10.00"):
        return Product.objects.create(
            owner=self.owner,
            code=code,
            name=code,
            sku=code,
            base_uom=self.base_uom,
            volume=Decimal(volume),
            price=Decimal(price),
            batch_control=False,
            expiry_control=False,
        )

    def seed_inventory(self, product, qty, *, location=None, allocated="0.0000"):
        location = location or self.pick_location
        qty_decimal = Decimal(str(qty))
        allocated_decimal = Decimal(str(allocated))
        detail = InventoryDetail.objects.create(
            owner=self.owner,
            product=product,
            warehouse=self.warehouse,
            location=location,
            onhand_qty=qty_decimal,
            allocated_qty=allocated_decimal,
            locked_qty=Decimal("0.0000"),
            damaged_qty=Decimal("0.0000"),
            base_unit=self.base_uom.code,
        )
        summary_totals = InventoryDetail.objects.filter(
            owner=self.owner,
            product=product,
            is_active=True,
        ).aggregate(
            onhand_qty=Sum("onhand_qty"),
            allocated_qty=Sum("allocated_qty"),
            locked_qty=Sum("locked_qty"),
            damaged_qty=Sum("damaged_qty"),
        )
        summary_onhand = summary_totals["onhand_qty"] or Decimal("0.0000")
        summary_allocated = summary_totals["allocated_qty"] or Decimal("0.0000")
        summary_locked = summary_totals["locked_qty"] or Decimal("0.0000")
        summary_damaged = summary_totals["damaged_qty"] or Decimal("0.0000")
        InventorySummary.objects.update_or_create(
            owner=self.owner,
            product=product,
            defaults={
                "base_unit": self.base_uom.code,
                "onhand_qty": summary_onhand,
                "allocated_qty": summary_allocated,
                "locked_qty": summary_locked,
                "damaged_qty": summary_damaged,
                "available_qty": summary_onhand - summary_allocated - summary_locked - summary_damaged,
            },
        )
        return detail

    def create_outbound_order(self, product, qty):
        order = OutboundOrder.objects.create(
            owner=self.owner,
            customer=self.customer,
            warehouse=self.warehouse,
            created_by=self.owner_user,
            submit_status="SUBMITTED",
        )
        OutboundOrderLine.objects.create(
            order=order,
            product=product,
            base_qty=Decimal(str(qty)),
            base_price=Decimal("10.00"),
            line_no=10,
        )
        return order

    def create_billed_period(self, *, product_code="BILLSKU", service_date=None):
        service_date = service_date or datetime.date(2026, 3, 31)
        product = self.create_product(product_code, volume="0.500000")
        self.seed_inventory(product, "10.0000", location=self.pick_location)

        generate_inventory_snapshot_for_date(
            service_date,
            owner_id=self.owner.id,
            warehouse_id=self.warehouse.id,
            bootstrap=True,
        )

        BillingRule.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.STORAGE,
            calc_method=CalcMethod.PER_CBM_DAY,
            unit_price=Decimal("2.00"),
            currency="CNY",
        )

        generate_metrics_for_date(self.owner.id, self.warehouse.id, service_date)
        accrue_metrics_for_date(self.owner.id, self.warehouse.id, service_date, by_user=self.owner_user)
        period = lock_period(
            self.owner.id,
            self.warehouse.id,
            service_date.strftime("%Y-%m"),
            service_date,
            service_date,
        )
        bill = generate_invoice_for_period(period, invoice_no=f"INV-{service_date.isoformat()}")
        return product, period, bill

    def assign_task_line(self, task, line, user):
        return TaskAssignment.objects.create(
            task=task,
            line=line,
            assignee=user,
            accepted_at=timezone.now(),
        )

    def create_formal_inbound_order(self, product, qty, *, base_price="8.0000"):
        inbound_order = InboundOrder.objects.create(
            owner=self.owner,
            supplier=self.supplier,
            warehouse=self.warehouse,
            created_by=self.owner_user,
        )
        InboundOrderLine.objects.create(
            order=inbound_order,
            product=product,
            base_qty=Decimal(str(qty)),
            base_price=Decimal(base_price),
        )

        inbound_order.submit_by_owner_buyers(self.owner_user)
        inbound_order.owner_approve(self.owner_user)
        inbound_order.wh_confirm(self.reviewer_user)
        return inbound_order

    def complete_formal_receive(self, inbound_order, product, qty, *, operator=None):
        operator = operator or self.reviewer_user
        qty_decimal = Decimal(str(qty))

        receive_task = WmsTask.objects.get(
            task_type=WmsTask.TaskType.RECEIVE,
            source_app="inbound",
            source_model="InboundOrder",
            source_pk=str(inbound_order.id),
        )
        receive_task.release()
        receive_line = receive_task.lines.get(product=product)
        self.assign_task_line(receive_task, receive_line, operator)

        scan_result = scan_task(
            task_id=receive_task.id,
            barcode=f"SKU:{product.code}",
            qty=qty_decimal,
            location_id=self.receive_location.id,
            by_user=operator,
        )
        self.assertFalse(scan_result["idempotent"])

        receive_extra = receive_line.receivelineextra
        receive_extra.qty_ok = qty_decimal
        receive_extra._by_user = operator
        receive_extra.save()

        approve_task(receive_task.id, by_user=self.superuser, note="formal inbound approve")
        _run_posting_handler(receive_task.id, by_user=self.superuser, note="formal inbound post")
        receive_task.refresh_from_db()
        return receive_task

    def complete_putaway(self, product, qty, *, to_location, operator=None):
        operator = operator or self.picker_user
        qty_decimal = Decimal(str(qty))

        putaway_task = (
            WmsTask.objects
            .filter(
                task_type=WmsTask.TaskType.PUTAWAY,
                owner=self.owner,
                warehouse=self.warehouse,
            )
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(putaway_task)
        self.assertEqual(putaway_task.status, WmsTask.Status.DRAFT)

        putaway_task.release()
        putaway_line = putaway_task.lines.get(product=product)
        WmsTaskLine.objects.filter(pk=putaway_line.id).update(to_location_id=to_location.id)
        putaway_line.refresh_from_db()
        self.assign_task_line(putaway_task, putaway_line, operator)

        scan_result = scan_task(
            task_id=putaway_task.id,
            barcode=f"SKU:{product.code}",
            qty=qty_decimal,
            location_id=to_location.id,
            by_user=operator,
        )
        self.assertFalse(scan_result["idempotent"])

        putaway_extra = PutawayLineExtra(
            line=putaway_line,
            to_location=to_location,
            qty_moved=qty_decimal,
        )
        putaway_extra._by_user = operator
        putaway_extra.save()

        putaway_task.refresh_from_db()
        self.assertEqual(putaway_task.status, WmsTask.Status.COMPLETED)
        self.assertEqual(putaway_task.review_status, WmsTask.ReviewStatus.PENDING)

        approve_task(putaway_task.id, by_user=self.superuser, note="putaway approve")
        _run_posting_handler(putaway_task.id, by_user=self.superuser, note="putaway post")
        putaway_task.refresh_from_db()
        return putaway_task

    def test_flow_1_receive_without_order_inventory_visible(self):
        product = self.create_product("RCVSKU")
        client = self.api_client_for(self.owner_user)

        response = client.post(
            "/api/inbound/receive_without_order/",
            {
                "owner_id": self.owner.id,
                "warehouse_id": self.warehouse.id,
                "location_id": self.receive_location.id,
                "items": [
                    {
                        "product_id": product.id,
                        "qty": "5.000",
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201, response.json())
        detail = InventoryDetail.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            location=self.receive_location,
            product=product,
        )
        summary = InventorySummary.objects.get(owner=self.owner, product=product)

        self.assertEqual(detail.onhand_qty, Decimal("5.0000"))
        self.assertEqual(detail.available_qty, Decimal("5.0000"))
        self.assertEqual(summary.onhand_qty, Decimal("5.0000"))
        self.assertTrue(
            InventoryTransaction.objects.filter(
                owner=self.owner,
                warehouse=self.warehouse,
                location=self.receive_location,
                product=product,
                tx_type="RECEIVE",
            ).exists()
        )

    def test_flow_2_outbound_approve_scan_post(self):
        product = self.create_product("PICKSKU")
        detail = self.seed_inventory(product, "10.0000")
        owner_client = self.api_client_for(self.owner_user)

        create_response = owner_client.post(
            "/api/outbound/orders/",
            {
                "customer_id": self.customer.id,
                "items": [
                    {
                        "product_id": product.id,
                        "qty": "3.000",
                        "price": "10.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, 201, create_response.json())
        order_id = create_response.json()["id"]

        approve_response = owner_client.post(
            f"/api/outbound/orders/{order_id}/owner-approve/",
            {},
            format="json",
        )
        self.assertEqual(approve_response.status_code, 200, approve_response.json())

        order = OutboundOrder.objects.get(pk=order_id)
        task = WmsTask.objects.get(task_type=WmsTask.TaskType.PICK, source_pk=str(order.id))
        promote_reserved_pick(order, new_status=WmsTask.Status.RELEASED)

        picker_client = self.api_client_for(self.picker_user)
        scan_response = picker_client.post(
            f"/api/pda/pick-tasks/{task.id}/scan/",
            {"barcode": f"SKU:{product.code}", "qty": "3.000"},
            format="json",
        )
        self.assertEqual(scan_response.status_code, 200, scan_response.json())

        review_response = picker_client.post(
            f"/api/pda/pick-tasks/{task.id}/create-review-task/",
            {},
            format="json",
        )
        self.assertEqual(review_response.status_code, 200, review_response.json())

        reviewer_client = self.api_client_for(self.reviewer_user)
        post_response = reviewer_client.post(
            f"/api/pda/pick-tasks/{task.id}/post/",
            {},
            format="json",
        )
        self.assertEqual(post_response.status_code, 200, post_response.json())

        detail.refresh_from_db()
        task.refresh_from_db()

        self.assertEqual(detail.onhand_qty, Decimal("7.0000"))
        self.assertEqual(detail.allocated_qty, Decimal("0.0000"))
        self.assertEqual(detail.available_qty, Decimal("7.0000"))
        self.assertEqual(task.status, WmsTask.Status.COMPLETED)
        self.assertEqual(task.posting_status, WmsTask.PostingStatus.POSTED)
        self.assertTrue(
            InventoryTransaction.objects.filter(
                src_model="WmsTask",
                src_id=task.id,
                tx_type="ISSUE",
            ).exists()
        )

    def test_flow_3_unapprove_or_cancel_releases_reservation(self):
        product = self.create_product("RELSKU")
        detail = self.seed_inventory(product, "6.0000")
        order = self.create_outbound_order(product, "4.000")

        allocate_inventory(order, by_user=self.owner_user, allow_backorder=False)
        detail.refresh_from_db()
        self.assertEqual(detail.allocated_qty, Decimal("4.0000"))

        released = unallocate_for_order(order)

        detail.refresh_from_db()
        task = WmsTask.objects.get(task_type=WmsTask.TaskType.PICK, source_pk=str(order.id))

        self.assertEqual(released, Decimal("4.000"))
        self.assertEqual(detail.allocated_qty, Decimal("0.0000"))
        self.assertEqual(detail.available_qty, Decimal("6.0000"))
        self.assertEqual(task.status, WmsTask.Status.CANCELLED)
        self.assertFalse(task.lines.exists())

    def test_flow_4_quick_adjust_updates_inventory_and_company_report(self):
        product = self.create_product("ADJSKU")
        quick_adjust_via_post_task(
            QuickAdjustInput(
                user=self.reviewer_user,
                owner=self.owner,
                warehouse=self.warehouse,
                location=self.adjust_location,
                product=product,
                qty_base_delta=Decimal("5.0000"),
                reason="SMOKE_ADJUST",
                remark="smoke adjust",
            )
        )

        detail = InventoryDetail.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            location=self.adjust_location,
            product=product,
        )
        self.assertEqual(detail.onhand_qty, Decimal("5.0000"))

        client = self.api_client_for(self.superuser)
        response = client.get(
            "/api/inventory/company-summary/",
            {
                "mode": "warehouse",
                "warehouse_id": self.warehouse.id,
                "owner_id": self.owner.id,
            },
        )
        self.assertEqual(response.status_code, 200, response.json())
        rows = self.response_rows(response)

        self.assertTrue(
            any(
                row["product_id"] == product.id
                and Decimal(str(row["onhand_qty"])) == Decimal("5.0000")
                for row in rows
            )
        )

    def test_flow_5_snapshot_metrics_accrual_lock_invoice(self):
        _, period, bill = self.create_billed_period(product_code="METRICSKU")

        metric = BillingMetricDaily.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=period.start_date,
            metric_type=MetricType.CBM,
        )
        accrual = BillingAccrual.objects.get(period=period)

        self.assertEqual(metric.value, Decimal("5.0000"))
        self.assertEqual(accrual.status, AccrualStatus.INVOICED)
        self.assertEqual(bill.subtotal, Decimal("10.00"))
        self.assertEqual(bill.total, Decimal("10.00"))
        self.assertEqual(period.status, "INVOICED")

    def test_flow_6_owner_portal_can_view_inventory_bill_and_export(self):
        product, period, bill = self.create_billed_period(product_code="PORTALSKU")
        client = self.api_client_for(self.owner_user)

        inventory_response = client.get("/api/inventory/summary/")
        self.assertEqual(inventory_response.status_code, 200, inventory_response.json())
        inventory_rows = self.response_rows(inventory_response)
        self.assertTrue(any(row["product_id"] == product.id for row in inventory_rows))

        periods_response = client.get("/api/billing/periods/")
        self.assertEqual(periods_response.status_code, 200, periods_response.json())
        period_rows = self.response_rows(periods_response)
        self.assertTrue(any(row["id"] == period.id for row in period_rows))

        preview_response = client.get(f"/api/billing/periods/{period.id}/preview/")
        self.assertEqual(preview_response.status_code, 200, preview_response.json())
        self.assertEqual(preview_response.json()["accrual_count"], 1)

        bill_response = client.get(f"/api/billing/bills/{bill.id}/")
        self.assertEqual(bill_response.status_code, 200, bill_response.json())
        self.assertEqual(bill_response.json()["invoice_no"], bill.invoice_no)

        export_response = client.get(f"/api/billing/bills/{bill.id}/export/")
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(io.BytesIO(export_response.content))
        self.assertEqual(workbook.sheetnames, ["Bill", "Lines"])

    def test_flow_7_pda_pick_scan_and_state_transition(self):
        product = self.create_product("PDASKU")
        self.seed_inventory(product, "2.0000")
        order = self.create_outbound_order(product, "2.000")
        allocate_inventory(order, by_user=self.owner_user, allow_backorder=False)
        task = promote_reserved_pick(order, new_status=WmsTask.Status.RELEASED)

        client = self.api_client_for(self.picker_user)

        list_response = client.get("/api/pda/pick-tasks/")
        self.assertEqual(list_response.status_code, 200, list_response.json())
        tasks = self.response_rows(list_response)
        self.assertTrue(any(item["id"] == task.id for item in tasks))

        lines_response = client.get(f"/api/pda/pick-tasks/{task.id}/lines/")
        self.assertEqual(lines_response.status_code, 200, lines_response.json())
        self.assertEqual(len(lines_response.json()), 1)

        scan_response = client.post(
            f"/api/pda/pick-tasks/{task.id}/scan/",
            {"barcode": f"SKU:{product.code}", "qty": "2.000"},
            format="json",
        )
        self.assertEqual(scan_response.status_code, 200, scan_response.json())
        self.assertEqual(Decimal(str(scan_response.json()["line"]["qty_done"])), Decimal("2.000"))

        review_response = client.post(
            f"/api/pda/pick-tasks/{task.id}/create-review-task/",
            {},
            format="json",
        )
        self.assertEqual(review_response.status_code, 200, review_response.json())
        task.refresh_from_db()

        self.assertEqual(task.status, WmsTask.Status.COMPLETED)
        self.assertEqual(task.review_status, WmsTask.ReviewStatus.PENDING)

    def test_flow_8_console_billing_overview_and_detail(self):
        _, period, bill = self.create_billed_period(product_code="CONSOLESKU")
        client = self.web_client_for(self.superuser)

        overview_response = client.get(
            reverse("console:billing_overview"),
            {
                "owner": self.owner.id,
                "warehouse": self.warehouse.id,
                "period": period.id,
            },
        )
        self.assertEqual(overview_response.status_code, 200)
        self.assertEqual(overview_response.context["current_bill"].id, bill.id)
        self.assertContains(overview_response, bill.invoice_no)

        detail_response = client.get(
            reverse("console:billing_bill_detail", args=[bill.id]),
            {
                "owner": self.owner.id,
                "warehouse": self.warehouse.id,
                "period": period.id,
            },
        )
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.context["bill"].id, bill.id)
        self.assertContains(detail_response, bill.invoice_no)
        self.assertContains(detail_response, "STORAGE")
        self.assertEqual(detail_response.context["overall_line_count"], Bill.objects.get(pk=bill.id).lines.count())

    def test_flow_9_formal_inbound_order_to_outbound_full_chain(self):
        product = self.create_product("FULLSKU")

        inbound_order = InboundOrder.objects.create(
            owner=self.owner,
            supplier=self.supplier,
            warehouse=self.warehouse,
            created_by=self.owner_user,
        )
        InboundOrderLine.objects.create(
            order=inbound_order,
            product=product,
            base_qty=Decimal("5.000"),
            base_price=Decimal("8.0000"),
        )

        inbound_order.submit_by_owner_buyers(self.owner_user)
        inbound_order.owner_approve(self.owner_user)
        inbound_order.wh_confirm(self.reviewer_user)

        receive_task = WmsTask.objects.get(
            task_type=WmsTask.TaskType.RECEIVE,
            source_app="inbound",
            source_model="InboundOrder",
            source_pk=str(inbound_order.id),
        )
        self.assertEqual(receive_task.status, WmsTask.Status.DRAFT)

        receive_task.release()
        receive_line = receive_task.lines.get()
        TaskAssignment.objects.create(
            task=receive_task,
            line=receive_line,
            assignee=self.reviewer_user,
            accepted_at=timezone.now(),
        )

        scan_result = scan_task(
            task_id=receive_task.id,
            barcode=f"SKU:{product.code}",
            qty=Decimal("5.000"),
            location_id=self.receive_location.id,
            by_user=self.reviewer_user,
        )
        self.assertFalse(scan_result["idempotent"])

        receive_extra = receive_line.receivelineextra
        receive_extra.qty_ok = Decimal("5.000")
        receive_extra._by_user = self.reviewer_user
        receive_extra.save()

        receive_task.refresh_from_db()
        receive_line.refresh_from_db()
        self.assertEqual(receive_line.qty_done, Decimal("5.000"))
        self.assertEqual(receive_task.status, WmsTask.Status.COMPLETED)
        self.assertEqual(receive_task.review_status, WmsTask.ReviewStatus.PENDING)

        approve_task(receive_task.id, by_user=self.superuser, note="formal inbound approve")
        _run_posting_handler(receive_task.id, by_user=self.superuser, note="formal inbound post")

        receive_task.refresh_from_db()
        inbound_detail = InventoryDetail.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            location=self.receive_location,
            product=product,
        )
        self.assertEqual(receive_task.posting_status, WmsTask.PostingStatus.POSTED)
        self.assertEqual(inbound_detail.onhand_qty, Decimal("5.0000"))
        self.assertEqual(inbound_detail.available_qty, Decimal("5.0000"))
        self.assertTrue(
            WmsTask.objects.filter(
                task_type=WmsTask.TaskType.PUTAWAY,
                owner=self.owner,
                warehouse=self.warehouse,
            ).exists()
        )

        owner_client = self.api_client_for(self.owner_user)
        outbound_create = owner_client.post(
            "/api/outbound/orders/",
            {
                "customer_id": self.customer.id,
                "items": [
                    {
                        "product_id": product.id,
                        "qty": "3.000",
                        "price": "12.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(outbound_create.status_code, 201, outbound_create.json())
        outbound_id = outbound_create.json()["id"]

        outbound_approve = owner_client.post(
            f"/api/outbound/orders/{outbound_id}/owner-approve/",
            {},
            format="json",
        )
        self.assertEqual(outbound_approve.status_code, 200, outbound_approve.json())

        outbound_order = OutboundOrder.objects.get(pk=outbound_id)
        pick_task = promote_reserved_pick(outbound_order, new_status=WmsTask.Status.RELEASED)

        picker_client = self.api_client_for(self.picker_user)
        pick_scan = picker_client.post(
            f"/api/pda/pick-tasks/{pick_task.id}/scan/",
            {"barcode": f"SKU:{product.code}", "qty": "3.000"},
            format="json",
        )
        self.assertEqual(pick_scan.status_code, 200, pick_scan.json())

        pick_review = picker_client.post(
            f"/api/pda/pick-tasks/{pick_task.id}/create-review-task/",
            {},
            format="json",
        )
        self.assertEqual(pick_review.status_code, 200, pick_review.json())

        reviewer_client = self.api_client_for(self.reviewer_user)
        pick_post = reviewer_client.post(
            f"/api/pda/pick-tasks/{pick_task.id}/post/",
            {},
            format="json",
        )
        self.assertEqual(pick_post.status_code, 200, pick_post.json())

        pick_task.refresh_from_db()
        inbound_detail.refresh_from_db()
        self.assertEqual(pick_task.posting_status, WmsTask.PostingStatus.POSTED)
        self.assertEqual(inbound_detail.onhand_qty, Decimal("2.0000"))
        self.assertEqual(inbound_detail.available_qty, Decimal("2.0000"))
        self.assertTrue(
            InventoryTransaction.objects.filter(
                src_model="WmsTask",
                src_id=receive_task.id,
                tx_type="RECEIVE",
            ).exists()
        )
        self.assertTrue(
            InventoryTransaction.objects.filter(
                src_model="WmsTask",
                src_id=pick_task.id,
                tx_type="ISSUE",
            ).exists()
        )

    def test_flow_10_formal_inbound_putaway_to_outbound_full_chain(self):
        product = self.create_product("PUTSKU")

        inbound_order = self.create_formal_inbound_order(product, "5.000")
        receive_task = self.complete_formal_receive(inbound_order, product, "5.000")

        receive_detail = InventoryDetail.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            location=self.receive_location,
            product=product,
        )
        self.assertEqual(receive_task.posting_status, WmsTask.PostingStatus.POSTED)
        self.assertEqual(receive_detail.onhand_qty, Decimal("5.0000"))

        putaway_task = self.complete_putaway(product, "5.000", to_location=self.pick_location)

        receive_detail.refresh_from_db()
        putaway_detail = InventoryDetail.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            location=self.pick_location,
            product=product,
        )
        self.assertEqual(putaway_task.posting_status, WmsTask.PostingStatus.POSTED)
        self.assertEqual(receive_detail.onhand_qty, Decimal("0.0000"))
        self.assertEqual(receive_detail.available_qty, Decimal("0.0000"))
        self.assertEqual(putaway_detail.onhand_qty, Decimal("5.0000"))
        self.assertEqual(putaway_detail.available_qty, Decimal("5.0000"))
        self.assertTrue(
            InventoryTransaction.objects.filter(
                src_model="WmsTask",
                src_id=putaway_task.id,
                tx_type="ISSUE",
                location=self.receive_location,
            ).exists()
        )
        self.assertTrue(
            InventoryTransaction.objects.filter(
                src_model="WmsTask",
                src_id=putaway_task.id,
                tx_type="RECEIVE",
                location=self.pick_location,
            ).exists()
        )

        owner_client = self.api_client_for(self.owner_user)
        outbound_create = owner_client.post(
            "/api/outbound/orders/",
            {
                "customer_id": self.customer.id,
                "items": [
                    {
                        "product_id": product.id,
                        "qty": "3.000",
                        "price": "12.0000",
                    }
                ],
            },
            format="json",
        )
        self.assertEqual(outbound_create.status_code, 201, outbound_create.json())
        outbound_id = outbound_create.json()["id"]

        outbound_approve = owner_client.post(
            f"/api/outbound/orders/{outbound_id}/owner-approve/",
            {},
            format="json",
        )
        self.assertEqual(outbound_approve.status_code, 200, outbound_approve.json())

        outbound_order = OutboundOrder.objects.get(pk=outbound_id)
        pick_task = promote_reserved_pick(outbound_order, new_status=WmsTask.Status.RELEASED)
        pick_line = pick_task.lines.get()
        self.assertEqual(pick_line.from_location_id, self.pick_location.id)

        picker_client = self.api_client_for(self.picker_user)
        pick_scan = picker_client.post(
            f"/api/pda/pick-tasks/{pick_task.id}/scan/",
            {"barcode": f"SKU:{product.code}", "qty": "3.000"},
            format="json",
        )
        self.assertEqual(pick_scan.status_code, 200, pick_scan.json())

        pick_review = picker_client.post(
            f"/api/pda/pick-tasks/{pick_task.id}/create-review-task/",
            {},
            format="json",
        )
        self.assertEqual(pick_review.status_code, 200, pick_review.json())

        reviewer_client = self.api_client_for(self.reviewer_user)
        pick_post = reviewer_client.post(
            f"/api/pda/pick-tasks/{pick_task.id}/post/",
            {},
            format="json",
        )
        self.assertEqual(pick_post.status_code, 200, pick_post.json())

        pick_task.refresh_from_db()
        putaway_detail.refresh_from_db()
        self.assertEqual(pick_task.posting_status, WmsTask.PostingStatus.POSTED)
        self.assertEqual(putaway_detail.onhand_qty, Decimal("2.0000"))
        self.assertEqual(putaway_detail.available_qty, Decimal("2.0000"))

    def test_flow_11_multiline_multilocation_outbound_full_chain(self):
        product_a = self.create_product("MULTIA")
        product_b = self.create_product("MULTIB")
        overflow_location = Location.objects.create(
            warehouse=self.warehouse,
            code="FLOWSW-01-01-04",
            name="Overflow Location",
        )

        detail_a_primary = self.seed_inventory(product_a, "4.0000", location=self.pick_location)
        detail_a_overflow = self.seed_inventory(product_a, "3.0000", location=overflow_location)
        detail_b = self.seed_inventory(product_b, "4.0000", location=self.adjust_location)

        owner_client = self.api_client_for(self.owner_user)
        create_response = owner_client.post(
            "/api/outbound/orders/",
            {
                "customer_id": self.customer.id,
                "items": [
                    {
                        "product_id": product_a.id,
                        "qty": "6.000",
                        "price": "10.0000",
                    },
                    {
                        "product_id": product_b.id,
                        "qty": "4.000",
                        "price": "8.0000",
                    },
                ],
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, 201, create_response.json())
        order_id = create_response.json()["id"]

        approve_response = owner_client.post(
            f"/api/outbound/orders/{order_id}/owner-approve/",
            {},
            format="json",
        )
        self.assertEqual(approve_response.status_code, 200, approve_response.json())

        order = OutboundOrder.objects.get(pk=order_id)
        pick_task = promote_reserved_pick(order, new_status=WmsTask.Status.RELEASED)
        task_lines = list(pick_task.lines.order_by("id"))
        product_a_lines = [line for line in task_lines if line.product_id == product_a.id]
        product_b_lines = [line for line in task_lines if line.product_id == product_b.id]

        self.assertEqual(len(task_lines), 3)
        self.assertEqual(len(product_a_lines), 2)
        self.assertEqual(sum(line.qty_plan for line in product_a_lines), Decimal("6.000"))
        self.assertEqual({line.from_location_id for line in product_a_lines}, {self.pick_location.id, overflow_location.id})
        self.assertEqual(len(product_b_lines), 1)
        self.assertEqual(product_b_lines[0].from_location_id, self.adjust_location.id)

        picker_client = self.api_client_for(self.picker_user)
        for line in task_lines:
            adjust_response = picker_client.post(
                f"/api/pda/pick-tasks/{pick_task.id}/adjust-line-qty/",
                {
                    "line_id": line.id,
                    "final_qty_done": str(line.qty_plan),
                },
                format="json",
            )
            self.assertEqual(adjust_response.status_code, 200, adjust_response.json())

        review_response = picker_client.post(
            f"/api/pda/pick-tasks/{pick_task.id}/create-review-task/",
            {},
            format="json",
        )
        self.assertEqual(review_response.status_code, 200, review_response.json())

        reviewer_client = self.api_client_for(self.reviewer_user)
        post_response = reviewer_client.post(
            f"/api/pda/pick-tasks/{pick_task.id}/post/",
            {},
            format="json",
        )
        self.assertEqual(post_response.status_code, 200, post_response.json())

        detail_a_primary.refresh_from_db()
        detail_a_overflow.refresh_from_db()
        detail_b.refresh_from_db()

        self.assertEqual(detail_a_primary.onhand_qty, Decimal("0.0000"))
        self.assertEqual(detail_a_primary.available_qty, Decimal("0.0000"))
        self.assertEqual(detail_a_overflow.onhand_qty, Decimal("1.0000"))
        self.assertEqual(detail_a_overflow.available_qty, Decimal("1.0000"))
        self.assertEqual(detail_b.onhand_qty, Decimal("0.0000"))
        self.assertEqual(detail_b.available_qty, Decimal("0.0000"))
        self.assertEqual(
            InventorySummary.objects.get(owner=self.owner, product=product_a).onhand_qty,
            Decimal("1.0000"),
        )
        self.assertEqual(
            InventorySummary.objects.get(owner=self.owner, product=product_b).onhand_qty,
            Decimal("0.0000"),
        )
        self.assertEqual(
            InventoryTransaction.objects.filter(
                src_model="WmsTask",
                src_id=pick_task.id,
                tx_type="ISSUE",
            ).count(),
            3,
        )

    def test_flow_12_operational_inventory_to_billing_invoice_chain(self):
        service_date = timezone.now().date()
        product = self.create_product("OPSBILL", volume="0.250000")

        inbound_order = self.create_formal_inbound_order(product, "5.000")
        self.complete_formal_receive(inbound_order, product, "5.000")
        putaway_task = self.complete_putaway(product, "5.000", to_location=self.pick_location)

        BillingRule.objects.create(
            owner=self.owner,
            warehouse=self.warehouse,
            charge_type=ChargeType.STORAGE,
            calc_method=CalcMethod.PER_CBM_DAY,
            unit_price=Decimal("2.00"),
            currency="CNY",
        )

        scheduler_summary = run_scheduled_metric_generation_for_date(
            service_date,
            owner_id=self.owner.id,
            warehouse_id=self.warehouse.id,
            force=True,
        )
        self.assertEqual(scheduler_summary["success"], 1)
        self.assertEqual(
            BillingJobRun.objects.get(
                job_name=BillingJobRun.JobName.DAILY_METRIC_GENERATION,
                owner=self.owner,
                warehouse=self.warehouse,
                service_date=service_date,
            ).status,
            BillingJobRun.Status.SUCCESS,
        )

        cbm_metric = BillingMetricDaily.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            service_date=service_date,
            metric_type=MetricType.CBM,
        )
        self.assertEqual(cbm_metric.value, Decimal("1.250000"))

        accrue_metrics_for_date(self.owner.id, self.warehouse.id, service_date, by_user=self.owner_user)
        period = lock_period(
            self.owner.id,
            self.warehouse.id,
            service_date.strftime("%Y-%m-%d"),
            service_date,
            service_date,
        )
        bill = generate_invoice_for_period(period, invoice_no=f"INV-OPS-{service_date.isoformat()}")

        accrual = BillingAccrual.objects.get(
            owner=self.owner,
            warehouse=self.warehouse,
            period=period,
            charge_type=ChargeType.STORAGE,
            service_date=service_date,
        )
        self.assertEqual(accrual.amount, Decimal("2.50"))
        self.assertEqual(accrual.status, AccrualStatus.INVOICED)
        self.assertEqual(putaway_task.posting_status, WmsTask.PostingStatus.POSTED)
        self.assertEqual(bill.subtotal, Decimal("2.50"))
        self.assertEqual(bill.total, Decimal("2.50"))
        self.assertEqual(bill.lines.count(), 1)
